"""Catálogo SICRO sintético: os bytes .xlsx que `croquito_valuation.sicro` sabe ler.

A tabela SICRO real é de distribuição pública, mas o arquivo real não existe neste
repositório — os bytes gerados aqui não pretendem reproduzir o layout real (que é DADO,
fechado em `SicroCatalogLayout` a partir de um arquivo baixado) e só exercitam o leitor
.xlsx mínimo de `croquito_valuation.sicro`.

Diferente do par EMOP (`emop_fixture.py`, que só grava bytes, com o gabarito e os
adulteradores morando à parte em `tests/valuation/emop_fixture.py`), este módulo já
concentra gerador, layout casado, gabarito (`expected_sicro_entries`) e a parametrização
mínima que os testes de recusa precisam para adulterar o arquivo — o Task Contract F-026/T2
pediu essa forma mais simples, porque o layout SICRO não tem o par leitor
DBF/DBF-de-teste que justificou o split do EMOP. Espelho de `sinapi_fixture.py`, deliberadamente
duplicado em vez de compartilhado (ADR-0039: um importador por fonte).
"""

from __future__ import annotations

from collections.abc import Sequence
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Final

from openpyxl import Workbook

from croquito_valuation.models import PriceCatalogEntry, PriceOrigin
from croquito_valuation.sicro import SICRO_FAMILY_CODE, SicroCatalogLayout

SICRO_CODE_PATTERN: Final = r"^\d{4}\.\d{2}\.\d{2}$"
"""Padrão sintético do exercício: um código SICRO real é dado do layout do arquivo
baixado, não constante de código — este padrão só existe para a fixture ter algo
consistente a casar (grupo.subgrupo.item, prefixo distinto do numérico puro da SINAPI)."""

SICRO_SOURCE_LABEL: Final = "TABELA SICRO SINTETICA (fixture)"
SICRO_REFERENCE_MONTH: Final = "2026-06"
SICRO_SHEET_NAME: Final = "SICRO"
SICRO_HEADER_ROW: Final = 1
SICRO_CODE_COLUMN: Final = "A"
SICRO_DESCRIPTION_COLUMN: Final = "B"
SICRO_UNIT_COLUMN: Final = "C"
SICRO_PRICE_COLUMN: Final = "D"

_HEADER_CELLS: Final[tuple[str, str, str, str]] = ("CODIGO", "DESCRICAO", "UNIDADE", "PRECO")


@dataclass(frozen=True, slots=True)
class SicroFixtureRow:
    """Uma linha sintética do catálogo SICRO: o que é gravado no .xlsx e o que se espera.

    `blank=True` é a linha que exercita a linha inteiramente em branco: pulada e contada,
    nunca importada — nada é escrito nas quatro colunas dela. `price` aceita `str` só para
    os testes de recusa escreverem um texto adulterado direto na célula (ver
    `write_sicro_xlsx`); a linha "feliz" da fixture sempre usa `Decimal`.
    """

    code: str
    description: str
    unit: str
    price: Decimal | str
    blank: bool = False


SICRO_FIXTURE_ROWS: Final[tuple[SicroFixtureRow, ...]] = (
    SicroFixtureRow(
        code="1234.01.02",
        description="TERRAPLANAGEM SINTETICA SICRO TIPO 1",
        unit="M3",
        price=Decimal("45.67"),
    ),
    SicroFixtureRow(
        code="5678.03.04",
        description="TERRAPLANAGEM SINTETICA SICRO TIPO 2",
        unit="M3",
        price=Decimal("52.10"),
    ),
    SicroFixtureRow(code="", description="", unit="", price=Decimal("0"), blank=True),
    SicroFixtureRow(
        code="9012.05.06",
        description="SINALIZACAO SINTETICA SICRO",
        unit="UN",
        price=Decimal("310.00"),
    ),
)
"""Três linhas ativas e uma linha em branco de propósito, entre as duas primeiras e a
última — precisa ficar no MEIO: uma linha em branco como última linha do intervalo escrito
não entraria na extensão usada da planilha e o leitor nunca a veria."""


def sicro_fixture_layout() -> SicroCatalogLayout:
    """Layout da fixture: a mesma aba, cabeçalho e colunas que `write_sicro_xlsx` grava."""
    return SicroCatalogLayout(
        source_label=SICRO_SOURCE_LABEL,
        reference_month=SICRO_REFERENCE_MONTH,
        sheet_name=SICRO_SHEET_NAME,
        header_row=SICRO_HEADER_ROW,
        code_column=SICRO_CODE_COLUMN,
        description_column=SICRO_DESCRIPTION_COLUMN,
        unit_column=SICRO_UNIT_COLUMN,
        price_column=SICRO_PRICE_COLUMN,
        code_pattern=SICRO_CODE_PATTERN,
    )


def write_sicro_xlsx(
    path: Path,
    rows: Sequence[SicroFixtureRow] = SICRO_FIXTURE_ROWS,
    *,
    sheet_name: str = SICRO_SHEET_NAME,
    header_row: int = SICRO_HEADER_ROW,
    price_as_float: bool = False,
) -> Path:
    """Grava os bytes .xlsx da fixture; a MESMA `rows` também gera o gabarito.

    `price_as_float` existe só para o teste de recusa "célula float não converte": grava o
    preço como número binário do openpyxl em vez de texto, em TODAS as linhas não vazias.
    Uma linha cujo `price` já seja `str` (adulteração de teste) é sempre gravada como texto,
    mesmo com `price_as_float=True` — a adulteração explícita da linha manda.
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    for offset, text in enumerate(_HEADER_CELLS):
        worksheet.cell(row=header_row, column=1 + offset, value=text)

    for row_offset, row in enumerate(rows, start=1):
        row_number = header_row + row_offset
        if row.blank:
            continue
        price_value: object
        if isinstance(row.price, str):
            price_value = row.price
        elif price_as_float:
            price_value = float(row.price)
        else:
            price_value = format(row.price, "f")
        worksheet.cell(row=row_number, column=1, value=row.code)
        worksheet.cell(row=row_number, column=2, value=row.description)
        worksheet.cell(row=row_number, column=3, value=row.unit)
        worksheet.cell(row=row_number, column=4, value=price_value)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def expected_sicro_entries(
    rows: Sequence[SicroFixtureRow] = SICRO_FIXTURE_ROWS,
) -> list[PriceCatalogEntry]:
    """Gabarito das entradas que `read_sicro_catalog` deve devolver: em branco fica fora."""
    return [
        PriceCatalogEntry(
            code=row.code,
            description=row.description,
            unit=row.unit,
            unit_price=Decimal(row.price) if isinstance(row.price, str) else row.price,
            family_code=SICRO_FAMILY_CODE,
            family_name=SICRO_SOURCE_LABEL,
            subgroup_code=SICRO_FAMILY_CODE,
            subgroup_name=SICRO_SOURCE_LABEL,
            origin=PriceOrigin.SICRO,
        )
        for row in rows
        if not row.blank
    ]
