"""Fixture sintética da medição: mini-MAPÃO e medição de obras inventadas.

Nada aqui vem de cliente. Os nomes dizem "SINTETICA" de propósito, os preços são
inventados e o conjunto foi escolhido para exercitar as fronteiras que quebram planilha
de medição no mundo real:

- item cuja fórmula viva reproduz o valor exato (`=TRUNC(...)` fica na célula);
- item cujo produto em ponto flutuante divergiria (o valor é fixado na célula);
- o par 1,15 x 10,30, que vale 11,84 truncado e 11,85 arredondado;
- o par 4,35 x 13,30, que vale 57,855 exato e só fecha em `Decimal`;
- variantes `(A)`/`(B)` do mesmo código base com preços diferentes;
- unidades escritas de formas diferentes na planilha de preços (`M2`, `m²`, `UNID.`).

A partir do M2 o módulo também produz o **MAPÃO anterior completo** — catálogo, PLANILHA
GERAL de duas medições já lançadas e aba de RE-RA — e a medição multi-obra do período
seguinte, que é o insumo da cadeia `import-workbook` → `export-valuation`.

O M4 fecha a outra ponta: o contrato sintético passou a contratar também os códigos que a
legenda da prancha mede, e o módulo devolve as decisões humanas da obra que nasce dela —
revisão do takeoff, confirmação de código e plano de memória de cálculo. Nenhuma dessas
decisões é inferida: cada uma cita o rótulo impresso na legenda e falha alto se ele mudar.
"""

from __future__ import annotations

from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path
from types import MappingProxyType
from typing import Any, Final, Literal
from uuid import NAMESPACE_URL, uuid5

from openpyxl import Workbook
from openpyxl.utils import column_index_from_string

from croquito_valuation.assignment import CodeAssignmentBatch, CodeAssignmentInput
from croquito_valuation.calc import CalcBlockPlan, CalcBuildResult, CalcPlan, ItemCalcPlan
from croquito_valuation.catalog import parse_money
from croquito_valuation.contract import ContractWorkbook
from croquito_valuation.errors import ValuationValidationError
from croquito_valuation.models import (
    BulletinLine,
    CalcBlock,
    CalcOperand,
    CalcRecipe,
    CalcSheet,
    PriceCatalog,
    ReviewerDecision,
    Valuation,
    ValuationApproval,
    WorksiteBulletin,
)
from croquito_valuation.rounding import money_trunc
from croquito_valuation.takeoff import (
    TakeoffDecisionBatch,
    TakeoffDecisionInput,
    TakeoffItem,
    TakeoffPacket,
)
from croquito_valuation.template import (
    AmendmentLayout,
    CatalogLayout,
    GeneralLayout,
    default_template,
)

SYNTHETIC_CATALOG_LABEL: Final = "MAPÃO SINTÉTICO FGV06 (fixture)"
SYNTHETIC_REFERENCE_MONTH: Final = "2026-01"
SYNTHETIC_WORKSITE_KEY: Final = "praca-sintetica-norte"
SYNTHETIC_WORKSITE_NAME: Final = "PRACA SINTETICA NORTE"
_SYNTHETIC_NAMESPACE: Final = uuid5(NAMESPACE_URL, "https://croquito.local/valuation/synthetic")
_FIXED_TIMESTAMP: Final = datetime(2026, 1, 1, tzinfo=UTC)
_COVER_SHEET: Final = "CAPA"
_CATALOG_SHEET: Final = "CATALOGO"

_CATALOG_ROWS: Final[tuple[tuple[str, str, str, str], ...]] = (
    ("AD", "PAVIMENTACAO SINTETICA", "", ""),
    ("AD0405", "PISO INTERTRAVADO SINTETICO", "", ""),
    ("AD04050050(/)", "PISO INTERTRAVADO SINTETICO 6CM CINZA", "M2", "89,30"),
    ("AD04050055(A)", "PISO INTERTRAVADO SINTETICO 8CM CINZA", "M2", "102,45"),
    ("AD04050055(B)", "PISO INTERTRAVADO SINTETICO 8CM COLORIDO", "m²", "118,60"),
    ("AD04050060(/)", "PISO INTERTRAVADO SINTETICO 10CM TRAFEGO", "m²", "131,20"),
    ("AD0410", "PISO CIMENTADO SINTETICO", "", ""),
    ("AD04100010(/)", "PISO CIMENTADO SINTETICO DESEMPENADO", "M2", "61,75"),
    ("AD04100015(/)", "PISO CIMENTADO SINTETICO ANTIDERRAPANTE", "M2", "74,90"),
    ("AD04100020(A)", "PISO CIMENTADO SINTETICO COM JUNTA PLASTICA", "M2", "82,15"),
    ("AD0415", "PISO EMBORRACHADO SINTETICO", "", ""),
    ("AD04150010(/)", "PISO EMBORRACHADO SINTETICO 40MM AMORTECEDOR", "M2", "210,40"),
    ("AD0420", "MEIO-FIO SINTETICO", "", ""),
    ("AD04200010(/)", "MEIO-FIO SINTETICO 15X30 ASSENTADO", "M", "58,40"),
    ("AD04200015(/)", "MEIO-FIO SINTETICO 12X25 ASSENTADO", "M", "49,85"),
    ("AD04200020(A)", "MEIO-FIO SINTETICO 15X30 REBAIXADO", "M", "63,70"),
    ("CE", "CERCAMENTO SINTETICO", "", ""),
    ("CE0210", "ALAMBRADO SINTETICO", "", ""),
    ("CE02100010(/)", "ALAMBRADO SINTETICO TELA GALVANIZADA 2POL", "M2", "128,35"),
    ("CE02100015(A)", "ALAMBRADO SINTETICO TELA GALVANIZADA 3POL H=2M", "M2", "142,75"),
    ("CE02100015(B)", "ALAMBRADO SINTETICO TELA REVESTIDA H=2M", "M2", "168,90"),
    ("CE02100020(/)", "PORTAO SINTETICO DE ALAMBRADO 2 FOLHAS", "UN", "1.580,00"),
    ("CE0220", "MURETA SINTETICA", "", ""),
    ("CE02200010(/)", "MURETA SINTETICA H=1,00M EM BLOCO", "M", "236,55"),
    ("CE02200015(/)", "MURETA SINTETICA H=1,50M EM BLOCO", "M", "289,40"),
    ("CE02200020(A)", "MURETA SINTETICA H=1,00M COM GRADIL", "M", "341,25"),
    ("MB", "MOBILIARIO URBANO SINTETICO", "", ""),
    ("MB0110", "BANCOS SINTETICOS", "", ""),
    ("MB01100010(/)", "BANCO SINTETICO DE CONCRETO PRE-MOLDADO", "UN", "486,20"),
    ("MB01100020(/)", "BANCO SINTETICO DE MADEIRA PLASTICA", "UNID.", "612,40"),
    ("MB01100030(A)", "BANCO SINTETICO COM ENCOSTO", "UNID.", "738,95"),
    ("MB0120", "LIXEIRAS E PLACAS SINTETICAS", "", ""),
    ("MB01200010(/)", "LIXEIRA SINTETICA 60L COM SUPORTE", "UN", "342,15"),
    ("MB01200020(/)", "LIXEIRA SINTETICA DUPLA PARA COLETA", "UN", "528,70"),
    ("MB01200030(/)", "PLACA SINTETICA DE IDENTIFICACAO DA OBRA", "UN", "20,15"),
    ("MB01200040(/)", "BEBEDOURO SINTETICO ACESSIVEL", "UNID", "1.245,00"),
    ("MB0130", "ILUMINACAO SINTETICA", "", ""),
    ("MB01300010(/)", "LUMINARIA SINTETICA DUPLA EM POSTE DE 4M", "UN", "2.150,00"),
    ("SP", "SERVICOS PRELIMINARES SINTETICOS", "", ""),
    ("SP0105", "MOVIMENTO DE TERRA SINTETICO", "", ""),
    ("SP01050010(/)", "ESCAVACAO SINTETICA MANUAL EM SOLO", "M3", "10,30"),
    ("SP01050015(A)", "ATERRO SINTETICO COMPACTADO MANUAL", "M³", "13,30"),
    ("SP01050020(/)", "REGULARIZACAO SINTETICA DE SUBLEITO", "M2", "18,45"),
    ("SP0210", "CANTEIRO DE OBRAS SINTETICO", "", ""),
    ("SP02100010(/)", "CONTAINER SINTETICO ESCRITORIO 6M", "MES", "1.850,00"),
    ("SP02100015(/)", "CONTAINER SINTETICO SANITARIO 6M", "MÊS", "1.620,00"),
    ("SP02100020(/)", "PLACA SINTETICA DE OBRA 2,00X1,00M", "UN", "480,00"),
)

_CATALOG_HEADER: Final = ("CODIGO", "DESCRICAO", "UNIDADE", "PRECO UNITARIO")


def build_synthetic_catalog_workbook(output_path: Path) -> Path:
    """Grava o mini-MAPÃO sintético, com a aba de preços oculta como no formato real."""
    workbook = Workbook()
    workbook.properties.created = _FIXED_TIMESTAMP
    cover = workbook.active
    cover.title = _COVER_SHEET
    cover["A1"] = "MAPAO SINTETICO — FIXTURE DE TESTE"
    cover["A2"] = "Nenhum dado real de cliente. Precos inventados."
    cover["A3"] = f"Referencia: {SYNTHETIC_REFERENCE_MONTH}"
    catalog = workbook.create_sheet(_CATALOG_SHEET)
    catalog.sheet_state = "hidden"
    for column, label in enumerate(_CATALOG_HEADER, start=1):
        catalog.cell(row=1, column=column, value=label)
    for index, (code, description, unit, price) in enumerate(_CATALOG_ROWS, start=2):
        catalog.cell(row=index, column=1, value=code)
        catalog.cell(row=index, column=2, value=description)
        if unit:
            catalog.cell(row=index, column=3, value=unit)
        if price:
            catalog.cell(row=index, column=4, value=price)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def _direct(label: str, name: str, value: str, unit: str) -> CalcBlock:
    quantity = Decimal(value)
    return CalcBlock(
        label=label,
        recipe=CalcRecipe.DIRECT_QUANTITY,
        operands=[CalcOperand(name=name, value=quantity, unit=unit)],
        subtotal=quantity,
    )


def _synthetic_calc_sheets() -> dict[str, CalcSheet]:
    """Memórias de cálculo da obra sintética, uma por item do boletim."""
    area_blocks = [
        CalcBlock(
            label="PASSEIO NORTE",
            recipe=CalcRecipe.LENGTH_TIMES_WIDTH,
            operands=[
                CalcOperand(name="COMPRIMENTO", value=Decimal("12.50"), unit="m"),
                CalcOperand(name="LARGURA", value=Decimal("8.40"), unit="m"),
            ],
            subtotal=Decimal("105.00"),
        ),
        CalcBlock(
            label="PASSEIO LESTE",
            recipe=CalcRecipe.LENGTH_TIMES_WIDTH,
            operands=[
                CalcOperand(name="COMPRIMENTO", value=Decimal("6.00"), unit="m"),
                CalcOperand(name="LARGURA", value=Decimal("3.50"), unit="m"),
            ],
            subtotal=Decimal("21.00"),
        ),
    ]
    fence_block = CalcBlock(
        label="ALAMBRADO DO CAMPO",
        recipe=CalcRecipe.PERIM_HEIGHT_MINUS_OPENINGS,
        operands=[
            CalcOperand(name="PERÍMETRO", value=Decimal("45.60"), unit="m"),
            CalcOperand(name="ALTURA", value=Decimal("2.00"), unit="m"),
        ],
        deductions=[CalcOperand(name="VÃOS", value=Decimal("6.00"), unit="m2")],
        subtotal=Decimal("85.20"),
    )
    gate_block = CalcBlock(
        label="FECHAMENTO DO DEPOSITO",
        recipe=CalcRecipe.PERIMETER_TIMES_HEIGHT,
        operands=[
            CalcOperand(name="PERÍMETRO", value=Decimal("4.00"), unit="m"),
            CalcOperand(name="ALTURA", value=Decimal("2.50"), unit="m"),
        ],
        subtotal=Decimal("10.00"),
    )
    container_block = CalcBlock(
        label="ESCRITORIO DE OBRA",
        recipe=CalcRecipe.QTY_TIMES_MONTHS,
        operands=[
            CalcOperand(name="QUANTIDADE", value=Decimal("2"), unit="un"),
            CalcOperand(name="MESES", value=Decimal("6"), unit="mes"),
        ],
        subtotal=Decimal("12.00"),
    )
    sheets = [
        CalcSheet(
            worksite_key=SYNTHETIC_WORKSITE_KEY,
            item_number="1",
            blocks=area_blocks,
            total_quantity=Decimal("126.00"),
        ),
        CalcSheet(
            worksite_key=SYNTHETIC_WORKSITE_KEY,
            item_number="2",
            blocks=[fence_block],
            total_quantity=Decimal("85.20"),
        ),
        CalcSheet(
            worksite_key=SYNTHETIC_WORKSITE_KEY,
            item_number="3",
            blocks=[gate_block],
            total_quantity=Decimal("10.00"),
        ),
        CalcSheet(
            worksite_key=SYNTHETIC_WORKSITE_KEY,
            item_number="4",
            blocks=[_direct("MURETA DA DIVISA", "EXTENSÃO", "18.40", "m")],
            total_quantity=Decimal("18.40"),
        ),
        CalcSheet(
            worksite_key=SYNTHETIC_WORKSITE_KEY,
            item_number="5",
            blocks=[_direct("BANCOS DA PRACA", "QUANTIDADE", "12.00", "un")],
            total_quantity=Decimal("12.00"),
        ),
        CalcSheet(
            worksite_key=SYNTHETIC_WORKSITE_KEY,
            item_number="6",
            blocks=[_direct("VALA DE DRENAGEM", "VOLUME", "1.15", "m3")],
            total_quantity=Decimal("1.15"),
        ),
        CalcSheet(
            worksite_key=SYNTHETIC_WORKSITE_KEY,
            item_number="7",
            blocks=[container_block],
            total_quantity=Decimal("12.00"),
        ),
        CalcSheet(
            worksite_key=SYNTHETIC_WORKSITE_KEY,
            item_number="8",
            blocks=[_direct("ATERRO DO TALUDE", "VOLUME", "4.35", "m3")],
            total_quantity=Decimal("4.35"),
        ),
    ]
    return {sheet.item_number: sheet for sheet in sheets}


_LINE_CODES: Final[tuple[tuple[str, str], ...]] = (
    ("1", "AD04050050(/)"),
    ("2", "CE02100015(A)"),
    ("3", "CE02100015(B)"),
    ("4", "CE02200010(/)"),
    ("5", "MB01100020(/)"),
    ("6", "SP01050010(/)"),
    ("7", "SP02100010(/)"),
    ("8", "SP01050015(A)"),
)


def build_synthetic_valuation(catalog: PriceCatalog) -> Valuation:
    """Constrói a medição sintética a partir do catálogo importado."""
    calc_sheets = _synthetic_calc_sheets()
    lines: list[BulletinLine] = []
    for item_number, code in _LINE_CODES:
        entry = catalog.entry_for(code)
        quantity = calc_sheets[item_number].total_quantity
        lines.append(
            BulletinLine(
                item_number=item_number,
                code=entry.code,
                description=entry.description,
                unit=entry.unit,
                unit_price=entry.unit_price,
                quantity=quantity,
                total=money_trunc(quantity * entry.unit_price),
            )
        )
    bulletin = WorksiteBulletin(
        worksite_key=SYNTHETIC_WORKSITE_KEY,
        worksite_name=SYNTHETIC_WORKSITE_NAME,
        address="RUA SINTETICA 100 — BAIRRO SINTETICO",
        contract_label="CONTRATO SINTETICO 001/2026",
        lines=lines,
        total_amount=sum((line.total for line in lines), Decimal("0.00")),
    )
    return Valuation(
        id=uuid5(_SYNTHETIC_NAMESPACE, f"{SYNTHETIC_WORKSITE_KEY}/{SYNTHETIC_REFERENCE_MONTH}/1"),
        period_number=1,
        reference_label="JANEIRO/2026",
        bulletins=[bulletin],
        calc_sheets=[calc_sheets[item_number] for item_number, _ in _LINE_CODES],
    )


# ---------------------------------------------------------------------------------------
# M2: o MAPÃO anterior completo e a medição multi-obra do período seguinte.
# ---------------------------------------------------------------------------------------

SYNTHETIC_CONTRACT_SOURCE_LABEL: Final = "MAPÃO SINTÉTICO 2ª MEDIÇÃO (fixture)"
SYNTHETIC_CONTRACT_LABEL: Final = "CONTRATO SINTETICO 001/2026"
SYNTHETIC_PREVIOUS_PERIOD_COUNT: Final = 2
SYNTHETIC_PERIOD_REFERENCE_LABEL: Final = "MARÇO/2026"
SYNTHETIC_REVIEWER_ID: Final = "orcamentista-sintetico"

_ZERO: Final = Decimal("0.00")
_PAVING_GROUP: Final = "PAVIMENTACAO SINTETICA"
_SERVICES_GROUP: Final = "SERVICOS PRELIMINARES E MOBILIARIO"
_FENCING_GROUP: Final = "CERCAMENTO E ILUMINACAO SINTETICOS"
_AMENDMENT_LABEL: Final = "1ª RE-RA"
_MULTI_NAMESPACE: Final = uuid5(NAMESPACE_URL, "https://croquito.local/valuation/synthetic-multi")

_CATALOG_BY_CODE: Final[dict[str, tuple[str, str, str]]] = {
    code: (description, unit, price) for code, description, unit, price in _CATALOG_ROWS if price
}


def _catalog_row(code: str) -> tuple[str, str, Decimal]:
    """Descrição, unidade escrita e preço do código, direto das linhas do catálogo.

    O MAPÃO anterior nasce dos mesmos valores que a aba de preços declara, de modo que o
    consolidado importado e o catálogo importado não possam divergir em preço ou unidade.
    """
    row = _CATALOG_BY_CODE.get(code)
    if row is None:  # pragma: no cover - protege a fixture contra código digitado errado
        raise ValuationValidationError(
            "SYNTHETIC_CODE_UNKNOWN",
            "a fixture sintética cita código que não existe no catálogo do módulo",
            {"code": code},
        )
    description, unit, price = row
    return description, unit, parse_money(price)


@dataclass(frozen=True, slots=True)
class _ContractItem:
    """Uma linha do MAPÃO anterior: quantidades declaradas e efeito da RE-RA.

    `period_quantities` com `None` é célula vazia — item que não mediu naquele período.
    O par QUANTIDADE|VALOR, o acumulado e o saldo são derivados aqui e escritos na
    planilha, para que o arquivo e o consolidado esperado não possam divergir.
    """

    group_label: str
    item_number: str
    code: str
    contract_quantity: Decimal
    period_quantities: tuple[Decimal | None, ...]
    reduced: Decimal | None = None
    new_item: Decimal | None = None

    @property
    def description(self) -> str:
        return _catalog_row(self.code)[0]

    @property
    def unit_label(self) -> str:
        """Unidade como a planilha a escreve (`M2`, `M³`, `UN`), antes de normalizar."""
        return _catalog_row(self.code)[1]

    @property
    def unit_price(self) -> Decimal:
        return _catalog_row(self.code)[2]

    @property
    def amended_quantity(self) -> Decimal:
        """Vigente: contratual mais o efeito líquido das RE-RA sobre o código."""
        delta = _ZERO
        if self.reduced is not None:
            delta -= self.reduced
        if self.new_item is not None:
            delta += self.new_item
        return self.contract_quantity + delta

    @property
    def period_amounts(self) -> tuple[Decimal, ...]:
        """Valor de cada período: quantidade x preço truncado, como a planilha grava."""
        return tuple(
            money_trunc((_ZERO if quantity is None else quantity) * self.unit_price)
            for quantity in self.period_quantities
        )

    @property
    def accumulated_quantity(self) -> Decimal:
        return sum(
            ((_ZERO if quantity is None else quantity) for quantity in self.period_quantities),
            _ZERO,
        )

    @property
    def accumulated_amount(self) -> Decimal:
        return sum(self.period_amounts, _ZERO)

    @property
    def balance_quantity(self) -> Decimal:
        return self.amended_quantity - self.accumulated_quantity


_PREVIOUS_ITEMS: Final[tuple[_ContractItem, ...]] = (
    _ContractItem(
        group_label=_PAVING_GROUP,
        item_number="1",
        code="AD04050050(/)",
        contract_quantity=Decimal("40.00"),
        period_quantities=(Decimal("8.00"), Decimal("6.00")),
        reduced=Decimal("5.00"),
    ),
    _ContractItem(
        group_label=_PAVING_GROUP,
        item_number="2",
        code="AD04050055(A)",
        contract_quantity=Decimal("10.00"),
        # 4,50 x 102,45 = 461,025: dinheiro trunca em 461,02 e o saldo fica em 0,50.
        period_quantities=(Decimal("5.00"), Decimal("4.50")),
    ),
    _ContractItem(
        group_label=_PAVING_GROUP,
        item_number="3",
        code="AD04100010(/)",
        contract_quantity=Decimal("20.00"),
        period_quantities=(None, None),
    ),
    _ContractItem(
        group_label=_PAVING_GROUP,
        item_number="4",
        code="AD04050060(/)",
        contract_quantity=Decimal("80.00"),
        period_quantities=(None, None),
    ),
    _ContractItem(
        group_label=_PAVING_GROUP,
        item_number="5",
        code="AD04150010(/)",
        contract_quantity=Decimal("25.00"),
        period_quantities=(None, None),
    ),
    _ContractItem(
        group_label=_SERVICES_GROUP,
        item_number="6",
        code="SP01050010(/)",
        contract_quantity=Decimal("30.00"),
        # 4,15 x 10,30 = 42,745, outro truncamento que arredondar mudaria de centavo.
        period_quantities=(Decimal("10.00"), Decimal("4.15")),
    ),
    _ContractItem(
        group_label=_SERVICES_GROUP,
        item_number="7",
        code="SP01050015(A)",
        contract_quantity=Decimal("8.00"),
        period_quantities=(Decimal("2.00"), Decimal("1.00")),
    ),
    _ContractItem(
        group_label=_SERVICES_GROUP,
        item_number="8",
        code="MB01100010(/)",
        contract_quantity=_ZERO,
        period_quantities=(None, None),
        # Vigente 12,00 cobre o mesmo código medido em duas obras: 4,00 (leste) + 4,00
        # (obra do takeoff).
        new_item=Decimal("12.00"),
    ),
    _ContractItem(
        group_label=_FENCING_GROUP,
        item_number="9",
        code="CE02100010(/)",
        contract_quantity=Decimal("70.00"),
        period_quantities=(None, None),
    ),
    _ContractItem(
        group_label=_FENCING_GROUP,
        item_number="10",
        code="MB01300010(/)",
        contract_quantity=Decimal("10.00"),
        period_quantities=(None, None),
    ),
)


def _period_columns(layout: GeneralLayout) -> tuple[list[int], int, int]:
    """Colunas derivadas do template: pares de medição, acumulado e saldo."""
    first = column_index_from_string(layout.first_period_column)
    pairs = [first + 2 * index for index in range(SYNTHETIC_PREVIOUS_PERIOD_COUNT)]
    accumulated = first + 2 * SYNTHETIC_PREVIOUS_PERIOD_COUNT
    return pairs, accumulated, accumulated + 2


def _write_catalog_sheet(worksheet: Any, layout: CatalogLayout) -> None:
    """Aba de preços do MAPÃO, nas colunas que o template declara."""
    header_row = layout.first_row - 1
    for letter, label in (
        (layout.code_column, "CODIGO"),
        (layout.description_column, "DESCRICAO"),
        (layout.unit_column, "UNIDADE"),
        (layout.price_column, "PRECO UNITARIO"),
    ):
        if header_row >= 1:
            worksheet[f"{letter}{header_row}"] = label
    row = layout.first_row
    for code, description, unit, price in _CATALOG_ROWS:
        worksheet[f"{layout.code_column}{row}"] = code
        worksheet[f"{layout.description_column}{row}"] = description
        if unit:
            worksheet[f"{layout.unit_column}{row}"] = unit
        if price:
            worksheet[f"{layout.price_column}{row}"] = price
        row += 1


def _write_general_item(
    worksheet: Any, layout: GeneralLayout, item: _ContractItem, row: int
) -> None:
    pairs, accumulated, balance = _period_columns(layout)
    worksheet[f"{layout.item_column}{row}"] = item.item_number
    worksheet[f"{layout.code_column}{row}"] = item.code
    worksheet[f"{layout.description_column}{row}"] = item.description
    worksheet[f"{layout.unit_column}{row}"] = item.unit_label
    worksheet[f"{layout.contract_quantity_column}{row}"] = item.contract_quantity
    worksheet[f"{layout.unit_price_column}{row}"] = item.unit_price
    worksheet[f"{layout.amended_quantity_column}{row}"] = item.amended_quantity
    for quantity, column, amount in zip(
        item.period_quantities, pairs, item.period_amounts, strict=True
    ):
        if quantity is None:
            continue
        worksheet.cell(row=row, column=column, value=quantity)
        worksheet.cell(row=row, column=column + 1, value=amount)
    worksheet.cell(row=row, column=accumulated, value=item.accumulated_quantity)
    worksheet.cell(row=row, column=accumulated + 1, value=item.accumulated_amount)
    worksheet.cell(row=row, column=balance, value=item.balance_quantity)


def _write_general_sheet(
    worksheet: Any, layout: GeneralLayout, items: Sequence[_ContractItem]
) -> None:
    """PLANILHA GERAL do contrato depois de duas medições, só com posições do template."""
    pairs, accumulated, balance = _period_columns(layout)
    worksheet[f"{layout.group_column}1"] = layout.title
    for letter, label in (
        (layout.group_column, "GRUPO"),
        (layout.item_column, "ITEM"),
        (layout.code_column, "CODIGO"),
        (layout.description_column, "DESCRICAO"),
        (layout.unit_column, "UN"),
        (layout.contract_quantity_column, "QUANT. CONTRATO"),
        (layout.unit_price_column, "VALOR UNIT"),
        (layout.amended_quantity_column, "QUANT. VIGENTE"),
    ):
        worksheet[f"{letter}{layout.header_row}"] = label
    for number, column in enumerate(pairs, start=1):
        worksheet.cell(row=layout.header_row, column=column, value=layout.period_label(number))
    worksheet.cell(row=layout.header_row, column=accumulated, value=layout.accumulated_label)
    worksheet.cell(row=layout.header_row, column=balance, value=layout.balance_label)
    sublabel_row = layout.pair_sublabel_row
    if sublabel_row is not None:
        for column in (*pairs, accumulated):
            worksheet.cell(row=sublabel_row, column=column, value=layout.quantity_pair_label)
            worksheet.cell(row=sublabel_row, column=column + 1, value=layout.amount_pair_label)
        worksheet.cell(row=sublabel_row, column=balance, value=layout.quantity_pair_label)

    row = layout.data_first_row
    group_label = ""
    for item in items:
        if item.group_label != group_label:
            worksheet[f"{layout.group_column}{row}"] = item.group_label
            group_label = item.group_label
            row += 1
        _write_general_item(worksheet, layout, item, row)
        row += 1
    row += 1  # linha totalmente vazia, ignorada pela varredura
    worksheet[f"{layout.group_column}{row}"] = layout.total_label
    worksheet.cell(
        row=row,
        column=accumulated + 1,
        value=sum((item.accumulated_amount for item in items), _ZERO),
    )


def _write_amendment_sheet(
    worksheet: Any, layout: AmendmentLayout, items: Sequence[_ContractItem]
) -> None:
    """Aba da prefeitura: uma RE-RA com uma redução e um item novo de contratual zero."""
    block = layout.blocks[0]
    worksheet[f"{layout.code_column}{layout.header_row}"] = "CODIGO"
    if layout.amended_quantity_column is not None:
        worksheet[f"{layout.amended_quantity_column}{layout.header_row}"] = "QUANT. VIGENTE"
    for letter, label in (
        (block.reduced_column, f"{block.label} REDUZIDA"),
        (block.added_column, f"{block.label} ACRESCIDA"),
        (block.new_item_column, f"{block.label} ITEM NOVO"),
    ):
        if letter is not None:
            worksheet[f"{letter}{layout.header_row}"] = label
    row = layout.data_first_row
    for item in items:
        worksheet[f"{layout.code_column}{row}"] = item.code
        if layout.amended_quantity_column is not None:
            worksheet[f"{layout.amended_quantity_column}{row}"] = item.amended_quantity
        for letter, value in (
            (block.reduced_column, item.reduced),
            (block.new_item_column, item.new_item),
        ):
            if letter is not None and value is not None:
                worksheet[f"{letter}{row}"] = value
        row += 1


def build_synthetic_previous_mapao(path: Path) -> Path:
    """Grava o MAPÃO sintético da medição anterior: catálogo, GERAL e RE-RA num arquivo só.

    É o formato do arquivo real do cliente: a aba de preços e a PLANILHA GERAL nascem
    ocultas e a aba da prefeitura carrega a RE-RA. Nenhuma posição de célula é constante
    aqui — linhas, colunas e rótulos saem todos do `default_template()`.
    """
    template = default_template()
    if template.amendment is None:  # pragma: no cover - o template padrão declara a aba
        raise ValuationValidationError(
            "TEMPLATE_AMENDMENT_MISSING",
            "o template padrão precisa declarar a aba de RE-RA para a fixture do M2",
            {"template": template.label},
        )
    workbook = Workbook()
    workbook.properties.created = _FIXED_TIMESTAMP
    cover = workbook.active
    cover.title = _COVER_SHEET
    cover["A1"] = SYNTHETIC_CONTRACT_SOURCE_LABEL
    cover["A2"] = "Nenhum dado real de cliente. Precos e quantidades inventados."
    cover["A3"] = f"Referencia do catalogo: {SYNTHETIC_REFERENCE_MONTH}"
    catalog = workbook.create_sheet(template.catalog.sheet_name)
    catalog.sheet_state = "hidden"
    _write_catalog_sheet(catalog, template.catalog)
    general = workbook.create_sheet(template.general.sheet_name)
    general.sheet_state = "hidden"
    _write_general_sheet(general, template.general, _PREVIOUS_ITEMS)
    _write_amendment_sheet(
        workbook.create_sheet(template.amendment.sheet_name), template.amendment, _PREVIOUS_ITEMS
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


@dataclass(frozen=True, slots=True)
class _MeasuredItem:
    """Um código medido numa obra; com dedução, a memória exercita o desconto de vãos."""

    code: str
    quantity: Decimal
    label: str
    deduction: Decimal | None = None


@dataclass(frozen=True, slots=True)
class _MeasuredWorksite:
    """Uma obra do período e o que ela mediu."""

    worksite_key: str
    worksite_name: str
    address: str
    items: tuple[_MeasuredItem, ...]


_MULTI_WORKSITE_MEASUREMENTS: Final[tuple[_MeasuredWorksite, ...]] = (
    _MeasuredWorksite(
        worksite_key="praca-sintetica-norte",
        worksite_name="PRACA SINTETICA NORTE",
        address="RUA SINTETICA 100 — BAIRRO SINTETICO",
        items=(
            # O bloco com dedução mantém viva a forma `=ROUND(PRODUCT(...),2)-<ref>`.
            _MeasuredItem(
                code="AD04050050(/)",
                quantity=Decimal("2.00"),
                label="PASSEIO NORTE",
                deduction=Decimal("1.00"),
            ),
            # 0,50 é exatamente o saldo do código: a medição encosta no limite sem passar.
            _MeasuredItem(
                code="AD04050055(A)", quantity=Decimal("0.50"), label="REMENDO DA ENTRADA"
            ),
        ),
    ),
    _MeasuredWorksite(
        worksite_key="praca-sintetica-sul",
        worksite_name="PRACA SINTETICA SUL",
        address="RUA SINTETICA 200 — BAIRRO SINTETICO",
        items=(
            # Mesmo código da obra do norte: a GERAL consolida 2,00 + 3,00 numa linha só, e
            # TRUNC(5,00 x 89,30) é igual a TRUNC(2,00 x 89,30) + TRUNC(3,00 x 89,30).
            _MeasuredItem(code="AD04050050(/)", quantity=Decimal("3.00"), label="PASSEIO SUL"),
            _MeasuredItem(code="SP01050010(/)", quantity=Decimal("1.25"), label="VALA DE DRENAGEM"),
        ),
    ),
    _MeasuredWorksite(
        worksite_key="praca-sintetica-leste",
        worksite_name="PRACA SINTETICA LESTE",
        address="RUA SINTETICA 300 — BAIRRO SINTETICO",
        items=(
            _MeasuredItem(code="MB01100010(/)", quantity=Decimal("4.00"), label="BANCOS DA PRACA"),
        ),
    ),
)


def _measured_block(item: _MeasuredItem, unit: str) -> CalcBlock:
    """Bloco de cálculo cuja quantidade recomputada é exatamente a medida."""
    if item.deduction is None:
        return CalcBlock(
            label=item.label,
            recipe=CalcRecipe.DIRECT_QUANTITY,
            operands=[CalcOperand(name="QUANTIDADE", value=item.quantity, unit=unit)],
            subtotal=item.quantity,
        )
    return CalcBlock(
        label=item.label,
        recipe=CalcRecipe.PERIM_HEIGHT_MINUS_OPENINGS,
        operands=[
            CalcOperand(name="PERÍMETRO", value=item.quantity + item.deduction, unit="m"),
            CalcOperand(name="ALTURA", value=Decimal("1.00"), unit="m"),
        ],
        deductions=[CalcOperand(name="VÃOS", value=item.deduction, unit=unit)],
        subtotal=item.quantity,
    )


def build_synthetic_multi_valuation(
    contract: ContractWorkbook,
    catalog: PriceCatalog,
    *,
    takeoff: CalcBuildResult | None = None,
) -> Valuation:
    """Medição do período seguinte ao consolidado, com três obras e um código repetido.

    As quantidades cabem no saldo de cada código e foram escolhidas para que o valor
    consolidado do código medido em duas obras não derive de centavo: `TRUNC(Σq x preço)`
    é igual a `Σ TRUNC(q_i x preço)`. Preço, unidade e descrição saem do catálogo
    importado, que é o mesmo arquivo do consolidado.

    Com `takeoff`, a quarta obra é a que nasceu da cadeia do M4 — takeoff revisado,
    códigos confirmados e plano de cálculo — e entra na medição como qualquer outra: o
    boletim e as memórias vêm prontos do builder do domínio, e a chave da obra participa
    do id determinístico da medição.
    """
    period_number = contract.next_period_number
    bulletins: list[WorksiteBulletin] = []
    calc_sheets: list[CalcSheet] = []
    for worksite in _MULTI_WORKSITE_MEASUREMENTS:
        lines: list[BulletinLine] = []
        for index, item in enumerate(worksite.items, start=1):
            entry = catalog.entry_for(item.code)
            item_number = str(index)
            lines.append(
                BulletinLine(
                    item_number=item_number,
                    code=entry.code,
                    description=entry.description,
                    unit=entry.unit,
                    unit_price=entry.unit_price,
                    quantity=item.quantity,
                    total=money_trunc(item.quantity * entry.unit_price),
                )
            )
            calc_sheets.append(
                CalcSheet(
                    worksite_key=worksite.worksite_key,
                    item_number=item_number,
                    blocks=[_measured_block(item, entry.unit)],
                    total_quantity=item.quantity,
                )
            )
        bulletins.append(
            WorksiteBulletin(
                worksite_key=worksite.worksite_key,
                worksite_name=worksite.worksite_name,
                address=worksite.address,
                contract_label=SYNTHETIC_CONTRACT_LABEL,
                lines=lines,
                total_amount=sum((line.total for line in lines), _ZERO),
            )
        )
    worksite_keys = [worksite.worksite_key for worksite in _MULTI_WORKSITE_MEASUREMENTS]
    if takeoff is not None:
        bulletins.append(takeoff.bulletin)
        calc_sheets.extend(takeoff.calc_sheets)
        worksite_keys.append(takeoff.bulletin.worksite_key)
    keys = "/".join(worksite_keys)
    return Valuation(
        id=uuid5(_MULTI_NAMESPACE, f"{keys}/{period_number}"),
        period_number=period_number,
        reference_label=SYNTHETIC_PERIOD_REFERENCE_LABEL,
        bulletins=bulletins,
        calc_sheets=calc_sheets,
    )


def build_synthetic_approval(valuation: Valuation) -> Valuation:
    """Devolve a medição com a aprovação nominal sintética amarrada ao digest do conteúdo.

    A decisão é fixa (id, revisor e horário) para que a demonstração continue
    determinística; qualquer edição posterior da medição muda o digest e invalida a
    aprovação no portão de exportação.
    """
    approval = ValuationApproval(
        decision=ReviewerDecision(
            decision_id="vd_5ec0a1d2c3b4e5f6",
            action="confirm",
            reviewer_id=SYNTHETIC_REVIEWER_ID,
            reviewer_role="orcamentista",
            decided_at=datetime(2026, 3, 31, 12, 0, tzinfo=UTC),
            note="medição sintética conferida contra o consolidado do contrato",
        ),
        valuation_digest=valuation.content_digest(),
    )
    payload = valuation.model_dump()
    payload["approval"] = approval.model_dump()
    return Valuation.model_validate(payload)


# ---------------------------------------------------------------------------------------
# M3/M4: a obra que nasce da prancha — revisão do takeoff, código confirmado e cálculo.
# ---------------------------------------------------------------------------------------

SYNTHETIC_TAKEOFF_REVIEWER: Final = "orcamentista-sintetico"
SYNTHETIC_TAKEOFF_DECIDED_AT: Final = datetime(2026, 2, 1, 12, 0, tzinfo=UTC)
SYNTHETIC_TAKEOFF_AMBIGUOUS_QUANTITY: Final = Decimal("18.40")
"""Quantidade que o orçamentista sintético informa para a linha ilegível da prancha. Ela
não é lida de lugar nenhum de propósito: é o ato humano suprindo o que a extração não
conseguiu ler."""
SYNTHETIC_TAKEOFF_AMBIGUOUS_NOTE: Final = (
    "quantidade ilegível na prancha; informada pelo orçamentista sintético"
)

SYNTHETIC_TAKEOFF_WORKSITE_KEY: Final = "praca-sintetica-oeste"
SYNTHETIC_TAKEOFF_WORKSITE_NAME: Final = "PRACA SINTETICA OESTE"
SYNTHETIC_TAKEOFF_WORKSITE_ADDRESS: Final = "RUA SINTETICA 400 — BAIRRO SINTETICO"
SYNTHETIC_CODE_DECIDED_AT: Final = datetime(2026, 3, 30, 12, 0, tzinfo=UTC)

_PAVEMENT_LABEL: Final = "PISO INTERTRAVADO SINTETICO"
_LAWN_LABEL: Final = "GRAMADO SINTETICO"
_FENCE_LABEL: Final = "ALAMBRADO SINTETICO"
_BENCH_LABEL: Final = "BANCO DE CONCRETO SINTETICO"
_LAMP_LABEL: Final = "LUMINARIA DUPLA SINTETICA"
_INTERVENTION_LABEL: Final = "AREA DE INTERVENCAO SINTETICA"
_RUBBER_LABEL: Final = "PISO EMBORRACHADO SINTETICO"

_FENCE_LENGTH: Final = Decimal("48.75")
_FENCE_HEIGHT: Final = Decimal("1.20")
_FENCE_AREA: Final = Decimal("58.50")
"""48,75 m x 1,20 m: o projetista entrega o alambrado em metro linear com a altura
anotada e o orçamentista mede em m². É a correção de revisão que o `TakeoffDecisionInput`
existe para registrar, e é ela que o plano de cálculo imprime na memória."""


def _written(value: Decimal) -> str:
    """Número como a nota do revisor o escreve: vírgula decimal.

    A nota cita os mesmos operandos do plano de cálculo em vez de repeti-los em texto:
    número escrito duas vezes é número que pode divergir.
    """
    return f"{value}".replace(".", ",")


@dataclass(frozen=True, slots=True)
class _DemoTakeoffDecision:
    """Uma decisão do orçamentista sintético sobre a linha da legenda de mesmo rótulo."""

    label: str
    action: Literal["confirm", "reject"]
    quantity: Decimal | None = None
    unit: str | None = None
    note: str | None = None


_DEMO_TAKEOFF_DECISIONS: Final[tuple[_DemoTakeoffDecision, ...]] = (
    _DemoTakeoffDecision(label=_PAVEMENT_LABEL, action="confirm"),
    _DemoTakeoffDecision(label=_LAWN_LABEL, action="confirm"),
    _DemoTakeoffDecision(
        label=_FENCE_LABEL,
        action="confirm",
        quantity=_FENCE_AREA,
        unit="m2",
        note=(f"convertido para área: {_written(_FENCE_LENGTH)} m x h={_written(_FENCE_HEIGHT)} m"),
    ),
    _DemoTakeoffDecision(label=_BENCH_LABEL, action="confirm"),
    _DemoTakeoffDecision(label=_LAMP_LABEL, action="confirm"),
    _DemoTakeoffDecision(
        label=_INTERVENTION_LABEL,
        action="reject",
        note="área de referência da prancha; não é serviço medido",
    ),
    _DemoTakeoffDecision(
        label=_RUBBER_LABEL,
        action="confirm",
        quantity=SYNTHETIC_TAKEOFF_AMBIGUOUS_QUANTITY,
        note=SYNTHETIC_TAKEOFF_AMBIGUOUS_NOTE,
    ),
)


@dataclass(frozen=True, slots=True)
class _DemoCodeAssignment:
    """Decisão de código do orçamentista sintético sobre o item de mesmo rótulo."""

    label: str
    action: Literal["confirm", "reject"]
    code: str | None = None
    note: str | None = None


_DEMO_CODE_ASSIGNMENTS: Final[tuple[_DemoCodeAssignment, ...]] = (
    _DemoCodeAssignment(label=_PAVEMENT_LABEL, action="confirm", code="AD04050060(/)"),
    _DemoCodeAssignment(
        label=_LAWN_LABEL,
        action="reject",
        note="sem cotação aplicável no contrato sintético",
    ),
    _DemoCodeAssignment(label=_FENCE_LABEL, action="confirm", code="CE02100010(/)"),
    _DemoCodeAssignment(label=_BENCH_LABEL, action="confirm", code="MB01100010(/)"),
    _DemoCodeAssignment(label=_LAMP_LABEL, action="confirm", code="MB01300010(/)"),
    _DemoCodeAssignment(label=_RUBBER_LABEL, action="confirm", code="AD04150010(/)"),
)


DEMO_EXPECTED_CODE_BY_LABEL: Final[Mapping[str, str]] = MappingProxyType(
    {
        assignment.label: assignment.code
        for assignment in _DEMO_CODE_ASSIGNMENTS
        if assignment.code is not None
    }
)
"""Gabarito público rótulo → código confirmado pelo orçamentista sintético.

É a MESMA fonte de `build_demo_code_assignments`, exposta como visão só-leitura para que a
eval de extração meça a sugestão contra o gabarito sem importar o privado — e sem que um
segundo mapa possa divergir dele. O item cujo código o demo rejeita (o gramado) fica de
fora: onde o gabarito não tem código, não existe acerto a medir."""


def _item_for_label(packet: TakeoffPacket, label: str) -> TakeoffItem:
    """Item da prancha pelo rótulo da legenda; rótulo ausente falha alto.

    A fixture cita o rótulo impresso, não a posição: mexer na legenda da prancha sem
    mexer aqui recusa em vez de decidir sobre o item errado.
    """
    for item in packet.items:
        if item.label == label:
            return item
    raise ValuationValidationError(
        "SYNTHETIC_LABEL_UNKNOWN",
        "a fixture sintética cita rótulo de legenda que a prancha não imprime",
        {"label": label, "labels": [item.label for item in packet.items]},
    )


def build_demo_takeoff_decisions(packet: TakeoffPacket) -> TakeoffDecisionBatch:
    """Revisão do orçamentista sintético sobre a legenda inteira da prancha.

    Não é confirmação em massa: a área de intervenção é **rejeitada** porque é referência
    de prancha e não serviço medido, o alambrado é confirmado com a conversão de metro
    linear para m² anotada, e a linha ilegível só fecha porque o revisor informa a
    quantidade que a extração não conseguiu ler.
    """
    return TakeoffDecisionBatch(
        decisions=[
            TakeoffDecisionInput(
                item_id=_item_for_label(packet, decision.label).id,
                action=decision.action,
                reviewer_id=SYNTHETIC_TAKEOFF_REVIEWER,
                reviewer_role="orcamentista",
                decided_at=SYNTHETIC_TAKEOFF_DECIDED_AT,
                quantity=decision.quantity,
                unit=decision.unit,
                note=decision.note,
            )
            for decision in _DEMO_TAKEOFF_DECISIONS
        ]
    )


def build_demo_code_assignments(packet: TakeoffPacket) -> CodeAssignmentBatch:
    """Confirmação de código do orçamentista sintético sobre o pacote JÁ REVISADO.

    O gramado é rejeitado: o contrato sintético não tem cotação para ele, e um item sem
    código não vira linha de boletim. Os demais recebem o código do catálogo que a
    sugestão lexical propôs — a confirmação continua sendo o ato humano.
    """
    return CodeAssignmentBatch(
        assignments=[
            CodeAssignmentInput(
                item_id=_item_for_label(packet, assignment.label).id,
                action=assignment.action,
                code=assignment.code,
                reviewer_id=SYNTHETIC_TAKEOFF_REVIEWER,
                reviewer_role="orcamentista",
                decided_at=SYNTHETIC_CODE_DECIDED_AT,
                note=assignment.note,
            )
            for assignment in _DEMO_CODE_ASSIGNMENTS
        ]
    )


def build_demo_calc_plan(packet: TakeoffPacket) -> CalcPlan:
    """Plano de memória do alambrado: a única quantidade que se decompõe na prancha.

    O subtotal recomputado (48,75 x 1,20 = 58,50) tem de fechar com a quantidade que o
    revisor confirmou; os demais itens ficam com o bloco de quantidade direta que o
    builder do domínio gera sozinho.
    """
    return CalcPlan(
        plans=[
            ItemCalcPlan(
                item_id=_item_for_label(packet, _FENCE_LABEL).id,
                blocks=[
                    CalcBlockPlan(
                        label="ALAMBRADO DA QUADRA SINTETICA",
                        recipe=CalcRecipe.PERIMETER_TIMES_HEIGHT,
                        operands=[
                            CalcOperand(name="EXTENSÃO", value=_FENCE_LENGTH, unit="m"),
                            CalcOperand(name="ALTURA", value=_FENCE_HEIGHT, unit="m"),
                        ],
                    )
                ],
            )
        ]
    )
