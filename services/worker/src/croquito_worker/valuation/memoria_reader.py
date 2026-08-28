"""Abre uma aba de memória de cálculo real (`.xlsx`) para o modo `--memoria` do
`precedent-eval` (F-044 T1, escopo ampliado): terceira entrada, para quando a praça
real chega como planilha de orçamento em vez de rodada gravada no sistema.

Ferramenta LOCAL de diagnóstico, da mesma família de `parity`/`bulletin_compare`: fora
da cadeia de medição e fora do CI. O arquivo do cliente nunca é copiado para o
repositório — muito menos para `tests/`, cujas fixtures continuam sintéticas — e nunca é
versionado; só entra por `--memoria` e só produz artefato dentro de `--output`.

`croquito_valuation.precedent.scan_memoria_rows` é quem interpreta o formato (puro,
sobre linhas já lidas); este módulo só abre o `.xlsx` real e entrega as linhas — o mesmo
desenho de `parity._open`/`bulletin_compare._open` (`read_only=True, data_only=True`:
valor em cache, não fórmula, e leitura em streaming em vez de carregar a pasta inteira
na árvore de objetos do openpyxl).
"""

from __future__ import annotations

from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from croquito_valuation.errors import ValuationValidationError
from croquito_valuation.precedent import MemoriaScan, scan_memoria_rows


def read_memoria_sheet(path: Path, sheet_name: str) -> MemoriaScan:
    """Abre `path`, lê `sheet_name` em cache e devolve os blocos que `scan_memoria_rows`
    encontrar. Recusa fechada e nomeada quando o arquivo não abre ou a aba não existe —
    nunca uma exceção crua do openpyxl subindo até o CLI."""
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (OSError, ValueError, KeyError, BadZipFile, InvalidFileException) as error:
        raise ValuationValidationError(
            "PRECEDENT_MEMORIA_WORKBOOK_UNREADABLE",
            "arquivo de memória de cálculo não pôde ser aberto como planilha",
            {"path": str(path), "reason": type(error).__name__},
        ) from error
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValuationValidationError(
                "PRECEDENT_MEMORIA_SHEET_NOT_FOUND",
                "aba não existe no arquivo de memória de cálculo",
                {
                    "path": str(path),
                    "sheet": sheet_name,
                    "available": list(workbook.sheetnames),
                },
            )
        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    return scan_memoria_rows(rows)
