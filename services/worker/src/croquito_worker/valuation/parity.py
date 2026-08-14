"""Paridade entre fórmula e valor em cache de uma pasta que o sistema não gerou.

Ferramenta **local** de diagnóstico, fora da cadeia de medição e fora do CI. Ela responde
uma pergunta só sobre a planilha que chega do cliente: o número que a célula mostra é o
número que a fórmula dela calcula? A planilha guarda os dois — a fórmula e o último valor
calculado pelo Excel — e uma pasta editada à mão pode ter os dois discordando sem que
nada acuse.

O que esta ferramenta **não** faz, de propósito:

- não recusa nem bloqueia nada: uma divergência aqui é achado, não invariante violado;
- não escreve no arquivo analisado e não copia conteúdo dele para o relatório além da
  fórmula, da referência da célula e dos números envolvidos;
- não interpreta fórmula fora da gramática fechada da auditoria (`GRAMMAR_PATTERNS`, de
  `croquito_valuation.canonical`): `VLOOKUP`, `IF`, referência a outra aba ou a outro
  arquivo são classificados e contados, nunca adivinhados.

A gramática é importada do módulo canônico em vez de reescrita: uma segunda cópia dela
poderia divergir em silêncio, e é justamente divergência silenciosa que se está caçando.
Já o `canonical_number` **não** é usado — ele recusaria número com mais de duas casas, e
aqui número com três casas é dado a reportar, não motivo para parar.

A aba que o template declara como catálogo é insumo de preço, não medição, e fica fora da
varredura pelo mesmo critério que `canonicalize_workbook` usa.
"""

from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Final, Literal
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import BaseModel, ConfigDict, Field

from croquito_valuation.canonical import CELL_REF_RE, GRAMMAR_PATTERNS, expand_range
from croquito_valuation.catalog import file_sha256
from croquito_valuation.errors import ValuationValidationError
from croquito_valuation.rounding import money_trunc, quantity_round
from croquito_valuation.template import WorkbookTemplate

ParityCode = Literal[
    "PARITY_VALUE_DIVERGENT",
    "PARITY_CACHE_MISSING",
    "PARITY_FORMULA_FOREIGN",
]
ParityCategory = Literal[
    "vlookup",
    "conditional",
    "external_reference",
    "cross_sheet",
    "other",
]

_ZERO: Final = Decimal("0.00")


class _ParityModel(BaseModel):
    """Configuração comum dos modelos do relatório de paridade."""

    model_config = ConfigDict(extra="forbid", validate_assignment=True)


class ParityFinding(_ParityModel):
    """Um achado sobre uma célula de fórmula do arquivo analisado."""

    code: ParityCode
    sheet: str
    ref: str
    formula: str
    cached: str | None = None
    recomputed: str | None = None
    missing_refs: list[str] = Field(default_factory=list)
    category: ParityCategory | None = None


class ParitySheetReport(_ParityModel):
    """Contagem de uma aba: quantas fórmulas e como cada uma terminou."""

    name: str
    formulas: int = Field(ge=0)
    verified: int = Field(ge=0)
    divergent: int = Field(ge=0)
    foreign: int = Field(ge=0)
    cache_missing: int = Field(ge=0)


class ParityReport(_ParityModel):
    """Resultado da varredura de paridade do arquivo inteiro."""

    workbook_sha256: str = Field(pattern=r"^[a-f0-9]{64}$")
    sheets: list[ParitySheetReport] = Field(default_factory=list)
    findings: list[ParityFinding] = Field(default_factory=list)

    @property
    def formulas(self) -> int:
        """Quantas células de fórmula foram encontradas."""
        return sum(sheet.formulas for sheet in self.sheets)

    @property
    def verified(self) -> int:
        """Quantas fórmulas da gramática reproduzem o valor em cache."""
        return sum(sheet.verified for sheet in self.sheets)

    @property
    def divergent(self) -> int:
        """Quantas fórmulas da gramática discordam do valor em cache."""
        return sum(sheet.divergent for sheet in self.sheets)

    @property
    def foreign(self) -> int:
        """Quantas fórmulas estão fora da gramática fechada."""
        return sum(sheet.foreign for sheet in self.sheets)

    @property
    def cache_missing(self) -> int:
        """Quantas fórmulas não puderam ser comparadas por falta de valor em cache."""
        return sum(sheet.cache_missing for sheet in self.sheets)


@dataclass(frozen=True, slots=True)
class _SheetCells:
    """As duas leituras da mesma aba: fórmulas e valores calculados em cache."""

    formulas: dict[str, str]
    cached: dict[str, object]


def _numeric(value: object) -> Decimal | None:
    """Valor numérico da célula em `Decimal`, ou `None` quando é texto, erro ou vazio.

    A conversão passa uma única vez por `Decimal(str(...))`, como no resto do módulo.
    Diferente da leitura canônica, aqui número com mais de duas casas é aceito: o objetivo
    é reportar o que o arquivo diz, não normalizá-lo.
    """
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int | float | Decimal):
        try:
            return Decimal(str(value))
        except InvalidOperation:  # pragma: no cover - float infinito não chega da planilha
            return None
    return None


def _required_values(cached: dict[str, object], refs: list[str]) -> tuple[list[Decimal], list[str]]:
    """Valores dos operandos obrigatórios e as refs que não têm número em cache."""
    values: list[Decimal] = []
    missing: list[str] = []
    for ref in refs:
        number = _numeric(cached.get(ref))
        if number is None:
            missing.append(ref)
        else:
            values.append(number)
    return values, missing


def _range_values(cached: dict[str, object], start: str, end: str) -> list[Decimal]:
    """Números do intervalo; como no Excel, texto e célula vazia não entram na conta."""
    values: list[Decimal] = []
    for ref in expand_range(start, end):
        number = _numeric(cached.get(ref))
        if number is not None:
            values.append(number)
    return values


def _product(values: list[Decimal]) -> Decimal:
    if not values:
        return _ZERO
    result = Decimal(1)
    for value in values:
        result *= value
    return result


def _match_grammar(formula: str) -> tuple[str, Any] | None:
    """Primeira forma da gramática que casa, na ordem de precedência do avaliador."""
    for shape, pattern in GRAMMAR_PATTERNS:
        match = pattern.fullmatch(formula)
        if match is not None:
            return shape, match
    return None


def _not_grammar(formula: str) -> ValuationValidationError:
    return ValuationValidationError(
        "FORMULA_UNSUPPORTED",
        "fórmula parecida com a gramática, mas com argumento que ela não aceita",
        {"formula": formula},
    )


def _recompute(
    shape: str, match: Any, formula: str, cached: dict[str, object]
) -> tuple[Decimal | None, list[str]]:
    """Recomputa a fórmula em `Decimal` sobre os valores em cache dos operandos.

    Devolve `(valor, [])` quando a conta fecha, `(None, refs)` quando falta valor em cache
    de um operando obrigatório, e levanta `ValuationValidationError` quando a expressão
    casou com a forma mas não é a gramática (intervalo invertido, lista com literal).
    A aritmética é a mesma do avaliador da auditoria: dinheiro trunca, quantidade
    arredonda, e dedução ausente vale zero.
    """
    if shape == "trunc_product":
        values, missing = _required_values(cached, [match.group(1), match.group(2)])
        if missing:
            return None, missing
        return money_trunc(values[0] * values[1]), []
    if shape == "round_product_minus":
        product = _product(_range_values(cached, match.group(1), match.group(2)))
        deduction = _numeric(cached.get(match.group(3)))
        return quantity_round(product) - (_ZERO if deduction is None else deduction), []
    if shape == "round_product":
        return quantity_round(_product(_range_values(cached, match.group(1), match.group(2)))), []
    if shape == "sum_range":
        return sum(_range_values(cached, match.group(1), match.group(2)), _ZERO), []
    if shape == "sum_list":
        tokens: list[str] = match.group(1).split(",")
        if len(tokens) < 2 or any(CELL_REF_RE.fullmatch(token) is None for token in tokens):
            raise _not_grammar(formula)
        listed = [_numeric(cached.get(token)) for token in tokens]
        return sum((number for number in listed if number is not None), _ZERO), []
    values, missing = _required_values(cached, [match.group(1), match.group(2)])
    if missing:
        return None, missing
    return values[0] - values[1], []


def _category(formula: str) -> ParityCategory:
    """Por que a fórmula está fora da gramática, na ordem em que o risco importa."""
    if "[" in formula:
        return "external_reference"
    if "!" in formula:
        return "cross_sheet"
    if formula.startswith("=VLOOKUP"):
        return "vlookup"
    if formula.startswith("=IF"):
        return "conditional"
    return "other"


def _read_cells(worksheet: Any) -> dict[str, object]:
    """Todas as células não vazias da aba, por referência A1, na ordem da planilha."""
    values: dict[str, object] = {}
    for row in worksheet.iter_rows():
        for cell in row:
            value = cell.value
            if value is None:
                continue
            values[str(cell.coordinate)] = value
    return values


def _scan_sheet(formula_sheet: Any, value_sheet: Any) -> _SheetCells:
    formulas = {
        ref: value
        for ref, value in _read_cells(formula_sheet).items()
        if isinstance(value, str) and value.startswith("=")
    }
    return _SheetCells(formulas=formulas, cached=_read_cells(value_sheet))


def _check_sheet(name: str, cells: _SheetCells) -> tuple[ParitySheetReport, list[ParityFinding]]:
    findings: list[ParityFinding] = []
    verified = divergent = foreign = cache_missing = 0
    for ref, formula in cells.formulas.items():
        matched = _match_grammar(formula)
        recomputed: Decimal | None = None
        missing: list[str] = []
        if matched is not None:
            try:
                recomputed, missing = _recompute(matched[0], matched[1], formula, cells.cached)
            except ValuationValidationError:
                matched = None
        if matched is None:
            foreign += 1
            findings.append(
                ParityFinding(
                    code="PARITY_FORMULA_FOREIGN",
                    sheet=name,
                    ref=ref,
                    formula=formula,
                    category=_category(formula),
                )
            )
            continue
        own = _numeric(cells.cached.get(ref))
        if own is None and ref not in missing:
            missing = [*missing, ref]
        if missing or recomputed is None or own is None:
            cache_missing += 1
            findings.append(
                ParityFinding(
                    code="PARITY_CACHE_MISSING",
                    sheet=name,
                    ref=ref,
                    formula=formula,
                    missing_refs=missing,
                )
            )
            continue
        if own != recomputed:
            divergent += 1
            findings.append(
                ParityFinding(
                    code="PARITY_VALUE_DIVERGENT",
                    sheet=name,
                    ref=ref,
                    formula=formula,
                    cached=str(own),
                    recomputed=str(recomputed),
                )
            )
            continue
        verified += 1
    report = ParitySheetReport(
        name=name,
        formulas=len(cells.formulas),
        verified=verified,
        divergent=divergent,
        foreign=foreign,
        cache_missing=cache_missing,
    )
    return report, findings


def _open(workbook_path: Path, *, data_only: bool) -> Any:
    try:
        return load_workbook(workbook_path, read_only=True, data_only=data_only)
    except (OSError, ValueError, KeyError, BadZipFile, InvalidFileException) as error:
        raise ValuationValidationError(
            "PARITY_WORKBOOK_UNREADABLE",
            "arquivo não pôde ser aberto como planilha",
            {"path": str(workbook_path), "reason": type(error).__name__},
        ) from error


def check_parity(workbook_path: Path, template: WorkbookTemplate) -> ParityReport:
    """Abre a pasta duas vezes — fórmulas e caches — e confere uma contra a outra."""
    digest = file_sha256(workbook_path)
    formula_workbook = _open(workbook_path, data_only=False)
    try:
        value_workbook = _open(workbook_path, data_only=True)
        try:
            sheets: list[ParitySheetReport] = []
            findings: list[ParityFinding] = []
            for worksheet in formula_workbook.worksheets:
                name = str(worksheet.title)
                if name == template.catalog.sheet_name:
                    continue
                cells = _scan_sheet(worksheet, value_workbook[name])
                sheet_report, sheet_findings = _check_sheet(name, cells)
                sheets.append(sheet_report)
                findings.extend(sheet_findings)
        finally:
            value_workbook.close()
    finally:
        formula_workbook.close()
    return ParityReport(workbook_sha256=digest, sheets=sheets, findings=findings)
