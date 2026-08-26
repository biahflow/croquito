"""Extrai a tabela de derivação de transporte da memória de cálculo da prefeitura.

O capítulo de transporte, carga e bota-fora de um orçamento não é medido na prancha: ele é
**derivado** do resto do orçamento. Na memória do Campo do Toca isso é uma tabela em que
cada linha diz `quantidade(outro serviço) x massa específica x espessura x distância`, com
a origem resolvida por `=IFERROR(VLOOKUP(<item>; B:Q; 16; FALSO); 0)`.

A tabela é a mesma para toda obra do contrato — densidade de concreto e espessura de camada
são propriedade do material, não da praça —, e hoje é redigitada a cada orçamento. Este
script a transforma em seed versionado (`sco-haulage-v1.json`), que é o que permite gerar o
capítulo inteiro em vez de copiá-lo.

Duas decisões de leitura, ambas forçadas pelo arquivo real:

- **A chave é o código SCO, nunca o número do item.** A memória referencia a origem por
  número (`04.11`), que é posicional: no arquivo real, 330 dos 433 itens têm código
  diferente entre a `PLANILHA ORÇAMENTÁRIA` e a `PLANILHA PADRÃO ORDENADA` para o mesmo
  número. A resolução usa a `PLANILHA ORÇAMENTÁRIA`, que é a numeração de onde a própria
  memória tira o código de cada bloco.
- **Os fatores são lidos pelo cabeçalho, não por posição.** A memória declara a fórmula de
  cada bloco numa linha de cabeçalho (`ÁREA x P.ESP x ESP x DAM`, `VOLUME x EMP`,
  `ÁREA x P.ESP`), e a forma muda com o destino: o transporte horizontal tem distância, a
  carga e descarga não, e a retirada de entulho usa empolamento. Ler por posição assumiria
  uma forma única que não existe.

O que não resolve é **declarado**, não descartado em silêncio: linhas que apontam itens
fora do contrato desta obra (resíduo de um template mais amplo) saem em `unmapped_labels`.

Uso:

    uv run python scripts/extract_haulage_table.py \\
        --input "<orçamento>.xlsx" \\
        --source-label "Memória do Campo do Toca — contrato SMH/Rio (Out/2023)" \\
        --output packages/valuation/src/croquito_valuation/data/sco-haulage-v1.json
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections.abc import Sequence
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any, Final

from openpyxl import load_workbook

MEMORY_SHEET: Final = "PRAÇA CAMPO DO TOCA"
BUDGET_SHEET: Final = "PLANILHA ORÇAMENTÁRIA"
BUDGET_FIRST_ROW: Final = 10
ITEM_RE: Final = re.compile(r"^\d{1,2}\.\d{1,3}$")
CROSS_ITEM_RE: Final = re.compile(r"VLOOKUP\(.*B:Q", re.IGNORECASE)

ITEM_COLUMN: Final = 2
CODE_COLUMN: Final = 3
LABEL_COLUMN: Final = 4
ORIGIN_QUANTITY_COLUMN: Final = 6
FACTOR_COLUMNS: Final = (8, 10, 12)

#: Unidade de cada fator, quando ela é inequívoca. `P.ESP` não entra: ele é t/m³ quando a
#: origem é volume e t/m² quando é área, e declarar uma só seria inventar. `EMP` é
#: adimensional.
FACTOR_UNITS: Final = {"ESP": "m", "DAM": "dam"}

#: Coluna onde a memória imprime a quantidade do item de origem — é o multiplicando, não um
#: fator.
ORIGIN_HEADER_COLUMN: Final = 6

#: Rótulos de cabeçalho que nomeiam o RESULTADO, não um fator. A largura do bloco varia, e
#: em `VOLUME x EMP = TOTAL` a coluna do total cai dentro da faixa dos fatores; sem esta
#: lista o total entraria multiplicando a si mesmo.
RESULT_HEADERS: Final = frozenset({"TOTAL"})


@dataclass(frozen=True, slots=True)
class RawDerivation:
    """Uma linha da tabela, ainda com a origem em número de item."""

    target_item: str
    origin_item: str
    label: str
    factors: tuple[tuple[str, Decimal], ...]
    row: int


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def read_item_codes(workbook_path: Path) -> dict[str, str]:
    """Mapa número do item → código SCO, pela planilha orçamentária."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook[BUDGET_SHEET]
        codes: dict[str, str] = {}
        for row in sheet.iter_rows(min_row=BUDGET_FIRST_ROW, max_col=2, values_only=True):
            item, code = _text(row[0]), _text(row[1])
            if ITEM_RE.match(item) and code:
                codes[item] = code
        return codes
    finally:
        workbook.close()


def _header_names(sheet: Any, row_number: int) -> dict[int, str]:
    """Mapa coluna → nome do fator, se esta linha for um cabeçalho de fórmula.

    A memória declara a fórmula do bloco numa linha própria (`ÁREA x P.ESP x ESP x DAM`,
    `VOLUME x EMP`), e o cabeçalho vale até o próximo — o bloco do transporte horizontal
    troca de forma no meio, ao passar dos itens medidos em área para os medidos em volume.
    Por isso o cabeçalho é rastreado enquanto se percorre, e não procurado para trás.
    """
    return {
        column: text
        for column in FACTOR_COLUMNS
        if (text := _text(sheet.cell(row_number, column).value))
        and not _is_number(sheet.cell(row_number, column).value)
        and text.upper() not in RESULT_HEADERS
    }


def _is_number(value: object) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def read_raw_derivations(workbook_path: Path) -> list[RawDerivation]:
    """Lê as linhas da memória que derivam a quantidade de OUTRO item."""
    formulas = load_workbook(workbook_path, data_only=False)
    values = load_workbook(workbook_path, data_only=True)
    try:
        formula_sheet = formulas[MEMORY_SHEET]
        value_sheet = values[MEMORY_SHEET]
        derivations: list[RawDerivation] = []
        current: str | None = None
        names: dict[int, str] = {}
        for row_number in range(1, formula_sheet.max_row + 1):
            item = _text(formula_sheet.cell(row_number, ITEM_COLUMN).value)
            code = _text(formula_sheet.cell(row_number, CODE_COLUMN).value)
            if ITEM_RE.match(item) and code:
                current = item
                names = {}
                continue
            if current is None:
                continue
            if header := _header_names(value_sheet, row_number):
                names = header
                continue
            origin_formula = formula_sheet.cell(row_number, ORIGIN_QUANTITY_COLUMN).value
            if not isinstance(origin_formula, str) or not CROSS_ITEM_RE.search(origin_formula):
                continue
            factors = tuple(
                (names[column], Decimal(str(raw)))
                for column in FACTOR_COLUMNS
                if column in names and _is_number(raw := value_sheet.cell(row_number, column).value)
            )
            if not factors:
                continue
            derivations.append(
                RawDerivation(
                    target_item=current,
                    origin_item=_text(value_sheet.cell(row_number, CODE_COLUMN).value),
                    label=_text(value_sheet.cell(row_number, LABEL_COLUMN).value),
                    factors=factors,
                    row=row_number,
                )
            )
        return derivations
    finally:
        formulas.close()
        values.close()


def build_seed(
    derivations: Sequence[RawDerivation], codes: dict[str, str], *, source_label: str
) -> tuple[dict[str, Any], list[str]]:
    """Monta o seed chaveado por código, declarando o que não resolveu."""
    entries: list[dict[str, Any]] = []
    unmapped: list[str] = []
    seen: set[tuple[str, str]] = set()
    for derivation in derivations:
        target_code = codes.get(derivation.target_item)
        origin_code = codes.get(derivation.origin_item)
        if target_code is None or origin_code is None:
            unmapped.append(derivation.label)
            continue
        pair = (target_code, origin_code)
        if pair in seen:
            raise SystemExit(
                f"HAULAGE_DUPLICATE_PAIR: {origin_code} deriva {target_code} mais de uma "
                f"vez (linha {derivation.row}); a fonte precisa ser curada antes"
            )
        seen.add(pair)
        entries.append(
            {
                "target_code": target_code,
                "origin_code": origin_code,
                "label": derivation.label,
                "factors": [
                    {"name": name, "value": str(value), "unit": FACTOR_UNITS.get(name)}
                    for name, value in derivation.factors
                ],
            }
        )
    seed = {
        "version": "sco-haulage-v1",
        "source_label": source_label,
        "derivations": entries,
        "unmapped_labels": sorted(set(unmapped)),
    }
    return seed, unmapped


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, required=True, help="planilha do orçamento")
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--source-label", required=True, help="de onde a tabela veio")
    args = parser.parse_args(argv)

    codes = read_item_codes(args.input)
    derivations = read_raw_derivations(args.input)
    seed, unmapped = build_seed(derivations, codes, source_label=args.source_label)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(seed, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(
        json.dumps(
            {
                "input": str(args.input),
                "output": str(args.output),
                "rows_read": len(derivations),
                "derivations": len(seed["derivations"]),
                "unmapped": len(unmapped),
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
