"""Importa o catálogo de preço CONTRATADO da planilha de orçamento da prefeitura.

O catálogo publicado (aba `FGV06`) traz o preço de referência do SCO. Não é ele que a obra
aplica: a planilha do contrato traz, na aba `PLANILHA GERAL`, o custo unitário efetivamente
contratado — o preço do SCO com o desconto da licitação. Medido no contrato do Campo do
Toca, a razão é constante em 0,99845 sobre os 427 códigos contratados, e é dessa coluna que
a `PLANILHA ORÇAMENTÁRIA` puxa o preço de cada linha.

Este script costura as duas abas porque nenhuma delas basta sozinha:

- **preço** vem da `PLANILHA GERAL`, que é o que a obra aplica;
- **descrição, unidade e hierarquia** vêm da `FGV06`, porque a `PLANILHA GERAL` grafa o
  texto sem acento (376 das 427 descrições do arquivo real) e o braço semântico do matcher
  preserva acento de propósito (`normalize_query_text`, `sco_matching.py`) — importar o
  texto achatado degradaria o retrieval sem que nada acusasse.

O `source_sha256` **não** é o digest do arquivo: as duas importações saem da mesma planilha,
e repetir o digest tornaria os dois catálogos indistinguíveis na cascata e na citação de
fonte de cada confirmação. Aqui ele é o digest do conteúdo canônico das entradas, que é
determinístico e distingue os dois.

Uso:

    uv run python scripts/import_contract_catalog.py \\
        --input "<orçamento>.xlsx" \\
        --catalog output/toca-2023-10/catalog.json \\
        --reference-month 2023-10 \\
        --output output/toca-2023-10/contract-catalog.json
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections.abc import Sequence
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Final

from openpyxl import load_workbook

from croquito_valuation.models import PriceCatalog, PriceCatalogEntry, PriceOrigin

SHEET_NAME: Final = "PLANILHA GERAL"
FIRST_DATA_ROW: Final = 5
CODE_COLUMN: Final = 5
UNIT_COLUMN: Final = 7
PRICE_COLUMN: Final = 9
SCO_CODE_RE: Final = re.compile(r"^[A-Z]{2}\d{8}\(")


@dataclass(frozen=True, slots=True)
class ContractPrice:
    """Custo unitário contratado de um código, e a unidade como o contrato a grafa."""

    code: str
    unit: str
    unit_price: Decimal
    row: int


def read_contract_prices(workbook_path: Path) -> list[ContractPrice]:
    """Lê o custo contratado de cada código, na primeira ocorrência.

    O mesmo código aparece em vários grupos do contrato (o arquivo real repete 392 vezes)
    sempre com o MESMO custo unitário — o que muda entre grupos é a quantidade contratada.
    A planilha resolve isso com `VLOOKUP`, que devolve a primeira ocorrência; fazemos o
    mesmo, mas recusando se duas ocorrências divergirem, porque aí a escolha deixaria de
    ser indiferente.
    """
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook[SHEET_NAME]
        seen: dict[str, ContractPrice] = {}
        conflicts: list[str] = []
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=FIRST_DATA_ROW, values_only=True), start=FIRST_DATA_ROW
        ):
            code = str(row[CODE_COLUMN - 1] or "").strip()
            if not SCO_CODE_RE.match(code):
                continue
            raw_price = row[PRICE_COLUMN - 1]
            if not isinstance(raw_price, (int, float, Decimal)):
                continue
            price = ContractPrice(
                code=code,
                unit=str(row[UNIT_COLUMN - 1] or "").strip(),
                unit_price=_money(Decimal(str(raw_price)), code=code, row=row_number),
                row=row_number,
            )
            previous = seen.get(code)
            if previous is None:
                seen[code] = price
            elif previous.unit_price != price.unit_price:
                conflicts.append(
                    f"{code}: linha {previous.row} custa {previous.unit_price}, "
                    f"linha {price.row} custa {price.unit_price}"
                )
        if conflicts:
            raise SystemExit(
                "CONTRACT_PRICE_DIVERGENT: o mesmo código tem custos diferentes no "
                "contrato; a primeira ocorrência deixaria de ser indiferente:\n  "
                + "\n  ".join(sorted(conflicts))
            )
        return sorted(seen.values(), key=lambda price: price.code)
    finally:
        workbook.close()


def _money(value: Decimal, *, code: str, row: int) -> Decimal:
    """Fixa a escala em centavos, recusando o que não couber.

    A célula do Excel guarda `0.2`, não `"0,20"`: sem isto o mesmo preço entraria com
    escala diferente da do catálogo publicado, que é lido de texto. Nenhum custo do
    arquivo real tem mais de duas casas, e um que tivesse seria dado novo — arredondar
    em silêncio decidiria centavo por conta própria, então recusa.
    """
    exponent = value.as_tuple().exponent
    if not isinstance(exponent, int) or -exponent > 2:
        raise SystemExit(
            f"CONTRACT_PRICE_SUBCENT: {code} (linha {row}) custa {value}, que não cabe em "
            "centavos; arredondar aqui decidiria centavo sem mandato"
        )
    return value.quantize(Decimal("0.01"))


def content_sha256(entries: Sequence[PriceCatalogEntry]) -> str:
    """Digest do conteúdo canônico; distingue este catálogo do publicado da mesma planilha."""
    payload = json.dumps(
        [entry.model_dump(mode="json") for entry in entries],
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=False,
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def build_contract_catalog(
    prices: Sequence[ContractPrice],
    published: PriceCatalog,
    *,
    source_label: str,
    reference_month: str,
) -> tuple[PriceCatalog, list[str], list[str]]:
    """Casa preço contratado com o texto do catálogo publicado.

    Devolve o catálogo, os códigos sem correspondência no publicado e as divergências de
    unidade entre as duas abas — nenhuma das duas listas interrompe: são observações para
    quem confere, e a unidade que vale é a do catálogo publicado, que é a acentuada.
    """
    published_by_code = {entry.code: entry for entry in published.entries}
    entries: list[PriceCatalogEntry] = []
    unknown: list[str] = []
    unit_divergences: list[str] = []
    for price in prices:
        reference = published_by_code.get(price.code)
        if reference is None:
            unknown.append(price.code)
            continue
        if _fold(price.unit) != _fold(reference.unit):
            unit_divergences.append(
                f"{price.code}: contrato diz {price.unit!r}, catálogo diz {reference.unit!r}"
            )
        entries.append(reference.model_copy(update={"unit_price": price.unit_price}))
    catalog = PriceCatalog(
        source_label=source_label,
        reference_month=reference_month,
        source_sha256=content_sha256(entries),
        entries=entries,
        origin=PriceOrigin.SCO,
    )
    return catalog, unknown, unit_divergences


def _fold(text: str) -> str:
    """Compara unidade sem acento e sem caixa: a `PLANILHA GERAL` grafa `un.mes`."""
    import unicodedata

    decomposed = unicodedata.normalize("NFKD", text.casefold())
    return "".join(char for char in decomposed if not unicodedata.combining(char))


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, required=True, help="planilha do orçamento")
    parser.add_argument(
        "--catalog",
        type=Path,
        required=True,
        help="catalog.json publicado da mesma planilha (saída de `import-catalog`)",
    )
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--reference-month", required=True, help="data-base, no formato AAAA-MM")
    parser.add_argument(
        "--source-label",
        default="Custo contratado (PLANILHA GERAL)",
        help="rótulo do catálogo contratado",
    )
    args = parser.parse_args(argv)

    published = PriceCatalog.model_validate_json(args.catalog.read_text(encoding="utf-8"))
    prices = read_contract_prices(args.input)
    catalog, unknown, unit_divergences = build_contract_catalog(
        prices,
        published,
        source_label=args.source_label,
        reference_month=args.reference_month,
    )
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(catalog.model_dump_json(indent=2), encoding="utf-8")

    report = {
        "input": str(args.input),
        "published_catalog": str(args.catalog),
        "output": str(args.output),
        "contract_codes": len(prices),
        "catalog_entries": len(catalog.entries),
        "source_sha256": catalog.source_sha256,
        "reference_month": catalog.reference_month,
        "codes_without_published_entry": unknown,
        "unit_divergences": unit_divergences,
    }
    print(json.dumps(report, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    sys.exit(main())
