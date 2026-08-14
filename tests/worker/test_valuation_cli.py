"""Wiring da CLI de medição: `import-workbook`, `export-valuation` e `demo`.

Cobre os dois lados de cada comando — o caminho feliz e a recusa fechada —, e o que
importa nas recusas não é a mensagem, é o que **não** aparece no diretório de saída:
importação que falha não deixa JSON pela metade e medição sem aprovação não deixa
planilha nenhuma.

Tudo aqui é sintético: a fixture do MAPÃO anterior é a mesma da demonstração.
"""

from __future__ import annotations

import json
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from croquitodxf_valuation.catalog import file_sha256
from croquitodxf_valuation.contract import ContractWorkbook
from croquitodxf_valuation.models import PriceCatalog, Valuation
from croquitodxf_valuation.template import WorkbookTemplate, default_template
from croquitodxf_worker.valuation.cli import (
    AUDIT_FILENAME,
    CATALOG_FILENAME,
    CATALOG_IMPORT_REPORT_FILENAME,
    CONTRACT_FILENAME,
    IMPORT_DIAGNOSIS_FILENAME,
    IMPORT_REPORT_FILENAME,
    WORKBOOK_FILENAME,
    main,
)
from croquitodxf_worker.valuation.synthetic import (
    SYNTHETIC_PREVIOUS_PERIOD_COUNT,
    build_synthetic_approval,
    build_synthetic_multi_valuation,
    build_synthetic_previous_mapao,
)
from tests.valuation.builders import (
    columnar_catalog_expectations,
    columnar_catalog_template,
    tamper_cell,
    write_columnar_catalog_workbook,
)

_MEASURED_TWICE_CODE = "AD04050050(/)"
_VALUATION_FILENAME = "valuation.json"


def _stdout(capsys: pytest.CaptureFixture[str]) -> dict[str, object]:
    """Último JSON impresso: cada comando imprime uma linha só, e o preparo imprime antes."""
    lines = [line for line in capsys.readouterr().out.splitlines() if line.strip()]
    return dict(json.loads(lines[-1]))


def _accumulated_quantity_ref(path: Path, template: WorkbookTemplate, code: str) -> str:
    """Célula de acumulado do código na PLANILHA GERAL, com a coluna derivada do template.

    O par ACUMULADO vem depois do último par de medição, então a coluna sai do número de
    períodos declarado — nenhuma letra é constante no teste, como no leitor.
    """
    layout = template.general
    column = column_index_from_string(layout.first_period_column) + 2 * (
        SYNTHETIC_PREVIOUS_PERIOD_COUNT
    )
    workbook = load_workbook(path)
    try:
        worksheet = workbook[layout.sheet_name]
        for row in range(layout.data_first_row, worksheet.max_row + 1):
            if worksheet[f"{layout.code_column}{row}"].value == code:
                return f"{get_column_letter(column)}{row}"
    finally:
        workbook.close()
    raise AssertionError(f"código ausente da planilha geral sintética: {code}")


def _float_noise(path: Path, sheet_name: str, ref: str) -> float:
    """O mesmo número da célula com o ruído de cache que a planilha real carrega.

    O desvio é sub-centavo de propósito: é exatamente o que a leitura absorve e registra
    nas notas, sem mexer no valor que o modelo vai conferir.
    """
    workbook = load_workbook(path)
    try:
        value = workbook[sheet_name][ref].value
    finally:
        workbook.close()
    assert isinstance(value, int | float | Decimal)
    return float(value) - 1e-9


def _import(previous_path: Path, output_dir: Path) -> int:
    return main(
        [
            "import-workbook",
            "--input",
            str(previous_path),
            "--output",
            str(output_dir),
        ]
    )


def test_import_workbook_reads_catalog_and_contract_from_the_same_file(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    previous_path = build_synthetic_previous_mapao(tmp_path / "previous-mapao.xlsx")
    before = file_sha256(previous_path)
    output_dir = tmp_path / "import"

    exit_code = _import(previous_path, output_dir)

    assert exit_code == 0
    payload = _stdout(capsys)
    assert payload["catalog_entries"] == 32
    assert payload["contract_lines"] == 10
    assert payload["period_count"] == SYNTHETIC_PREVIOUS_PERIOD_COUNT
    assert payload["amendments"] == 1
    assert payload["unknown_units"] == []
    for filename in (CATALOG_FILENAME, CONTRACT_FILENAME, IMPORT_REPORT_FILENAME):
        assert (output_dir / filename).is_file()
    # O original é insumo: a importação lê e não reescreve o arquivo do cliente.
    assert file_sha256(previous_path) == before

    contract = ContractWorkbook.model_validate_json(
        (output_dir / CONTRACT_FILENAME).read_text(encoding="utf-8")
    )
    catalog = PriceCatalog.model_validate_json(
        (output_dir / CATALOG_FILENAME).read_text(encoding="utf-8")
    )
    assert contract.source_sha256 == before
    assert (
        catalog.entry_for(_MEASURED_TWICE_CODE).unit_price
        == contract.line_for_code(_MEASURED_TWICE_CODE).unit_price
    )


def test_import_report_carries_the_reading_notes(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    """O ruído de ponto flutuante do cache é absorvido e aparece no relatório, não some."""
    template = default_template()
    previous_path = build_synthetic_previous_mapao(tmp_path / "previous-mapao.xlsx")
    sheet = template.general.sheet_name
    ref = _accumulated_quantity_ref(previous_path, template, _MEASURED_TWICE_CODE)
    tamper_cell(previous_path, sheet, ref, _float_noise(previous_path, sheet, ref))
    output_dir = tmp_path / "import"

    assert _import(previous_path, output_dir) == 0
    _stdout(capsys)

    report = json.loads((output_dir / IMPORT_REPORT_FILENAME).read_text(encoding="utf-8"))
    notes = report["notes"]
    assert notes["period_numbers"] == [1, 2]
    assert notes["period_gaps"] == []
    assert notes["separator_rows"] == {"rows": [], "count": 0}
    assert notes["normalized"]["count"] == 1
    assert notes["normalized"]["truncated"] is False
    assert [(cell["sheet"], cell["ref"]) for cell in notes["normalized"]["cells"]] == [(sheet, ref)]


def test_import_workbook_refuses_a_tampered_accumulated_with_a_dossier_and_no_artifacts(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    """Recusa semântica publica o dossiê e mais nada: nem catálogo, nem consolidado."""
    template = default_template()
    previous_path = build_synthetic_previous_mapao(tmp_path / "previous-mapao.xlsx")
    ref = _accumulated_quantity_ref(previous_path, template, _MEASURED_TWICE_CODE)
    tamper_cell(previous_path, template.general.sheet_name, ref, 99)
    output_dir = tmp_path / "import"

    exit_code = _import(previous_path, output_dir)

    assert exit_code == 2
    payload = _stdout(capsys)
    assert payload["refused"] == "CONTRACT_SEMANTICS_DIVERGENT"
    summary = payload["summary"]
    assert isinstance(summary, dict)
    assert summary["CONTRACT_ACCUMULATED_MISMATCH"] == 1
    diagnosis_path = output_dir / IMPORT_DIAGNOSIS_FILENAME
    assert payload["diagnosis"] == str(diagnosis_path)
    dossier = json.loads(diagnosis_path.read_text(encoding="utf-8"))
    assert dossier["refused"] == "CONTRACT_SEMANTICS_DIVERGENT"
    assert dossier["source_sha256"] == file_sha256(previous_path)
    assert dossier["template_label"] == template.label
    assert dossier["summary"] == summary
    assert dossier["finding_count"] == payload["finding_count"] == len(dossier["findings"])
    accumulated = [
        finding
        for finding in dossier["findings"]
        if finding["finding_code"] == "CONTRACT_ACCUMULATED_MISMATCH"
    ]
    assert [finding["ref"] for finding in accumulated] == [ref]
    assert accumulated[0]["code"] == _MEASURED_TWICE_CODE
    assert not (output_dir / CATALOG_FILENAME).exists()
    assert not (output_dir / CONTRACT_FILENAME).exists()
    assert not (output_dir / IMPORT_REPORT_FILENAME).exists()


def test_import_workbook_refuses_a_broken_layout_without_a_dossier(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    """Sem entender a planilha não há o que recomputar: recusa de layout não gera dossiê."""
    template = default_template()
    previous_path = build_synthetic_previous_mapao(tmp_path / "previous-mapao.xlsx")
    layout = template.general
    tamper_cell(
        previous_path,
        layout.sheet_name,
        f"{layout.first_period_column}{layout.header_row}",
        "SEGUNDA MEDICAO",
    )
    output_dir = tmp_path / "import"

    exit_code = _import(previous_path, output_dir)

    assert exit_code == 2
    assert _stdout(capsys)["refused"] == "PERIOD_HEADER_UNPARSEABLE"
    assert not (output_dir / IMPORT_DIAGNOSIS_FILENAME).exists()
    assert not output_dir.exists() or list(output_dir.iterdir()) == []


def _import_catalog(input_path: Path, output_dir: Path, *, template: Path | None = None) -> int:
    args = ["import-catalog", "--input", str(input_path), "--output", str(output_dir)]
    if template is not None:
        args.extend(["--template", str(template)])
    return main(args)


def test_import_catalog_publishes_only_the_catalog(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    """`import-catalog` nunca produz contract-workbook.json nem import-report.json: só o
    catálogo e o relatório próprio, com o consolidado explicitamente não importado."""
    previous_path = build_synthetic_previous_mapao(tmp_path / "previous-mapao.xlsx")
    before = file_sha256(previous_path)
    output_dir = tmp_path / "import-catalog"

    exit_code = _import_catalog(previous_path, output_dir)

    assert exit_code == 0
    payload = _stdout(capsys)
    assert payload["consolidado"] == "not_imported"
    note = payload["note"]
    assert isinstance(note, str)
    assert "contract-workbook.json" in note
    assert "acumulado" in note.lower()
    assert payload["catalog_entries"] == 32
    assert payload["template_label"] == default_template().label
    assert (output_dir / CATALOG_FILENAME).is_file()
    assert (output_dir / CATALOG_IMPORT_REPORT_FILENAME).is_file()
    # A recusa semântica do consolidado nunca é vista por este caminho: nem o consolidado,
    # nem o relatório de import-workbook, nem o dossiê de recusa semântica existem aqui.
    assert not (output_dir / CONTRACT_FILENAME).exists()
    assert not (output_dir / IMPORT_REPORT_FILENAME).exists()
    assert not (output_dir / IMPORT_DIAGNOSIS_FILENAME).exists()
    # O original é insumo: import-catalog também só lê, nunca reescreve.
    assert file_sha256(previous_path) == before

    catalog = PriceCatalog.model_validate_json(
        (output_dir / CATALOG_FILENAME).read_text(encoding="utf-8")
    )
    assert catalog.source_sha256 == before
    assert payload["family_count"] == len({entry.family_code for entry in catalog.entries})
    assert payload["subgroup_count"] == len({entry.subgroup_code for entry in catalog.entries})

    report = json.loads((output_dir / CATALOG_IMPORT_REPORT_FILENAME).read_text(encoding="utf-8"))
    for key, value in report.items():
        assert payload[key] == value


def test_import_catalog_refuses_a_broken_sheet_without_artifacts(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    """Recusa de layout do catálogo é fechada e nunca vê o portão semântico do consolidado:
    o arquivo tem a PLANILHA GERAL e a RE-RA em ordem, só a aba de catálogo é que some."""
    previous_path = build_synthetic_previous_mapao(tmp_path / "previous-mapao.xlsx")
    payload = default_template().model_dump()
    payload["catalog"]["sheet_name"] = "CATALOGO_AUSENTE"
    broken_template = WorkbookTemplate.model_validate(payload)
    template_path = tmp_path / "template-broken.json"
    template_path.write_text(broken_template.model_dump_json(indent=2), encoding="utf-8")
    output_dir = tmp_path / "import-catalog"

    exit_code = _import_catalog(previous_path, output_dir, template=template_path)

    assert exit_code == 2
    stdout = _stdout(capsys)
    assert stdout["refused"] == "CATALOG_SHEET_MISSING"
    assert not (output_dir / CATALOG_FILENAME).exists()
    assert not (output_dir / CATALOG_IMPORT_REPORT_FILENAME).exists()
    assert not (output_dir / IMPORT_DIAGNOSIS_FILENAME).exists()
    assert not output_dir.exists() or list(output_dir.iterdir()) == []


def test_import_catalog_applies_a_custom_template(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    """`--template` é aplicado de verdade: o catálogo colunar publicado só lê certo com o
    layout que o template declara, nunca com o `default_template()`."""
    catalog_path = write_columnar_catalog_workbook(tmp_path / "catalogo-publicado.xlsx")
    template = columnar_catalog_template()
    template_path = tmp_path / "template-colunar.json"
    template_path.write_text(template.model_dump_json(indent=2), encoding="utf-8")
    output_dir = tmp_path / "import-catalog"

    exit_code = _import_catalog(catalog_path, output_dir, template=template_path)

    assert exit_code == 0
    payload = _stdout(capsys)
    expectations = columnar_catalog_expectations()
    assert payload["catalog_entries"] == len(expectations)
    assert payload["family_count"] == len({entry.family_code for entry in expectations})
    assert payload["subgroup_count"] == len({entry.subgroup_code for entry in expectations})
    assert payload["template_label"] == template.label

    catalog = PriceCatalog.model_validate_json(
        (output_dir / CATALOG_FILENAME).read_text(encoding="utf-8")
    )
    assert [entry.code for entry in catalog.entries] == [expected.code for expected in expectations]


def _export_inputs(
    tmp_path: Path, *, approved: bool, period_shift: int = 0
) -> tuple[Path, Path, Path]:
    """Importa o consolidado sintético e grava a medição do período seguinte em JSON.

    `period_shift` desloca o número do período antes da aprovação, para exercitar o
    portão sem invalidar o digest: a medição fora da sequência continua sendo uma medição
    legitimamente aprovada.
    """
    previous_path = build_synthetic_previous_mapao(tmp_path / "previous-mapao.xlsx")
    imported = tmp_path / "import"
    assert _import(previous_path, imported) == 0
    contract_path = imported / CONTRACT_FILENAME
    catalog_path = imported / CATALOG_FILENAME
    contract = ContractWorkbook.model_validate_json(contract_path.read_text(encoding="utf-8"))
    catalog = PriceCatalog.model_validate_json(catalog_path.read_text(encoding="utf-8"))
    valuation = build_synthetic_multi_valuation(contract, catalog)
    if period_shift:
        payload = valuation.model_dump()
        payload["period_number"] = valuation.period_number + period_shift
        valuation = Valuation.model_validate(payload)
    if approved:
        valuation = build_synthetic_approval(valuation)
    valuation_path = tmp_path / _VALUATION_FILENAME
    valuation_path.write_text(valuation.model_dump_json(indent=2), encoding="utf-8")
    return valuation_path, contract_path, catalog_path


def _export(paths: tuple[Path, Path, Path], output_dir: Path) -> int:
    valuation_path, contract_path, catalog_path = paths
    return main(
        [
            "export-valuation",
            "--valuation",
            str(valuation_path),
            "--contract",
            str(contract_path),
            "--catalog",
            str(catalog_path),
            "--output",
            str(output_dir),
        ]
    )


def test_export_valuation_without_approval_publishes_nothing(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    paths = _export_inputs(tmp_path, approved=False)
    output_dir = tmp_path / "export"

    exit_code = _export(paths, output_dir)

    assert exit_code == 2
    payload = _stdout(capsys)
    assert payload["refused"] == "VALUATION_EXPORT_BLOCKED"
    assert payload["errors"] == ["VALUATION_NOT_APPROVED"]
    assert not output_dir.exists() or list(output_dir.glob("*.xlsx")) == []


def test_export_valuation_publishes_only_after_the_audit_approves(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    paths = _export_inputs(tmp_path, approved=True)
    output_dir = tmp_path / "export"

    exit_code = _export(paths, output_dir)

    assert exit_code == 0
    payload = _stdout(capsys)
    assert payload["status"] == "ok"
    assert payload["general_sheet"] == "PLANILHA GERAL"
    assert payload["worksites"] == [
        "praca-sintetica-norte",
        "praca-sintetica-sul",
        "praca-sintetica-leste",
    ]
    workbook_path = output_dir / WORKBOOK_FILENAME
    assert workbook_path.is_file()
    assert file_sha256(workbook_path) == payload["workbook_sha256"]
    audit = json.loads((output_dir / AUDIT_FILENAME).read_text(encoding="utf-8"))
    assert audit["status"] == "ok"
    assert audit["findings"] == []
    # Nenhum arquivo pendente sobrevive à publicação.
    assert sorted(path.name for path in output_dir.iterdir()) == [AUDIT_FILENAME, WORKBOOK_FILENAME]


def test_export_valuation_refuses_a_valuation_that_is_not_the_next_period(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    paths = _export_inputs(tmp_path, approved=True, period_shift=1)
    output_dir = tmp_path / "export"

    exit_code = _export(paths, output_dir)

    assert exit_code == 2
    payload = _stdout(capsys)
    assert payload["refused"] == "VALUATION_EXPORT_BLOCKED"
    assert payload["errors"] == ["PERIOD_NOT_SEQUENTIAL:3:4"]
    assert not (output_dir / WORKBOOK_FILENAME).exists()


def test_demo_runs_the_whole_chain(tmp_path: Path, capsys: pytest.CaptureFixture[str]) -> None:
    """A demonstração vai da prancha à pasta auditada, e o bloco `takeoff` presta contas.

    O total cresceu com a quarta obra, a que nasce da legenda revisada: a linha ilegível
    entra com a quantidade que o revisor informou, a área de intervenção é rejeitada no
    takeoff e o gramado sai por falta de cotação — daí sete itens, um rejeitado e um
    excluído do boletim.
    """
    output_dir = tmp_path / "demo"

    exit_code = main(["demo", "--output", str(output_dir)])

    assert exit_code == 0
    payload = _stdout(capsys)
    assert payload["status"] == "ok"
    assert payload["period_number"] == SYNTHETIC_PREVIOUS_PERIOD_COUNT + 1
    assert payload["total_amount"] == "38859.46"
    takeoff = payload["takeoff"]
    assert isinstance(takeoff, dict)
    assert takeoff["items"] == 7
    assert takeoff["confirmed"] == 6
    assert takeoff["rejected"] == 1
    assert takeoff["pending"] == 0
    assert len(takeoff["excluded"]) == 1
    for key, filename in (
        ("suggestions", "code-suggestions.json"),
        ("assignment_decisions", "code-assignment-decisions.json"),
        ("assignments", "code-assignments.json"),
        ("calc_plan", "calc-plan.json"),
    ):
        assert takeoff[key] == str(output_dir / filename)
        assert (output_dir / filename).is_file()
    valuation = Valuation.model_validate_json(
        (output_dir / _VALUATION_FILENAME).read_text(encoding="utf-8")
    )
    assert valuation.approval is not None
    assert valuation.approval.valuation_digest == valuation.content_digest()
