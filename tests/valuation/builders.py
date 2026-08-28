"""Helpers de teste sobre a fixture sintética de medição."""

from __future__ import annotations

from collections.abc import Callable, Mapping, Sequence
from dataclasses import dataclass, field, replace
from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any
from uuid import NAMESPACE_URL, uuid5

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from croquito_valuation.canonical import canonicalize_workbook
from croquito_valuation.catalog import file_sha256, read_price_catalog
from croquito_valuation.contract import (
    Amendment,
    AmendmentLine,
    ContractLine,
    ContractWorkbook,
    PeriodProgress,
    contract_workbook_id_for,
)
from croquito_valuation.models import (
    BulletinLine,
    CalcBlock,
    CalcOperand,
    CalcRecipe,
    CalcSheet,
    PriceCatalog,
    ReviewerDecision,
    Valuation,
    ValuationApproval,
    WorksiteBulletin,
)
from croquito_valuation.rounding import money_trunc
from croquito_valuation.sco import is_sco_code, parse_sco_code
from croquito_valuation.template import (
    AmendmentLayout,
    CatalogLayout,
    EstimateTemplateColumns,
    EstimateTemplateLayout,
    EstimateTemplateRow,
    GeneralLayout,
    SheetColumn,
    WorkbookTemplate,
    default_template,
)
from croquito_valuation.workbook_writer import WriteReport, write_valuation_workbook
from croquito_worker.valuation.synthetic import (
    SYNTHETIC_CATALOG_LABEL,
    SYNTHETIC_REFERENCE_MONTH,
    build_synthetic_catalog_workbook,
    build_synthetic_valuation,
)

CatalogRow = tuple[str, str, str, str]


@dataclass(frozen=True, slots=True)
class ValuationFixture:
    """Catálogo sintético importado e a medição construída sobre ele."""

    catalog_path: Path
    template: WorkbookTemplate
    catalog: PriceCatalog
    valuation: Valuation


def build_fixture(tmp_path: Path) -> ValuationFixture:
    """Grava o mini-MAPÃO sintético, importa o catálogo e monta a medição."""
    template = default_template()
    catalog_path = build_synthetic_catalog_workbook(tmp_path / "synthetic-mapao.xlsx")
    catalog = read_price_catalog(
        catalog_path,
        template,
        source_label=SYNTHETIC_CATALOG_LABEL,
        reference_month=SYNTHETIC_REFERENCE_MONTH,
    )
    return ValuationFixture(
        catalog_path=catalog_path,
        template=template,
        catalog=catalog,
        valuation=build_synthetic_valuation(catalog),
    )


def write_fixture_workbook(fixture: ValuationFixture, output_path: Path) -> WriteReport:
    """Escreve a planilha da medição sintética."""
    return write_valuation_workbook(
        fixture.valuation, fixture.catalog, fixture.template, output_path
    )


PREVIOUS_MAPAO_SOURCE_LABEL = "MAPÃO SINTÉTICO 2ª MEDIÇÃO (fixture)"
PREVIOUS_MAPAO_CONTRACT_LABEL = "CONTRATO SINTETICO 001/2026"
PREVIOUS_MAPAO_PERIOD_NUMBERS: tuple[int, ...] = (1, 2)
PREVIOUS_MAPAO_COVER_SHEET = "CAPA"
_ZERO = Decimal("0.00")


@dataclass(frozen=True, slots=True)
class _MapaoItem:
    """Uma linha do MAPÃO sintético: fonte única da planilha e do gabarito.

    A planilha é escrita a partir destes valores e o `ContractWorkbook` esperado é
    construído dos mesmos valores, de modo que gabarito e arquivo não possam divergir.
    `period_quantities` com `None` é célula vazia — item que ainda não mediu no período,
    e casa posicionalmente com os números de medição que a planilha declara.
    """

    group_label: str
    item_number: str
    numeric_item_number: bool
    code: str
    description: str
    unit_label: str
    unit: str
    unit_price: Decimal
    contract_quantity: Decimal
    amended_quantity: Decimal
    period_quantities: tuple[Decimal | None, ...]
    reduced: Decimal | None = None
    added: Decimal | None = None
    new_item: Decimal | None = None

    @property
    def measured_quantities(self) -> tuple[Decimal, ...]:
        """Quantidade de cada período como a planilha a mostra: célula vazia vale 0,00."""
        return tuple(_ZERO if quantity is None else quantity for quantity in self.period_quantities)

    def periods(self, period_numbers: Sequence[int]) -> tuple[PeriodProgress, ...]:
        """Pares QUANTIDADE|VALOR nos números de medição que a planilha declara.

        O item novo da RE-RA tem os pares vazios e mesmo assim emite os `PeriodProgress`
        zerados, porque é isso que a planilha fisicamente mostra —
        `CONTRACT_PERIOD_NUMBERS_MISMATCH` aceita a mesma lista de números do consolidado.
        """
        return tuple(
            PeriodProgress(
                period_number=number,
                quantity=measured,
                amount=money_trunc(measured * self.unit_price),
            )
            for number, measured in zip(period_numbers, self.measured_quantities, strict=True)
        )

    @property
    def accumulated_quantity(self) -> Decimal:
        return sum(self.measured_quantities, _ZERO)

    @property
    def accumulated_amount(self) -> Decimal:
        return sum(
            (money_trunc(measured * self.unit_price) for measured in self.measured_quantities),
            _ZERO,
        )

    @property
    def balance_quantity(self) -> Decimal:
        return self.amended_quantity - self.accumulated_quantity

    def to_contract_line(self, period_numbers: Sequence[int]) -> ContractLine:
        """Linha esperada do consolidado para este item."""
        return ContractLine(
            group_label=self.group_label,
            item_number=self.item_number,
            code=self.code,
            description=self.description,
            unit=self.unit,
            unit_price=self.unit_price,
            contract_quantity=self.contract_quantity,
            amended_quantity=self.amended_quantity,
            periods=list(self.periods(period_numbers)),
            accumulated_quantity=self.accumulated_quantity,
            accumulated_amount=self.accumulated_amount,
            balance_quantity=self.balance_quantity,
        )

    def amendment_lines(self) -> list[AmendmentLine]:
        """Efeito declarado deste item na 1ª RE-RA, na ordem das colunas do bloco."""
        lines: list[AmendmentLine] = []
        if self.reduced is not None:
            lines.append(AmendmentLine(code=self.code, quantity_delta=-self.reduced))
        if self.added is not None:
            lines.append(AmendmentLine(code=self.code, quantity_delta=self.added))
        if self.new_item is not None:
            lines.append(
                AmendmentLine(code=self.code, quantity_delta=self.new_item, is_new_item=True)
            )
        return lines


_PAVING = "PAVIMENTACAO SINTETICA"
_SERVICES = "SERVICOS PRELIMINARES E MOBILIARIO"

_PREVIOUS_MAPAO_ITEMS: tuple[_MapaoItem, ...] = (
    _MapaoItem(
        group_label=_PAVING,
        item_number="1",
        numeric_item_number=True,
        code="AD04050050(/)",
        description="PISO INTERTRAVADO SINTETICO 6CM",
        unit_label="M2",
        unit="m2",
        unit_price=Decimal("10.00"),
        contract_quantity=Decimal("20.00"),
        amended_quantity=Decimal("16.00"),
        period_quantities=(Decimal("3.00"), Decimal("2.00")),
        reduced=Decimal("4.00"),
    ),
    _MapaoItem(
        group_label=_PAVING,
        item_number="2",
        numeric_item_number=True,
        code="AD04050055(A)",
        description="PISO INTERTRAVADO SINTETICO 8CM VARIANTE A",
        unit_label="M2",
        unit="m2",
        unit_price=Decimal("12.50"),
        contract_quantity=Decimal("10.00"),
        amended_quantity=Decimal("10.00"),
        # 5,99 x 12,50 = 74,875: dinheiro trunca em 74,87 e o saldo fecha em 0,01.
        period_quantities=(Decimal("4.00"), Decimal("5.99")),
    ),
    _MapaoItem(
        group_label=_PAVING,
        item_number="3",
        numeric_item_number=True,
        code="AD04050055(B)",
        description="PISO INTERTRAVADO SINTETICO 8CM VARIANTE B",
        unit_label="M2",
        unit="m2",
        unit_price=Decimal("15.00"),
        contract_quantity=Decimal("6.00"),
        amended_quantity=Decimal("6.00"),
        period_quantities=(Decimal("1.50"), None),
    ),
    _MapaoItem(
        group_label=_SERVICES,
        item_number="4",
        numeric_item_number=False,
        code="SP01050010(/)",
        description="ESCAVACAO SINTETICA MANUAL EM SOLO",
        unit_label="M3",
        unit="m3",
        unit_price=Decimal("20.00"),
        contract_quantity=Decimal("8.00"),
        amended_quantity=Decimal("8.00"),
        period_quantities=(Decimal("1.00"), Decimal("1.50")),
    ),
    _MapaoItem(
        group_label=_SERVICES,
        item_number="5",
        numeric_item_number=False,
        code="SP01050015(A)",
        description="REATERRO SINTETICO COMPACTADO",
        unit_label="M3",
        unit="m3",
        unit_price=Decimal("8.00"),
        contract_quantity=Decimal("3.00"),
        amended_quantity=Decimal("0.00"),
        period_quantities=(None, None),
        reduced=Decimal("3.00"),
    ),
    _MapaoItem(
        group_label=_SERVICES,
        item_number="6",
        numeric_item_number=False,
        code="MB01100010(/)",
        description="BANCO SINTETICO DE CONCRETO PRE-MOLDADO",
        unit_label="UN",
        unit="un",
        unit_price=Decimal("5.00"),
        contract_quantity=Decimal("0.00"),
        amended_quantity=Decimal("6.00"),
        period_quantities=(None, None),
        new_item=Decimal("6.00"),
    ),
)


@dataclass(frozen=True, slots=True)
class PreviousMapaoFixture:
    """MAPÃO sintético do contrato e o consolidado que o leitor deve devolver."""

    workbook_path: Path
    expected: ContractWorkbook
    separator_rows: tuple[int, ...] = ()


def _mapao_items(*, with_amendments: bool) -> tuple[_MapaoItem, ...]:
    """Sem RE-RA, o contratual passa a ser o vigente: a GERAL fecha sem aditivo."""
    if with_amendments:
        return _PREVIOUS_MAPAO_ITEMS
    return tuple(
        replace(
            item,
            contract_quantity=item.amended_quantity,
            reduced=None,
            added=None,
            new_item=None,
        )
        for item in _PREVIOUS_MAPAO_ITEMS
    )


def _period_columns(layout: GeneralLayout, period_count: int) -> tuple[list[int], int, int]:
    """Colunas derivadas do template: pares de medição, acumulado e saldo."""
    first = column_index_from_string(layout.first_period_column)
    pairs = [first + 2 * index for index in range(period_count)]
    accumulated = first + 2 * period_count
    return pairs, accumulated, accumulated + 2


def _general_header_labels(layout: GeneralLayout) -> tuple[tuple[str, str], ...]:
    """Rótulos das colunas fixas; sem coluna de vigente, ela simplesmente não é escrita."""
    labels = [
        (layout.group_column, "GRUPO"),
        (layout.item_column, "ITEM"),
        (layout.code_column, "CODIGO"),
        (layout.description_column, "DESCRICAO"),
        (layout.unit_column, "UN"),
        (layout.contract_quantity_column, "QUANT. CONTRATO"),
        (layout.unit_price_column, "VALOR UNIT"),
    ]
    if layout.amended_quantity_column is not None:
        labels.append((layout.amended_quantity_column, "QUANT. VIGENTE"))
    return tuple(labels)


def _write_separator_row(
    worksheet: Any, layout: GeneralLayout, row: int, period_count: int
) -> None:
    """Linha separadora do arquivo real: sem texto nenhum, só o zero em cache das fórmulas."""
    pairs, accumulated, balance = _period_columns(layout, period_count)
    for column in (*pairs, accumulated):
        worksheet.cell(row=row, column=column, value=0)
        worksheet.cell(row=row, column=column + 1, value=0)
    worksheet.cell(row=row, column=balance, value=0)


def _write_general_sheet(
    worksheet: Any,
    layout: GeneralLayout,
    items: Sequence[_MapaoItem],
    *,
    period_numbers: Sequence[int],
    period_label_suffixes: Mapping[int, str],
    separator_rows: bool,
) -> tuple[int, ...]:
    """Escreve a PLANILHA GERAL como o cliente a entrega, usando só o template.

    Devolve os números das linhas separadoras escritas, para que o gabarito do teste
    nasça da mesma fonte que a planilha.
    """
    period_count = len(period_numbers)
    pairs, accumulated, balance = _period_columns(layout, period_count)
    worksheet[f"{layout.group_column}1"] = layout.title
    for letter, label in _general_header_labels(layout):
        worksheet[f"{letter}{layout.header_row}"] = label
    sublabel_row = layout.pair_sublabel_row
    for number, column in zip(period_numbers, pairs, strict=True):
        worksheet.cell(
            row=layout.header_row,
            column=column,
            value=layout.period_label(number) + period_label_suffixes.get(number, ""),
        )
    worksheet.cell(row=layout.header_row, column=accumulated, value=layout.accumulated_label)
    worksheet.cell(row=layout.header_row, column=balance, value=layout.balance_label)
    if sublabel_row is not None:
        for column in (*pairs, accumulated):
            worksheet.cell(row=sublabel_row, column=column, value=layout.quantity_pair_label)
            worksheet.cell(row=sublabel_row, column=column + 1, value=layout.amount_pair_label)
        worksheet.cell(row=sublabel_row, column=balance, value=layout.quantity_pair_label)

    row = layout.data_first_row
    group_label = ""
    separators: list[int] = []
    for item in items:
        if item.group_label != group_label:
            if separator_rows and group_label:
                _write_separator_row(worksheet, layout, row, period_count)
                separators.append(row)
                row += 1
            worksheet[f"{layout.group_column}{row}"] = item.group_label
            group_label = item.group_label
            row += 1
        _write_general_item(worksheet, layout, item, row, period_numbers)
        row += 1
    if separator_rows and group_label:
        _write_separator_row(worksheet, layout, row, period_count)
        separators.append(row)
        row += 1
    row += 1  # linha totalmente vazia, ignorada pela varredura
    worksheet[f"{layout.group_column}{row}"] = layout.total_label
    worksheet.cell(
        row=row,
        column=accumulated + 1,
        value=sum((item.accumulated_amount for item in items), _ZERO),
    )
    # Rodapé depois do total: se a varredura não parasse na linha de total, esta linha
    # viraria item de código ilegível.
    worksheet[f"{layout.code_column}{row + 1}"] = "RESUMO"
    worksheet[f"{layout.description_column}{row + 1}"] = "VALORES APOS A ULTIMA MEDICAO"
    return tuple(separators)


def _write_general_item(
    worksheet: Any,
    layout: GeneralLayout,
    item: _MapaoItem,
    row: int,
    period_numbers: Sequence[int],
) -> None:
    pairs, accumulated, balance = _period_columns(layout, len(period_numbers))
    number: object = int(item.item_number) if item.numeric_item_number else item.item_number
    worksheet[f"{layout.item_column}{row}"] = number
    worksheet[f"{layout.code_column}{row}"] = item.code
    worksheet[f"{layout.description_column}{row}"] = item.description
    worksheet[f"{layout.unit_column}{row}"] = item.unit_label
    worksheet[f"{layout.contract_quantity_column}{row}"] = item.contract_quantity
    worksheet[f"{layout.unit_price_column}{row}"] = item.unit_price
    if layout.amended_quantity_column is not None:
        worksheet[f"{layout.amended_quantity_column}{row}"] = item.amended_quantity
    for quantity, column, period in zip(
        item.period_quantities, pairs, item.periods(period_numbers), strict=True
    ):
        if quantity is None:
            continue
        worksheet.cell(row=row, column=column, value=quantity)
        worksheet.cell(row=row, column=column + 1, value=period.amount)
    worksheet.cell(row=row, column=accumulated, value=item.accumulated_quantity)
    worksheet.cell(row=row, column=accumulated + 1, value=item.accumulated_amount)
    worksheet.cell(row=row, column=balance, value=item.balance_quantity)


def _first_item_by_code(items: Sequence[_MapaoItem]) -> tuple[_MapaoItem, ...]:
    """Um item por código, na ordem da planilha: a aba da prefeitura fala por código."""
    seen: set[str] = set()
    unique: list[_MapaoItem] = []
    for item in items:
        if item.code in seen:
            continue
        seen.add(item.code)
        unique.append(item)
    return tuple(unique)


def _write_amendment_sheet(
    worksheet: Any, layout: AmendmentLayout, items: Sequence[_MapaoItem]
) -> None:
    """Aba da prefeitura: um bloco de RE-RA e o vigente declarado para cada código."""
    block = layout.blocks[0]
    worksheet[f"{layout.code_column}{layout.header_row}"] = "CODIGO"
    if layout.amended_quantity_column is not None:
        worksheet[f"{layout.amended_quantity_column}{layout.header_row}"] = "QUANT. VIGENTE"
    for letter, label in (
        (block.reduced_column, f"{block.label} REDUZIDA"),
        (block.added_column, f"{block.label} ACRESCIDA"),
        (block.new_item_column, f"{block.label} ITEM NOVO"),
    ):
        if letter is not None:
            worksheet[f"{letter}{layout.header_row}"] = label
    row = layout.data_first_row
    for item in _first_item_by_code(items):
        worksheet[f"{layout.code_column}{row}"] = item.code
        if layout.amended_quantity_column is not None:
            worksheet[f"{layout.amended_quantity_column}{row}"] = item.amended_quantity
        for letter, value in (
            (block.reduced_column, item.reduced),
            (block.added_column, item.added),
            (block.new_item_column, item.new_item),
        ):
            if letter is not None and value is not None:
                worksheet[f"{letter}{row}"] = value
        row += 1


@dataclass(frozen=True, slots=True)
class WrittenMapao:
    """Arquivo do MAPÃO anterior gravado, sem gabarito: para fixture que não importa."""

    workbook_path: Path
    separator_rows: tuple[int, ...]


def write_previous_mapao_workbook(
    path: Path,
    *,
    with_amendments: bool = True,
    template: WorkbookTemplate | None = None,
    items: Sequence[_MapaoItem] | None = None,
    period_numbers: Sequence[int] = PREVIOUS_MAPAO_PERIOD_NUMBERS,
    period_label_suffixes: Mapping[int, str] | None = None,
    separator_rows: bool = False,
) -> WrittenMapao:
    """Grava o MAPÃO anterior sintético; os defaults reproduzem a fixture padrão.

    A GERAL nasce oculta, como no arquivo real, e nenhuma posição de célula é escrita à
    mão: linhas, colunas e rótulos saem todos do template. Esta função grava e não monta
    gabarito, porque há layout real que o leitor precisa recusar — e recusa não tem
    consolidado esperado.
    """
    workbook_template = default_template() if template is None else template
    mapao_items = _mapao_items(with_amendments=with_amendments) if items is None else tuple(items)
    workbook = Workbook()
    cover = workbook.active
    cover.title = PREVIOUS_MAPAO_COVER_SHEET
    cover["A1"] = PREVIOUS_MAPAO_SOURCE_LABEL
    general = workbook.create_sheet(workbook_template.general.sheet_name)
    general.sheet_state = "hidden"
    separators = _write_general_sheet(
        general,
        workbook_template.general,
        mapao_items,
        period_numbers=period_numbers,
        period_label_suffixes={} if period_label_suffixes is None else period_label_suffixes,
        separator_rows=separator_rows,
    )
    if with_amendments and workbook_template.amendment is not None:
        _write_amendment_sheet(
            workbook.create_sheet(workbook_template.amendment.sheet_name),
            workbook_template.amendment,
            mapao_items,
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return WrittenMapao(workbook_path=path, separator_rows=separators)


def build_previous_mapao_workbook(
    path: Path,
    *,
    with_amendments: bool = True,
    template: WorkbookTemplate | None = None,
    items: Sequence[_MapaoItem] | None = None,
    period_numbers: Sequence[int] = PREVIOUS_MAPAO_PERIOD_NUMBERS,
    period_label_suffixes: Mapping[int, str] | None = None,
    separator_rows: bool = False,
) -> PreviousMapaoFixture:
    """Grava o MAPÃO anterior sintético e devolve o consolidado esperado da leitura."""
    workbook_template = default_template() if template is None else template
    mapao_items = _mapao_items(with_amendments=with_amendments) if items is None else tuple(items)
    written = write_previous_mapao_workbook(
        path,
        with_amendments=with_amendments,
        template=workbook_template,
        items=mapao_items,
        period_numbers=period_numbers,
        period_label_suffixes=period_label_suffixes,
        separator_rows=separator_rows,
    )

    amendment_lines: list[AmendmentLine] = []
    for item in _first_item_by_code(mapao_items):
        amendment_lines.extend(item.amendment_lines())
    amendments: list[Amendment] = []
    if amendment_lines and with_amendments and workbook_template.amendment is not None:
        amendments.append(
            Amendment(label=workbook_template.amendment.blocks[0].label, lines=amendment_lines)
        )
    expected = ContractWorkbook(
        id=contract_workbook_id_for(file_sha256(path)),
        source_label=PREVIOUS_MAPAO_SOURCE_LABEL,
        source_sha256=file_sha256(path),
        contract_label=PREVIOUS_MAPAO_CONTRACT_LABEL,
        period_numbers=list(period_numbers),
        lines=[item.to_contract_line(period_numbers) for item in mapao_items],
        amendments=amendments,
    )
    return PreviousMapaoFixture(
        workbook_path=written.workbook_path,
        expected=expected,
        separator_rows=written.separator_rows,
    )


def tamper_cell(path: Path, sheet_name: str, ref: str, value: object) -> None:
    """Altera uma célula do arquivo já gravado, como faria uma planilha adulterada."""
    workbook = load_workbook(path)
    workbook[sheet_name][ref] = value
    workbook.save(path)


@dataclass(frozen=True, slots=True)
class GeneralColumns:
    """Letras das colunas derivadas do template, por medição declarada.

    Existe para que o teste adultere célula sem escrever letra de coluna à mão: a posição
    do par é derivada do número de medições, exatamente como o leitor a deriva.
    """

    period_quantity: tuple[str, ...]
    period_amount: tuple[str, ...]
    accumulated_quantity: str
    accumulated_amount: str
    balance: str


def general_columns(layout: GeneralLayout, period_count: int) -> GeneralColumns:
    """Colunas da PLANILHA GERAL para um número de medições declarado no cabeçalho."""
    pairs, accumulated, balance = _period_columns(layout, period_count)
    return GeneralColumns(
        period_quantity=tuple(get_column_letter(column) for column in pairs),
        period_amount=tuple(get_column_letter(column + 1) for column in pairs),
        accumulated_quantity=get_column_letter(accumulated),
        accumulated_amount=get_column_letter(accumulated + 1),
        balance=get_column_letter(balance),
    )


def general_item_rows(path: Path, layout: GeneralLayout) -> dict[tuple[str, str], int]:
    """Linha de cada par grupo+código na GERAL gravada, na ordem em que a planilha os traz."""
    workbook = load_workbook(path)
    try:
        worksheet = workbook[layout.sheet_name]
        rows: dict[tuple[str, str], int] = {}
        group_label = ""
        for row in range(layout.data_first_row, worksheet.max_row + 1):
            group = worksheet[f"{layout.group_column}{row}"].value
            code = worksheet[f"{layout.code_column}{row}"].value
            if isinstance(group, str) and group.strip() and not isinstance(code, str):
                group_label = group.strip()
                continue
            if isinstance(code, str) and is_sco_code(code.strip()):
                rows[(group_label, code.strip())] = row
        return rows
    finally:
        workbook.close()


def amendment_code_rows(path: Path, layout: AmendmentLayout) -> dict[str, int]:
    """Linha de cada código na aba da prefeitura gravada."""
    workbook = load_workbook(path)
    try:
        worksheet = workbook[layout.sheet_name]
        rows: dict[str, int] = {}
        for row in range(layout.data_first_row, worksheet.max_row + 1):
            code = worksheet[f"{layout.code_column}{row}"].value
            if isinstance(code, str) and is_sco_code(code.strip()):
                rows.setdefault(code.strip(), row)
        return rows
    finally:
        workbook.close()


REAL_LAYOUT_PERIOD_NUMBERS: tuple[int, ...] = (1, 2, 4)
"""Numeração com buraco, como no arquivo do cliente: a 3ª medição não existe."""

REAL_LAYOUT_PERIOD_SUFFIXES: dict[int, str] = {
    1: " ",
    2: " - COMPLEMENTAR",
    4: " (COMPLEMENTAR",
}
"""Sufixos reais do cabeçalho, inclusive o parêntese que o cliente não fecha."""

REAL_LAYOUT_DUPLICATED_CODE = "AD04050050(/)"
"""Código que o MAPÃO real repete em grupos diferentes."""


def real_layout_template() -> WorkbookTemplate:
    """Template da GERAL como o cliente a entrega, verificado por sonda no arquivo real.

    Três diferenças em relação ao template padrão, todas declaradas como dado: não há
    coluna de quantidade vigente, há uma coluna entre o preço e o primeiro par que o
    leitor não lê (o custo parcial) e a quantidade tem quatro casas decimais.
    """
    template = default_template()
    payload = template.model_dump()
    payload["general"].update(
        {
            "contract_quantity_column": "H",
            "unit_price_column": "I",
            "amended_quantity_column": None,
            "first_period_column": "K",
            "quantity_decimal_scale": 4,
        }
    )
    return WorkbookTemplate.model_validate(payload)


def template_without_amended_column() -> WorkbookTemplate:
    """Template padrão sem a coluna de vigente; tudo o mais fica onde estava."""
    template = default_template()
    payload = template.model_dump()
    payload["general"]["amended_quantity_column"] = None
    return WorkbookTemplate.model_validate(payload)


EXTRA_CODE_PATTERN = r"^IE\d{8}$"
"""Padrão único de código nu usado nos testes de padrão extra: fonte única do valor."""

EXTRA_CODE_SAMPLE = "IE00040849"
"""Código nu de exemplo: fora da tabela SCO, sem variante, como o contrato real mede."""


def template_with_extra_code_pattern(
    *,
    template: WorkbookTemplate | None = None,
    patterns: Sequence[str] = (EXTRA_CODE_PATTERN,),
) -> WorkbookTemplate:
    """Template (padrão ou informado) com o padrão extra de código declarado.

    Sem esse padrão declarado, o leitor não aceita o código nu — é o template quem decide
    qual forma nua vale para a importação, não o superset estrutural sozinho.
    """
    base = default_template() if template is None else template
    payload = base.model_dump()
    payload["extra_code_patterns"] = list(patterns)
    return WorkbookTemplate.model_validate(payload)


def extra_code_items(*, with_amendment: bool = False) -> tuple[_MapaoItem, ...]:
    """Um único item de código nu; com `with_amendment`, a 1ª RE-RA reduz a quantidade."""
    return (
        _MapaoItem(
            group_label=_SERVICES,
            item_number="1",
            numeric_item_number=True,
            code=EXTRA_CODE_SAMPLE,
            description="SERVICO CONTRATADO FORA DA TABELA SCO",
            unit_label="UN",
            unit="un",
            unit_price=Decimal("5.00"),
            contract_quantity=Decimal("10.00"),
            amended_quantity=Decimal("9.00") if with_amendment else Decimal("10.00"),
            period_quantities=(Decimal("2.00"), Decimal("1.00")),
            reduced=Decimal("1.00") if with_amendment else None,
        ),
    )


def template_with_section_group_subtotal(
    *, template: WorkbookTemplate | None = None
) -> WorkbookTemplate:
    """Template (padrão ou informado) com a flag de subtotal de grupo na linha de seção.

    Layout observado no arquivo real: a linha que abre um grupo na aba da prefeitura
    carrega o subtotal do grupo na própria coluna de vigente, em vez de ficar em branco.
    """
    base = default_template() if template is None else template
    payload = base.model_dump()
    assert payload["amendment"] is not None, "template sem aba de RE-RA declarada"
    payload["amendment"]["section_rows_carry_group_subtotal"] = True
    return WorkbookTemplate.model_validate(payload)


_REAL_LAYOUT_ITEMS: tuple[_MapaoItem, ...] = (
    _MapaoItem(
        group_label=_PAVING,
        item_number="1",
        numeric_item_number=True,
        code=REAL_LAYOUT_DUPLICATED_CODE,
        description="PISO INTERTRAVADO SINTETICO 6CM",
        unit_label="M2",
        unit="m2",
        unit_price=Decimal("10.00"),
        # Quantidade de quatro casas, como a planilha do cliente escreve.
        contract_quantity=Decimal("100459.5525"),
        amended_quantity=Decimal("100459.5525"),
        period_quantities=(Decimal("1.3200"), Decimal("2.0000"), None),
    ),
    _MapaoItem(
        group_label=_PAVING,
        item_number="2",
        numeric_item_number=True,
        code="AD04050055(A)",
        description="PISO INTERTRAVADO SINTETICO 8CM VARIANTE A",
        unit_label="M2",
        unit="m2",
        unit_price=Decimal("12.50"),
        contract_quantity=Decimal("10.0000"),
        amended_quantity=Decimal("10.0000"),
        period_quantities=(Decimal("4.0000"), Decimal("5.9900"), None),
    ),
    _MapaoItem(
        # A numeração de item reinicia no grupo e o código do primeiro grupo se repete.
        group_label=_SERVICES,
        item_number="1",
        numeric_item_number=True,
        code=REAL_LAYOUT_DUPLICATED_CODE,
        description="PISO INTERTRAVADO SINTETICO 6CM EM CALCADA",
        unit_label="M2",
        unit="m2",
        unit_price=Decimal("10.00"),
        contract_quantity=Decimal("8.0000"),
        amended_quantity=Decimal("8.0000"),
        period_quantities=(None, Decimal("1.5000"), Decimal("0.2500")),
    ),
    _MapaoItem(
        group_label=_SERVICES,
        item_number="2",
        numeric_item_number=True,
        code="SP01050010(/)",
        description="ESCAVACAO SINTETICA MANUAL EM SOLO",
        unit_label="M3",
        unit="m3",
        unit_price=Decimal("20.00"),
        contract_quantity=Decimal("8.0000"),
        amended_quantity=Decimal("8.0000"),
        period_quantities=(Decimal("1.0000"), Decimal("1.5000"), None),
    ),
)


def real_layout_items(*, reduce_duplicated_code: bool = False) -> tuple[_MapaoItem, ...]:
    """Itens do layout real; com `reduce_duplicated_code`, a RE-RA altera o código repetido."""
    if not reduce_duplicated_code:
        return _REAL_LAYOUT_ITEMS
    return tuple(
        replace(
            item,
            contract_quantity=item.amended_quantity + Decimal("4.0000"),
            reduced=Decimal("4.0000"),
        )
        if item.code == REAL_LAYOUT_DUPLICATED_CODE
        else item
        for item in _REAL_LAYOUT_ITEMS
    )


MULTI_WORKSITE_REFERENCE_LABEL = "MARCO/2026"
CONTRACT_CATALOG_LABEL = "CATALOGO SINTETICO DO CONTRATO (fixture)"
_VALUATION_NAMESPACE = uuid5(NAMESPACE_URL, "https://croquito.local/valuation/multi-worksite")


@dataclass(frozen=True, slots=True)
class MeasuredItem:
    """Um código medido numa obra; com dedução, a memória exercita o desconto de vãos."""

    code: str
    quantity: Decimal
    deduction: Decimal | None = None


@dataclass(frozen=True, slots=True)
class MeasuredWorksite:
    """Uma obra do período e o que ela mediu."""

    worksite_key: str
    worksite_name: str
    items: tuple[MeasuredItem, ...] = field(default_factory=tuple)


_MULTI_WORKSITE_MEASUREMENTS: tuple[MeasuredWorksite, ...] = (
    MeasuredWorksite(
        worksite_key="praca-sintetica-norte",
        worksite_name="PRACA SINTETICA NORTE",
        items=(
            # O bloco com dedução mantém a forma `=ROUND(PRODUCT(...),2)-<ref>` viva.
            MeasuredItem(code="AD04050050(/)", quantity=Decimal("2.00"), deduction=Decimal("1.00")),
            MeasuredItem(code="AD04050055(B)", quantity=Decimal("1.50")),
        ),
    ),
    MeasuredWorksite(
        worksite_key="praca-sintetica-sul",
        worksite_name="PRACA SINTETICA SUL",
        items=(
            # Mesmo código da obra do norte: a GERAL consolida 2,00 + 3,00 numa linha só.
            MeasuredItem(code="AD04050050(/)", quantity=Decimal("3.00")),
            MeasuredItem(code="SP01050010(/)", quantity=Decimal("1.25")),
        ),
    ),
    MeasuredWorksite(
        worksite_key="praca-sintetica-leste",
        worksite_name="PRACA SINTETICA LESTE",
        items=(MeasuredItem(code="MB01100010(/)", quantity=Decimal("4.00")),),
    ),
)


def _calc_block(item: MeasuredItem, unit: str) -> CalcBlock:
    """Bloco de cálculo cuja quantidade recomputada é exatamente a medida."""
    if item.deduction is None:
        return CalcBlock(
            label="TRECHO MEDIDO",
            recipe=CalcRecipe.DIRECT_QUANTITY,
            operands=[CalcOperand(name="QUANTIDADE", value=item.quantity, unit=unit)],
            subtotal=item.quantity,
        )
    return CalcBlock(
        label="TRECHO MEDIDO COM VAO",
        recipe=CalcRecipe.PERIM_HEIGHT_MINUS_OPENINGS,
        operands=[
            CalcOperand(name="PERÍMETRO", value=item.quantity + item.deduction, unit="m"),
            CalcOperand(name="ALTURA", value=Decimal("1.00"), unit="m"),
        ],
        deductions=[CalcOperand(name="VÃOS", value=item.deduction, unit=unit)],
        subtotal=item.quantity,
    )


def _build_valuation(
    period_number: int,
    worksites: Sequence[MeasuredWorksite],
    describe: Callable[[str], tuple[str, str, Decimal]],
) -> Valuation:
    """Monta a medição a partir do que cada obra mediu, com memória 1:1 por item."""
    bulletins: list[WorksiteBulletin] = []
    calc_sheets: list[CalcSheet] = []
    for worksite in worksites:
        lines: list[BulletinLine] = []
        for index, item in enumerate(worksite.items, start=1):
            description, unit, unit_price = describe(item.code)
            item_number = str(index)
            lines.append(
                BulletinLine(
                    item_number=item_number,
                    code=item.code,
                    description=description,
                    unit=unit,
                    unit_price=unit_price,
                    quantity=item.quantity,
                    total=money_trunc(item.quantity * unit_price),
                )
            )
            calc_sheets.append(
                CalcSheet(
                    worksite_key=worksite.worksite_key,
                    item_number=item_number,
                    blocks=[_calc_block(item, unit)],
                    total_quantity=item.quantity,
                )
            )
        bulletins.append(
            WorksiteBulletin(
                worksite_key=worksite.worksite_key,
                worksite_name=worksite.worksite_name,
                address=f"RUA SINTETICA {len(bulletins) + 1}00 — BAIRRO SINTETICO",
                contract_label=PREVIOUS_MAPAO_CONTRACT_LABEL,
                lines=lines,
                total_amount=sum((line.total for line in lines), _ZERO),
            )
        )
    keys = "/".join(worksite.worksite_key for worksite in worksites)
    return Valuation(
        id=uuid5(_VALUATION_NAMESPACE, f"{keys}/{period_number}"),
        period_number=period_number,
        reference_label=MULTI_WORKSITE_REFERENCE_LABEL,
        bulletins=bulletins,
        calc_sheets=calc_sheets,
    )


def build_multi_worksite_valuation(contract: ContractWorkbook) -> Valuation:
    """Medição do período seguinte ao consolidado, com três obras e um código repetido.

    As quantidades foram escolhidas para caber no saldo e para que o valor consolidado do
    código medido em duas obras não derive de centavo: `TRUNC(Σq x preço)` é igual a
    `Σ TRUNC(q_i x preço)`.
    """

    def describe(code: str) -> tuple[str, str, Decimal]:
        line = contract.line_for_code(code)
        return line.description, line.unit, line.unit_price

    return _build_valuation(contract.next_period_number, _MULTI_WORKSITE_MEASUREMENTS, describe)


def build_valuation_from_catalog(
    catalog: PriceCatalog, period_number: int, worksites: Sequence[MeasuredWorksite]
) -> Valuation:
    """Medição arbitrária cujo preço, unidade e descrição saem do catálogo importado."""

    def describe(code: str) -> tuple[str, str, Decimal]:
        entry = catalog.entry_for(code)
        return entry.description, entry.unit, entry.unit_price

    return _build_valuation(period_number, worksites, describe)


def build_catalog_for_contract(
    path: Path, contract: ContractWorkbook, *, extra_rows: Sequence[CatalogRow] = ()
) -> PriceCatalog:
    """Grava e importa um catálogo com uma linha por código do consolidado contratual."""
    rows: list[CatalogRow] = []
    family = ""
    subgroup = ""
    for line in contract.lines:
        parts = parse_sco_code(line.code)
        if parts.family != family:
            family = parts.family
            subgroup = ""
            rows.append((family, f"FAMILIA SINTETICA {family}", "", ""))
        if parts.subgroup_row_code != subgroup:
            subgroup = parts.subgroup_row_code
            rows.append((subgroup, f"SUBGRUPO SINTETICO {subgroup}", "", ""))
        rows.append((line.code, line.description, line.unit, str(line.unit_price)))
    rows.extend(extra_rows)
    write_catalog_workbook(path, rows)
    return read_price_catalog(
        path,
        default_template(),
        source_label=CONTRACT_CATALOG_LABEL,
        reference_month=SYNTHETIC_REFERENCE_MONTH,
    )


def build_approval(valuation: Valuation) -> Valuation:
    """Devolve a medição com a aprovação nominal do orçamentista amarrada ao digest."""
    approval = ValuationApproval(
        decision=ReviewerDecision(
            decision_id="vd_0123456789abcdef",
            action="confirm",
            reviewer_id="orcamentista-sintetico",
            reviewer_role="orcamentista",
            decided_at=datetime(2026, 3, 1, 12, 0, tzinfo=UTC),
            note="medição sintética conferida contra o consolidado",
        ),
        valuation_digest=valuation.content_digest(),
    )
    payload = valuation.model_dump()
    payload["approval"] = approval.model_dump()
    return Valuation.model_validate(payload)


def bake_formula_values(source: Path, target: Path, template: WorkbookTemplate) -> Path:
    """Copia a pasta com cada fórmula já substituída pelo valor que ela representa.

    O `openpyxl` não grava o valor calculado de uma fórmula: reabrir o arquivo gerado com
    `data_only=True` devolveria célula vazia onde há fórmula, e o leitor do consolidado só
    lê valores. Esta cópia repõe o que uma planilha recalculada mostraria, usando o
    avaliador da própria auditoria — que recomputa cada fórmula em `Decimal` a partir do
    arquivo, não do plano.
    """
    canonical = canonicalize_workbook(source, template)
    sheets = canonical["sheets"]
    assert isinstance(sheets, list)
    workbook = Workbook()
    for index, sheet in enumerate(sheets):
        assert isinstance(sheet, dict)
        name = str(sheet["name"])
        worksheet = workbook.active if index == 0 else workbook.create_sheet(name)
        worksheet.title = name
        cells = sheet["cells"]
        assert isinstance(cells, list)
        for cell in cells:
            value = str(cell["value"])
            worksheet[str(cell["ref"])] = value if cell["kind"] == "text" else Decimal(value)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    return target


COLUMNAR_CATALOG_SHEET = "FGV06 SINTETICO"
COLUMNAR_CATALOG_FIRST_ROW = 4
COLUMNAR_CATALOG_TITLE_ROWS: tuple[str, ...] = (
    "PCRJ  SCO-Sistema de Custos de Obras (fixture)",
    "CATALOGO SINTETICO DE PRECOS",
    "REFERENCIA: Janeiro/2026",
)
COLUMNAR_NOT_APPLICABLE = "NA"

# Família, subgrupo intermediário, código, descrição, unidade e preço, na ordem das colunas.
ColumnarCatalogRow = tuple[str, str, str, str, str, str | float]
_BLANK_COLUMNAR_ROW: ColumnarCatalogRow = ("", "", "", "", "", "")


@dataclass(frozen=True, slots=True)
class ColumnarCatalogItem:
    """Item do catálogo publicado: como a planilha escreve e como o catálogo fica."""

    code: str
    description: str
    unit_label: str
    unit: str
    raw_price: str | float
    unit_price: Decimal


@dataclass(frozen=True, slots=True)
class _ColumnarSubgroup:
    """Cabeçalho de subgrupo escrito na coluna de código, com os itens abaixo dele."""

    header_text: str
    code: str
    name: str
    items: tuple[ColumnarCatalogItem, ...]


@dataclass(frozen=True, slots=True)
class _ColumnarIntermediate:
    """Nível entre família e subgrupo, escrito na coluna própria com preço `NA`."""

    header_text: str
    subgroups: tuple[_ColumnarSubgroup, ...]


@dataclass(frozen=True, slots=True)
class _ColumnarFamily:
    """Família na coluna mais à esquerda; `noise` reproduz o texto sem valor das demais."""

    header_text: str
    code: str
    name: str
    noise: tuple[str, str, str, str, str]
    intermediates: tuple[_ColumnarIntermediate, ...]


@dataclass(frozen=True, slots=True)
class ExpectedColumnarEntry:
    """Entrada esperada do catálogo colunar, derivada da mesma árvore que gera a planilha."""

    code: str
    description: str
    unit: str
    unit_price: Decimal
    family_code: str
    family_name: str
    subgroup_code: str
    subgroup_name: str


_COLUMNAR_FAMILIES: tuple[_ColumnarFamily, ...] = (
    _ColumnarFamily(
        header_text="AD PAVIMENTACAO SINTETICA",
        code="AD",
        name="PAVIMENTACAO SINTETICA",
        noise=("...2", "...3", "...4", "...5", "...6"),
        intermediates=(
            _ColumnarIntermediate(
                header_text="AD 04 SERVICOS DE PAVIMENTACAO",
                subgroups=(
                    _ColumnarSubgroup(
                        header_text="AD 04.05 PISO INTERTRAVADO SINTETICO",
                        code="AD0405",
                        name="PISO INTERTRAVADO SINTETICO",
                        items=(
                            ColumnarCatalogItem(
                                code="AD04050050(/)",
                                description="PISO INTERTRAVADO SINTETICO 6CM",
                                unit_label="M2",
                                unit="m2",
                                raw_price="   217,69",
                                unit_price=Decimal("217.69"),
                            ),
                            ColumnarCatalogItem(
                                code="AD04050055(B)",
                                description="PISO INTERTRAVADO SINTETICO 8CM",
                                unit_label="M²",
                                unit="m2",
                                raw_price=1625.02,
                                unit_price=Decimal("1625.02"),
                            ),
                        ),
                    ),
                    _ColumnarSubgroup(
                        header_text="AD0410MEIOFIODEGRANITOSINTETICO",
                        code="AD0410",
                        name="MEIOFIODEGRANITOSINTETICO",
                        items=(
                            ColumnarCatalogItem(
                                code="AD04100015(/)",
                                description="MEIO FIO DE GRANITO SINTETICO",
                                unit_label="ML",
                                unit="m",
                                raw_price="  89,30",
                                unit_price=Decimal("89.30"),
                            ),
                        ),
                    ),
                ),
            ),
        ),
    ),
    _ColumnarFamily(
        header_text="CE SERVICOS PRELIMINARES E MOBILIARIO",
        code="CE",
        name="SERVICOS PRELIMINARES E MOBILIARIO",
        noise=("...2", "...3", "...4", "...5", "...6"),
        intermediates=(
            _ColumnarIntermediate(
                header_text="CE 02 MOBILIARIO URBANO",
                subgroups=(
                    _ColumnarSubgroup(
                        header_text="CE 02.10 BANCOS E LIXEIRAS",
                        code="CE0210",
                        name="BANCOS E LIXEIRAS",
                        items=(
                            ColumnarCatalogItem(
                                code="CE02100015(A)",
                                description="BANCO DE CONCRETO SINTETICO",
                                unit_label="UN",
                                unit="un",
                                raw_price="142,75",
                                unit_price=Decimal("142.75"),
                            ),
                            ColumnarCatalogItem(
                                code="CE02100020(/)",
                                description="LIXEIRA SINTETICA",
                                unit_label="UNID.",
                                unit="un",
                                raw_price="1.580,00",
                                unit_price=Decimal("1580.00"),
                            ),
                        ),
                    ),
                ),
            ),
        ),
    ),
)


def columnar_catalog_rows() -> tuple[ColumnarCatalogRow, ...]:
    """Área de dados do catálogo colunar sintético, na ordem em que a planilha a mostra.

    Reproduz o que o arquivo publicado tem de incômodo: a linha de família com texto sem
    valor nas outras colunas, `NA` no preço dos cabeçalhos, subgrupo escrito com e sem
    separadores, preço em texto pt-BR com espaços à esquerda e linha separadora vazia.
    """
    rows: list[ColumnarCatalogRow] = []
    for family in _COLUMNAR_FAMILIES:
        rows.append((family.header_text, *family.noise))
        for intermediate in family.intermediates:
            rows.append(("", intermediate.header_text, "", "", "", COLUMNAR_NOT_APPLICABLE))
            for subgroup in intermediate.subgroups:
                rows.append(("", "", subgroup.header_text, "", "", COLUMNAR_NOT_APPLICABLE))
                for item in subgroup.items:
                    rows.append(
                        ("", "", item.code, item.description, item.unit_label, item.raw_price)
                    )
        rows.append(_BLANK_COLUMNAR_ROW)
    return tuple(rows)


def columnar_catalog_expectations() -> tuple[ExpectedColumnarEntry, ...]:
    """Gabarito do catálogo colunar, achatado da mesma árvore que gera a planilha."""
    expected: list[ExpectedColumnarEntry] = []
    for family in _COLUMNAR_FAMILIES:
        for intermediate in family.intermediates:
            for subgroup in intermediate.subgroups:
                for item in subgroup.items:
                    expected.append(
                        ExpectedColumnarEntry(
                            code=item.code,
                            description=item.description,
                            unit=item.unit,
                            unit_price=item.unit_price,
                            family_code=family.code,
                            family_name=family.name,
                            subgroup_code=subgroup.code,
                            subgroup_name=subgroup.name,
                        )
                    )
    return tuple(expected)


def columnar_catalog_template(
    *,
    sheet_name: str = COLUMNAR_CATALOG_SHEET,
    note_prefixes: Sequence[str] = (),
    unpriced_markers: Sequence[str] = (),
) -> WorkbookTemplate:
    """Template padrão com o catálogo lido no modo de hierarquia por colunas.

    `note_prefixes` e `unpriced_markers` são o que o catálogo publicado obriga a declarar:
    a nota editorial entre cabeçalhos e o item sem preço publicado.
    """
    template = default_template()
    layout = CatalogLayout(
        sheet_name=sheet_name,
        first_row=COLUMNAR_CATALOG_FIRST_ROW,
        family_column="A",
        subgroup_intermediate_column="B",
        code_column="C",
        description_column="D",
        unit_column="E",
        price_column="F",
        note_prefixes=list(note_prefixes),
        unpriced_markers=list(unpriced_markers),
    )
    return WorkbookTemplate.model_validate(
        {**template.model_dump(), "catalog": layout.model_dump()}
    )


def write_columnar_rows_workbook(
    path: Path,
    rows: Sequence[ColumnarCatalogRow],
    *,
    sheet_name: str = COLUMNAR_CATALOG_SHEET,
) -> Path:
    """Grava um catálogo com hierarquia por colunas a partir das linhas informadas.

    As três primeiras linhas são títulos da planilha, fora da área de dados declarada no
    template — é assim que o arquivo publicado começa.
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    for index, title in enumerate(COLUMNAR_CATALOG_TITLE_ROWS, start=1):
        worksheet.cell(row=index, column=1, value=title)
    for offset, row in enumerate(rows):
        for column, value in enumerate(row, start=1):
            if value == "":
                continue
            worksheet.cell(row=COLUMNAR_CATALOG_FIRST_ROW + offset, column=column, value=value)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def write_columnar_catalog_workbook(
    path: Path, *, sheet_name: str = COLUMNAR_CATALOG_SHEET
) -> Path:
    """Grava o catálogo colunar sintético completo, com títulos, ruído e cabeçalhos `NA`."""
    return write_columnar_rows_workbook(path, columnar_catalog_rows(), sheet_name=sheet_name)


def write_catalog_workbook(
    path: Path,
    rows: Sequence[CatalogRow],
    *,
    sheet_name: str = "CATALOGO",
    hidden: bool = False,
) -> Path:
    """Grava uma planilha de catálogo mínima com as linhas informadas."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.cell(row=1, column=1, value="CODIGO")
    worksheet.cell(row=1, column=2, value="DESCRICAO")
    worksheet.cell(row=1, column=3, value="UNIDADE")
    worksheet.cell(row=1, column=4, value="PRECO UNITARIO")
    for index, (code, description, unit, price) in enumerate(rows, start=2):
        worksheet.cell(row=index, column=1, value=code)
        worksheet.cell(row=index, column=2, value=description)
        if unit:
            worksheet.cell(row=index, column=3, value=unit)
        if price:
            worksheet.cell(row=index, column=4, value=price)
    if hidden:
        worksheet.sheet_state = "hidden"
        workbook.create_sheet("CAPA")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


ESTIMATE_GRID_SHEET_NAME = "PLANILHA ORÇAMENTÁRIA"
ESTIMATE_GRID_MEMORY_SHEET_NAME = "MEMÓRIA ORÇAMENTO"
ESTIMATE_GRID_REVISION_LABEL = "GABARITO SINTÉTICO REV. 0"
"""Identificação da revisão do gabarito sintético; o arquivo publicado a imprime."""


def estimate_grid_rows() -> tuple[EstimateTemplateRow, ...]:
    """Gabarito sintético com as três propriedades que importam no documento real.

    1. **Mais de um grupo** e **lacuna de grupo**: 01, 02 e 04 — não existe grupo 03, do
       mesmo jeito que o gabarito do cliente não tem os grupos 5, 15 e 22.
    2. **Numeração `GG.N` dentro do grupo**, preservada como texto (o escritor nunca a
       recomputa).
    3. **Mais linhas do que o orçamento preenche**: cinco linhas casam com os cinco códigos
       da demo sintética e as outras cinco saem zeradas — três com preço declarado no
       gabarito e duas sem preço nenhum.

    A ordem NÃO é a de `estimate.lines`: os códigos da demo aparecem espalhados, para que
    um escritor que voltasse ao cursor sequencial reprovasse na primeira asserção.
    """
    return (
        EstimateTemplateRow(
            group="01",
            item="01.1",
            code="SIN.PLACA.001",
            description="PLACA DE OBRA SINTETICA",
            unit="m2",
            unit_price=Decimal("612.50"),
        ),
        EstimateTemplateRow(
            group="01",
            item="01.2",
            code="EMOP.AD.001",
            description="PAVIMENTACAO SINTETICA EMOP TIPO 1",
            unit="M2",
        ),
        EstimateTemplateRow(
            group="01",
            item="01.3",
            code="SIN.TAPUME.001",
            description="TAPUME SINTETICO DE CANTEIRO",
            unit="m",
        ),
        EstimateTemplateRow(
            group="02",
            item="02.1",
            code="CE02100010(/)",
            description="ALAMBRADO SINTETICO TELA GALVANIZADA",
            unit="m2",
        ),
        EstimateTemplateRow(
            group="02",
            item="02.2",
            code="SIN.BANCO.001",
            description="BANCO SINTETICO DE CONCRETO",
            unit="UN",
            unit_price=Decimal("880.00"),
        ),
        EstimateTemplateRow(
            group="02",
            item="02.3",
            code="AD04050060(/)",
            description="PISO INTERTRAVADO SINTETICO 10CM",
            unit="m2",
        ),
        EstimateTemplateRow(
            group="04",
            item="04.1",
            code="COMP.GRAMADO.001",
            description="EXECUCAO DE GRAMADO SINTETICO EM PLACAS",
            unit="m2",
        ),
        EstimateTemplateRow(
            group="04",
            item="04.2",
            code="SIN.MUDA.001",
            description="MUDA SINTETICA DE ARVORE",
            unit="UN",
        ),
        EstimateTemplateRow(
            group="04",
            item="04.3",
            code="EMOP.CE.001",
            description="MOBILIARIO SINTETICO EMOP",
            unit="UN",
        ),
        EstimateTemplateRow(
            group="04",
            item="04.4",
            code="SIN.LIXEIRA.001",
            description="LIXEIRA SINTETICA DUPLA",
            unit="UN",
            unit_price=Decimal("445.30"),
        ),
    )


def estimate_grid_layout(
    *,
    rows: Sequence[EstimateTemplateRow] | None = None,
    revision_label: str = ESTIMATE_GRID_REVISION_LABEL,
) -> EstimateTemplateLayout:
    """Layout do gabarito sintético: oito colunas e o bloco de identificação acima delas.

    `header_row=7` porque o bloco tem quatro pares (INTERVENÇÃO, ENDEREÇO, BDI e a revisão
    do gabarito) e o escritor exige folga para eles.
    """
    return EstimateTemplateLayout(
        sheet_name=ESTIMATE_GRID_SHEET_NAME,
        title="PLANILHA ORÇAMENTÁRIA SINTÉTICA",
        revision_label=revision_label,
        memory_sheet_name=ESTIMATE_GRID_MEMORY_SHEET_NAME,
        header_row=7,
        columns=EstimateTemplateColumns(
            group=SheetColumn(letter="A", label="GRUPO", width=8),
            item=SheetColumn(letter="B", label="ITEM", width=8),
            code=SheetColumn(letter="C", label="CÓDIGO", width=18),
            description=SheetColumn(letter="D", label="ESPECIFICAÇÃO", width=56),
            unit=SheetColumn(letter="E", label="UN", width=8),
            quantity=SheetColumn(letter="F", label="QUANT", width=12),
            unit_price=SheetColumn(letter="G", label="VALOR UNIT", width=14),
            total=SheetColumn(letter="H", label="TOTAL", width=16),
        ),
        rows=list(estimate_grid_rows() if rows is None else rows),
    )


def estimate_grid_template(
    *,
    rows: Sequence[EstimateTemplateRow] | None = None,
    revision_label: str = ESTIMATE_GRID_REVISION_LABEL,
) -> WorkbookTemplate:
    """Template padrão acrescido do gabarito sintético; o resto do layout não muda."""
    template = default_template()
    payload = template.model_dump()
    payload["estimate_grid"] = estimate_grid_layout(
        rows=rows, revision_label=revision_label
    ).model_dump()
    return WorkbookTemplate.model_validate(payload)
