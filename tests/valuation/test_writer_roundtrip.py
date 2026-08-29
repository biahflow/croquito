"""A planilha gerada é reaberta, recomputada e conferida centavo a centavo."""

from __future__ import annotations

import re
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from croquito_valuation.canonical import audit_workbook, canonicalize_workbook
from croquito_valuation.contract import ContractWorkbook
from croquito_valuation.errors import ValuationValidationError
from croquito_valuation.models import (
    BulletinLine,
    CalcBlock,
    CalcOperand,
    CalcRecipe,
    PriceCatalog,
    Valuation,
)
from croquito_valuation.rounding import money_trunc
from croquito_valuation.template import WorkbookTemplate, default_template
from croquito_valuation.workbook_reader import read_contract_workbook
from croquito_valuation.workbook_writer import (
    PlannedCell,
    consolidate_by_code,
    plan_workbook,
    write_valuation_workbook,
)
from tests.valuation.builders import (
    PREVIOUS_MAPAO_CONTRACT_LABEL,
    PREVIOUS_MAPAO_SOURCE_LABEL,
    MeasuredItem,
    MeasuredWorksite,
    ValuationFixture,
    bake_formula_values,
    build_approval,
    build_catalog_for_contract,
    build_fixture,
    build_multi_worksite_valuation,
    build_previous_mapao_workbook,
    build_valuation_from_catalog,
    tamper_cell,
    template_without_amended_column,
    write_fixture_workbook,
)


def _cells(canonical: dict[str, object], sheet_name: str) -> dict[str, dict[str, object]]:
    sheets = canonical["sheets"]
    assert isinstance(sheets, list)
    for sheet in sheets:
        assert isinstance(sheet, dict)
        if sheet["name"] == sheet_name:
            cells = sheet["cells"]
            assert isinstance(cells, list)
            return {str(cell["ref"]): cell for cell in cells}
    raise AssertionError(f"aba ausente no canônico: {sheet_name}")


def _all_cells(canonical: dict[str, object]) -> list[dict[str, object]]:
    sheets = canonical["sheets"]
    assert isinstance(sheets, list)
    cells: list[dict[str, object]] = []
    for sheet in sheets:
        assert isinstance(sheet, dict)
        sheet_cells = sheet["cells"]
        assert isinstance(sheet_cells, list)
        cells.extend(sheet_cells)
    return cells


def _write(fixture: ValuationFixture, tmp_path: Path) -> Path:
    workbook_path = tmp_path / "medicao.xlsx"
    write_fixture_workbook(fixture, workbook_path)
    return workbook_path


def test_roundtrip_matches_the_valuation_cent_by_cent(tmp_path: Path) -> None:
    fixture = build_fixture(tmp_path)
    workbook_path = _write(fixture, tmp_path)

    report = audit_workbook(workbook_path, fixture.valuation, fixture.catalog, fixture.template)

    assert report.status == "ok"
    assert report.findings == []
    assert report.total_amount == fixture.valuation.total_amount

    canonical = canonicalize_workbook(workbook_path, fixture.template)
    bulletin_cells = _cells(canonical, report.worksites[0].bulletin_sheet)
    totals = [
        cell["value"]
        for ref, cell in bulletin_cells.items()
        if ref.startswith("G") and cell["kind"] in {"number", "formula"}
    ]
    for line in fixture.valuation.bulletins[0].lines:
        assert str(line.total) in totals
    assert str(fixture.valuation.total_amount) in totals
    assert "11.84" in totals, "o par 1,15 x 10,30 prova que dinheiro trunca"


def test_a_block_of_four_operands_fits_and_round_trips(tmp_path: Path) -> None:
    """A sapata do alambrado (`0,6 x 0,6 x 0,6 x 58 postes`) tem quatro operandos.

    Antes da T7 o escritor recusava com `MEMORY_BLOCK_TOO_WIDE`: só havia três colunas de
    operando. Agora imprime nas colunas C..F, com o subtotal recuado para I, e o auditor
    reabre e fecha.
    """
    fixture = build_fixture(tmp_path)
    quantity = fixture.valuation.calc_sheets[0].total_quantity
    wide_block = CalcBlock(
        label="SAPATA DO ALAMBRADO",
        recipe=CalcRecipe.DECLARED_PRODUCT,
        operands=[
            CalcOperand(name="QUANTIDADE", value=quantity, unit="m3"),
            CalcOperand(name="FATOR B", value=Decimal("1"), unit="m"),
            CalcOperand(name="FATOR C", value=Decimal("1"), unit="m"),
            CalcOperand(name="POSTES", value=Decimal("1"), unit="un"),
        ],
        subtotal=quantity,
    )
    payload = fixture.valuation.model_dump()
    payload["calc_sheets"][0]["blocks"] = [wide_block.model_dump()]
    valuation = Valuation.model_validate(payload)

    workbook_path = tmp_path / "quatro-operandos.xlsx"
    # Não levanta `MEMORY_BLOCK_TOO_WIDE`: o bloco de quatro operandos agora cabe.
    write_valuation_workbook(valuation, fixture.catalog, fixture.template, workbook_path)
    report = audit_workbook(workbook_path, valuation, fixture.catalog, fixture.template)
    assert report.status == "ok"
    assert report.findings == []

    canonical = canonicalize_workbook(workbook_path, fixture.template)
    memory = _cells(canonical, report.worksites[0].memory_sheet)
    # O subtotal do bloco é `=ROUND(PRODUCT(C..:F..),2)` na coluna I: os quatro operandos
    # ocupam C..F e o subtotal recuou para além deles.
    subtotal = next(
        (ref, cell)
        for ref, cell in memory.items()
        if cell["kind"] == "formula"
        and re.fullmatch(r"=ROUND\(PRODUCT\(C\d+:F\d+\),2\)", str(cell["formula"]))
    )
    assert subtotal[0].startswith("I")
    assert str(subtotal[1]["value"]) == str(quantity)


def test_emitted_formulas_stay_inside_the_closed_grammar(tmp_path: Path) -> None:
    fixture = build_fixture(tmp_path)
    workbook_path = _write(fixture, tmp_path)

    canonical = canonicalize_workbook(workbook_path, fixture.template)
    formulas = [str(cell["formula"]) for cell in _all_cells(canonical) if cell["kind"] == "formula"]

    assert any(formula.startswith("=TRUNC(") for formula in formulas)
    assert any(formula.startswith("=SUM(") for formula in formulas)
    assert any(formula.startswith("=ROUND(PRODUCT(") for formula in formulas)
    assert any("),2)-" in formula for formula in formulas)
    for formula in formulas:
        assert formula.startswith(("=TRUNC(", "=SUM(", "=ROUND(PRODUCT("))


def test_pinned_cell_is_written_as_a_literal_and_listed_in_the_audit(tmp_path: Path) -> None:
    fixture = build_fixture(tmp_path)
    workbook_path = tmp_path / "medicao.xlsx"
    write_report = write_fixture_workbook(fixture, workbook_path)

    assert write_report.pinned_cells, "a fixture precisa exercitar o ramo de célula fixada"
    report = audit_workbook(workbook_path, fixture.valuation, fixture.catalog, fixture.template)
    assert [pinned.ref for pinned in report.pinned_cells] == [
        pinned.ref for pinned in write_report.pinned_cells
    ]

    canonical = canonicalize_workbook(workbook_path, fixture.template)
    for pinned in write_report.pinned_cells:
        cell = _cells(canonical, pinned.sheet)[pinned.ref]
        assert cell["kind"] == "number"
        assert cell["value"] == str(pinned.value)
        assert pinned.reason == "TRUNC_DOUBLE_DIVERGENCE"


@pytest.mark.parametrize(
    "formula",
    [
        "=IF(F8>0,F8*E8,0)",
        "='BM X'!G12",
        "=H9-J9-K9",
        # `SUM` de argumento único fica fora da forma 5: a lista começa em duas refs.
        "=SUM(H9)",
    ],
)
def test_formula_outside_the_grammar_is_refused(tmp_path: Path, formula: str) -> None:
    fixture = build_fixture(tmp_path)
    workbook_path = _write(fixture, tmp_path)

    workbook = load_workbook(workbook_path)
    worksheet = workbook[fixture.template.bulletin_sheet_name("PRACA SINTETICA NORTE")]
    worksheet["G8"] = formula
    workbook.save(workbook_path)

    with pytest.raises(ValuationValidationError) as raised:
        canonicalize_workbook(workbook_path, fixture.template)

    assert raised.value.code == "FORMULA_UNSUPPORTED"
    assert raised.value.details["ref"] == "G8"


def test_audit_detects_a_tampered_quantity(tmp_path: Path) -> None:
    fixture = build_fixture(tmp_path)
    workbook_path = _write(fixture, tmp_path)

    workbook = load_workbook(workbook_path)
    worksheet = workbook[fixture.template.bulletin_sheet_name("PRACA SINTETICA NORTE")]
    worksheet["F8"] = Decimal("999.00")
    workbook.save(workbook_path)

    report = audit_workbook(workbook_path, fixture.valuation, fixture.catalog, fixture.template)

    assert report.status == "divergent"
    codes = {finding.code for finding in report.findings}
    assert "CELL_VALUE_MISMATCH" in codes
    assert {finding.ref for finding in report.findings} >= {"F8", "G8", "G16"}


def test_writer_refuses_a_price_that_is_not_in_the_catalog(tmp_path: Path) -> None:
    fixture = build_fixture(tmp_path)
    lines = list(fixture.valuation.bulletins[0].lines)
    tampered = BulletinLine(
        item_number=lines[0].item_number,
        code=lines[0].code,
        description=lines[0].description,
        unit=lines[0].unit,
        unit_price=Decimal("99.99"),
        quantity=lines[0].quantity,
        total=Decimal("12598.74"),
    )
    payload = fixture.valuation.model_dump()
    payload["bulletins"][0]["lines"][0] = tampered.model_dump()
    payload["bulletins"][0]["total_amount"] = sum(
        (line["total"] for line in payload["bulletins"][0]["lines"]), Decimal("0.00")
    )
    valuation = Valuation.model_validate(payload)

    with pytest.raises(ValuationValidationError) as raised:
        write_fixture_workbook(
            ValuationFixture(
                catalog_path=fixture.catalog_path,
                template=fixture.template,
                catalog=fixture.catalog,
                valuation=valuation,
            ),
            tmp_path / "divergente.xlsx",
        )

    assert raised.value.code == "LINE_PRICE_NOT_IN_CATALOG"


def test_the_writer_refuses_a_catalog_whose_origin_is_not_sco(tmp_path: Path) -> None:
    """Segunda linha de defesa do guardrail da licitada: o escritor recebe catálogo
    próprio na hora de gravar a pasta, e ele também nunca aceita origem fora do SCO."""
    fixture = build_fixture(tmp_path)
    payload = fixture.catalog.model_dump()
    payload["origin"] = "emop"
    for entry in payload["entries"]:
        entry["origin"] = "emop"
    emop_catalog = PriceCatalog.model_validate(payload)

    with pytest.raises(ValuationValidationError) as raised:
        plan_workbook(fixture.valuation, emop_catalog, fixture.template)

    assert raised.value.code == "BULLETIN_PRICE_ORIGIN_FORBIDDEN"


@pytest.mark.parametrize("origin", ["sinapi", "sicro"])
def test_the_writer_refuses_a_catalog_whose_origin_is_sinapi_or_sicro(
    tmp_path: Path, origin: str
) -> None:
    """As origens novas do ADR-0039 (F-026) caem na mesma recusa: o escritor da
    medição licitada nunca aceita catálogo fora do SCO, nem os recém-chegados."""
    fixture = build_fixture(tmp_path)
    payload = fixture.catalog.model_dump()
    payload["origin"] = origin
    for entry in payload["entries"]:
        entry["origin"] = origin
    non_sco_catalog = PriceCatalog.model_validate(payload)

    with pytest.raises(ValuationValidationError) as raised:
        plan_workbook(fixture.valuation, non_sco_catalog, fixture.template)

    assert raised.value.code == "BULLETIN_PRICE_ORIGIN_FORBIDDEN"


@dataclass(frozen=True, slots=True)
class _MultiWorksiteFixture:
    """Consolidado importado, catálogo do contrato e a medição de três obras."""

    template: WorkbookTemplate
    contract: ContractWorkbook
    catalog: PriceCatalog
    valuation: Valuation


def _multi_worksite(tmp_path: Path) -> _MultiWorksiteFixture:
    template = default_template()
    previous = build_previous_mapao_workbook(tmp_path / "mapao-anterior.xlsx")
    contract = read_contract_workbook(
        previous.workbook_path,
        template,
        source_label=PREVIOUS_MAPAO_SOURCE_LABEL,
        contract_label=PREVIOUS_MAPAO_CONTRACT_LABEL,
    )
    catalog = build_catalog_for_contract(tmp_path / "catalogo-contrato.xlsx", contract)
    return _MultiWorksiteFixture(
        template=template,
        contract=contract,
        catalog=catalog,
        valuation=build_multi_worksite_valuation(contract),
    )


def _write_multi(fixture: _MultiWorksiteFixture, tmp_path: Path) -> Path:
    output = tmp_path / "medicao-consolidada.xlsx"
    write_valuation_workbook(
        fixture.valuation,
        fixture.catalog,
        fixture.template,
        output,
        contract=fixture.contract,
    )
    return output


def _planned_cell(
    fixture: _MultiWorksiteFixture, role: str, item_number: str | None = None
) -> PlannedCell:
    plan = plan_workbook(fixture.valuation, fixture.catalog, fixture.template, fixture.contract)
    for sheet in plan.sheets:
        if sheet.name != fixture.template.general.sheet_name:
            continue
        for cell in sheet.cells:
            if cell.role == role and (item_number is None or cell.item_number == item_number):
                return cell
    raise AssertionError(f"célula ausente do plano da geral: {role}/{item_number}")


# Código: (vigente, acumulado depois desta medição, saldo depois desta medição).
_AFTER_MEASUREMENT: dict[str, tuple[str, str, str]] = {
    "AD04050050(/)": ("16.00", "10.00", "6.00"),
    "AD04050055(A)": ("10.00", "9.99", "0.01"),
    "AD04050055(B)": ("6.00", "3.00", "3.00"),
    "SP01050010(/)": ("8.00", "3.75", "4.25"),
    "SP01050015(A)": ("0.00", "0.00", "0.00"),
    "MB01100010(/)": ("6.00", "4.00", "2.00"),
}


def test_multi_worksite_workbook_carries_general_amendments_and_every_pair(
    tmp_path: Path,
) -> None:
    fixture = _multi_worksite(tmp_path)
    output = _write_multi(fixture, tmp_path)

    report = audit_workbook(
        output,
        fixture.valuation,
        fixture.catalog,
        fixture.template,
        fixture.contract,
    )

    assert report.status == "ok"
    assert report.findings == []
    assert report.total_amount == fixture.valuation.total_amount == Decimal("117.50")
    assert report.general_sheet == fixture.template.general.sheet_name
    assert fixture.template.amendment is not None
    assert report.amendment_sheet == fixture.template.amendment.sheet_name
    assert [worksite.worksite_key for worksite in report.worksites] == [
        bulletin.worksite_key for bulletin in fixture.valuation.bulletins
    ]

    workbook = load_workbook(output)
    try:
        names = list(workbook.sheetnames)
    finally:
        workbook.close()
    assert names[0] == fixture.template.general.sheet_name
    assert names[1] == fixture.template.amendment.sheet_name
    for worksite in report.worksites:
        assert worksite.bulletin_sheet in names
        assert worksite.memory_sheet in names

    # O portão de exportação vê a mesma medição fechar contra o consolidado.
    assert build_approval(fixture.valuation).export_errors(fixture.contract) == []


def test_the_six_accepted_forms_show_up_in_the_consolidated_workbook(tmp_path: Path) -> None:
    fixture = _multi_worksite(tmp_path)
    output = _write_multi(fixture, tmp_path)

    canonical = canonicalize_workbook(output, fixture.template)
    formulas = [str(cell["formula"]) for cell in _all_cells(canonical) if cell["kind"] == "formula"]

    ref = r"[A-Z]+\d+"
    shapes = (
        rf"=TRUNC\({ref}\*{ref},2\)",
        rf"=ROUND\(PRODUCT\({ref}:{ref}\),2\)",
        rf"=ROUND\(PRODUCT\({ref}:{ref}\),2\)-{ref}",
        rf"=SUM\({ref}:{ref}\)",
        rf"=SUM\({ref}(,{ref})+\)",
        rf"={ref}-{ref}",
    )
    for shape in shapes:
        assert any(re.fullmatch(shape, formula) for formula in formulas), shape
    for formula in formulas:
        assert any(re.fullmatch(shape, formula) for shape in shapes), formula


def test_the_generated_general_can_be_imported_back_as_the_next_consolidation(
    tmp_path: Path,
) -> None:
    fixture = _multi_worksite(tmp_path)
    output = _write_multi(fixture, tmp_path)
    recalculated = bake_formula_values(
        output, tmp_path / "medicao-recalculada.xlsx", fixture.template
    )

    reimported = read_contract_workbook(
        recalculated,
        fixture.template,
        source_label="MAPÃO GERADO PELA MEDIÇÃO",
        contract_label=PREVIOUS_MAPAO_CONTRACT_LABEL,
    )

    assert reimported.period_numbers == [*fixture.contract.period_numbers, 3]
    assert [line.code for line in reimported.lines] == [
        line.code for line in fixture.contract.lines
    ]
    for line in reimported.lines:
        amended, accumulated, balance = _AFTER_MEASUREMENT[line.code]
        assert line.amended_quantity == Decimal(amended), line.code
        assert line.accumulated_quantity == Decimal(accumulated), line.code
        assert line.balance_quantity == Decimal(balance), line.code
        assert len(line.periods) == reimported.period_count
    assert reimported.amendments == fixture.contract.amendments


def test_a_tampered_general_quantity_is_caught_with_its_cascade(tmp_path: Path) -> None:
    fixture = _multi_worksite(tmp_path)
    output = _write_multi(fixture, tmp_path)
    quantity = _planned_cell(fixture, "general_current_quantity", "1")

    tamper_cell(output, fixture.template.general.sheet_name, quantity.ref, Decimal("999.00"))

    report = audit_workbook(
        output,
        fixture.valuation,
        fixture.catalog,
        fixture.template,
        fixture.contract,
    )

    assert report.status == "divergent"
    assert {finding.code for finding in report.findings} == {"CELL_VALUE_MISMATCH"}
    divergent = {finding.ref for finding in report.findings}
    cascade = {
        quantity.ref,
        _planned_cell(fixture, "general_current_amount", "1").ref,
        _planned_cell(fixture, "general_accumulated_quantity", "1").ref,
        _planned_cell(fixture, "general_accumulated_amount", "1").ref,
        _planned_cell(fixture, "general_balance", "1").ref,
        _planned_cell(fixture, "general_total").ref,
    }
    assert cascade <= divergent


def _drifting_valuation(fixture: _MultiWorksiteFixture) -> Valuation:
    """Duas obras medindo o MESMO código, com o truncamento por linha perdendo um centavo.

    1,15 e 2,15 a 12,50: 14,37 + 26,87 = 41,24 linha a linha e 41,25 consolidado.
    """
    return build_valuation_from_catalog(
        fixture.catalog,
        fixture.contract.next_period_number,
        [
            MeasuredWorksite(
                worksite_key="praca-sintetica-norte",
                worksite_name="PRACA SINTETICA NORTE",
                items=(MeasuredItem(code="AD04050055(A)", quantity=Decimal("1.15")),),
            ),
            MeasuredWorksite(
                worksite_key="praca-sintetica-sul",
                worksite_name="PRACA SINTETICA SUL",
                items=(MeasuredItem(code="AD04050055(A)", quantity=Decimal("2.15")),),
            ),
        ],
    )


def test_consolidation_drift_of_one_cent_is_declared_and_the_workbook_is_generated(
    tmp_path: Path,
) -> None:
    """ADR-0062 completa a decisão (c) do ADR-0018: a deriva de centavo não recusa mais a
    pasta. O valor da GERAL (`TRUNC(Σq x p)`) governa e a pasta é gerada; a diferença contra
    a soma dos boletins (`Σ TRUNC(qᵢ x p)`) vira `ConsolidationDrift` declarado no plano, no
    relatório de gravação e na auditoria de round-trip — sem ajustar a linha de nenhum
    boletim, que continua truncando só o que ela mede.
    """
    fixture = _multi_worksite(tmp_path)
    drifting = _drifting_valuation(fixture)

    plan = plan_workbook(drifting, fixture.catalog, fixture.template, fixture.contract)

    assert [drift.code for drift in plan.consolidation_drifts] == ["AD04050055(A)"]
    drift = plan.consolidation_drifts[0]
    assert drift.reason == "TRUNC_CONSOLIDATION_DRIFT"
    assert drift.quantity == Decimal("3.30")
    assert drift.general == Decimal("41.25")
    assert drift.bulletins == Decimal("41.24")
    assert drift.difference == Decimal("0.01")

    # A linha de cada boletim continua com o próprio truncamento, sem ajuste.
    for bulletin in drifting.bulletins:
        for line in bulletin.lines:
            if line.code == "AD04050055(A)":
                assert line.total == money_trunc(line.quantity * line.unit_price)

    output = tmp_path / "medicao-com-deriva.xlsx"
    write_report = write_valuation_workbook(
        drifting, fixture.catalog, fixture.template, output, contract=fixture.contract
    )
    assert output.is_file()
    assert write_report.consolidation_drifts == plan.consolidation_drifts

    audit = audit_workbook(output, drifting, fixture.catalog, fixture.template, fixture.contract)
    assert audit.status == "ok"
    assert audit.findings == []
    assert audit.consolidation_drifts == plan.consolidation_drifts


def _assert_consolidacao_bate_com_a_geral(fixture: _MultiWorksiteFixture, output: Path) -> None:
    """A consolidação servida e a coluna corrente da GERAL, célula a célula.

    A auditoria de round-trip entra ANTES da comparação porque é ela que faz do plano um
    oráculo do ARQUIVO: laudo aprovado significa que toda célula planejada está no `.xlsx`
    reaberto com o valor planejado. Sem esse elo, comparar com o plano provaria só que duas
    chamadas da mesma função concordam.
    """
    audit = audit_workbook(
        output, fixture.valuation, fixture.catalog, fixture.template, fixture.contract
    )
    assert audit.status == "ok"
    assert audit.findings == []

    consolidacao = {
        item.code: item
        for item in consolidate_by_code(
            fixture.valuation,
            unit_prices={line.code: line.unit_price for line in fixture.contract.lines},
        )
    }
    assert consolidacao, "medição sem código consolidado não provaria nada"

    plan = plan_workbook(fixture.valuation, fixture.catalog, fixture.template, fixture.contract)
    geral = next(
        sheet for sheet in plan.sheets if sheet.name == fixture.template.general.sheet_name
    )
    impresso = {(cell.role, cell.item_number): cell for cell in geral.cells}
    conferidos = 0
    for line in fixture.contract.lines:
        item = consolidacao.get(line.code)
        if item is None:
            continue
        assert impresso[("general_current_quantity", line.item_number)].number == item.quantity
        assert impresso[("general_current_amount", line.item_number)].number == item.amount
        conferidos += 1
    assert conferidos == len(consolidacao), "código medido que a GERAL não imprime"

    # E a deriva declarada é a mesma lista, pelos mesmos dois valores.
    assert [
        (drift.code, drift.general, drift.bulletins, drift.difference)
        for drift in plan.consolidation_drifts
    ] == [
        (item.code, item.amount, item.bulletins_amount, item.difference)
        for item in consolidacao.values()
        if item.has_drift
    ]


def test_a_consolidacao_por_codigo_e_o_mesmo_numero_que_a_pasta_gravada_imprime(
    tmp_path: Path,
) -> None:
    """F-046 T4e, critério 2: as duas derivações não podem divergir — prova, não promessa.

    A consolidação por código que a `/v1` serve à praça sai de `consolidate_by_code`, a mesma
    função que planeja a coluna corrente da PLANILHA GERAL. Este teste amarra as duas pontas
    no artefato: a pasta é gravada com o consolidado contratual, reaberta e auditada, e cada
    par (quantidade, valor) da consolidação é confrontado com a célula que a GERAL imprime.

    O caminho COM deriva entra junto de propósito: é justamente quando `TRUNC(Σq x p)` deixa
    de ser `Σ TRUNC(qᵢ x p)` que uma segunda derivação silenciosa apareceria como um centavo
    de diferença entre o que a tela mostra e o que a prefeitura lê.
    """
    fixture = _multi_worksite(tmp_path)
    _assert_consolidacao_bate_com_a_geral(fixture, _write_multi(fixture, tmp_path))

    com_deriva = _MultiWorksiteFixture(
        template=fixture.template,
        contract=fixture.contract,
        catalog=fixture.catalog,
        valuation=_drifting_valuation(fixture),
    )
    saida = tmp_path / "medicao-com-deriva-servida.xlsx"
    write_valuation_workbook(
        com_deriva.valuation,
        com_deriva.catalog,
        com_deriva.template,
        saida,
        contract=com_deriva.contract,
    )
    _assert_consolidacao_bate_com_a_geral(com_deriva, saida)
    # E a rodada com deriva tem MESMO deriva: um teste que passasse sem ela não provaria nada.
    assert [
        item.code
        for item in consolidate_by_code(
            com_deriva.valuation,
            unit_prices={line.code: line.unit_price for line in com_deriva.contract.lines},
        )
        if item.has_drift
    ] == ["AD04050055(A)"]


def test_a_code_outside_the_consolidation_refuses_the_workbook(tmp_path: Path) -> None:
    template = default_template()
    previous = build_previous_mapao_workbook(tmp_path / "mapao-anterior.xlsx")
    contract = read_contract_workbook(
        previous.workbook_path,
        template,
        source_label=PREVIOUS_MAPAO_SOURCE_LABEL,
        contract_label=PREVIOUS_MAPAO_CONTRACT_LABEL,
    )
    outside = "CE02100010(/)"
    catalog = build_catalog_for_contract(
        tmp_path / "catalogo-contrato.xlsx",
        contract,
        extra_rows=(
            ("CE", "CERCAMENTO SINTETICO", "", ""),
            ("CE0210", "ALAMBRADO SINTETICO", "", ""),
            (outside, "ALAMBRADO SINTETICO TELA GALVANIZADA", "m2", "128.35"),
        ),
    )
    valuation = build_valuation_from_catalog(
        catalog,
        contract.next_period_number,
        [
            MeasuredWorksite(
                worksite_key="praca-sintetica-norte",
                worksite_name="PRACA SINTETICA NORTE",
                items=(MeasuredItem(code=outside, quantity=Decimal("1.00")),),
            )
        ],
    )

    with pytest.raises(ValuationValidationError) as raised:
        plan_workbook(valuation, catalog, template, contract)

    assert raised.value.code == "GENERAL_CONSOLIDATION_MISMATCH"
    assert raised.value.details["reason"] == "CODE_NOT_IN_CONTRACT"
    assert raised.value.details["codes"] == [outside]


def test_a_template_without_the_amended_column_refuses_to_plan_the_general(
    tmp_path: Path,
) -> None:
    """A GERAL gerada declara o vigente; sem coluna para ele, o escritor recusa fechado."""
    fixture = _multi_worksite(tmp_path)
    template = template_without_amended_column()

    with pytest.raises(ValuationValidationError) as raised:
        plan_workbook(fixture.valuation, fixture.catalog, template, fixture.contract)

    assert raised.value.code == "PLAN_GENERAL_NOT_REPRESENTABLE"
    assert raised.value.details["reason"] == "AMENDED_QUANTITY_COLUMN_MISSING"


def test_two_sheets_with_the_same_name_refuse_the_plan(tmp_path: Path) -> None:
    fixture = _multi_worksite(tmp_path)
    collided = fixture.template.bulletin_sheet_name("PRACA SINTETICA NORTE")
    payload = fixture.template.model_dump()
    payload["general"]["sheet_name"] = collided
    template = WorkbookTemplate.model_validate(payload)

    with pytest.raises(ValuationValidationError) as raised:
        plan_workbook(fixture.valuation, fixture.catalog, template, fixture.contract)

    assert raised.value.code == "PLAN_SHEET_NAME_COLLISION"
    assert raised.value.details["names"] == [collided]


def test_two_worksites_that_shorten_to_the_same_label_refuse_the_plan(tmp_path: Path) -> None:
    """Nome encurtado que colide não vira uma aba só: as duas obras somariam em silêncio.

    Duas obras que só diferem na partícula caem no mesmo rótulo quando a forma curta entra
    (`Praça do Sol` e `Praça de Sol`), e é exatamente aí que o cheque de colisão do plano
    precisa continuar valendo.
    """
    fixture = _multi_worksite(tmp_path)
    colididos = ["Praça do Sol Poente Azul", "Praça de Sol Poente Azul"]
    renamed = [
        bulletin.model_copy(update={"worksite_name": colididos[index]})
        if index < len(colididos)
        else bulletin
        for index, bulletin in enumerate(fixture.valuation.bulletins)
    ]
    valuation = fixture.valuation.model_copy(update={"bulletins": renamed})

    with pytest.raises(ValuationValidationError) as raised:
        plan_workbook(valuation, fixture.catalog, fixture.template, fixture.contract)

    assert raised.value.code == "PLAN_SHEET_NAME_COLLISION"
    assert raised.value.details["names"] == [
        "BM Praça Sol Poente Azul",
        "MEMÓRIA Praça Sol Poente Azul",
    ]


def test_the_first_measurement_of_a_contract_accumulates_a_single_pair(tmp_path: Path) -> None:
    fixture = _multi_worksite(tmp_path)
    payload = fixture.contract.model_dump()
    payload["period_numbers"] = []
    for line in payload["lines"]:
        line["periods"] = []
        line["accumulated_quantity"] = Decimal("0.00")
        line["accumulated_amount"] = Decimal("0.00")
        line["balance_quantity"] = line["amended_quantity"]
    contract = ContractWorkbook.model_validate(payload)
    valuation = build_multi_worksite_valuation(contract)
    output = tmp_path / "primeira-medicao.xlsx"

    write_valuation_workbook(
        valuation, fixture.catalog, fixture.template, output, contract=contract
    )
    report = audit_workbook(output, valuation, fixture.catalog, fixture.template, contract)

    assert valuation.period_number == 1
    assert report.status == "ok"
    assert report.findings == []
    plan = plan_workbook(valuation, fixture.catalog, fixture.template, contract)
    general = next(
        sheet for sheet in plan.sheets if sheet.name == fixture.template.general.sheet_name
    )
    accumulated = [
        cell.formula for cell in general.cells if cell.role == "general_accumulated_quantity"
    ]
    # Sem histórico existe um par só: a lista de uma ref sairia da gramática.
    assert accumulated == ["=SUM(I8:I8)", "=SUM(I9:I9)", "=SUM(I10:I10)"] + [
        f"=SUM(I{row}:I{row})" for row in (12, 13, 14)
    ]
