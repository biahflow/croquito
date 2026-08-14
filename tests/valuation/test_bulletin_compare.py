"""Domínio do `compare-bulletin`: leitura do BM real e comparação centavo a centavo.

O lado gerado vem de `build_fixture` (mesma fixture da demonstração do M1); o lado real
é um xlsx pequeno escrito com `openpyxl` no próprio teste, no layout do
`default_template()` — sem nenhum documento de cliente envolvido. O rótulo de total vai
na coluna de rótulo (`label_column`), como o BM real observado traz — diferente da
coluna de descrição que o gerador do sistema usa (`workbook_writer.py`).
"""

from __future__ import annotations

from collections.abc import Sequence
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook

from croquito_valuation.bulletin_compare import (
    BulletinLineDiff,
    ReferenceBulletin,
    compare_bulletin,
    read_bulletin_lines,
)
from croquito_valuation.errors import ValuationValidationError
from croquito_valuation.models import Valuation
from croquito_valuation.template import WorkbookTemplate
from tests.valuation.builders import ValuationFixture, build_fixture

SHEET_NAME = "BM PRACA SINTETICA NORTE"


@dataclass(frozen=True, slots=True)
class _RefLine:
    """Uma linha do BM real a escrever no xlsx do teste."""

    code: str
    description: str
    unit: str
    quantity: Decimal
    unit_price: Decimal
    total: Decimal


def _reference_lines_from_valuation(valuation: Valuation, worksite_key: str) -> list[_RefLine]:
    """Espelha o boletim gerado como linhas de referência, sem nenhuma divergência."""
    bulletin = next(item for item in valuation.bulletins if item.worksite_key == worksite_key)
    return [
        _RefLine(
            code=line.code,
            description=line.description,
            unit=line.unit,
            quantity=line.quantity,
            unit_price=line.unit_price,
            total=line.total,
        )
        for line in bulletin.lines
    ]


def _write_reference_bm(
    path: Path,
    template: WorkbookTemplate,
    lines: Sequence[_RefLine],
    total: Decimal,
    *,
    sheet_name: str = SHEET_NAME,
    separator_before_total: bool = False,
) -> Path:
    """Grava um BM real pequeno no layout do template, sem depender de documento real."""
    layout = template.bulletin
    columns = layout.columns
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet[f"{layout.label_column}1"] = layout.title
    row = layout.header_row + 1
    for line in lines:
        worksheet[f"{columns.code.letter}{row}"] = line.code
        worksheet[f"{columns.description.letter}{row}"] = line.description
        worksheet[f"{columns.unit.letter}{row}"] = line.unit
        worksheet[f"{columns.quantity.letter}{row}"] = line.quantity
        worksheet[f"{columns.unit_price.letter}{row}"] = line.unit_price
        worksheet[f"{columns.total.letter}{row}"] = line.total
        row += 1
    if separator_before_total:
        row += 1  # linha em branco: sem código, sem número — separadora
    worksheet[f"{layout.label_column}{row}"] = layout.total_label
    worksheet[f"{columns.total.letter}{row}"] = total
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def _fixture_and_worksite(tmp_path: Path) -> tuple[ValuationFixture, str]:
    fixture = build_fixture(tmp_path)
    worksite_key = fixture.valuation.bulletins[0].worksite_key
    return fixture, worksite_key


def _read_reference(path: Path, template: WorkbookTemplate) -> ReferenceBulletin:
    return read_bulletin_lines(path, template, sheet_name=SHEET_NAME)


def test_generated_matches_itself_is_zero_cent(tmp_path: Path) -> None:
    fixture, worksite_key = _fixture_and_worksite(tmp_path)
    bulletin = fixture.valuation.bulletins[0]
    lines = _reference_lines_from_valuation(fixture.valuation, worksite_key)
    path = _write_reference_bm(
        tmp_path / "bm-real.xlsx", fixture.template, lines, bulletin.total_amount
    )

    reference = _read_reference(path, fixture.template)
    report = compare_bulletin(fixture.valuation, worksite_key, reference)

    assert reference.skipped_rows == []
    assert report.zero_cent is True
    assert report.missing_in_reference == []
    assert report.missing_in_generated == []
    assert report.quantity_diffs == []
    assert report.unit_price_diffs == []
    assert report.line_total_diffs == []
    assert report.bulletin_total_diff is None
    assert report.unit_notes == []


def test_one_cent_more_on_a_line_total_is_a_line_total_diff(tmp_path: Path) -> None:
    fixture, worksite_key = _fixture_and_worksite(tmp_path)
    bulletin = fixture.valuation.bulletins[0]
    lines = _reference_lines_from_valuation(fixture.valuation, worksite_key)
    tampered_code = lines[0].code
    tampered = [
        line
        if line.code != tampered_code
        else _RefLine(
            code=line.code,
            description=line.description,
            unit=line.unit,
            quantity=line.quantity,
            unit_price=line.unit_price,
            total=line.total + Decimal("0.01"),
        )
        for line in lines
    ]
    # O total declarado no arquivo continua o mesmo do gerado: só a linha diverge.
    path = _write_reference_bm(
        tmp_path / "bm-real.xlsx", fixture.template, tampered, bulletin.total_amount
    )

    reference = _read_reference(path, fixture.template)
    report = compare_bulletin(fixture.valuation, worksite_key, reference)

    assert report.zero_cent is False
    generated_line = next(line for line in bulletin.lines if line.code == tampered_code)
    reference_line = next(item for item in reference.lines if item.code == tampered_code)
    assert report.line_total_diffs == [
        BulletinLineDiff(
            code=tampered_code,
            generated=str(generated_line.total),
            reference=str(generated_line.total + Decimal("0.01")),
            cell=reference_line.total_ref,
        )
    ]
    assert report.quantity_diffs == []
    assert report.unit_price_diffs == []
    assert report.bulletin_total_diff is None


def test_missing_code_on_each_side_is_reported(tmp_path: Path) -> None:
    fixture, worksite_key = _fixture_and_worksite(tmp_path)
    bulletin = fixture.valuation.bulletins[0]
    lines = _reference_lines_from_valuation(fixture.valuation, worksite_key)
    dropped_code = lines[0].code
    extra_code = "SP09090090(/)"
    kept = [line for line in lines if line.code != dropped_code]
    extra = _RefLine(
        code=extra_code,
        description="ITEM QUE NAO EXISTE NO GERADO",
        unit="un",
        quantity=Decimal("1.00"),
        unit_price=Decimal("1.00"),
        total=Decimal("1.00"),
    )
    path = _write_reference_bm(
        tmp_path / "bm-real.xlsx",
        fixture.template,
        [*kept, extra],
        bulletin.total_amount,
    )

    reference = _read_reference(path, fixture.template)
    report = compare_bulletin(fixture.valuation, worksite_key, reference)

    assert report.missing_in_reference == [dropped_code]
    assert report.missing_in_generated == [extra_code]
    assert report.zero_cent is False


def test_quantity_and_unit_price_diverge_in_their_own_classes(tmp_path: Path) -> None:
    fixture, worksite_key = _fixture_and_worksite(tmp_path)
    bulletin = fixture.valuation.bulletins[0]
    lines = _reference_lines_from_valuation(fixture.valuation, worksite_key)
    quantity_code = lines[0].code
    price_code = lines[1].code
    tampered = []
    for line in lines:
        if line.code == quantity_code:
            tampered.append(
                _RefLine(
                    code=line.code,
                    description=line.description,
                    unit=line.unit,
                    quantity=line.quantity + Decimal("1.00"),
                    unit_price=line.unit_price,
                    total=line.total,
                )
            )
        elif line.code == price_code:
            tampered.append(
                _RefLine(
                    code=line.code,
                    description=line.description,
                    unit=line.unit,
                    quantity=line.quantity,
                    unit_price=line.unit_price + Decimal("1.00"),
                    total=line.total,
                )
            )
        else:
            tampered.append(line)
    path = _write_reference_bm(
        tmp_path / "bm-real.xlsx", fixture.template, tampered, bulletin.total_amount
    )

    reference = _read_reference(path, fixture.template)
    report = compare_bulletin(fixture.valuation, worksite_key, reference)

    assert [diff.code for diff in report.quantity_diffs] == [quantity_code]
    assert [diff.code for diff in report.unit_price_diffs] == [price_code]
    assert report.line_total_diffs == []
    assert report.zero_cent is False


def test_unit_mismatch_with_equal_numbers_is_a_note_not_a_diff(tmp_path: Path) -> None:
    fixture, worksite_key = _fixture_and_worksite(tmp_path)
    bulletin = fixture.valuation.bulletins[0]
    lines = _reference_lines_from_valuation(fixture.valuation, worksite_key)
    noted_code = lines[0].code
    tampered = [
        line
        if line.code != noted_code
        else _RefLine(
            code=line.code,
            description=line.description,
            unit="m2 " + line.unit,
            quantity=line.quantity,
            unit_price=line.unit_price,
            total=line.total,
        )
        for line in lines
    ]
    path = _write_reference_bm(
        tmp_path / "bm-real.xlsx", fixture.template, tampered, bulletin.total_amount
    )

    reference = _read_reference(path, fixture.template)
    report = compare_bulletin(fixture.valuation, worksite_key, reference)

    assert [note.code for note in report.unit_notes] == [noted_code]
    assert report.quantity_diffs == []
    assert report.unit_price_diffs == []
    assert report.line_total_diffs == []
    # Nota de unidade não é diff de centavo: o resto bate, então zero_cent continua True.
    assert report.zero_cent is True


def test_separator_row_is_counted_and_duplicate_code_refuses(tmp_path: Path) -> None:
    fixture, worksite_key = _fixture_and_worksite(tmp_path)
    bulletin = fixture.valuation.bulletins[0]
    lines = _reference_lines_from_valuation(fixture.valuation, worksite_key)

    separator_path = _write_reference_bm(
        tmp_path / "bm-real-separator.xlsx",
        fixture.template,
        lines,
        bulletin.total_amount,
        separator_before_total=True,
    )
    reference = _read_reference(separator_path, fixture.template)
    assert len(reference.skipped_rows) == 1

    duplicate_path = tmp_path / "bm-real-duplicate.xlsx"
    _write_reference_bm(duplicate_path, fixture.template, [*lines, lines[0]], bulletin.total_amount)
    with pytest.raises(ValuationValidationError) as excinfo:
        _read_reference(duplicate_path, fixture.template)
    assert excinfo.value.code == "COMPARE_DUPLICATE_CODE"


def test_duplicate_code_in_generated_bulletin_refuses(tmp_path: Path) -> None:
    """O modelo permite o mesmo código em item_numbers distintos; indexar colapsaria um.

    `model_copy` não revalida de propósito: o objetivo é exatamente um boletim que o
    domínio aceita (códigos repetidos com item_number único) chegando ao comparador.
    """
    fixture, worksite_key = _fixture_and_worksite(tmp_path)
    bulletin = fixture.valuation.bulletins[0]
    lines = _reference_lines_from_valuation(fixture.valuation, worksite_key)
    reference = _read_reference(
        _write_reference_bm(
            tmp_path / "bm-real-ok.xlsx", fixture.template, lines, bulletin.total_amount
        ),
        fixture.template,
    )

    repeated = bulletin.lines[0].model_copy(update={"item_number": "99"})
    doubled = bulletin.model_copy(
        update={
            "lines": [*bulletin.lines, repeated],
            "total_amount": bulletin.total_amount + repeated.total,
        }
    )
    valuation = fixture.valuation.model_copy(
        update={
            "bulletins": [
                doubled if item.worksite_key == worksite_key else item
                for item in fixture.valuation.bulletins
            ]
        }
    )
    with pytest.raises(ValuationValidationError) as excinfo:
        compare_bulletin(valuation, worksite_key, reference)
    assert excinfo.value.code == "COMPARE_DUPLICATE_CODE"
    assert excinfo.value.details["codes"] == [bulletin.lines[0].code]


def test_missing_sheet_and_missing_worksite_key_refuse(tmp_path: Path) -> None:
    fixture, worksite_key = _fixture_and_worksite(tmp_path)
    bulletin = fixture.valuation.bulletins[0]
    lines = _reference_lines_from_valuation(fixture.valuation, worksite_key)
    path = _write_reference_bm(
        tmp_path / "bm-real.xlsx", fixture.template, lines, bulletin.total_amount
    )

    with pytest.raises(ValuationValidationError) as missing_sheet:
        read_bulletin_lines(path, fixture.template, sheet_name="ABA QUE NAO EXISTE")
    assert missing_sheet.value.code == "BULLETIN_SHEET_MISSING"

    reference = _read_reference(path, fixture.template)
    with pytest.raises(ValuationValidationError) as missing_worksite:
        compare_bulletin(fixture.valuation, "obra-que-nao-existe", reference)
    assert missing_worksite.value.code == "COMPARE_WORKSITE_NOT_FOUND"


def test_row_without_code_or_missing_total_row_are_closed_refusals(tmp_path: Path) -> None:
    fixture, worksite_key = _fixture_and_worksite(tmp_path)
    bulletin = fixture.valuation.bulletins[0]
    lines = _reference_lines_from_valuation(fixture.valuation, worksite_key)
    layout = fixture.template.bulletin
    columns = layout.columns

    # Linha sem código mas com número: não é separadora, e não é item legível.
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_NAME
    row = layout.header_row + 1
    for line in lines:
        worksheet[f"{columns.code.letter}{row}"] = line.code
        worksheet[f"{columns.description.letter}{row}"] = line.description
        worksheet[f"{columns.unit.letter}{row}"] = line.unit
        worksheet[f"{columns.quantity.letter}{row}"] = line.quantity
        worksheet[f"{columns.unit_price.letter}{row}"] = line.unit_price
        worksheet[f"{columns.total.letter}{row}"] = line.total
        row += 1
    worksheet[f"{columns.quantity.letter}{row}"] = Decimal("1.00")  # sem código
    row += 1
    worksheet[f"{layout.label_column}{row}"] = layout.total_label
    worksheet[f"{columns.total.letter}{row}"] = bulletin.total_amount
    unparseable_path = tmp_path / "bm-real-unparseable.xlsx"
    workbook.save(unparseable_path)
    with pytest.raises(ValuationValidationError) as unparseable:
        read_bulletin_lines(unparseable_path, fixture.template, sheet_name=SHEET_NAME)
    assert unparseable.value.code == "BULLETIN_ROW_UNPARSEABLE"

    # Planilha sem nenhuma linha de total: a leitura não tem o que devolver.
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_NAME
    worksheet[f"{layout.label_column}1"] = layout.title
    no_total_path = tmp_path / "bm-real-no-total.xlsx"
    workbook.save(no_total_path)
    with pytest.raises(ValuationValidationError) as no_total:
        read_bulletin_lines(no_total_path, fixture.template, sheet_name=SHEET_NAME)
    assert no_total.value.code == "BULLETIN_TOTAL_ROW_MISSING"
