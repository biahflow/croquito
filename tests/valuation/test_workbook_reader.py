"""Importação do consolidado contratual: o leitor lê, o modelo recusa o drift.

Drift semântico não sobe mais como violação isolada: a recusa é
`CONTRACT_SEMANTICS_DIVERGENT` com o dossiê do recomputo pendurado nela, e é lá que o
código estável de cada divergência aparece (`tests/valuation/test_contract_diagnosis.py`
cobre o dossiê em si). Falha de layout continua recusando sem dossiê.
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from croquitodxf_valuation.catalog import file_sha256
from croquitodxf_valuation.contract import ContractWorkbook, contract_workbook_id_for
from croquitodxf_valuation.contract_diagnosis import (
    ContractDiagnosis,
    ContractSemanticsError,
    SemanticFinding,
)
from croquitodxf_valuation.errors import ValuationValidationError
from croquitodxf_valuation.template import WorkbookTemplate, default_template
from croquitodxf_valuation.workbook_reader import (
    ContractImportNotes,
    read_contract_workbook,
    read_contract_workbook_with_notes,
)
from tests.valuation.builders import (
    EXTRA_CODE_SAMPLE,
    PREVIOUS_MAPAO_CONTRACT_LABEL,
    PREVIOUS_MAPAO_PERIOD_NUMBERS,
    PREVIOUS_MAPAO_SOURCE_LABEL,
    REAL_LAYOUT_DUPLICATED_CODE,
    REAL_LAYOUT_PERIOD_NUMBERS,
    REAL_LAYOUT_PERIOD_SUFFIXES,
    PreviousMapaoFixture,
    amendment_code_rows,
    build_previous_mapao_workbook,
    extra_code_items,
    real_layout_items,
    real_layout_template,
    tamper_cell,
    template_with_extra_code_pattern,
    template_with_section_group_subtotal,
    template_without_amended_column,
    write_previous_mapao_workbook,
)

_GENERAL_SHEET = "PLANILHA GERAL"
_AMENDMENT_SHEET = "MAPÃO - PREFEITURA"
_FIRST_ITEM_ROW = 8
_SECOND_ITEM_ROW = 9
_FIRST_AMENDMENT_ROW = 5
_REDUCED_CODE = "AD04050050(/)"
_NEW_ITEM_CODE = "MB01100010(/)"
_PAVING_GROUP = "PAVIMENTACAO SINTETICA"
_SERVICES_GROUP = "SERVICOS PRELIMINARES E MOBILIARIO"


def _read(path: Path, template: WorkbookTemplate | None = None) -> ContractWorkbook:
    return read_contract_workbook(
        path,
        default_template() if template is None else template,
        source_label=PREVIOUS_MAPAO_SOURCE_LABEL,
        contract_label=PREVIOUS_MAPAO_CONTRACT_LABEL,
    )


def _read_with_notes(
    path: Path, template: WorkbookTemplate | None = None
) -> tuple[ContractWorkbook, ContractImportNotes]:
    return read_contract_workbook_with_notes(
        path,
        default_template() if template is None else template,
        source_label=PREVIOUS_MAPAO_SOURCE_LABEL,
        contract_label=PREVIOUS_MAPAO_CONTRACT_LABEL,
    )


def _without_amendment_layout(template: WorkbookTemplate | None = None) -> WorkbookTemplate:
    base = default_template() if template is None else template
    return WorkbookTemplate.model_validate({**base.model_dump(), "amendment": None})


def _real_layout_fixture(path: Path, *, separator_rows: bool = False) -> PreviousMapaoFixture:
    """MAPÃO no layout real: sem coluna de vigente, com salto de numeração e sem RE-RA."""
    return build_previous_mapao_workbook(
        path,
        with_amendments=False,
        template=real_layout_template(),
        items=real_layout_items(),
        period_numbers=REAL_LAYOUT_PERIOD_NUMBERS,
        period_label_suffixes=REAL_LAYOUT_PERIOD_SUFFIXES,
        separator_rows=separator_rows,
    )


def _diagnosis_of(path: Path, template: WorkbookTemplate | None = None) -> ContractDiagnosis:
    """Lê esperando recusa semântica e devolve o dossiê que veio com ela."""
    with pytest.raises(ContractSemanticsError) as raised:
        _read(path, template)
    assert raised.value.code == "CONTRACT_SEMANTICS_DIVERGENT"
    assert raised.value.details["finding_count"] == len(raised.value.diagnosis.findings)
    return raised.value.diagnosis


def _findings(diagnosis: ContractDiagnosis, finding_code: str) -> list[SemanticFinding]:
    return [finding for finding in diagnosis.findings if finding.finding_code == finding_code]


def _only(diagnosis: ContractDiagnosis, finding_code: str) -> SemanticFinding:
    found = _findings(diagnosis, finding_code)
    assert len(found) == 1, [finding.finding_code for finding in diagnosis.findings]
    return found[0]


def _rename_sheet(path: Path, sheet_name: str, new_name: str) -> None:
    workbook = load_workbook(path)
    workbook[sheet_name].title = new_name
    workbook.save(path)


def _delete_rows(path: Path, sheet_name: str, first_row: int, amount: int) -> None:
    workbook = load_workbook(path)
    workbook[sheet_name].delete_rows(first_row, amount)
    workbook.save(path)


def test_reads_the_previous_mapao_exactly_as_the_fixture_declares(tmp_path: Path) -> None:
    fixture = build_previous_mapao_workbook(tmp_path / "mapao-anterior.xlsx")

    workbook = _read(fixture.workbook_path)

    assert workbook.model_dump() == fixture.expected.model_dump()
    assert workbook == fixture.expected
    assert workbook.period_count == 2
    assert [line.item_number for line in workbook.lines] == ["1", "2", "3", "4", "5", "6"]
    assert workbook.line_for_code("AD04050055(A)").balance_quantity == Decimal("0.01")
    assert workbook.line_for_code("AD04050055(A)").periods[1].amount == Decimal("74.87")


def test_the_consolidated_id_comes_from_the_file_content(tmp_path: Path) -> None:
    fixture = build_previous_mapao_workbook(tmp_path / "mapao-anterior.xlsx")

    workbook = _read(fixture.workbook_path)

    assert workbook.source_sha256 == file_sha256(fixture.workbook_path)
    assert workbook.id == contract_workbook_id_for(file_sha256(fixture.workbook_path))
    assert _read(fixture.workbook_path).id == workbook.id


def test_the_hidden_general_sheet_is_read_like_any_other(tmp_path: Path) -> None:
    fixture = build_previous_mapao_workbook(tmp_path / "mapao-anterior.xlsx")

    assert load_workbook(fixture.workbook_path)[_GENERAL_SHEET].sheet_state == "hidden"
    assert len(_read(fixture.workbook_path).lines) == 6


def test_an_item_with_empty_period_cells_declares_the_periods_as_zero(tmp_path: Path) -> None:
    fixture = build_previous_mapao_workbook(tmp_path / "mapao-anterior.xlsx")

    new_item = _read(fixture.workbook_path).line_for_code(_NEW_ITEM_CODE)

    assert [period.period_number for period in new_item.periods] == [1, 2]
    assert all(period.quantity == Decimal("0.00") for period in new_item.periods)
    assert all(period.amount == Decimal("0.00") for period in new_item.periods)
    assert new_item.contract_quantity == Decimal("0.00")
    assert new_item.balance_quantity == Decimal("6.00")


def test_a_tampered_accumulated_is_refused_by_the_aggregate(tmp_path: Path) -> None:
    fixture = build_previous_mapao_workbook(tmp_path / "mapao-anterior.xlsx")
    tamper_cell(fixture.workbook_path, _GENERAL_SHEET, f"M{_FIRST_ITEM_ROW}", Decimal("5.50"))

    diagnosis = _diagnosis_of(fixture.workbook_path)

    finding = _only(diagnosis, "CONTRACT_ACCUMULATED_MISMATCH")
    assert finding.code == _REDUCED_CODE
    assert finding.ref == f"M{_FIRST_ITEM_ROW}"
    assert (finding.declared, finding.expected) == ("5.50", "5.00")


def test_a_tampered_balance_is_refused_by_the_aggregate(tmp_path: Path) -> None:
    fixture = build_previous_mapao_workbook(tmp_path / "mapao-anterior.xlsx")
    tamper_cell(fixture.workbook_path, _GENERAL_SHEET, f"O{_FIRST_ITEM_ROW}", Decimal("10.00"))

    diagnosis = _diagnosis_of(fixture.workbook_path)

    finding = _only(diagnosis, "CONTRACT_BALANCE_MISMATCH")
    assert finding.expected == "11.00"
    assert finding.declared == "10.00"
    assert finding.ref == f"O{_FIRST_ITEM_ROW}"


def test_the_amended_quantity_of_the_two_sheets_must_agree(tmp_path: Path) -> None:
    fixture = build_previous_mapao_workbook(tmp_path / "mapao-anterior.xlsx")
    tamper_cell(
        fixture.workbook_path, _AMENDMENT_SHEET, f"H{_FIRST_AMENDMENT_ROW}", Decimal("15.00")
    )

    diagnosis = _diagnosis_of(fixture.workbook_path)

    finding = _only(diagnosis, "GENERAL_AMENDED_DIVERGENT")
    assert finding.code == _REDUCED_CODE
    assert (finding.declared, finding.expected) == ("15.00", "16.00")
    assert finding.detail["amendment_sheet"] == _AMENDMENT_SHEET


def test_one_unreadable_row_fails_with_the_sheet_and_the_row(tmp_path: Path) -> None:
    fixture = build_previous_mapao_workbook(tmp_path / "mapao-anterior.xlsx")
    tamper_cell(fixture.workbook_path, _GENERAL_SHEET, f"C{_FIRST_ITEM_ROW}", "AD0405005(/)")

    with pytest.raises(ValuationValidationError) as raised:
        _read(fixture.workbook_path)

    assert raised.value.code == "ROW_UNPARSEABLE"
    assert raised.value.details["sheet"] == _GENERAL_SHEET
    assert raised.value.details["row"] == _FIRST_ITEM_ROW


def test_two_unreadable_rows_come_back_together(tmp_path: Path) -> None:
    fixture = build_previous_mapao_workbook(tmp_path / "mapao-anterior.xlsx")
    tamper_cell(fixture.workbook_path, _GENERAL_SHEET, f"C{_FIRST_ITEM_ROW}", "AD0405005(/)")
    tamper_cell(fixture.workbook_path, _GENERAL_SHEET, f"G{_SECOND_ITEM_ROW}", "sob consulta")

    with pytest.raises(ValuationValidationError) as raised:
        _read(fixture.workbook_path)

    assert raised.value.code == "ROWS_UNPARSEABLE"
    failures = raised.value.details["failures"]
    assert isinstance(failures, list)
    assert [failure["row"] for failure in failures] == [_FIRST_ITEM_ROW, _SECOND_ITEM_ROW]
    assert [failure["code"] for failure in failures] == [
        "SCO_CODE_UNPARSEABLE",
        "MONEY_UNPARSEABLE",
    ]
    assert raised.value.details["truncated"] is False


def test_a_general_sheet_with_another_name_fails(tmp_path: Path) -> None:
    fixture = build_previous_mapao_workbook(tmp_path / "mapao-anterior.xlsx")
    _rename_sheet(fixture.workbook_path, _GENERAL_SHEET, "GERAL 2026")

    with pytest.raises(ValuationValidationError) as raised:
        _read(fixture.workbook_path)

    assert raised.value.code == "CONTRACT_SHEET_MISSING"
    assert raised.value.details["sheet"] == _GENERAL_SHEET


def test_a_period_header_outside_the_template_fails(tmp_path: Path) -> None:
    fixture = build_previous_mapao_workbook(tmp_path / "mapao-anterior.xlsx")
    tamper_cell(fixture.workbook_path, _GENERAL_SHEET, "K5", "SEGUNDA MEDICAO")

    with pytest.raises(ValuationValidationError) as raised:
        _read(fixture.workbook_path)

    assert raised.value.code == "PERIOD_HEADER_UNPARSEABLE"
    assert raised.value.details["column"] == "K"
    assert raised.value.details["row"] == 5
    assert raised.value.details["expected"] == "2ª MEDIÇÃO ou ACUMULADO"


def test_a_pair_sublabel_outside_the_template_fails(tmp_path: Path) -> None:
    fixture = build_previous_mapao_workbook(tmp_path / "mapao-anterior.xlsx")
    tamper_cell(fixture.workbook_path, _GENERAL_SHEET, "I6", "QTDE")

    with pytest.raises(ValuationValidationError) as raised:
        _read(fixture.workbook_path)

    assert raised.value.code == "PERIOD_HEADER_UNPARSEABLE"
    assert raised.value.details["row"] == 6
    assert raised.value.details["expected"] == "QUANTIDADE"


def test_a_balance_column_outside_the_template_fails(tmp_path: Path) -> None:
    fixture = build_previous_mapao_workbook(tmp_path / "mapao-anterior.xlsx")
    tamper_cell(fixture.workbook_path, _GENERAL_SHEET, "O5", "SALDO CONTRATUAL")

    with pytest.raises(ValuationValidationError) as raised:
        _read(fixture.workbook_path)

    assert raised.value.code == "GENERAL_COLUMN_MISSING"
    assert raised.value.details["column"] == "O"
    assert raised.value.details["expected"] == "SALDO"


def test_a_general_sheet_without_item_rows_fails(tmp_path: Path) -> None:
    fixture = build_previous_mapao_workbook(tmp_path / "mapao-anterior.xlsx")
    _delete_rows(fixture.workbook_path, _GENERAL_SHEET, 7, 9)

    with pytest.raises(ValuationValidationError) as raised:
        _read(fixture.workbook_path)

    assert raised.value.code == "CONTRACT_EMPTY"
    assert raised.value.details["sheet"] == _GENERAL_SHEET


def test_a_template_without_amendment_layout_reads_no_amendment(tmp_path: Path) -> None:
    fixture = build_previous_mapao_workbook(tmp_path / "mapao-sem-rera.xlsx", with_amendments=False)

    workbook = _read(fixture.workbook_path, _without_amendment_layout())

    assert workbook.amendments == []
    assert workbook.model_dump() == fixture.expected.model_dump()
    assert workbook.line_for_code(_REDUCED_CODE).contract_quantity == Decimal("16.00")


def test_period_labels_with_suffix_and_a_gap_keep_the_numbers_the_sheet_declares(
    tmp_path: Path,
) -> None:
    """`11ª MEDIÇÃO - COMPLEMENTAR` é a 11ª, e o contrato real pula uma medição."""
    fixture = _real_layout_fixture(tmp_path / "mapao-real.xlsx")

    workbook, notes = _read_with_notes(
        fixture.workbook_path, _without_amendment_layout(real_layout_template())
    )

    assert workbook.period_numbers == list(REAL_LAYOUT_PERIOD_NUMBERS)
    assert workbook.next_period_number == 5
    assert notes.period_numbers == REAL_LAYOUT_PERIOD_NUMBERS
    assert notes.period_gaps == (3,)
    assert [period.period_number for period in workbook.lines[0].periods] == [1, 2, 4]
    assert workbook.model_dump() == fixture.expected.model_dump()


def test_a_measurement_number_out_of_order_in_the_header_is_refused(tmp_path: Path) -> None:
    fixture = build_previous_mapao_workbook(tmp_path / "mapao-anterior.xlsx")
    tamper_cell(fixture.workbook_path, _GENERAL_SHEET, "K5", "1ª MEDIÇÃO - COMPLEMENTAR")

    with pytest.raises(ValuationValidationError) as raised:
        _read(fixture.workbook_path)

    assert raised.value.code == "PERIOD_HEADER_UNPARSEABLE"
    assert raised.value.details["reason"] == "número de medição fora de ordem no cabeçalho"
    assert raised.value.details["found"] == "1ª MEDIÇÃO - COMPLEMENTAR"
    assert raised.value.details["number"] == 1
    assert raised.value.details["previous"] == 1


def test_a_general_without_the_amended_column_derives_the_amended_from_the_amendments(
    tmp_path: Path,
) -> None:
    """Sem coluna de vigente, o vigente é contratual mais as RE-RA — e bate com a prefeitura."""
    template = template_without_amended_column()
    fixture = build_previous_mapao_workbook(tmp_path / "mapao-sem-vigente.xlsx", template=template)

    workbook = _read(fixture.workbook_path, template)

    assert workbook.model_dump() == fixture.expected.model_dump()
    reduced = workbook.line_for_code(_REDUCED_CODE)
    assert reduced.contract_quantity == Decimal("20.00")
    assert reduced.amended_quantity == Decimal("16.00")
    assert workbook.line_for_code(_NEW_ITEM_CODE).amended_quantity == Decimal("6.00")


def test_a_general_without_the_amended_column_and_without_amendments_keeps_the_contracted(
    tmp_path: Path,
) -> None:
    fixture = _real_layout_fixture(tmp_path / "mapao-real.xlsx")

    workbook = _read(fixture.workbook_path, _without_amendment_layout(real_layout_template()))

    assert workbook.amendments == []
    for line in workbook.lines:
        assert line.amended_quantity == line.contract_quantity, line.code


def test_a_quantity_with_four_decimals_is_read_in_the_declared_scale(tmp_path: Path) -> None:
    fixture = _real_layout_fixture(tmp_path / "mapao-real.xlsx")

    workbook = _read(fixture.workbook_path, _without_amendment_layout(real_layout_template()))

    line = workbook.line_for_code(REAL_LAYOUT_DUPLICATED_CODE, group_label=_PAVING_GROUP)
    assert line.contract_quantity == Decimal("100459.5525")
    assert line.balance_quantity == Decimal("100456.2325")


def test_float_noise_from_the_formula_cache_is_normalized_and_recorded(tmp_path: Path) -> None:
    """O cache da planilha real traz 4,999999999999 onde a conta fecha em 5,00."""
    fixture = build_previous_mapao_workbook(tmp_path / "mapao-anterior.xlsx")
    tamper_cell(fixture.workbook_path, _GENERAL_SHEET, f"M{_FIRST_ITEM_ROW}", 4.999999999999)

    workbook, notes = _read_with_notes(fixture.workbook_path)

    assert workbook.line_for_code(_REDUCED_CODE).accumulated_quantity == Decimal("5.00")
    assert notes.normalized_count == 1
    assert notes.normalized_truncated is False
    assert [
        (cell.sheet, cell.ref, cell.raw, cell.normalized) for cell in notes.normalized_cells
    ] == [(_GENERAL_SHEET, f"M{_FIRST_ITEM_ROW}", "4.999999999999", "5.00")]


def test_a_quantity_above_the_declared_scale_is_refused(tmp_path: Path) -> None:
    """Um milionésimo é ruído de cache; dois milésimos e meio são dado com casa a mais."""
    fixture = build_previous_mapao_workbook(tmp_path / "mapao-anterior.xlsx")
    tamper_cell(fixture.workbook_path, _GENERAL_SHEET, f"F{_FIRST_ITEM_ROW}", 100459.5525)
    tamper_cell(fixture.workbook_path, _GENERAL_SHEET, f"F{_SECOND_ITEM_ROW}", 10.12345)

    with pytest.raises(ValuationValidationError) as raised:
        _read(fixture.workbook_path)

    assert raised.value.code == "ROWS_UNPARSEABLE"
    failures = raised.value.details["failures"]
    assert isinstance(failures, list)
    assert [failure["code"] for failure in failures] == [
        "NUMBER_SCALE_UNSUPPORTED",
        "NUMBER_SCALE_UNSUPPORTED",
    ]
    assert [failure["row"] for failure in failures] == [_FIRST_ITEM_ROW, _SECOND_ITEM_ROW]


def test_separator_rows_are_skipped_counted_and_do_not_shift_the_row_numbers(
    tmp_path: Path,
) -> None:
    fixture = _real_layout_fixture(tmp_path / "mapao-real.xlsx", separator_rows=True)
    template = _without_amendment_layout(real_layout_template())

    workbook, notes = _read_with_notes(fixture.workbook_path, template)

    assert fixture.separator_rows, "a fixture precisa escrever linha separadora"
    assert notes.separator_rows == fixture.separator_rows
    assert len(workbook.lines) == len(fixture.expected.lines)
    assert workbook.model_dump() == fixture.expected.model_dump()

    # A linha que vem depois da separadora continua sendo reportada com o número dela.
    after_separator = fixture.separator_rows[0] + 2
    tamper_cell(fixture.workbook_path, _GENERAL_SHEET, f"C{after_separator}", "AD0405005(/)")
    with pytest.raises(ValuationValidationError) as raised:
        _read(fixture.workbook_path, template)

    assert raised.value.code == "ROW_UNPARSEABLE"
    assert raised.value.details["row"] == after_separator


def test_the_same_code_in_two_groups_is_imported_without_choosing_a_line(tmp_path: Path) -> None:
    fixture = _real_layout_fixture(tmp_path / "mapao-real.xlsx")

    workbook = _read(fixture.workbook_path, _without_amendment_layout(real_layout_template()))

    repeated = workbook.lines_for_code(REAL_LAYOUT_DUPLICATED_CODE)
    assert [line.group_label for line in repeated] == [_PAVING_GROUP, _SERVICES_GROUP]
    assert [line.item_number for line in repeated] == ["1", "1"]
    with pytest.raises(ValuationValidationError) as raised:
        workbook.line_for_code(REAL_LAYOUT_DUPLICATED_CODE)
    assert raised.value.code == "CODE_AMBIGUOUS_IN_CONTRACT"


def test_an_amendment_over_a_code_repeated_in_two_groups_is_refused(tmp_path: Path) -> None:
    """A aba da prefeitura fala por código; com o código em dois grupos, o leitor recusa."""
    written = write_previous_mapao_workbook(
        tmp_path / "mapao-real-com-rera.xlsx",
        template=real_layout_template(),
        items=real_layout_items(reduce_duplicated_code=True),
        period_numbers=REAL_LAYOUT_PERIOD_NUMBERS,
        period_label_suffixes=REAL_LAYOUT_PERIOD_SUFFIXES,
    )

    diagnosis = _diagnosis_of(written.workbook_path, real_layout_template())

    finding = _only(diagnosis, "CODE_AMBIGUOUS_IN_CONTRACT")
    assert finding.code == REAL_LAYOUT_DUPLICATED_CODE
    assert finding.detail["groups"] == [_PAVING_GROUP, _SERVICES_GROUP]


def test_a_negative_balance_becomes_a_finding_instead_of_an_opaque_validation_error(
    tmp_path: Path,
) -> None:
    """Saldo negativo existe no arquivo real: o leitor converte e o dossiê o classifica.

    Sem o dossiê, o número negativo morria na validação de **campo** do Pydantic
    (`ge=0`), sem código estável nenhum — a recusa dizia `balance_quantity` e não dizia
    saldo negativo.
    """
    fixture = build_previous_mapao_workbook(tmp_path / "mapao-anterior.xlsx")
    tamper_cell(fixture.workbook_path, _GENERAL_SHEET, f"H{_FIRST_ITEM_ROW}", Decimal("1.00"))
    tamper_cell(fixture.workbook_path, _GENERAL_SHEET, f"O{_FIRST_ITEM_ROW}", Decimal("-4.00"))

    diagnosis = _diagnosis_of(fixture.workbook_path)

    finding = _only(diagnosis, "CONTRACT_BALANCE_NEGATIVE")
    assert finding.code == _REDUCED_CODE
    assert finding.ref == f"O{_FIRST_ITEM_ROW}"
    assert finding.declared == "-4.00"
    assert finding.detail["declared_negative"] is True
    assert finding.detail["recomputed_negative"] is True
    # Saldo negativo consistente com vigente menos acumulado não é, além disso, drift.
    assert _findings(diagnosis, "CONTRACT_BALANCE_MISMATCH") == []


def test_an_accumulated_above_the_amended_is_refused_as_a_negative_balance(
    tmp_path: Path,
) -> None:
    fixture = build_previous_mapao_workbook(tmp_path / "mapao-anterior.xlsx")
    tamper_cell(fixture.workbook_path, _GENERAL_SHEET, f"H{_FIRST_ITEM_ROW}", Decimal("1.00"))
    tamper_cell(fixture.workbook_path, _GENERAL_SHEET, f"O{_FIRST_ITEM_ROW}", Decimal("0.00"))

    diagnosis = _diagnosis_of(fixture.workbook_path)

    negative = _only(diagnosis, "CONTRACT_BALANCE_NEGATIVE")
    assert negative.code == _REDUCED_CODE
    assert negative.detail["declared_negative"] is False
    assert negative.detail["recomputed_negative"] is True
    assert _only(diagnosis, "CONTRACT_BALANCE_MISMATCH").expected == "-4.00"


# --- M2.1 fase D: código contratual extra como dado + linhas de seção da prefeitura ---


def test_a_bare_code_declared_by_the_template_reaches_the_contract_line(tmp_path: Path) -> None:
    """`IE00040849` não é SCO, mas o template declara o padrão que o aceita."""
    template = _without_amendment_layout(template_with_extra_code_pattern())
    fixture = build_previous_mapao_workbook(
        tmp_path / "mapao-codigo-extra.xlsx",
        with_amendments=False,
        template=template,
        items=extra_code_items(),
        period_numbers=PREVIOUS_MAPAO_PERIOD_NUMBERS,
    )

    workbook, notes = _read_with_notes(fixture.workbook_path, template)

    assert workbook.line_for_code(EXTRA_CODE_SAMPLE).code == EXTRA_CODE_SAMPLE
    assert workbook.model_dump() == fixture.expected.model_dump()
    assert notes.extra_code_rows == ((_FIRST_ITEM_ROW, EXTRA_CODE_SAMPLE),)


def test_a_bare_code_without_the_declared_pattern_is_refused(tmp_path: Path) -> None:
    """O mesmo código nu, sem o template declarar o padrão, continua ilegível.

    Uma única linha ilegível sobe encapsulada em `ROW_UNPARSEABLE` (como qualquer outra
    linha ilegível, `test_one_unreadable_row_fails_with_the_sheet_and_the_row`); o motivo
    de domínio ("código fora do formato SCO") continua acessível em `details["reason"]`.
    """
    template = _without_amendment_layout(default_template())
    fixture = build_previous_mapao_workbook(
        tmp_path / "mapao-codigo-extra-sem-padrao.xlsx",
        with_amendments=False,
        template=template,
        items=extra_code_items(),
        period_numbers=PREVIOUS_MAPAO_PERIOD_NUMBERS,
    )

    with pytest.raises(ValuationValidationError) as raised:
        _read(fixture.workbook_path, template)

    assert raised.value.code == "ROW_UNPARSEABLE"
    assert raised.value.details["sheet"] == _GENERAL_SHEET
    assert raised.value.details["row"] == _FIRST_ITEM_ROW
    assert raised.value.details["reason"] == "código fora do formato SCO"


def test_a_bare_code_amendment_in_the_town_hall_sheet_is_applied(tmp_path: Path) -> None:
    """A RE-RA da aba da prefeitura aplica normalmente sobre um código nu declarado."""
    template = template_with_extra_code_pattern()
    fixture = build_previous_mapao_workbook(
        tmp_path / "mapao-codigo-extra-rera.xlsx",
        with_amendments=True,
        template=template,
        items=extra_code_items(with_amendment=True),
        period_numbers=PREVIOUS_MAPAO_PERIOD_NUMBERS,
    )

    workbook = _read(fixture.workbook_path, template)

    line = workbook.line_for_code(EXTRA_CODE_SAMPLE)
    assert line.contract_quantity == Decimal("10.00")
    assert line.amended_quantity == Decimal("9.00")
    assert workbook.model_dump() == fixture.expected.model_dump()


def test_a_town_hall_section_row_without_any_value_is_skipped_and_counted(
    tmp_path: Path,
) -> None:
    """`LAZER / PAISAGISMO` na coluna de código, sem nenhum valor: linha de seção."""
    fixture = build_previous_mapao_workbook(tmp_path / "mapao-secao.xlsx")
    layout = default_template().amendment
    assert layout is not None
    rows = amendment_code_rows(fixture.workbook_path, layout)
    section_row = max(rows.values()) + 1
    tamper_cell(
        fixture.workbook_path,
        _AMENDMENT_SHEET,
        f"{layout.code_column}{section_row}",
        "LAZER / PAISAGISMO",
    )
    # `tamper_cell` reescreve o arquivo: o digest muda mesmo sem alterar nenhuma linha de
    # item, e o gabarito precisa acompanhar o arquivo que o leitor de fato vai ler.
    tampered_sha256 = file_sha256(fixture.workbook_path)
    expected = fixture.expected.model_copy(
        update={"source_sha256": tampered_sha256, "id": contract_workbook_id_for(tampered_sha256)}
    )

    workbook, notes = _read_with_notes(fixture.workbook_path)

    assert notes.amendment_section_rows == ((section_row, None),)
    assert workbook.model_dump() == expected.model_dump()


def test_a_town_hall_section_row_with_a_declared_value_is_refused(tmp_path: Path) -> None:
    """A mesma linha de seção, agora com um valor num bloco de RE-RA, continua ilegível."""
    fixture = build_previous_mapao_workbook(tmp_path / "mapao-secao-com-valor.xlsx")
    layout = default_template().amendment
    assert layout is not None
    block = layout.blocks[0]
    assert block.added_column is not None
    rows = amendment_code_rows(fixture.workbook_path, layout)
    section_row = max(rows.values()) + 1
    tamper_cell(
        fixture.workbook_path,
        _AMENDMENT_SHEET,
        f"{layout.code_column}{section_row}",
        "LAZER / PAISAGISMO",
    )
    tamper_cell(
        fixture.workbook_path,
        _AMENDMENT_SHEET,
        f"{block.added_column}{section_row}",
        Decimal("1.00"),
    )

    with pytest.raises(ValuationValidationError) as raised:
        _read(fixture.workbook_path)

    assert raised.value.code == "ROW_UNPARSEABLE"
    assert raised.value.details["sheet"] == _AMENDMENT_SHEET
    assert raised.value.details["row"] == section_row


def test_a_town_hall_section_row_with_the_group_subtotal_flag_keeps_the_ignored_value(
    tmp_path: Path,
) -> None:
    """Com a flag ligada, o subtotal de grupo na coluna de vigente não bloqueia o skip —
    e o valor ignorado chega inteiro nas notas, nunca descartado em silêncio."""
    fixture = build_previous_mapao_workbook(tmp_path / "mapao-secao-subtotal.xlsx")
    template = template_with_section_group_subtotal()
    layout = template.amendment
    assert layout is not None
    assert layout.amended_quantity_column is not None
    rows = amendment_code_rows(fixture.workbook_path, layout)
    section_row = max(rows.values()) + 1
    tamper_cell(
        fixture.workbook_path,
        _AMENDMENT_SHEET,
        f"{layout.code_column}{section_row}",
        "LAZER / PAISAGISMO",
    )
    tamper_cell(
        fixture.workbook_path,
        _AMENDMENT_SHEET,
        f"{layout.amended_quantity_column}{section_row}",
        Decimal("171025043.01"),
    )
    tampered_sha256 = file_sha256(fixture.workbook_path)
    expected = fixture.expected.model_copy(
        update={"source_sha256": tampered_sha256, "id": contract_workbook_id_for(tampered_sha256)}
    )

    workbook, notes = _read_with_notes(fixture.workbook_path, template)

    assert notes.amendment_section_rows == ((section_row, Decimal("171025043.01")),)
    assert workbook.model_dump() == expected.model_dump()


def test_a_town_hall_section_row_with_a_block_value_is_refused_even_with_the_flag(
    tmp_path: Path,
) -> None:
    """A flag só isenta a coluna de vigente; um valor real de RE-RA continua recusando."""
    fixture = build_previous_mapao_workbook(tmp_path / "mapao-secao-subtotal-com-bloco.xlsx")
    template = template_with_section_group_subtotal()
    layout = template.amendment
    assert layout is not None
    block = layout.blocks[0]
    assert block.added_column is not None
    rows = amendment_code_rows(fixture.workbook_path, layout)
    section_row = max(rows.values()) + 1
    tamper_cell(
        fixture.workbook_path,
        _AMENDMENT_SHEET,
        f"{layout.code_column}{section_row}",
        "LAZER / PAISAGISMO",
    )
    tamper_cell(
        fixture.workbook_path,
        _AMENDMENT_SHEET,
        f"{block.added_column}{section_row}",
        Decimal("1.00"),
    )

    with pytest.raises(ValuationValidationError) as raised:
        _read(fixture.workbook_path, template)

    assert raised.value.code == "ROW_UNPARSEABLE"
    assert raised.value.details["sheet"] == _AMENDMENT_SHEET
    assert raised.value.details["row"] == section_row
