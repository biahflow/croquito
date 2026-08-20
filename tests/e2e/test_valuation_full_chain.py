"""Cadeia completa da medição de obra, ponta a ponta, pelos comandos reais do CLI.

Cada elo (`import-workbook`, `extract-legend`, `review-takeoff`, `suggest-codes`,
`confirm-codes`, `build-calc`, `export-valuation`) já tem cobertura própria em
`tests/valuation/` e `tests/worker/test_valuation_cli.py`. O que só este teste prova é o
encaixe: cada comando lê o artefato que o comando anterior gravou em disco — não o objeto
em memória —, e a cadeia inteira, de `import-workbook` até a reabertura do `.xlsx`
publicado, fecha sem nenhum atalho substituindo um comando real por chamada direta ao
domínio.

Não há stack HTTP aqui: a cadeia de medição é inteiramente CLI, offline e síncrona
(`croquito_worker.valuation.cli.main`), como em `tests/worker/test_valuation_cli.py`.
A fixture `chain` roda a cadeia cara uma vez por módulo — a extração da prancha sintética
é o passo caro —, devolvendo os artefatos e modelos de cada elo.
`test_full_chain_happy_path` faz as asserções de negócio sobre esse resultado, uma por
elo; os testes de recusa derivam cenários dos mesmos artefatos sem repetir os passos que
já correram com exit 0.
"""

from __future__ import annotations

import json
from collections.abc import Sequence
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Final, Literal

import pytest
from openpyxl import load_workbook

from croquito_valuation.amendment_dossier import AmendmentDossier
from croquito_valuation.assignment import (
    CodeAssignmentBatch,
    CodeAssignmentInput,
    CodeAssignmentSet,
    CodeSuggestionSet,
)
from croquito_valuation.contract import ContractWorkbook
from croquito_valuation.estimate import Estimate
from croquito_valuation.models import CalcRecipe, PriceCatalog, PriceOrigin, Valuation
from croquito_valuation.takeoff import TakeoffItemStatus, TakeoffPacket, load_takeoff_packet
from croquito_valuation.template import WorkbookTemplate, default_template
from croquito_worker.valuation.cli import (
    AMENDMENT_DOSSIER_FILENAME,
    AUDIT_FILENAME,
    CALC_PLAN_FILENAME,
    CATALOG_FILENAME,
    CODE_ASSIGNMENTS_FILENAME,
    CODE_SUGGESTIONS_FILENAME,
    CONTRACT_FILENAME,
    ESTIMATE_FILENAME,
    TAKEOFF_PACKET_FILENAME,
    VALUATION_FILENAME,
    WORKBOOK_FILENAME,
    main,
)
from croquito_worker.valuation.emop_fixture import emop_fixture_layout, write_emop_dbf
from croquito_worker.valuation.estimate_fixture import (
    SYNTHETIC_ESTIMATE_BDI_PERCENT,
    SYNTHETIC_ESTIMATE_WORKSITE_KEY,
    SYNTHETIC_ESTIMATE_WORKSITE_NAME,
    build_demo_estimate_assignments,
    build_synthetic_composition_set,
)
from croquito_worker.valuation.plate import SYNTHETIC_LEGEND_ROWS, SYNTHETIC_PLATE_ID
from croquito_worker.valuation.sicro_fixture import sicro_fixture_layout, write_sicro_xlsx
from croquito_worker.valuation.sinapi_fixture import sinapi_fixture_layout, write_sinapi_xlsx
from croquito_worker.valuation.synthetic import (
    SYNTHETIC_CODE_DECIDED_AT,
    SYNTHETIC_CONTRACT_LABEL,
    SYNTHETIC_PERIOD_REFERENCE_LABEL,
    SYNTHETIC_PREVIOUS_PERIOD_COUNT,
    SYNTHETIC_TAKEOFF_REVIEWER,
    SYNTHETIC_TAKEOFF_WORKSITE_ADDRESS,
    SYNTHETIC_TAKEOFF_WORKSITE_KEY,
    SYNTHETIC_TAKEOFF_WORKSITE_NAME,
    build_demo_calc_plan,
    build_demo_code_assignments,
    build_demo_takeoff_decisions,
    build_synthetic_approval,
    build_synthetic_previous_mapao,
)
from croquito_worker.valuation.takeoff_fixture import takeoff_item_id

# Rótulos e ids da legenda sintética: a MESMA fonte que `tests/valuation/test_chain_demo.py`
# usa para a obra que nasce da prancha, para que os dois testes não possam divergir sobre
# qual item é qual.
_PAVEMENT_ITEM_ID = takeoff_item_id(SYNTHETIC_PLATE_ID, SYNTHETIC_LEGEND_ROWS[0].label)
_LAWN_ITEM_ID = takeoff_item_id(SYNTHETIC_PLATE_ID, SYNTHETIC_LEGEND_ROWS[1].label)
_FENCE_ITEM_ID = takeoff_item_id(SYNTHETIC_PLATE_ID, SYNTHETIC_LEGEND_ROWS[2].label)
_BENCH_ITEM_ID = takeoff_item_id(SYNTHETIC_PLATE_ID, SYNTHETIC_LEGEND_ROWS[3].label)
_LAMP_ITEM_ID = takeoff_item_id(SYNTHETIC_PLATE_ID, SYNTHETIC_LEGEND_ROWS[4].label)
_RUBBER_ITEM_ID = takeoff_item_id(SYNTHETIC_PLATE_ID, SYNTHETIC_LEGEND_ROWS[6].label)

# Numeração dos itens no boletim da obra oeste depois que o gramado (código rejeitado) sai:
# pavimento, alambrado, banco, luminária, piso emborrachado — mesma ordem fixada em
# `tests/valuation/test_chain_demo.py`.
_FENCE_ITEM_NUMBER = "2"

_EXPECTED_PERIOD_NUMBER = SYNTHETIC_PREVIOUS_PERIOD_COUNT + 1
"""3 na fixture: duas medições já lançadas no MAPÃO anterior, a próxima é a terceira."""

_OUT_OF_CONTRACT_CATALOG_CODE = "AD04100015(/)"
"""Existe no catálogo sintético (32 entradas) e não está entre os 10 códigos do
consolidado — candidato citado pelo próprio spec da Fase C."""

_INCOMPATIBLE_UNIT_CONTRACT_CODE = "SP01050010(/)"
"""Está no contrato (item 6, unidade M3); o banco confirmado mede em UN."""

_MISSING_ASSIGNMENT_BATCH: Final[
    tuple[tuple[str, Literal["confirm", "reject"], str | None], ...]
] = (
    (_PAVEMENT_ITEM_ID, "confirm", "AD04050060(/)"),
    (_LAWN_ITEM_ID, "reject", None),
    (_FENCE_ITEM_ID, "confirm", "CE02100010(/)"),
    (_BENCH_ITEM_ID, "confirm", "MB01100010(/)"),
    (_RUBBER_ITEM_ID, "confirm", "AD04150010(/)"),
    # A luminária (`_LAMP_ITEM_ID`) fica de fora de propósito: é o item confirmado no
    # takeoff que não recebe nenhuma decisão de código neste lote — o que falta para o
    # build-calc acusar CALC_ASSIGNMENT_MISSING.
)


def _stdout(capsys: pytest.CaptureFixture[str]) -> dict[str, object]:
    """Último JSON impresso: cada comando imprime uma linha só."""
    lines = [line for line in capsys.readouterr().out.splitlines() if line.strip()]
    return dict(json.loads(lines[-1]))


def _write_json(path: Path, payload: str) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(payload, encoding="utf-8")
    return path


def _catalog_args(catalog_paths: Sequence[Path]) -> list[str]:
    """`--catalog` repetido na ORDEM da cascata; o comando não reordena o que recebe."""
    return [argument for path in catalog_paths for argument in ("--catalog", str(path))]


def _bulletin_quantity(
    workbook_path: Path, template: WorkbookTemplate, worksite_name: str, code: str
) -> float:
    """Quantidade de um código na aba BM da obra, localizada pela coluna do template.

    Nenhuma célula é fixa aqui: aba, linha e coluna vêm de `default_template()`, como no
    helper equivalente de `tests/worker/test_valuation_cli.py`.
    """
    workbook = load_workbook(workbook_path)
    try:
        sheet_name = template.bulletin_sheet_name(worksite_name)
        worksheet = workbook[sheet_name]
        layout = template.bulletin
        row = layout.header_row + 1
        while True:
            cell_code = worksheet[f"{layout.columns.code.letter}{row}"].value
            if cell_code is None:
                raise AssertionError(f"código {code} ausente da aba {sheet_name}")
            if cell_code == code:
                quantity = worksheet[f"{layout.columns.quantity.letter}{row}"].value
                assert isinstance(quantity, int | float)
                return float(quantity)
            row += 1
    finally:
        workbook.close()


@dataclass(frozen=True, slots=True)
class ChainArtifacts:
    """Artefatos e modelos produzidos por uma passada da cadeia inteira pelo CLI."""

    template: WorkbookTemplate
    catalog_path: Path
    contract_path: Path
    contract: ContractWorkbook
    extracted_packet_path: Path
    extracted_packet: TakeoffPacket
    reviewed_packet_path: Path
    reviewed_packet: TakeoffPacket
    suggestions_path: Path
    suggestions: CodeSuggestionSet
    assignment_decisions_path: Path
    assignments_path: Path
    assignments: CodeAssignmentSet
    valuation_path: Path
    valuation_unapproved: Valuation
    valuation_approved: Valuation
    workbook_path: Path
    audit_path: Path
    dossier_path: Path
    dossier: AmendmentDossier


@pytest.fixture(scope="module")
def chain(tmp_path_factory: pytest.TempPathFactory) -> ChainArtifacts:
    """Percorre a cadeia de medição inteira pelos comandos reais do CLI, uma vez só.

    Roda uma única vez por módulo porque gerar e extrair a prancha sintética é o passo
    caro da cadeia. Cada `assert main([...]) == 0` aqui é o "um comando por elo, exit 0
    em todos" do caminho feliz; as asserções de negócio (contagens, códigos, quantidades)
    ficam em `test_full_chain_happy_path`, que só lê o que esta fixture já produziu. Os
    testes de recusa também partem destes artefatos, sem repetir os passos caros.
    """
    root = tmp_path_factory.mktemp("valuation-chain")
    template = default_template()

    # 1. MAPÃO anterior sintético → import-workbook (catálogo + consolidado).
    previous_path = build_synthetic_previous_mapao(root / "previous-mapao.xlsx")
    import_dir = root / "import"
    assert (
        main(["import-workbook", "--input", str(previous_path), "--output", str(import_dir)]) == 0
    )
    catalog_path = import_dir / CATALOG_FILENAME
    contract_path = import_dir / CONTRACT_FILENAME
    contract = ContractWorkbook.model_validate_json(contract_path.read_text(encoding="utf-8"))

    # 2. extract-legend → pacote proposto (7 itens, 1 ambíguo).
    takeoff_dir = root / "takeoff"
    assert main(["extract-legend", "--output", str(takeoff_dir)]) == 0
    extracted_packet_path = takeoff_dir / TAKEOFF_PACKET_FILENAME
    extracted_packet = load_takeoff_packet(extracted_packet_path)

    # 3. Decisões do orçamentista sintético gravadas → review-takeoff.
    decisions = build_demo_takeoff_decisions(extracted_packet)
    decisions_path = _write_json(
        root / "inputs" / "takeoff-decisions.json", decisions.model_dump_json(indent=2)
    )
    review_dir = root / "review"
    assert (
        main(
            [
                "review-takeoff",
                "--packet",
                str(extracted_packet_path),
                "--decisions",
                str(decisions_path),
                "--output",
                str(review_dir),
            ]
        )
        == 0
    )
    reviewed_packet_path = review_dir / TAKEOFF_PACKET_FILENAME
    reviewed_packet = load_takeoff_packet(reviewed_packet_path)

    # 4. suggest-codes --contract → code-suggestions.json.
    suggest_dir = root / "suggest"
    assert (
        main(
            [
                "suggest-codes",
                "--packet",
                str(reviewed_packet_path),
                "--catalog",
                str(catalog_path),
                "--contract",
                str(contract_path),
                "--output",
                str(suggest_dir),
            ]
        )
        == 0
    )
    suggestions_path = suggest_dir / CODE_SUGGESTIONS_FILENAME
    suggestions = CodeSuggestionSet.model_validate_json(
        suggestions_path.read_text(encoding="utf-8")
    )

    # 5. Decisões de código do orçamentista sintético gravadas → confirm-codes.
    assignment_decisions = build_demo_code_assignments(reviewed_packet)
    assignment_decisions_path = _write_json(
        root / "inputs" / "code-assignment-decisions.json",
        assignment_decisions.model_dump_json(indent=2),
    )
    assign_dir = root / "assign"
    assert (
        main(
            [
                "confirm-codes",
                "--packet",
                str(reviewed_packet_path),
                "--decisions",
                str(assignment_decisions_path),
                "--catalog",
                str(catalog_path),
                "--contract",
                str(contract_path),
                "--output",
                str(assign_dir),
            ]
        )
        == 0
    )
    assignments_path = assign_dir / CODE_ASSIGNMENTS_FILENAME
    assignments = CodeAssignmentSet.model_validate_json(
        assignments_path.read_text(encoding="utf-8")
    )

    # 6. calc-plan.json (decompõe só o alambrado) → build-calc: obra oeste, período 3.
    calc_plan = build_demo_calc_plan(reviewed_packet)
    calc_plan_path = _write_json(
        root / "inputs" / CALC_PLAN_FILENAME, calc_plan.model_dump_json(indent=2)
    )
    calc_dir = root / "calc"
    assert (
        main(
            [
                "build-calc",
                "--packet",
                str(reviewed_packet_path),
                "--assignments",
                str(assignments_path),
                "--catalog",
                str(catalog_path),
                "--worksite-key",
                SYNTHETIC_TAKEOFF_WORKSITE_KEY,
                "--worksite-name",
                SYNTHETIC_TAKEOFF_WORKSITE_NAME,
                "--period-number",
                str(contract.next_period_number),
                "--reference-label",
                SYNTHETIC_PERIOD_REFERENCE_LABEL,
                "--address",
                SYNTHETIC_TAKEOFF_WORKSITE_ADDRESS,
                "--contract-label",
                SYNTHETIC_CONTRACT_LABEL,
                "--calc-plan",
                str(calc_plan_path),
                "--output",
                str(calc_dir),
            ]
        )
        == 0
    )
    valuation_path = calc_dir / VALUATION_FILENAME
    valuation_unapproved = Valuation.model_validate_json(valuation_path.read_text(encoding="utf-8"))

    # 6b. build-amendment-dossier: mesmos packet/assignments do build-calc, sem depender
    #     dele — o dossiê é o outro artefato de fechamento da rodada (aditivo, não boletim).
    dossier_dir = root / "dossier"
    assert (
        main(
            [
                "build-amendment-dossier",
                "--packet",
                str(reviewed_packet_path),
                "--assignments",
                str(assignments_path),
                "--output",
                str(dossier_dir),
            ]
        )
        == 0
    )
    dossier_path = dossier_dir / AMENDMENT_DOSSIER_FILENAME
    dossier = AmendmentDossier.model_validate_json(dossier_path.read_text(encoding="utf-8"))

    # 7. Aprovação em código sobre o Valuation lido do artefato, regravando o JSON, e só
    #    então export-valuation.
    valuation_approved = build_synthetic_approval(valuation_unapproved)
    valuation_path.write_text(valuation_approved.model_dump_json(indent=2), encoding="utf-8")

    export_dir = root / "export"
    assert (
        main(
            [
                "export-valuation",
                "--valuation",
                str(valuation_path),
                "--contract",
                str(contract_path),
                "--catalog",
                str(catalog_path),
                "--output",
                str(export_dir),
            ]
        )
        == 0
    )
    workbook_path = export_dir / WORKBOOK_FILENAME
    audit_path = export_dir / AUDIT_FILENAME

    return ChainArtifacts(
        template=template,
        catalog_path=catalog_path,
        contract_path=contract_path,
        contract=contract,
        extracted_packet_path=extracted_packet_path,
        extracted_packet=extracted_packet,
        reviewed_packet_path=reviewed_packet_path,
        reviewed_packet=reviewed_packet,
        suggestions_path=suggestions_path,
        suggestions=suggestions,
        assignment_decisions_path=assignment_decisions_path,
        assignments_path=assignments_path,
        assignments=assignments,
        valuation_path=valuation_path,
        valuation_unapproved=valuation_unapproved,
        valuation_approved=valuation_approved,
        workbook_path=workbook_path,
        audit_path=audit_path,
        dossier_path=dossier_path,
        dossier=dossier,
    )


def test_full_chain_happy_path(chain: ChainArtifacts) -> None:
    """Um comando por elo, exit 0 em todos (já verificado na fixture `chain`).

    Cada comentário numerado confere o que aquele elo deveria ter produzido em disco.
    """
    # 1. import-workbook: consolidado com o período seguinte igual a 3 na fixture.
    assert chain.contract.next_period_number == _EXPECTED_PERIOD_NUMBER

    # 2. extract-legend: pacote proposto com 7 itens, 1 ambíguo (nenhum nasce confirmado).
    assert len(chain.extracted_packet.items) == 7
    assert chain.extracted_packet.confirmed_items() == []
    ambiguous = [
        item for item in chain.extracted_packet.items if item.status is TakeoffItemStatus.AMBIGUOUS
    ]
    assert len(ambiguous) == 1

    # 3. review-takeoff: 6 confirmados, 1 rejeitado (área de intervenção sintética).
    assert len(chain.reviewed_packet.confirmed_items()) == 6
    rejected_takeoff = [
        item for item in chain.reviewed_packet.items if item.status is TakeoffItemStatus.REJECTED
    ]
    assert len(rejected_takeoff) == 1

    # 4. suggest-codes: todo item confirmado tem resposta, sugestão ou unmatched.
    confirmed_ids = {item.id for item in chain.reviewed_packet.confirmed_items()}
    answered = {suggestion.item_id for suggestion in chain.suggestions.suggestions} | set(
        chain.suggestions.unmatched_item_ids
    )
    assert answered == confirmed_ids

    # 5. confirm-codes: 5 confirmados, 1 rejeitado — o gramado, sem cotação no contrato.
    confirmed_assignments = [a for a in chain.assignments.assignments if a.status == "confirmed"]
    rejected_assignments = [a for a in chain.assignments.assignments if a.status == "rejected"]
    assert len(confirmed_assignments) == 5
    assert len(rejected_assignments) == 1
    assert rejected_assignments[0].item_id == _LAWN_ITEM_ID

    # 6. build-calc: obra oeste com 5 linhas; alambrado com bloco PERIMETER_TIMES_HEIGHT e
    #    subtotal 58,50 (48,75 m x 1,20 m), quantidade que o revisor confirmou.
    bulletin = chain.valuation_unapproved.bulletins[0]
    assert bulletin.worksite_key == SYNTHETIC_TAKEOFF_WORKSITE_KEY
    assert len(bulletin.lines) == 5
    fence_line = next(line for line in bulletin.lines if line.item_number == _FENCE_ITEM_NUMBER)
    fence_sheet = chain.valuation_unapproved.calc_sheet_for(
        SYNTHETIC_TAKEOFF_WORKSITE_KEY, _FENCE_ITEM_NUMBER
    )
    assert fence_line.quantity == fence_sheet.total_quantity
    assert [block.recipe for block in fence_sheet.blocks] == [CalcRecipe.PERIMETER_TIMES_HEIGHT]
    assert [operand.value for operand in fence_sheet.blocks[0].operands] == [
        Decimal("48.75"),
        Decimal("1.20"),
    ]
    assert fence_sheet.blocks[0].subtotal == Decimal("58.50")
    assert fence_sheet.total_quantity == Decimal("58.50")

    # 7. Aprovação em código + export-valuation: xlsx e audit publicados com status ok.
    assert chain.valuation_approved.approval is not None
    assert (
        chain.valuation_approved.approval.valuation_digest
        == chain.valuation_approved.content_digest()
    )
    audit = json.loads(chain.audit_path.read_text(encoding="utf-8"))
    assert audit["status"] == "ok"
    assert chain.workbook_path.is_file()

    # 8. Reabertura do medicao.xlsx: PLANILHA GERAL presente, par BM/MEMÓRIA da obra oeste
    #    presente, e a linha do alambrado na BM com quantidade 58,50 — coluna e aba
    #    derivadas do template, não número mágico.
    workbook = load_workbook(chain.workbook_path)
    try:
        assert chain.template.general.sheet_name in workbook.sheetnames
        bulletin_sheet = chain.template.bulletin_sheet_name(SYNTHETIC_TAKEOFF_WORKSITE_NAME)
        memory_sheet = chain.template.memory_sheet_name(SYNTHETIC_TAKEOFF_WORKSITE_NAME)
        assert bulletin_sheet in workbook.sheetnames
        assert memory_sheet in workbook.sheetnames
    finally:
        workbook.close()
    fence_quantity = _bulletin_quantity(
        chain.workbook_path, chain.template, SYNTHETIC_TAKEOFF_WORKSITE_NAME, fence_line.code
    )
    assert fence_quantity == 58.5

    # 9. build-amendment-dossier: o gramado (código rejeitado por falta de cotação) é o
    #    único item do dossiê, com a nota da rejeição como justificativa — nunca preço.
    assert [item.item_id for item in chain.dossier.items] == [_LAWN_ITEM_ID]
    lawn_item = chain.dossier.items[0]
    assert lawn_item.justification == "sem cotação aplicável no contrato sintético"
    assert lawn_item.decision.action == "reject"
    assert "unit_price" not in chain.dossier_path.read_text(encoding="utf-8")
    assert "total" not in chain.dossier_path.read_text(encoding="utf-8")


def test_export_before_approval_refuses(
    chain: ChainArtifacts, tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    """`export-valuation` antes da aprovação: VALUATION_EXPORT_BLOCKED com o motivo na lista."""
    unapproved_path = _write_json(
        tmp_path / "inputs" / VALUATION_FILENAME,
        chain.valuation_unapproved.model_dump_json(indent=2),
    )
    output_dir = tmp_path / "export"

    exit_code = main(
        [
            "export-valuation",
            "--valuation",
            str(unapproved_path),
            "--contract",
            str(chain.contract_path),
            "--catalog",
            str(chain.catalog_path),
            "--output",
            str(output_dir),
        ]
    )

    assert exit_code == 2
    payload = _stdout(capsys)
    assert payload["refused"] == "VALUATION_EXPORT_BLOCKED"
    errors = payload["errors"]
    assert isinstance(errors, list)
    assert "VALUATION_NOT_APPROVED" in errors
    assert not (output_dir / WORKBOOK_FILENAME).exists()


def test_confirm_codes_redecision_refuses(
    chain: ChainArtifacts, tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    """`confirm-codes` de novo com `--previous` e o MESMO lote: re-decisão é recusada."""
    output_dir = tmp_path / "redecision"

    exit_code = main(
        [
            "confirm-codes",
            "--packet",
            str(chain.reviewed_packet_path),
            "--decisions",
            str(chain.assignment_decisions_path),
            "--catalog",
            str(chain.catalog_path),
            "--contract",
            str(chain.contract_path),
            "--previous",
            str(chain.assignments_path),
            "--output",
            str(output_dir),
        ]
    )

    assert exit_code == 2
    payload = _stdout(capsys)
    assert payload["refused"] == "ASSIGNMENT_ITEM_ALREADY_DECIDED"
    assert not (output_dir / CODE_ASSIGNMENTS_FILENAME).exists()


def test_confirm_codes_code_out_of_contract_refuses(
    chain: ChainArtifacts, tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    """Código do catálogo fora dos 10 do consolidado: CODE_NOT_IN_CONTRACT."""
    batch = CodeAssignmentBatch(
        assignments=[
            CodeAssignmentInput(
                item_id=_PAVEMENT_ITEM_ID,
                action="confirm",
                code=_OUT_OF_CONTRACT_CATALOG_CODE,
                reviewer_id=SYNTHETIC_TAKEOFF_REVIEWER,
                reviewer_role="orcamentista",
                decided_at=SYNTHETIC_CODE_DECIDED_AT,
            )
        ]
    )
    decisions_path = _write_json(
        tmp_path / "inputs" / "decisions.json", batch.model_dump_json(indent=2)
    )
    output_dir = tmp_path / "out-of-contract"

    exit_code = main(
        [
            "confirm-codes",
            "--packet",
            str(chain.reviewed_packet_path),
            "--decisions",
            str(decisions_path),
            "--catalog",
            str(chain.catalog_path),
            "--contract",
            str(chain.contract_path),
            "--output",
            str(output_dir),
        ]
    )

    assert exit_code == 2
    payload = _stdout(capsys)
    assert payload["refused"] == "CODE_NOT_IN_CONTRACT"
    assert not (output_dir / CODE_ASSIGNMENTS_FILENAME).exists()


def test_confirm_codes_unit_incompatible_without_note_refuses(
    chain: ChainArtifacts, tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    """Código do contrato com unidade incompatível e sem nota: recusa fail-closed."""
    batch = CodeAssignmentBatch(
        assignments=[
            CodeAssignmentInput(
                item_id=_BENCH_ITEM_ID,
                action="confirm",
                code=_INCOMPATIBLE_UNIT_CONTRACT_CODE,
                reviewer_id=SYNTHETIC_TAKEOFF_REVIEWER,
                reviewer_role="orcamentista",
                decided_at=SYNTHETIC_CODE_DECIDED_AT,
            )
        ]
    )
    decisions_path = _write_json(
        tmp_path / "inputs" / "decisions.json", batch.model_dump_json(indent=2)
    )
    output_dir = tmp_path / "unit-incompatible"

    exit_code = main(
        [
            "confirm-codes",
            "--packet",
            str(chain.reviewed_packet_path),
            "--decisions",
            str(decisions_path),
            "--catalog",
            str(chain.catalog_path),
            "--contract",
            str(chain.contract_path),
            "--output",
            str(output_dir),
        ]
    )

    assert exit_code == 2
    payload = _stdout(capsys)
    assert payload["refused"] == "ASSIGNMENT_UNIT_INCOMPATIBLE_WITHOUT_NOTE"
    assert not (output_dir / CODE_ASSIGNMENTS_FILENAME).exists()


def test_build_calc_missing_assignment_refuses(
    chain: ChainArtifacts, tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    """Lote de código parcial (sem a luminária), montado à parte: CALC_ASSIGNMENT_MISSING."""
    batch = CodeAssignmentBatch(
        assignments=[
            CodeAssignmentInput(
                item_id=item_id,
                action=action,
                code=code,
                reviewer_id=SYNTHETIC_TAKEOFF_REVIEWER,
                reviewer_role="orcamentista",
                decided_at=SYNTHETIC_CODE_DECIDED_AT,
                note=None if action == "confirm" else "sem cotação aplicável no contrato sintético",
            )
            for item_id, action, code in _MISSING_ASSIGNMENT_BATCH
        ]
    )
    decisions_path = _write_json(
        tmp_path / "inputs" / "decisions.json", batch.model_dump_json(indent=2)
    )
    partial_assign_dir = tmp_path / "partial-assign"
    assert (
        main(
            [
                "confirm-codes",
                "--packet",
                str(chain.reviewed_packet_path),
                "--decisions",
                str(decisions_path),
                "--catalog",
                str(chain.catalog_path),
                "--contract",
                str(chain.contract_path),
                "--output",
                str(partial_assign_dir),
            ]
        )
        == 0
    )
    partial_assignments_path = partial_assign_dir / CODE_ASSIGNMENTS_FILENAME

    output_dir = tmp_path / "build-calc"
    exit_code = main(
        [
            "build-calc",
            "--packet",
            str(chain.reviewed_packet_path),
            "--assignments",
            str(partial_assignments_path),
            "--catalog",
            str(chain.catalog_path),
            "--worksite-key",
            SYNTHETIC_TAKEOFF_WORKSITE_KEY,
            "--worksite-name",
            SYNTHETIC_TAKEOFF_WORKSITE_NAME,
            "--period-number",
            str(chain.contract.next_period_number),
            "--reference-label",
            SYNTHETIC_PERIOD_REFERENCE_LABEL,
            "--output",
            str(output_dir),
        ]
    )

    assert exit_code == 2
    payload = _stdout(capsys)
    assert payload["refused"] == "CALC_ASSIGNMENT_MISSING"
    details = payload["details"]
    assert isinstance(details, dict)
    assert details["item_ids"] == [_LAMP_ITEM_ID]
    assert not (output_dir / VALUATION_FILENAME).exists()

    # O MESMO lote parcial (sem a luminária) exercita a recusa irmã do dossiê: ele é
    # artefato de FECHAMENTO da rodada e recusa foto parcial tanto quanto o boletim.
    dossier_output_dir = tmp_path / "build-amendment-dossier"
    dossier_exit_code = main(
        [
            "build-amendment-dossier",
            "--packet",
            str(chain.reviewed_packet_path),
            "--assignments",
            str(partial_assignments_path),
            "--output",
            str(dossier_output_dir),
        ]
    )

    assert dossier_exit_code == 2
    dossier_payload = _stdout(capsys)
    assert dossier_payload["refused"] == "AMENDMENT_DOSSIER_ASSIGNMENTS_INCOMPLETE"
    dossier_details = dossier_payload["details"]
    assert isinstance(dossier_details, dict)
    assert dossier_details["item_ids"] == [_LAMP_ITEM_ID]
    assert not (dossier_output_dir / AMENDMENT_DOSSIER_FILENAME).exists()


def test_suggest_codes_before_review_refuses(
    chain: ChainArtifacts, tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    """`suggest-codes` sobre o pacote ainda proposto (antes de `review-takeoff`)."""
    output_dir = tmp_path / "suggest-before-review"

    exit_code = main(
        [
            "suggest-codes",
            "--packet",
            str(chain.extracted_packet_path),
            "--catalog",
            str(chain.catalog_path),
            "--output",
            str(output_dir),
        ]
    )

    assert exit_code == 2
    payload = _stdout(capsys)
    assert payload["refused"] == "SUGGESTION_NO_CONFIRMED_ITEMS"
    assert not (output_dir / CODE_SUGGESTIONS_FILENAME).exists()


# --------------------------------------------------------------------------------------
# M8: a cadeia do ORÇAMENTO-BASE (pré-licitação) pelos mesmos comandos do CLI
# --------------------------------------------------------------------------------------


@dataclass(frozen=True, slots=True)
class EstimateChainArtifacts:
    """Cascata importada e o orçamento-base publicado por uma passada da cadeia nova."""

    catalog_paths: tuple[Path, ...]
    cascade: tuple[PriceCatalog, ...]
    assignments_path: Path
    assignments: CodeAssignmentSet
    estimate_path: Path
    estimate: Estimate


@pytest.fixture(scope="module")
def estimate_chain(
    chain: ChainArtifacts, tmp_path_factory: pytest.TempPathFactory
) -> EstimateChainArtifacts:
    """Percorre a cadeia do orçamento-base pelos comandos reais, sobre a MESMA prancha.

    O takeoff revisado e o catálogo SCO vêm da fixture `chain` — é a mesma obra, e repetir
    a extração da prancha só pagaria o passo caro de novo. O que este encaixe prova é o do
    M8: `import-emop` e `import-compositions` publicam catálogos que `confirm-codes` e
    `build-estimate` leem do disco, com a cascata declarada na ordem dos `--catalog`.
    """
    root = tmp_path_factory.mktemp("estimate-chain")

    # 1. import-emop: segunda fonte, a partir dos bytes .DBF sintéticos.
    emop_dir = root / "emop"
    dbf_path = write_emop_dbf(root / "inputs" / "emop-sintetico.dbf")
    layout_path = _write_json(
        root / "inputs" / "emop-layout.json", emop_fixture_layout().model_dump_json(indent=2)
    )
    assert (
        main(
            [
                "import-emop",
                "--input",
                str(dbf_path),
                "--layout",
                str(layout_path),
                "--output",
                str(emop_dir),
            ]
        )
        == 0
    )

    # 2. import-compositions: terceira fonte, compilada das composições do orçamentista.
    compositions_dir = root / "compositions"
    compositions_path = _write_json(
        root / "inputs" / "compositions.json",
        build_synthetic_composition_set().model_dump_json(indent=2),
    )
    assert (
        main(
            [
                "import-compositions",
                "--input",
                str(compositions_path),
                "--output",
                str(compositions_dir),
            ]
        )
        == 0
    )

    catalog_paths = (
        chain.catalog_path,
        emop_dir / CATALOG_FILENAME,
        compositions_dir / CATALOG_FILENAME,
    )
    cascade = tuple(
        PriceCatalog.model_validate_json(path.read_text(encoding="utf-8")) for path in catalog_paths
    )

    # 3. confirm-codes sobre a cascata: cada confirmação cita a fonte que precifica o item.
    decisions_path = _write_json(
        root / "inputs" / "estimate-code-decisions.json",
        build_demo_estimate_assignments(chain.reviewed_packet, cascade).model_dump_json(indent=2),
    )
    assign_dir = root / "assign"
    assert (
        main(
            [
                "confirm-codes",
                "--packet",
                str(chain.reviewed_packet_path),
                "--decisions",
                str(decisions_path),
                *_catalog_args(catalog_paths),
                "--output",
                str(assign_dir),
            ]
        )
        == 0
    )
    assignments_path = assign_dir / CODE_ASSIGNMENTS_FILENAME
    assignments = CodeAssignmentSet.model_validate_json(
        assignments_path.read_text(encoding="utf-8")
    )

    # 4. build-estimate: orçamento-base da obra, sem contrato, sem período e sem aprovação.
    calc_plan_path = _write_json(
        root / "inputs" / CALC_PLAN_FILENAME,
        build_demo_calc_plan(chain.reviewed_packet).model_dump_json(indent=2),
    )
    estimate_dir = root / "estimate"
    assert (
        main(
            [
                "build-estimate",
                "--packet",
                str(chain.reviewed_packet_path),
                "--assignments",
                str(assignments_path),
                *_catalog_args(catalog_paths),
                "--worksite-key",
                SYNTHETIC_ESTIMATE_WORKSITE_KEY,
                "--worksite-name",
                SYNTHETIC_ESTIMATE_WORKSITE_NAME,
                "--bdi",
                str(SYNTHETIC_ESTIMATE_BDI_PERCENT),
                "--calc-plan",
                str(calc_plan_path),
                "--output",
                str(estimate_dir),
            ]
        )
        == 0
    )
    estimate_path = estimate_dir / ESTIMATE_FILENAME
    estimate = Estimate.model_validate_json(estimate_path.read_text(encoding="utf-8"))

    return EstimateChainArtifacts(
        catalog_paths=catalog_paths,
        cascade=cascade,
        assignments_path=assignments_path,
        assignments=assignments,
        estimate_path=estimate_path,
        estimate=estimate,
    )


def test_estimate_chain_happy_path(estimate_chain: EstimateChainArtifacts) -> None:
    """Um comando por elo, exit 0 em todos (já verificado na fixture `estimate_chain`)."""
    # 1-2. As duas fontes novas nascem com a origem que a cadeia licitada recusa.
    assert [catalog.origin for catalog in estimate_chain.cascade] == [
        PriceOrigin.SCO,
        PriceOrigin.EMOP,
        PriceOrigin.COMPOSITION,
    ]

    # 3. confirm-codes: cada confirmação carrega a fonte citada; a rejeição não cita nenhuma.
    confirmed = [item for item in estimate_chain.assignments.assignments if item.code is not None]
    assert confirmed and all(item.catalog_sha256 is not None for item in confirmed)
    rejected = [item for item in estimate_chain.assignments.assignments if item.code is None]
    assert rejected and all(item.catalog_sha256 is None for item in rejected)

    # 4. build-estimate: as três origens precificam linhas e a proveniência fecha com a
    #    cascata declarada, catálogo por catálogo.
    estimate = estimate_chain.estimate
    digests = {catalog.origin: catalog.source_sha256 for catalog in estimate_chain.cascade}
    assert {line.price_origin for line in estimate.lines} == set(digests)
    for line in estimate.lines:
        assert line.catalog_sha256 == digests[line.price_origin]
    assert [source.origin for source in estimate.cascade] == list(digests)
    # O item cujo código foi rejeitado em toda a cascata sai declarado, nunca precificado.
    assert estimate.unpriced_item_ids == [_LAMP_ITEM_ID]
    assert estimate.total_amount == sum(line.total for line in estimate.lines)
    # Orçamento-base não é medição: nada aqui carrega contrato, período ou aprovação.
    assert estimate.plate_id == SYNTHETIC_PLATE_ID


def test_build_estimate_refuses_a_cited_source_missing_from_the_cascade(
    chain: ChainArtifacts,
    estimate_chain: EstimateChainArtifacts,
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
) -> None:
    """Rodar o orçamento sem a fonte que uma confirmação citou não vira preço de outra.

    A mesma confirmação, com a cascata sem o catálogo de composição: em vez de o gramado
    cair para o preço de algum item parecido do SCO, o comando recusa fechado e o
    diretório de saída fica intacto.
    """
    output_dir = tmp_path / "estimate-sem-composicao"

    exit_code = main(
        [
            "build-estimate",
            "--packet",
            str(chain.reviewed_packet_path),
            "--assignments",
            str(estimate_chain.assignments_path),
            *_catalog_args(estimate_chain.catalog_paths[:2]),
            "--worksite-key",
            SYNTHETIC_ESTIMATE_WORKSITE_KEY,
            "--worksite-name",
            SYNTHETIC_ESTIMATE_WORKSITE_NAME,
            "--bdi",
            str(SYNTHETIC_ESTIMATE_BDI_PERCENT),
            "--output",
            str(output_dir),
        ]
    )

    assert exit_code == 2
    assert _stdout(capsys)["refused"] == "ASSIGNMENT_CATALOG_UNKNOWN"
    assert not (output_dir / ESTIMATE_FILENAME).exists()


def test_estimate_chain_with_five_sources_including_sinapi_and_sicro(
    chain: ChainArtifacts,
    estimate_chain: EstimateChainArtifacts,
    tmp_path: Path,
) -> None:
    """F-026/T3: `import-sinapi` + `import-sicro` fecham a cascata do orçamento em cinco fontes.

    Reaproveita o SCO/EMOP/COMPOSIÇÃO já importados por `estimate_chain` (mesmos bytes,
    mesmos digests) e soma as duas fontes novas por cima, pelos comandos `import-sinapi`/
    `import-sicro` reais — nunca por chamada direta ao domínio, como o resto deste arquivo.
    A decisão do banco de concreto é redecidida, só nesta cascata, para citar a SINAPI (em
    vez da EMOP que `build_demo_estimate_assignments` decide por padrão): é essa
    redecisão que prova que a origem nova precifica uma linha de verdade
    (`price_origin == "sinapi"`) e não só ocupa um lugar vazio na cascata. Nenhuma fonte
    cita a SICRO — o critério aceito para este arquivo é a cascata de cinco fontes fechar
    no `build-estimate`, não que toda origem precifique algo — mas a SICRO ainda entra na
    cascata e é conferida pela origem, literalmente.

    `test_estimate_chain_happy_path` continua rodando, sem alteração, sobre a cascata de
    três fontes original de `estimate_chain`.
    """
    root = tmp_path / "estimate-chain-five-sources"

    sinapi_dir = root / "sinapi"
    sinapi_input_path = write_sinapi_xlsx(root / "inputs" / "sinapi-sintetico.xlsx")
    sinapi_layout_path = _write_json(
        root / "inputs" / "sinapi-layout.json",
        sinapi_fixture_layout().model_dump_json(indent=2),
    )
    assert (
        main(
            [
                "import-sinapi",
                "--input",
                str(sinapi_input_path),
                "--layout",
                str(sinapi_layout_path),
                "--output",
                str(sinapi_dir),
            ]
        )
        == 0
    )

    sicro_dir = root / "sicro"
    sicro_input_path = write_sicro_xlsx(root / "inputs" / "sicro-sintetico.xlsx")
    sicro_layout_path = _write_json(
        root / "inputs" / "sicro-layout.json",
        sicro_fixture_layout().model_dump_json(indent=2),
    )
    assert (
        main(
            [
                "import-sicro",
                "--input",
                str(sicro_input_path),
                "--layout",
                str(sicro_layout_path),
                "--output",
                str(sicro_dir),
            ]
        )
        == 0
    )

    catalog_paths = (
        *estimate_chain.catalog_paths,
        sinapi_dir / CATALOG_FILENAME,
        sicro_dir / CATALOG_FILENAME,
    )
    cascade = tuple(
        PriceCatalog.model_validate_json(path.read_text(encoding="utf-8")) for path in catalog_paths
    )
    assert [catalog.origin for catalog in cascade] == [
        PriceOrigin.SCO,
        PriceOrigin.EMOP,
        PriceOrigin.COMPOSITION,
        PriceOrigin.SINAPI,
        PriceOrigin.SICRO,
    ]
    sinapi_catalog = cascade[3]

    decisions = build_demo_estimate_assignments(chain.reviewed_packet, cascade)
    decisions = CodeAssignmentBatch(
        assignments=[
            assignment.model_copy(
                update={
                    "code": "0009012",
                    "catalog_sha256": sinapi_catalog.source_sha256,
                    "note": "mobiliario cotado pela SINAPI nesta pre-licitacao",
                }
            )
            if assignment.item_id == _BENCH_ITEM_ID
            else assignment
            for assignment in decisions.assignments
        ]
    )
    decisions_path = _write_json(
        root / "inputs" / "estimate-code-decisions.json", decisions.model_dump_json(indent=2)
    )

    assign_dir = root / "assign"
    assert (
        main(
            [
                "confirm-codes",
                "--packet",
                str(chain.reviewed_packet_path),
                "--decisions",
                str(decisions_path),
                *_catalog_args(catalog_paths),
                "--output",
                str(assign_dir),
            ]
        )
        == 0
    )
    assignments_path = assign_dir / CODE_ASSIGNMENTS_FILENAME

    calc_plan_path = _write_json(
        root / "inputs" / CALC_PLAN_FILENAME,
        build_demo_calc_plan(chain.reviewed_packet).model_dump_json(indent=2),
    )
    estimate_dir = root / "estimate"
    assert (
        main(
            [
                "build-estimate",
                "--packet",
                str(chain.reviewed_packet_path),
                "--assignments",
                str(assignments_path),
                *_catalog_args(catalog_paths),
                "--worksite-key",
                SYNTHETIC_ESTIMATE_WORKSITE_KEY,
                "--worksite-name",
                SYNTHETIC_ESTIMATE_WORKSITE_NAME,
                "--bdi",
                str(SYNTHETIC_ESTIMATE_BDI_PERCENT),
                "--calc-plan",
                str(calc_plan_path),
                "--output",
                str(estimate_dir),
            ]
        )
        == 0
    )
    estimate_path = estimate_dir / ESTIMATE_FILENAME
    estimate = Estimate.model_validate_json(estimate_path.read_text(encoding="utf-8"))

    assert [source.origin for source in estimate.cascade] == [
        PriceOrigin.SCO,
        PriceOrigin.EMOP,
        PriceOrigin.COMPOSITION,
        PriceOrigin.SINAPI,
        PriceOrigin.SICRO,
    ]
    sinapi_lines = [line for line in estimate.lines if line.price_origin == PriceOrigin.SINAPI]
    assert len(sinapi_lines) == 1
    sinapi_line = sinapi_lines[0]
    assert sinapi_line.price_origin.value == "sinapi"
    assert sinapi_line.code == "0009012"
    assert sinapi_line.catalog_sha256 == sinapi_catalog.source_sha256
