"""Importador offline do catálogo SICRO: leitor .xlsx mínimo + catálogo canônico.

A tabela SICRO (Sistema de Custos Referenciais de Obras, DNIT) é publicada periodicamente
e é de distribuição pública, mas o arquivo real não existe neste repositório — tudo aqui é
sintético (`croquito_worker.valuation.sicro_fixture`), e o layout real (aba, linha de
cabeçalho, colunas, padrão de código) fecha como DADO em `SicroCatalogLayout`, nunca como
constante de código (ADR-0039, decisão 2 — o mesmo desenho do importador EMOP,
`croquito_valuation.emop`, que é o molde deste módulo; ver também `sinapi.py`, o
irmão desta fonte, deliberadamente NÃO compartilhado por implementação — ADR-0039,
alternativas rejeitadas: "um importador único multi-formato").

Regra da orçamentista (M8, `docs/architecture/VALUATION_CONTEXT.md`): em obra LICITADA o
contrato manda — preço nunca vem da SICRO. Este módulo só importa o catálogo com
`origin=PriceOrigin.SICRO` (`models.py`); é o guardrail em
`calc.py::build_worksite_bulletin` e `workbook_writer.py::_validate_against_catalog`
(`BULLETIN_PRICE_ORIGIN_FORBIDDEN`) quem impede esse catálogo de entrar na cadeia da
medição. A cadeia SCO → EMOP → SINAPI → SICRO → composição só vale PRÉ-licitação
(orçamento-base, fase futura) e fora do escopo deste módulo.

## O leitor .xlsx

Implementação mínima sobre `openpyxl` (`read_only`, `data_only`), só do que o catálogo
precisa: uma aba declarada, uma linha de cabeçalho declarada e quatro colunas por letra
(código, descrição, unidade, preço) — qualquer aba ou coluna declarada fora da tabela real
recusa fechado (`SICRO_SHEET_MISSING`/`SICRO_FIELD_MISSING`). Diferente do leitor de
catálogo do MAPÃO (`catalog.py::parse_money`), que aceita `float` de célula numérica e
converte via `Decimal(str(...))`, o preço aqui é lido como TEXTO e convertido direto para
`Decimal` — nunca por `float` — coerente com `ExactDecimal` (`models.py`): uma célula de
preço que o openpyxl devolve como `float`/`bool` recusa a linha inteira
(`SICRO_ROW_UNPARSEABLE`) em vez de converter. Linha inteiramente em branco (as quatro
colunas vazias) é pulada e contada — nunca vira entrada do catálogo, o mesmo tratamento
que o registro deletado tem no EMOP.

## Família e subgrupo

O layout declarado no contrato desta fonte (`SicroCatalogLayout`) não tem coluna de
família/subgrupo — ao contrário do EMOP, cuja tabela paga já publica essa hierarquia em
campos próprios. `PriceCatalogEntry.family_code`/`family_name`/`subgroup_code`/
`subgroup_name` continuam obrigatórios no modelo canônico (não têm default), então este
importador preenche os quatro com um valor único e determinístico por catálogo
(`SICRO_FAMILY_CODE` e `layout.source_label`): a tabela inteira nasce em uma família só.
É uma decisão consciente de escopo mínimo, registrada no Task Contract (F-026/T2) como
ASSUNÇÃO — nada na cadeia da medição ou na cascata do orçamento-base (fase futura) lê
família/subgrupo de um catálogo `sicro`, então a simplificação não perde informação que
algum consumidor atual precise.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Final, NoReturn
from uuid import NAMESPACE_URL, UUID, uuid5
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import Field, ValidationError, model_validator

from croquito_valuation.catalog import file_sha256
from croquito_valuation.errors import ValuationValidationError
from croquito_valuation.models import (
    REFERENCE_MONTH_PATTERN,
    PriceCatalog,
    PriceCatalogEntry,
    PriceOrigin,
    ValuationContractModel,
)
from croquito_valuation.template import COLUMN_LETTER_PATTERN, MAX_SHEET_NAME_LENGTH

SICRO_FAMILY_CODE: Final = "SICRO"
"""Família/subgrupo únicos que todo catálogo `sicro` importado por este módulo recebe —
ver a seção "Família e subgrupo" no docstring do módulo."""

_SICRO_CATALOG_ID_NAMESPACE: Final = uuid5(
    NAMESPACE_URL, "https://croquito.local/valuation/catalog/sicro"
)


def sicro_catalog_id_for(source_sha256: str) -> UUID:
    """Id derivado do conteúdo do .xlsx: reimportar o mesmo arquivo devolve o mesmo catálogo.

    Namespace próprio, distinto do `catalog_id_for` (SCO), do `emop_catalog_id_for` e do
    `sinapi_catalog_id_for`: as fontes nunca deveriam colidir de id mesmo que o SHA-256 dos
    bytes coincidisse por acaso, porque são catálogos de naturezas diferentes.
    """
    return uuid5(_SICRO_CATALOG_ID_NAMESPACE, source_sha256)


class SicroCatalogLayout(ValuationContractModel):
    """Layout do catálogo SICRO como DADO: aba, cabeçalho e colunas do .xlsx.

    Espelha `EmopCatalogLayout` (`emop.py`) e `SinapiCatalogLayout` (`sinapi.py`): nenhuma
    posição é constante do código. `code_pattern` é o padrão REAL do código SICRO — regex
    como dado, validada na construção com o mesmo cinto e suspensório de
    `EmopCatalogLayout.validate_code_pattern` (compila e não casa a string vazia). Toda
    linha lida ainda revalida contra o superset estrutural de
    `PriceCatalogEntry.validate_code_for_origin` (`NON_SCO_CODE_PATTERN`): um `code_pattern`
    frouxo demais aqui não basta para injetar um código fora da estrutura.
    """

    source_label: str = Field(min_length=1, max_length=200)
    reference_month: str = Field(pattern=REFERENCE_MONTH_PATTERN)
    sheet_name: str = Field(min_length=1, max_length=MAX_SHEET_NAME_LENGTH)
    header_row: int = Field(ge=1)
    code_column: str = Field(pattern=COLUMN_LETTER_PATTERN)
    description_column: str = Field(pattern=COLUMN_LETTER_PATTERN)
    unit_column: str = Field(pattern=COLUMN_LETTER_PATTERN)
    price_column: str = Field(pattern=COLUMN_LETTER_PATTERN)
    code_pattern: str = Field(min_length=1, max_length=200)

    @model_validator(mode="after")
    def validate_code_pattern(self) -> SicroCatalogLayout:
        try:
            compiled = re.compile(self.code_pattern)
        except re.error as error:
            raise ValuationValidationError(
                "SICRO_LAYOUT_CODE_PATTERN_INVALID",
                "padrão de código do layout SICRO não compila como expressão regular",
                {"pattern": self.code_pattern, "reason": str(error)},
            ) from error
        if compiled.fullmatch("") is not None:
            raise ValuationValidationError(
                "SICRO_LAYOUT_CODE_PATTERN_INVALID",
                "padrão de código do layout SICRO casa a string vazia; frouxo demais "
                "para distinguir um código real",
                {"pattern": self.code_pattern},
            )
        return self

    @property
    def declared_columns(self) -> tuple[str, ...]:
        """Letras das colunas do .xlsx que o layout precisa encontrar na tabela."""
        return (self.code_column, self.description_column, self.unit_column, self.price_column)


@dataclass(frozen=True, slots=True)
class SicroImportNotes:
    """O que a importação do .xlsx observou sem recusar: só as linhas em branco.

    Espelho minúsculo de `EmopImportNotes` (`emop.py`): nada aqui muda o catálogo
    devolvido, é o que o relatório do comando precisa para declarar a contagem.
    """

    blank_row_count: int
    blank_rows: tuple[int, ...]


def _xlsx_unsupported(reason: str, **details: object) -> NoReturn:
    raise ValuationValidationError("SICRO_XLSX_UNSUPPORTED", reason, details)


def _row_error(row_number: int, reason: str, **details: object) -> NoReturn:
    raise ValuationValidationError(
        "SICRO_ROW_UNPARSEABLE",
        f"linha ilegível no catálogo SICRO (SICRO_ROW_UNPARSEABLE:{row_number}): {reason}",
        {"row": row_number, "reason": reason, **details},
    )


def _load_workbook(path: Path) -> Any:
    try:
        return load_workbook(path, read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, KeyError, OSError) as error:
        _xlsx_unsupported(
            "arquivo não pôde ser lido como .xlsx (dBASE III seria outro leitor; este é "
            "um leitor de planilha)",
            cause=str(error),
        )


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _row_price(row_number: int, code: str, raw_price: object) -> Decimal:
    """Converte a célula de preço em `Decimal` sem NUNCA passar por `float`.

    Célula numérica do openpyxl (`float`/`bool`) recusa a linha: o preço tem que chegar
    como texto (ou já como `Decimal`/`int`, que não perdem precisão), igual ao campo `N`
    do .DBF do EMOP, que é lido como texto e nunca convertido por binário.
    """
    if isinstance(raw_price, bool | float):
        _row_error(
            row_number,
            "preço viria de célula numérica binária (float) do xlsx; a coluna de preço "
            "precisa trazer texto",
            code=code,
            raw_price=repr(raw_price),
        )
    if isinstance(raw_price, str):
        try:
            return Decimal(raw_price.strip())
        except InvalidOperation:
            _row_error(row_number, "preço não numérico", code=code, raw_price=raw_price)
    if isinstance(raw_price, int):
        return Decimal(raw_price)
    if isinstance(raw_price, Decimal):
        return raw_price
    _row_error(
        row_number, "preço ausente ou de tipo inesperado", code=code, raw_price=repr(raw_price)
    )


def read_sicro_catalog_with_report(
    xlsx_path: Path, layout: SicroCatalogLayout
) -> tuple[PriceCatalog, SicroImportNotes]:
    """Lê a tabela .xlsx do catálogo SICRO e devolve o catálogo canônico e as notas da leitura.

    Fail-closed, na ordem: estrutura do .xlsx em si (`SICRO_XLSX_UNSUPPORTED`), aba
    declarada ausente (`SICRO_SHEET_MISSING`), coluna declarada fora da extensão real da
    tabela (`SICRO_FIELD_MISSING`), depois linha a linha — linha inteiramente em branco é
    pulada e contada (nunca importada), código fora do `code_pattern` do layout, preço não
    numérico ou negativo (nunca via `float`), descrição vazia e por fim qualquer outra
    violação estrutural do modelo (`SICRO_ROW_UNPARSEABLE` em todos os casos, com o
    número da linha). Tabela sem nenhuma entrada válida recusa (`SICRO_EMPTY`); código
    duplicado recusa pelo `CATALOG_DUPLICATE_CODE` que o próprio `PriceCatalog` já garante.
    O catálogo devolvido nasce inteiro com `origin=PriceOrigin.SICRO` — é esse dado que os
    guardrails de `calc.py` e `workbook_writer.py` usam para recusar este catálogo na
    cadeia da medição licitada.
    """
    source_sha256 = file_sha256(xlsx_path)
    workbook = _load_workbook(xlsx_path)
    try:
        if layout.sheet_name not in workbook.sheetnames:
            raise ValuationValidationError(
                "SICRO_SHEET_MISSING",
                "planilha SICRO não possui a aba declarada no layout",
                {"sheet": layout.sheet_name, "available": list(workbook.sheetnames)},
            )
        worksheet = workbook[layout.sheet_name]
        max_column = worksheet.max_column or 0
        missing = sorted(
            {
                letter
                for letter in layout.declared_columns
                if column_index_from_string(letter) > max_column
            }
        )
        if missing:
            raise ValuationValidationError(
                "SICRO_FIELD_MISSING",
                "coluna declarada no layout SICRO está fora da extensão real da tabela",
                {"columns": missing, "sheet": layout.sheet_name, "max_column": max_column},
            )

        code_index = column_index_from_string(layout.code_column) - 1
        description_index = column_index_from_string(layout.description_column) - 1
        unit_index = column_index_from_string(layout.unit_column) - 1
        price_index = column_index_from_string(layout.price_column) - 1
        needed = max(code_index, description_index, unit_index, price_index) + 1

        entries: list[PriceCatalogEntry] = []
        blank_rows: list[int] = []
        row_number = layout.header_row
        for row in worksheet.iter_rows(min_row=layout.header_row + 1, values_only=True):
            row_number += 1
            cells: list[object] = list(row) + [None] * (needed - len(row))
            code = _cell_text(cells[code_index])
            description = _cell_text(cells[description_index])
            unit = _cell_text(cells[unit_index])
            raw_price = cells[price_index]
            if not code and not description and not unit and raw_price is None:
                blank_rows.append(row_number)
                continue

            if not re.fullmatch(layout.code_pattern, code):
                _row_error(row_number, "código fora do padrão declarado no layout", code=code)

            unit_price = _row_price(row_number, code, raw_price)
            if unit_price < 0:
                _row_error(row_number, "preço negativo", code=code, raw_price=str(unit_price))

            if not description:
                _row_error(row_number, "linha sem descrição", code=code)

            try:
                entry = PriceCatalogEntry(
                    code=code,
                    description=description,
                    unit=unit,
                    unit_price=unit_price,
                    family_code=SICRO_FAMILY_CODE,
                    family_name=layout.source_label,
                    subgroup_code=SICRO_FAMILY_CODE,
                    subgroup_name=layout.source_label,
                    origin=PriceOrigin.SICRO,
                )
            except (ValidationError, ValuationValidationError) as error:
                _row_error(
                    row_number,
                    "linha não corresponde ao contrato de entrada do catálogo",
                    code=code,
                    cause=str(error),
                )
            entries.append(entry)
    finally:
        workbook.close()

    if not entries:
        raise ValuationValidationError(
            "SICRO_EMPTY",
            "tabela .xlsx não produziu nenhuma entrada válida para o catálogo",
            {"source": str(xlsx_path), "blank_rows": len(blank_rows)},
        )

    catalog = PriceCatalog(
        id=sicro_catalog_id_for(source_sha256),
        source_label=layout.source_label,
        reference_month=layout.reference_month,
        source_sha256=source_sha256,
        entries=entries,
        origin=PriceOrigin.SICRO,
    )
    notes = SicroImportNotes(blank_row_count=len(blank_rows), blank_rows=tuple(blank_rows))
    return catalog, notes


def read_sicro_catalog(xlsx_path: Path, layout: SicroCatalogLayout) -> PriceCatalog:
    """Lê o catálogo SICRO a partir do .xlsx, descartando as notas da leitura."""
    catalog, _ = read_sicro_catalog_with_report(xlsx_path, layout)
    return catalog
