"""Importação do consolidado contratual a partir do MAPÃO da prefeitura.

O leitor lê o que a planilha declara — contratado, vigente, o par QUANTIDADE|VALOR de
cada medição já lançada, acumulado e saldo — e não conserta nada. Quem recusa drift é o
modelo (`contract.py`): acumulado que não fecha com os períodos, saldo que não fecha com
vigente menos acumulado e RE-RA que não explica a quantidade vigente fazem a construção
do agregado falhar. É a mesma filosofia do M1: nenhum total informado vale sem recomputo,
e a importação falha fechada.

Falha semântica do agregado não sobe como violação isolada: o arquivo real tem centenas
de linhas de histórico já publicado, e recusar na primeira delas não diz ao orçamentista
o tamanho do problema. Toda recusa dessa natureza é convertida em `ContractSemanticsError`
(`CONTRACT_SEMANTICS_DIVERGENT`) carregando o dossiê completo do recomputo
(`contract_diagnosis`). A recusa continua fechada — o dossiê descreve, não aceita.

Este módulo só recusa por conta própria duas coisas, e nenhuma em silêncio: layout que
não é o declarado no template (aba ausente, cabeçalho de medição fora do padrão, coluna
de saldo no lugar errado) e linha ilegível. Linha ilegível não aborta a varredura: as
falhas são acumuladas com aba e linha e devolvidas juntas, porque o orçamentista corrige
a planilha inteira de uma vez.

O arquivo real do cliente obriga três coisas que este leitor faz explicitamente e
registra em `ContractImportNotes`, em vez de adivinhar:

1. o número da medição vem do rótulo lido (`11ª MEDIÇÃO - COMPLEMENTAR` é a 11ª) e a
   numeração pode ter buraco (13ª → 15ª); a posição do par no cabeçalho não numera nada;
2. a quantidade é normalizada na escala que o template declara e o ruído de ponto
   flutuante do cache (`107,44999999999996`) é absorvido só abaixo de um milionésimo —
   deriva de centavo sobrevive e vai morrer nos validadores do agregado;
3. linha separadora (sem código, descrição, unidade, item nem grupo, mas com zeros de
   fórmula nas colunas numéricas) é pulada e contada, não vira item ilegível.

O contrato real também traz item fora da tabela SCO (`IE00040849`, sem variante) e a aba
da prefeitura tem linha de seção cujo nome ocupa a coluna de código (`LAZER /
PAISAGISMO`). O primeiro entra por padrão declarado no template
(`WorkbookTemplate.extra_code_patterns`/`matches_extra_code`) e é revalidado contra o
superset estrutural (`is_contract_code`) antes de aceito — nenhum padrão frouxo demais
injeta identidade fora da estrutura; a leitura registra a linha e o código aceitos em
`ContractImportNotes.extra_code_rows`. O segundo é layout: linha cujo código não é aceito
e cujos blocos de RE-RA não carregam nenhum valor é pulada e contada em
`ContractImportNotes.amendment_section_rows`, junto com o subtotal de grupo ignorado
quando `AmendmentLayout.section_rows_carry_group_subtotal` declara que a linha que abre um
grupo carrega esse subtotal na coluna de vigente (opt-in, `False` por padrão); sem a flag,
qualquer valor na coluna de vigente também bloqueia o skip. Linha com valor num bloco de
RE-RA continua recusando como ilegível de qualquer forma.
"""

from __future__ import annotations

from collections.abc import Iterator, Mapping, Sequence
from dataclasses import dataclass
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import Any, Final

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from pydantic import ValidationError

from croquito_valuation.catalog import file_sha256, normalize_unit, parse_money
from croquito_valuation.contract import (
    Amendment,
    AmendmentLine,
    ContractLine,
    ContractWorkbook,
    PeriodProgress,
    contract_workbook_id_for,
)
from croquito_valuation.contract_diagnosis import (
    ContractParse,
    ContractSemanticsError,
    PairColumns,
    ParsedAmendmentBlock,
    ParsedAmendmentLine,
    ParsedContractLine,
    ParsedPeriod,
    diagnose_contract,
)
from croquito_valuation.errors import ValuationValidationError
from croquito_valuation.sco import is_contract_code, is_sco_code
from croquito_valuation.template import (
    AmendmentColumns,
    AmendmentLayout,
    GeneralLayout,
    WorkbookTemplate,
)

MAX_ROW_FAILURES: Final = 20
"""Quantas linhas ilegíveis são reportadas antes de truncar a lista."""

MAX_NORMALIZED_CELLS: Final = 50
"""Quantas células normalizadas são detalhadas nas notas antes de truncar a lista."""

MONEY_DECIMAL_SCALE: Final = 2
"""Dinheiro tem duas casas na planilha, qualquer que seja a escala das quantidades."""

NORMALIZATION_EPSILON: Final = Decimal("0.000001")
"""Ruído absorvido na leitura de um número da planilha.

Abaixo dele é erro de ponto flutuante do cache (`107,44999999999996`); a partir dele é
dado, e dado com casa a mais do que o template declara é recusado. O limite é
sub-centavo de propósito: deriva real de centavo tem de sobreviver à normalização e ir
morrer nos validadores de recomputo do agregado.
"""

_BLANK_CHARS: Final = " \t\r\n\xa0\u202f\u2007"
_ZERO: Final = Decimal("0.00")


@dataclass(frozen=True, slots=True)
class _RowFailure:
    """Uma linha ilegível: onde ela está e qual invariante ela quebrou."""

    sheet: str
    row: int
    code: str
    message: str

    def as_details(self) -> dict[str, object]:
        """Forma serializável da falha, para `details` do erro agregado."""
        return {"sheet": self.sheet, "row": self.row, "code": self.code, "message": self.message}


@dataclass(frozen=True, slots=True)
class NormalizedCell:
    """Célula cujo valor bruto foi normalizado na leitura, com o antes e o depois."""

    sheet: str
    ref: str
    raw: str
    normalized: str

    def as_details(self) -> dict[str, object]:
        """Forma serializável da normalização, para relatório e diagnóstico."""
        return {
            "sheet": self.sheet,
            "ref": self.ref,
            "raw": self.raw,
            "normalized": self.normalized,
        }


@dataclass(frozen=True, slots=True)
class ContractImportNotes:
    """O que a importação observou sem recusar: numeração, separadores e normalizações.

    Nada aqui muda o consolidado devolvido. São observações que o orçamentista precisa
    ver — a medição que falta na numeração, a linha que foi pulada, a célula cujo cache
    de fórmula chegou com ruído de ponto flutuante, o código fora da tabela SCO aceito por
    padrão extra do template e a linha de seção pulada na aba da prefeitura.

    `amendment_section_rows` carrega o número da linha e, quando
    `AmendmentLayout.section_rows_carry_group_subtotal` está ligada e a linha tinha um
    valor na coluna de vigente, o subtotal ignorado — `None` quando a coluna estava vazia
    ou a flag não está declarada. O valor nunca é descartado sem aparecer aqui.
    """

    period_numbers: tuple[int, ...]
    period_gaps: tuple[int, ...]
    separator_rows: tuple[int, ...]
    normalized_cells: tuple[NormalizedCell, ...]
    normalized_count: int
    normalized_truncated: bool
    extra_code_rows: tuple[tuple[int, str], ...]
    amendment_section_rows: tuple[tuple[int, Decimal | None], ...]


class _NormalizationLog:
    """Acumulador das células normalizadas; a lista detalhada é limitada, a contagem não."""

    def __init__(self) -> None:
        self.cells: list[NormalizedCell] = []
        self.count = 0
        self.truncated = False

    def record(self, sheet: str, ref: str, raw: Decimal, normalized: Decimal) -> None:
        """Registra a normalização de uma célula."""
        self.count += 1
        if len(self.cells) >= MAX_NORMALIZED_CELLS:
            self.truncated = True
            return
        self.cells.append(
            NormalizedCell(sheet=sheet, ref=ref, raw=str(raw), normalized=str(normalized))
        )


@dataclass(frozen=True, slots=True)
class _PairColumns:
    """Índices 1-based do par QUANTIDADE|VALOR de uma medição.

    `number` é a medição que o cabeçalho declara naquele par — o número lido, não a
    posição. O par ACUMULADO não é medição e fica com zero.
    """

    quantity: int
    amount: int
    number: int = 0


@dataclass(frozen=True, slots=True)
class _GeneralHeader:
    """Colunas da PLANILHA GERAL depois de descobrir que medições ela declara."""

    periods: tuple[_PairColumns, ...]
    accumulated: _PairColumns
    balance: int

    @property
    def period_numbers(self) -> tuple[int, ...]:
        """Números das medições declaradas, na ordem do cabeçalho."""
        return tuple(pair.number for pair in self.periods)


def _to_contract_line(parsed: ParsedContractLine, amended_quantity: Decimal) -> ContractLine:
    """Constrói a linha do consolidado; drift declarado falha aqui, no modelo.

    É a fronteira entre o que foi lido e o que é validado: até aqui número negativo e
    total que não fecha são dado; a partir daqui são recusa — e o dossiê descreve o que
    a recusa encontrou.
    """
    return ContractLine(
        group_label=parsed.group_label,
        item_number=parsed.item_number,
        code=parsed.code,
        description=parsed.description,
        unit=parsed.unit,
        unit_price=parsed.unit_price,
        contract_quantity=parsed.contract_quantity,
        amended_quantity=amended_quantity,
        periods=[
            PeriodProgress(
                period_number=period.number, quantity=period.quantity, amount=period.amount
            )
            for period in parsed.periods
        ],
        accumulated_quantity=parsed.accumulated_quantity,
        accumulated_amount=parsed.accumulated_amount,
        balance_quantity=parsed.balance_quantity,
    )


class _FailureLog:
    """Acumulador de linhas ilegíveis: a varredura não para na primeira."""

    def __init__(self) -> None:
        self.failures: list[_RowFailure] = []
        self.truncated = False

    def record(self, sheet: str, row: int, error: ValuationValidationError) -> None:
        """Registra a falha da linha; acima do limite só marca que houve truncamento."""
        if len(self.failures) >= MAX_ROW_FAILURES:
            self.truncated = True
            return
        self.failures.append(
            _RowFailure(sheet=sheet, row=row, code=error.code, message=error.message)
        )

    def raise_if_any(self) -> None:
        """Uma falha aponta aba e linha; várias voltam juntas para uma correção só."""
        if not self.failures:
            return
        if len(self.failures) == 1:
            failure = self.failures[0]
            raise ValuationValidationError(
                "ROW_UNPARSEABLE",
                f"linha ilegível no consolidado ({failure.sheet}:{failure.row}): {failure.message}",
                {"sheet": failure.sheet, "row": failure.row, "reason": failure.message},
            )
        raise ValuationValidationError(
            "ROWS_UNPARSEABLE",
            "o consolidado possui linhas ilegíveis",
            {
                "failures": [failure.as_details() for failure in self.failures],
                "truncated": self.truncated,
            },
        )


def _cell_text(value: object) -> str:
    """Texto da célula sem os espaços que a planilha carrega; célula vazia vira ``""``."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip(_BLANK_CHARS).strip()
    return str(value).strip()


def _value_at(cells: Sequence[object], column: int) -> object:
    """Valor da coluna 1-based; coluna além do fim da linha é célula vazia."""
    index = column - 1
    if 0 <= index < len(cells):
        return cells[index]
    return None


def _text_at(cells: Sequence[object], column: int) -> str:
    return _cell_text(_value_at(cells, column))


def _item_number_at(cells: Sequence[object], column: int) -> str:
    """Número do item como texto, mesmo quando a planilha o guarda como número."""
    value = _value_at(cells, column)
    if isinstance(value, bool):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    if isinstance(value, Decimal):
        return str(value.normalize())
    return _cell_text(value)


def _normalized(
    value: Decimal, *, scale: int, sheet_name: str, ref: str, log: _NormalizationLog
) -> Decimal:
    """Normaliza o número lido na escala declarada, absorvendo só o ruído do cache.

    A auditoria da pasta gerada continua estrita em duas casas (`canonical_number`);
    aqui a escala é a que o template declara para a planilha do cliente, e nada é
    arredondado em silêncio: o que passa do épsilon é recusado e o que é absorvido fica
    registrado nas notas da importação.
    """
    normalized = value.quantize(Decimal(1).scaleb(-scale), rounding=ROUND_HALF_UP)
    if abs(value - normalized) > NORMALIZATION_EPSILON:
        raise ValuationValidationError(
            "NUMBER_SCALE_UNSUPPORTED",
            "planilha traz número com mais casas decimais do que o template declara",
            {"sheet": sheet_name, "ref": ref, "value": str(value), "scale": scale},
        )
    if normalized != value:
        log.record(sheet_name, ref, value, normalized)
    return normalized


def _read_number_at(
    cells: Sequence[object],
    column: int,
    sheet_name: str,
    row: int,
    *,
    scale: int,
    log: _NormalizationLog,
) -> Decimal:
    """Número da célula em `Decimal`; célula vazia é 0,00 — item que não mediu no período.

    Fronteira de leitura declarada: a planilha devolve `float` e a conversão passa uma
    única vez por `Decimal(str(...))`. Número negativo é convertido normalmente — saldo
    negativo existe no arquivo real e quem o recusa é o modelo, não o leitor.
    """
    value = _value_at(cells, column)
    ref = f"{get_column_letter(column)}{row}"
    if value is None:
        return _ZERO
    if isinstance(value, bool):
        raise ValuationValidationError(
            "CELL_TYPE_UNSUPPORTED",
            "planilha traz booleano onde o consolidado só admite número",
            {"sheet": sheet_name, "ref": ref},
        )
    if isinstance(value, str):
        text = value.strip(_BLANK_CHARS).strip()
        if not text:
            return _ZERO
        return _normalized(parse_money(text), scale=scale, sheet_name=sheet_name, ref=ref, log=log)
    if isinstance(value, int | float | Decimal):
        return _normalized(
            Decimal(str(value)), scale=scale, sheet_name=sheet_name, ref=ref, log=log
        )
    raise ValuationValidationError(
        "CELL_TYPE_UNSUPPORTED",
        "tipo de célula não suportado pelo consolidado",
        {"sheet": sheet_name, "ref": ref, "type": type(value).__name__},
    )


def _read_money_at(
    cells: Sequence[object], column: int, sheet_name: str, row: int, log: _NormalizationLog
) -> Decimal:
    """Coluna de dinheiro: sempre duas casas, qualquer que seja a escala das quantidades."""
    return _read_number_at(cells, column, sheet_name, row, scale=MONEY_DECIMAL_SCALE, log=log)


def _read_quantity_at(
    cells: Sequence[object],
    column: int,
    sheet_name: str,
    row: int,
    scale: int,
    log: _NormalizationLog,
) -> Decimal:
    """Coluna de quantidade, na escala que o template declara para a planilha do cliente."""
    return _read_number_at(cells, column, sheet_name, row, scale=scale, log=log)


def _row_values(worksheet: Any, row_number: int) -> list[object]:
    """Uma linha isolada da aba, já materializada."""
    for row in worksheet.iter_rows(min_row=row_number, max_row=row_number, values_only=True):
        cells: list[object] = list(row)
        return cells
    return []


def _iter_rows(worksheet: Any, first_row: int) -> Iterator[tuple[int, list[object]]]:
    """Linhas da aba a partir de `first_row`, com o número da linha da planilha."""
    row_number = first_row - 1
    for row in worksheet.iter_rows(min_row=first_row, values_only=True):
        row_number += 1
        cells: list[object] = list(row)
        yield row_number, cells


def _require_sheet(workbook: Any, sheet_name: str) -> Any:
    """Aba declarada no template; aba oculta é lida normalmente, aba ausente falha alto."""
    names: list[str] = list(workbook.sheetnames)
    if sheet_name not in names:
        raise ValuationValidationError(
            "CONTRACT_SHEET_MISSING",
            "planilha do contrato não possui a aba declarada no template",
            {"sheet": sheet_name, "available": names},
        )
    return workbook[sheet_name]


def _read_general_header(worksheet: Any, layout: GeneralLayout) -> _GeneralHeader:
    """Descobre os pares de medição pelo cabeçalho: a posição vem do dado, não do código.

    O número de cada medição é o que o rótulo declara, não a posição do par: o arquivo
    real pula a 14ª medição e escreve sufixo depois do rótulo. A única regra de ordem é
    que a numeração cresça — cabeçalho fora de ordem é layout que o leitor não entende.
    """
    header = _row_values(worksheet, layout.header_row)
    first_column = column_index_from_string(layout.first_period_column)
    periods: list[_PairColumns] = []
    while True:
        previous = periods[-1].number if periods else 0
        quantity_column = first_column + 2 * len(periods)
        found = _text_at(header, quantity_column)
        number = layout.parse_period_label(found)
        if number is not None:
            if number <= previous:
                raise ValuationValidationError(
                    "PERIOD_HEADER_UNPARSEABLE",
                    "cabeçalho da planilha geral declara medição fora de ordem",
                    {
                        "sheet": layout.sheet_name,
                        "row": layout.header_row,
                        "column": get_column_letter(quantity_column),
                        "reason": "número de medição fora de ordem no cabeçalho",
                        "found": found,
                        "number": number,
                        "previous": previous,
                    },
                )
            periods.append(
                _PairColumns(quantity=quantity_column, amount=quantity_column + 1, number=number)
            )
            continue
        if found == layout.accumulated_label:
            accumulated = _PairColumns(quantity=quantity_column, amount=quantity_column + 1)
            break
        raise ValuationValidationError(
            "PERIOD_HEADER_UNPARSEABLE",
            "cabeçalho da planilha geral não declara o par de medição esperado",
            {
                "sheet": layout.sheet_name,
                "row": layout.header_row,
                "column": get_column_letter(quantity_column),
                "found": found,
                "expected": f"{layout.period_label(previous + 1)} ou {layout.accumulated_label}",
            },
        )
    balance_column = accumulated.amount + 1
    found_balance = _text_at(header, balance_column)
    if found_balance != layout.balance_label:
        raise ValuationValidationError(
            "GENERAL_COLUMN_MISSING",
            "planilha geral não declara a coluna de saldo depois do acumulado",
            {
                "sheet": layout.sheet_name,
                "row": layout.header_row,
                "column": get_column_letter(balance_column),
                "found": found_balance,
                "expected": layout.balance_label,
            },
        )
    _validate_pair_sublabels(worksheet, layout, periods)
    return _GeneralHeader(periods=tuple(periods), accumulated=accumulated, balance=balance_column)


def _validate_pair_sublabels(
    worksheet: Any, layout: GeneralLayout, periods: Sequence[_PairColumns]
) -> None:
    """Confere os sub-rótulos do primeiro par; uma amostra basta para provar o layout."""
    sublabel_row = layout.pair_sublabel_row
    if sublabel_row is None or not periods:
        return
    sublabels = _row_values(worksheet, sublabel_row)
    first = periods[0]
    for column, expected in (
        (first.quantity, layout.quantity_pair_label),
        (first.amount, layout.amount_pair_label),
    ):
        found = _text_at(sublabels, column)
        if found != expected:
            raise ValuationValidationError(
                "PERIOD_HEADER_UNPARSEABLE",
                "sub-rótulo do par de medição não é o declarado no template",
                {
                    "sheet": layout.sheet_name,
                    "row": sublabel_row,
                    "column": get_column_letter(column),
                    "found": found,
                    "expected": expected,
                },
            )


def _is_blank_row(cells: Sequence[object], layout: GeneralLayout, header: _GeneralHeader) -> bool:
    columns = [column_index_from_string(letter) for letter in layout.fixed_columns]
    for pair in (*header.periods, header.accumulated):
        columns.extend((pair.quantity, pair.amount))
    columns.append(header.balance)
    return all(not _text_at(cells, column) for column in columns)


def _has_no_text_in(cells: Sequence[object], letters: Sequence[str | None]) -> bool:
    """`True` quando nenhuma das colunas declaradas tem texto na linha."""
    return all(
        not _text_at(cells, column_index_from_string(letter))
        for letter in letters
        if letter is not None
    )


def _is_group_header_row(cells: Sequence[object], layout: GeneralLayout) -> bool:
    """Cabeçalho de grupo: só o rótulo do grupo preenchido, sem nenhum dado de item."""
    if not _text_at(cells, column_index_from_string(layout.group_column)):
        return False
    return _has_no_text_in(
        cells,
        (
            layout.item_column,
            layout.code_column,
            layout.description_column,
            layout.unit_column,
            layout.contract_quantity_column,
            layout.unit_price_column,
            layout.amended_quantity_column,
        ),
    )


def _is_separator_row(cells: Sequence[object], layout: GeneralLayout) -> bool:
    """Linha separadora: nenhuma coluna de texto preenchida, mas a linha não é vazia.

    O arquivo do cliente separa grupos com uma linha que só carrega o zero em cache das
    fórmulas dos pares. Ela não é item — linha com código nunca é separadora — e também
    não é cabeçalho de grupo, porque não traz o rótulo do grupo.
    """
    return _has_no_text_in(
        cells,
        (
            layout.group_column,
            layout.item_column,
            layout.code_column,
            layout.description_column,
            layout.unit_column,
        ),
    )


def _is_total_row(cells: Sequence[object], layout: GeneralLayout) -> bool:
    """Linha de total: a varredura para nela, mesmo que o rótulo esteja na coluna do item."""
    return layout.total_label in (
        _text_at(cells, column_index_from_string(layout.group_column)),
        _text_at(cells, column_index_from_string(layout.item_column)),
    )


def _accepts_contract_code(code: str, template: WorkbookTemplate) -> bool:
    """`True` quando o código é SCO completo ou casa um padrão extra declarado no template.

    Código nu só entra por decisão do template (`extra_code_patterns`); mesmo assim, ele
    precisa ter a estrutura do superset (`is_contract_code`) — um padrão extra frouxo
    demais nunca basta sozinho para aceitar texto sem forma de código.
    """
    if is_sco_code(code):
        return True
    return template.matches_extra_code(code) and is_contract_code(code)


def _parse_general_row(
    cells: Sequence[object],
    row: int,
    layout: GeneralLayout,
    header: _GeneralHeader,
    group_label: str,
    log: _NormalizationLog,
    template: WorkbookTemplate,
) -> ParsedContractLine:
    """Converte a linha de item; o que ela declara vai inteiro para o modelo."""
    sheet = layout.sheet_name
    scale = layout.quantity_decimal_scale
    code = _text_at(cells, column_index_from_string(layout.code_column))
    if not _accepts_contract_code(code, template):
        raise ValuationValidationError(
            "SCO_CODE_UNPARSEABLE",
            "código fora do formato SCO",
            {"sheet": sheet, "row": row, "code": code},
        )
    if not group_label:
        raise ValuationValidationError(
            "CONTRACT_GROUP_MISSING",
            "linha de item antes do primeiro grupo da planilha geral",
            {"sheet": sheet, "row": row, "code": code},
        )
    unit_price_column = column_index_from_string(layout.unit_price_column)
    unit_price = _normalized(
        parse_money(_value_at(cells, unit_price_column)),
        scale=MONEY_DECIMAL_SCALE,
        sheet_name=sheet,
        ref=f"{layout.unit_price_column}{row}",
        log=log,
    )
    periods = tuple(
        ParsedPeriod(
            number=pair.number,
            quantity=_read_quantity_at(cells, pair.quantity, sheet, row, scale, log),
            amount=_read_money_at(cells, pair.amount, sheet, row, log),
        )
        for pair in header.periods
    )
    amended_column = layout.amended_quantity_column
    return ParsedContractLine(
        row=row,
        group_label=group_label,
        item_number=_item_number_at(cells, column_index_from_string(layout.item_column)),
        code=code,
        description=_text_at(cells, column_index_from_string(layout.description_column)),
        unit=normalize_unit(_text_at(cells, column_index_from_string(layout.unit_column))),
        unit_price=unit_price,
        contract_quantity=_read_quantity_at(
            cells, column_index_from_string(layout.contract_quantity_column), sheet, row, scale, log
        ),
        amended_quantity=None
        if amended_column is None
        else _read_quantity_at(
            cells, column_index_from_string(amended_column), sheet, row, scale, log
        ),
        periods=periods,
        accumulated_quantity=_read_quantity_at(
            cells, header.accumulated.quantity, sheet, row, scale, log
        ),
        accumulated_amount=_read_money_at(cells, header.accumulated.amount, sheet, row, log),
        balance_quantity=_read_quantity_at(cells, header.balance, sheet, row, scale, log),
    )


def _read_general_lines(
    worksheet: Any,
    layout: GeneralLayout,
    header: _GeneralHeader,
    failures: _FailureLog,
    log: _NormalizationLog,
    template: WorkbookTemplate,
) -> tuple[list[ParsedContractLine], tuple[int, ...], tuple[tuple[int, str], ...]]:
    """Varre a PLANILHA GERAL do primeiro dado até a linha de total.

    Devolve as linhas convertidas, os números das linhas separadoras que foram puladas e
    a linha/código de cada item cujo código só foi aceito por um padrão extra do template
    (não é SCO). A construção do `ContractLine` fica fora daqui: sem a aba de RE-RA lida, o
    vigente do layout que não declara a coluna ainda não é conhecido.
    """
    parsed_lines: list[ParsedContractLine] = []
    separator_rows: list[int] = []
    extra_code_rows: list[tuple[int, str]] = []
    group_label = ""
    for row, cells in _iter_rows(worksheet, layout.data_first_row):
        if _is_blank_row(cells, layout, header):
            continue
        if _is_total_row(cells, layout):
            break
        if _is_group_header_row(cells, layout):
            group_label = _text_at(cells, column_index_from_string(layout.group_column))
            continue
        if _is_separator_row(cells, layout):
            separator_rows.append(row)
            continue
        try:
            parsed = _parse_general_row(cells, row, layout, header, group_label, log, template)
        except ValuationValidationError as error:
            failures.record(layout.sheet_name, row, error)
            continue
        if not is_sco_code(parsed.code):
            extra_code_rows.append((row, parsed.code))
        parsed_lines.append(parsed)
    return parsed_lines, tuple(separator_rows), tuple(extra_code_rows)


def _amendment_lines(
    cells: Sequence[object],
    row: int,
    sheet: str,
    code: str,
    block: AmendmentColumns,
    scale: int,
    log: _NormalizationLog,
) -> list[ParsedAmendmentLine]:
    """Efeito do bloco sobre o código; coluna vazia ou zero não vira linha de RE-RA.

    Um bloco que preenche duas colunas para o mesmo código devolve duas linhas de
    propósito: o leitor não escolhe qual delas vale, e o agregado recusa com
    `AMENDMENT_DUPLICATE_CODE`. A linha crua carrega o número da linha da aba, para que o
    dossiê aponte a célula em vez do código solto.
    """
    lines: list[ParsedAmendmentLine] = []
    for column_letter, sign, is_new_item in (
        (block.reduced_column, Decimal(-1), False),
        (block.added_column, Decimal(1), False),
        (block.new_item_column, Decimal(1), True),
    ):
        if column_letter is None:
            continue
        value = _read_quantity_at(
            cells, column_index_from_string(column_letter), sheet, row, scale, log
        )
        if value == 0:
            continue
        lines.append(
            ParsedAmendmentLine(
                row=row, code=code, quantity_delta=sign * value, is_new_item=is_new_item
            )
        )
    return lines


def _amendment_block_columns(layout: AmendmentLayout) -> tuple[str | None, ...]:
    """Colunas de RE-RA declaradas da aba da prefeitura: reduzida/acrescida/item novo de
    cada bloco. Precisam estar sempre vazias para uma linha contar como seção — são elas
    que carregariam um delta real sobre um código, ao contrário do subtotal de grupo que a
    coluna de vigente pode carregar (`AmendmentLayout.section_rows_carry_group_subtotal`).
    """
    columns: list[str | None] = []
    for block in layout.blocks:
        columns.extend((block.reduced_column, block.added_column, block.new_item_column))
    return tuple(columns)


def _is_amendment_section_row(cells: Sequence[object], layout: AmendmentLayout) -> bool:
    """Linha de seção da aba da prefeitura: título de grupo na coluna de código, sem valor.

    Espelha o cabeçalho de grupo da GERAL (`_is_group_header_row`): a diferença é que aqui
    o "cabeçalho" ocupa a própria coluna de código, porque esta aba não separa grupo e
    código em colunas distintas. Os blocos de RE-RA precisam estar sempre vazios — linha
    com qualquer delta continua sendo código ilegível, não layout. A coluna de vigente só
    entra nesse cheque quando o template não declara que a linha de seção carrega ali o
    subtotal do grupo; com a flag ligada, um valor na coluna de vigente não impede o skip
    — ele é registrado nas notas, não descartado (`_amendment_section_subtotal`).
    """
    if not _has_no_text_in(cells, _amendment_block_columns(layout)):
        return False
    if layout.section_rows_carry_group_subtotal:
        return True
    return _has_no_text_in(cells, (layout.amended_quantity_column,))


def _amendment_section_subtotal(
    cells: Sequence[object],
    row: int,
    layout: AmendmentLayout,
    scale: int,
    log: _NormalizationLog,
) -> Decimal | None:
    """Subtotal de grupo ignorado na coluna de vigente de uma linha de seção.

    `None` quando a flag não está ligada, a coluna não está declarada ou a célula está
    vazia — nesses casos não há nada para registrar.
    """
    column = layout.amended_quantity_column
    if not layout.section_rows_carry_group_subtotal or column is None:
        return None
    index = column_index_from_string(column)
    if not _text_at(cells, index):
        return None
    return _read_quantity_at(cells, index, layout.sheet_name, row, scale, log)


def _read_amendments(
    worksheet: Any,
    layout: AmendmentLayout,
    failures: _FailureLog,
    scale: int,
    log: _NormalizationLog,
    template: WorkbookTemplate,
) -> tuple[list[ParsedAmendmentBlock], dict[str, Decimal], tuple[tuple[int, Decimal | None], ...]]:
    """Lê a aba da prefeitura: um bloco cru por RE-RA declarada que altera algum código.

    Nenhum modelo é construído aqui de propósito. `Amendment` valida código repetido na
    própria construção, e uma RE-RA repetida abortaria a leitura antes do dossiê existir.
    Devolve também a linha e o subtotal ignorado (quando houver) de cada linha de seção
    que foi pulada — o número nunca desaparece em silêncio, mesmo quando a linha é layout.
    """
    lines_by_block: dict[int, list[ParsedAmendmentLine]] = {
        index: [] for index in range(len(layout.blocks))
    }
    declared: dict[str, Decimal] = {}
    section_rows: list[tuple[int, Decimal | None]] = []
    code_column = column_index_from_string(layout.code_column)
    for row, cells in _iter_rows(worksheet, layout.data_first_row):
        code = _text_at(cells, code_column)
        if not code:
            continue
        if not _accepts_contract_code(code, template):
            if _is_amendment_section_row(cells, layout):
                subtotal = _amendment_section_subtotal(cells, row, layout, scale, log)
                section_rows.append((row, subtotal))
                continue
            failures.record(
                layout.sheet_name,
                row,
                ValuationValidationError(
                    "SCO_CODE_UNPARSEABLE",
                    "código fora do formato SCO",
                    {"sheet": layout.sheet_name, "row": row, "code": code},
                ),
            )
            continue
        try:
            if layout.amended_quantity_column is not None:
                column = column_index_from_string(layout.amended_quantity_column)
                if _text_at(cells, column):
                    declared[code] = _read_quantity_at(
                        cells, column, layout.sheet_name, row, scale, log
                    )
            for index, block in enumerate(layout.blocks):
                lines_by_block[index].extend(
                    _amendment_lines(cells, row, layout.sheet_name, code, block, scale, log)
                )
        except ValuationValidationError as error:
            failures.record(layout.sheet_name, row, error)
            continue
    blocks = [
        ParsedAmendmentBlock(label=block.label, lines=tuple(lines_by_block[index]))
        for index, block in enumerate(layout.blocks)
        if lines_by_block[index]
    ]
    return blocks, declared, tuple(section_rows)


def _to_amendments(blocks: Sequence[ParsedAmendmentBlock]) -> list[Amendment]:
    """Constrói os modelos de RE-RA; código repetido no mesmo bloco falha aqui."""
    return [
        Amendment(
            label=block.label,
            lines=[
                AmendmentLine(
                    code=line.code,
                    quantity_delta=line.quantity_delta,
                    is_new_item=line.is_new_item,
                )
                for line in block.lines
            ],
        )
        for block in blocks
    ]


def _check_unambiguous_amendment_targets(
    parsed_lines: Sequence[ParsedContractLine],
    deltas: Mapping[str, Decimal],
    declared: Mapping[str, Decimal],
) -> None:
    """O mesmo código em dois grupos é aceito; escolher entre eles não é do leitor.

    Enquanto ninguém aponta o código, as duas linhas convivem. A partir do momento em que
    a aba da prefeitura fala daquele código — alterando a quantidade ou declarando o
    vigente —, aplicar ou conferir exigiria escolher uma das linhas, e o leitor recusa em
    vez de escolher.
    """
    groups_by_code: dict[str, list[str]] = {}
    for line in parsed_lines:
        groups_by_code.setdefault(line.code, []).append(line.group_label)
    for code, groups in groups_by_code.items():
        if len(groups) < 2:
            continue
        if code not in deltas and code not in declared:
            continue
        raise ValuationValidationError(
            "CODE_AMBIGUOUS_IN_CONTRACT",
            "a aba da prefeitura fala de um código que a planilha geral repete em mais "
            "de um grupo, e o leitor não escolhe a linha",
            {"code": code, "groups": groups},
        )


def _check_declared_amended(lines: Sequence[ContractLine], declared: Mapping[str, Decimal]) -> None:
    """A quantidade vigente da GERAL e a da aba da prefeitura têm de ser a mesma.

    O índice por código só é consultado para código que a prefeitura declara, e código
    declarado que a GERAL repete em mais de um grupo já foi recusado antes daqui
    (`_check_unambiguous_amendment_targets`) — nenhuma linha é escolhida em silêncio.
    """
    by_code = {line.code: line for line in lines}
    divergences: list[dict[str, object]] = []
    for code, quantity in declared.items():
        line = by_code.get(code)
        if line is None:
            continue
        if line.amended_quantity != quantity:
            divergences.append(
                {
                    "code": code,
                    "general": str(line.amended_quantity),
                    "declared": str(quantity),
                }
            )
    if divergences:
        raise ValuationValidationError(
            "GENERAL_AMENDED_DIVERGENT",
            "quantidade vigente da planilha geral diverge da declarada na aba da prefeitura",
            {"divergences": divergences},
        )


def _period_gaps(numbers: Sequence[int]) -> tuple[int, ...]:
    """Medições que faltam entre a primeira e a última declaradas no cabeçalho."""
    if not numbers:
        return ()
    declared = set(numbers)
    return tuple(number for number in range(numbers[0], numbers[-1] + 1) if number not in declared)


def _build_parse(
    layout: GeneralLayout,
    amendment_layout: AmendmentLayout | None,
    header: _GeneralHeader,
    parsed_lines: Sequence[ParsedContractLine],
    blocks: Sequence[ParsedAmendmentBlock],
    declared: Mapping[str, Decimal],
) -> ContractParse:
    """Reúne o que a leitura extraiu, com as letras de coluna que o dossiê usa nas refs."""
    return ContractParse(
        general_sheet=layout.sheet_name,
        amendment_sheet=None if amendment_layout is None else amendment_layout.sheet_name,
        period_numbers=header.period_numbers,
        period_columns={
            pair.number: PairColumns(
                quantity=get_column_letter(pair.quantity), amount=get_column_letter(pair.amount)
            )
            for pair in header.periods
        },
        accumulated_columns=PairColumns(
            quantity=get_column_letter(header.accumulated.quantity),
            amount=get_column_letter(header.accumulated.amount),
        ),
        balance_column=get_column_letter(header.balance),
        item_column=layout.item_column,
        code_column=layout.code_column,
        contract_quantity_column=layout.contract_quantity_column,
        unit_price_column=layout.unit_price_column,
        amended_column=layout.amended_quantity_column,
        lines=tuple(parsed_lines),
        amendment_blocks=tuple(blocks),
        declared_amended=dict(declared),
    )


def _build_contract(
    parse: ContractParse,
    *,
    source_sha256: str,
    source_label: str,
    contract_label: str | None,
) -> ContractWorkbook:
    """Caminho estrito: guardas do leitor, modelos de RE-RA e o agregado, nesta ordem."""
    deltas = parse.amendment_deltas()
    _check_unambiguous_amendment_targets(parse.lines, deltas, parse.declared_amended)
    amendments = _to_amendments(parse.amendment_blocks)
    lines = [
        _to_contract_line(parsed, parsed.resolved_amended_quantity(deltas))
        for parsed in parse.lines
    ]
    _check_declared_amended(lines, parse.declared_amended)
    return ContractWorkbook(
        id=contract_workbook_id_for(source_sha256),
        source_label=source_label,
        source_sha256=source_sha256,
        contract_label=contract_label,
        period_numbers=list(parse.period_numbers),
        lines=lines,
        amendments=amendments,
    )


def read_contract_parse(
    workbook_path: Path, template: WorkbookTemplate
) -> tuple[ContractParse, ContractImportNotes]:
    """Lê a pasta e devolve o material bruto da importação, sem construir o agregado.

    É o primeiro passo de `read_contract_workbook_with_notes` e a entrada do dossiê:
    quem quiser diagnosticar um arquivo sem importá-lo passa por aqui. Falha de layout e
    linha ilegível continuam recusando, porque sem entender a planilha não há material.
    """
    layout = template.general
    failures = _FailureLog()
    log = _NormalizationLog()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        general_sheet = _require_sheet(workbook, layout.sheet_name)
        header = _read_general_header(general_sheet, layout)
        parsed_lines, separator_rows, extra_code_rows = _read_general_lines(
            general_sheet, layout, header, failures, log, template
        )
        blocks: list[ParsedAmendmentBlock] = []
        declared: dict[str, Decimal] = {}
        amendment_section_rows: tuple[tuple[int, Decimal | None], ...] = ()
        if template.amendment is not None:
            amendment_sheet = _require_sheet(workbook, template.amendment.sheet_name)
            blocks, declared, amendment_section_rows = _read_amendments(
                amendment_sheet,
                template.amendment,
                failures,
                layout.quantity_decimal_scale,
                log,
                template,
            )
    finally:
        workbook.close()

    failures.raise_if_any()
    if not parsed_lines:
        raise ValuationValidationError(
            "CONTRACT_EMPTY",
            "planilha geral não possui nenhuma linha de item",
            {"sheet": layout.sheet_name},
        )
    parse = _build_parse(layout, template.amendment, header, parsed_lines, blocks, declared)
    notes = ContractImportNotes(
        period_numbers=header.period_numbers,
        period_gaps=_period_gaps(header.period_numbers),
        separator_rows=separator_rows,
        normalized_cells=tuple(log.cells),
        normalized_count=log.count,
        normalized_truncated=log.truncated,
        extra_code_rows=extra_code_rows,
        amendment_section_rows=amendment_section_rows,
    )
    return parse, notes


def read_contract_workbook_with_notes(
    workbook_path: Path,
    template: WorkbookTemplate,
    *,
    source_label: str,
    contract_label: str | None = None,
) -> tuple[ContractWorkbook, ContractImportNotes]:
    """Lê o consolidado contratual (PLANILHA GERAL + RE-RA) e o que a leitura observou.

    A ordem é obrigatória: a GERAL é convertida, a aba da prefeitura é lida e só então
    cada linha vira `ContractLine` — no layout sem coluna de vigente, o vigente é
    contratual mais as RE-RA do código e não existe antes das duas leituras.

    Falha de **layout** (aba ausente, cabeçalho fora do padrão, linha ilegível) recusa
    sem dossiê: sem entender a planilha não há o que recomputar. Falha **semântica** do
    agregado vira `ContractSemanticsError` com o dossiê inteiro do recomputo. Se o
    recomputo não encontrar nada — bug nosso, invariante que o dossiê não cobre —, o erro
    original sobe intacto em vez de ser mascarado por um dossiê vazio.
    """
    source_sha256 = file_sha256(workbook_path)
    parse, notes = read_contract_parse(workbook_path, template)
    try:
        contract = _build_contract(
            parse,
            source_sha256=source_sha256,
            source_label=source_label,
            contract_label=contract_label,
        )
    except (ValuationValidationError, ValidationError) as error:
        diagnosis = diagnose_contract(parse)
        if diagnosis.clean:
            raise
        raise ContractSemanticsError(diagnosis, source_sha256=source_sha256) from error
    return contract, notes


def read_contract_workbook(
    workbook_path: Path,
    template: WorkbookTemplate,
    *,
    source_label: str,
    contract_label: str | None = None,
) -> ContractWorkbook:
    """Lê o consolidado contratual do MAPÃO anterior, descartando as notas da leitura."""
    contract, _ = read_contract_workbook_with_notes(
        workbook_path,
        template,
        source_label=source_label,
        contract_label=contract_label,
    )
    return contract
