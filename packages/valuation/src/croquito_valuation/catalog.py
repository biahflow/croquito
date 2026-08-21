"""Importação do catálogo de preços a partir da planilha de referência.

A planilha de preços mistura tipos de linha diferentes: cabeçalhos de classificação e
itens com preço. Os cabeçalhos não viram entrada, mas classificam os itens abaixo deles.
Uma linha que não é nenhum dos tipos previstos não é ignorada em silêncio: falha alto com
a aba e a linha em `details`.

O template escolhe onde a hierarquia está escrita (`CatalogLayout.hierarchy_columns`):

- **na coluna de código**: `AD` e `AD0405` aparecem na mesma coluna dos itens;
- **em colunas próprias**: família e nível intermediário têm coluna cada um, com o nome
  junto do código, e o cabeçalho de subgrupo fica na coluna de código. Nesse modo a
  linha de família carrega texto sem valor nas demais colunas e é consumida inteira,
  e cada item é conferido contra o subgrupo corrente antes de virar entrada. O arquivo
  publicado também escreve, aqui e ali, o nível intermediário na coluna de código; ele é
  aceito ali com a mesma conferência de hierarquia da coluna própria.

O código do item é canonizado na leitura (`canonical_sco_code`): o catálogo publicado o
escreve com separadores (`AD 04.05.0050 (/)`) e a entrada guarda sempre a forma canônica
(`AD04050050(/)`). Reconhecer as duas formas é o que separa item de cabeçalho — um código
publicado lido como outra coisa some do catálogo em silêncio.
"""

from __future__ import annotations

import difflib
import hashlib
import math
import re
import unicodedata
from collections.abc import Sequence
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from importlib import resources
from pathlib import Path
from typing import Any, Final, NoReturn
from uuid import NAMESPACE_URL, UUID, uuid5

from openpyxl import load_workbook
from pydantic import Field, field_validator

from croquito_valuation.errors import ValuationValidationError
from croquito_valuation.models import PriceCatalog, PriceCatalogEntry, ValuationContractModel
from croquito_valuation.sco import (
    canonical_sco_code,
    is_family_row_code,
    is_subgroup_row_code,
    parse_family_header,
    parse_intermediate_header,
    parse_sco_code,
    parse_subgroup_header,
)
from croquito_valuation.template import CatalogLayout, WorkbookTemplate

_CATALOG_ID_NAMESPACE: Final = uuid5(NAMESPACE_URL, "https://croquito.local/valuation/catalog")
_BLANK_CHARS: Final = " \t\r\n\xa0\u202f\u2007"

UNIT_ALIASES: Final[dict[str, str]] = {
    "m2": "m2",
    "m²": "m2",
    "m3": "m3",
    "m³": "m3",
    "m": "m",
    "ml": "m",
    "un": "un",
    "und": "un",
    "unid": "un",
    "unid.": "un",
    "unidade": "un",
    "mes": "mes",
    "mês": "mes",
    "meses": "mes",
}


def normalize_unit(raw_unit: str) -> str:
    """Normaliza a unidade escrita na planilha para a tabela única do módulo.

    Unidade desconhecida não é erro no M1: passa em minúsculas e sem espaços, para o
    orçamentista revisar sem que a importação trave.
    """
    cleaned = raw_unit.strip(_BLANK_CHARS).lower()
    cleaned = " ".join(cleaned.split())
    return UNIT_ALIASES.get(cleaned, cleaned)


def parse_money(raw_value: object) -> Decimal:
    """Converte um valor de célula em `Decimal`.

    Fronteira de leitura declarada: a planilha devolve `float` para número, então o
    módulo usa `Decimal(str(...))` uma única vez, aqui. Texto aceita `R$`, espaço fino,
    separador de milhar e vírgula decimal.
    """
    if isinstance(raw_value, Decimal):
        return raw_value
    if isinstance(raw_value, bool):
        raise ValuationValidationError(
            "MONEY_UNPARSEABLE", "valor booleano não é preço", {"value": repr(raw_value)}
        )
    if isinstance(raw_value, int | float):
        return Decimal(str(raw_value))
    if isinstance(raw_value, str):
        text = raw_value
        for char in _BLANK_CHARS:
            text = text.replace(char, "")
        text = text.replace("R$", "").replace("r$", "")
        if "," in text:
            text = text.replace(".", "").replace(",", ".")
        try:
            return Decimal(text)
        except InvalidOperation as error:
            raise ValuationValidationError(
                "MONEY_UNPARSEABLE", "preço ilegível na planilha", {"value": raw_value}
            ) from error
    raise ValuationValidationError(
        "MONEY_UNPARSEABLE", "preço ausente ou de tipo inesperado", {"value": repr(raw_value)}
    )


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip(_BLANK_CHARS).strip()
    return str(value).strip()


def _column_index(letter: str) -> int:
    index = 0
    for char in letter:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index - 1


def _row_error(sheet_name: str, row_number: int, reason: str, **details: object) -> NoReturn:
    raise ValuationValidationError(
        "ROW_UNPARSEABLE",
        f"linha ilegível no catálogo (ROW_UNPARSEABLE:{sheet_name}:{row_number}): {reason}",
        {"sheet": sheet_name, "row": row_number, "reason": reason, **details},
    )


def _item_unit_price(
    sheet_name: str,
    row_number: int,
    code: str,
    raw_unit: str,
    raw_price: object,
) -> Decimal:
    """Unidade e preço de uma linha de item; qualquer defeito falha com aba e linha."""
    if not raw_unit:
        _row_error(sheet_name, row_number, "item sem unidade", code=code)
    try:
        unit_price = parse_money(raw_price)
    except ValuationValidationError as error:
        _row_error(sheet_name, row_number, "preço ilegível", code=code, price_error=error.code)
    if unit_price < 0:
        _row_error(sheet_name, row_number, "preço negativo", code=code)
    return unit_price


def _is_declared_note(layout: CatalogLayout, texts: Sequence[str]) -> bool:
    """Linha de nota editorial do catálogo publicado, declarada como tal no template.

    O catálogo real intercala texto livre entre os cabeçalhos ("Nota: As marcas indicadas
    servem apenas para definir o modelo..."). Ela é pulada porque o template diz que
    aquele texto é nota; texto não declarado continua sendo linha ilegível. Quem chama já
    garantiu que a linha não tem código SCO — nota nunca engole item com preço.
    """
    if not layout.note_prefixes:
        return False
    return any(text.startswith(prefix) for text in texts if text for prefix in layout.note_prefixes)


def _read_in_column_entries(worksheet: Any, layout: CatalogLayout) -> list[PriceCatalogEntry]:
    """Modo padrão: os cabeçalhos de família e subgrupo vivem na coluna de código."""
    code_index = _column_index(layout.code_column)
    description_index = _column_index(layout.description_column)
    unit_index = _column_index(layout.unit_column)
    price_index = _column_index(layout.price_column)
    needed = max(code_index, description_index, unit_index, price_index) + 1

    entries: list[PriceCatalogEntry] = []
    family_code = ""
    family_name = ""
    subgroup_code = ""
    subgroup_name = ""
    row_number = layout.first_row - 1
    for row in worksheet.iter_rows(min_row=layout.first_row, values_only=True):
        row_number += 1
        cells: list[object] = list(row) + [None] * (needed - len(row))
        code = _cell_text(cells[code_index])
        description = _cell_text(cells[description_index])
        raw_unit = _cell_text(cells[unit_index])
        raw_price = cells[price_index]
        if not code and not description and raw_price is None:
            continue
        item_code = canonical_sco_code(code)
        if item_code is None and _is_declared_note(layout, (code, description)):
            continue
        if is_family_row_code(code):
            if not description:
                _row_error(layout.sheet_name, row_number, "família sem descrição", code=code)
            family_code, family_name = code, description
            subgroup_code, subgroup_name = "", ""
            continue
        if is_subgroup_row_code(code):
            if not description:
                _row_error(layout.sheet_name, row_number, "subgrupo sem descrição", code=code)
            subgroup_code, subgroup_name = code, description
            continue
        if item_code is None:
            _row_error(layout.sheet_name, row_number, "código fora do formato SCO", code=code)
        if not description:
            _row_error(layout.sheet_name, row_number, "item sem descrição", code=item_code)
        if not family_code or not subgroup_code:
            _row_error(
                layout.sheet_name, row_number, "item antes de família ou subgrupo", code=item_code
            )
        if layout.is_unpriced(raw_price):
            continue
        unit_price = _item_unit_price(layout.sheet_name, row_number, item_code, raw_unit, raw_price)
        entries.append(
            PriceCatalogEntry(
                code=item_code,
                description=description,
                unit=normalize_unit(raw_unit),
                unit_price=unit_price,
                family_code=family_code,
                family_name=family_name,
                subgroup_code=subgroup_code,
                subgroup_name=subgroup_name,
            )
        )
    return entries


def _read_columnar_entries(
    worksheet: Any, layout: CatalogLayout, hierarchy: tuple[str, str]
) -> list[PriceCatalogEntry]:
    """Modo do catálogo publicado: família e nível intermediário em colunas próprias.

    A hierarquia é conferida em cada degrau — o intermediário precisa pertencer à família
    corrente, o cabeçalho de subgrupo ao intermediário e o item ao subgrupo. Prefixo que
    não bate é linha ilegível, nunca reclassificação silenciosa.
    """
    family_column, intermediate_column = hierarchy
    family_index = _column_index(family_column)
    intermediate_index = _column_index(intermediate_column)
    code_index = _column_index(layout.code_column)
    description_index = _column_index(layout.description_column)
    unit_index = _column_index(layout.unit_column)
    price_index = _column_index(layout.price_column)
    needed = (
        max(
            family_index,
            intermediate_index,
            code_index,
            description_index,
            unit_index,
            price_index,
        )
        + 1
    )

    entries: list[PriceCatalogEntry] = []
    family_code = ""
    family_name = ""
    intermediate_code = ""
    subgroup_code = ""
    subgroup_name = ""
    row_number = layout.first_row - 1
    for row in worksheet.iter_rows(min_row=layout.first_row, values_only=True):
        row_number += 1
        cells: list[object] = list(row) + [None] * (needed - len(row))
        family_text = _cell_text(cells[family_index])
        intermediate_text = _cell_text(cells[intermediate_index])
        code = _cell_text(cells[code_index])
        description = _cell_text(cells[description_index])
        raw_unit = _cell_text(cells[unit_index])
        raw_price = cells[price_index]
        if (
            not family_text
            and not intermediate_text
            and not code
            and not description
            and not raw_unit
            and raw_price is None
        ):
            continue
        item_code = canonical_sco_code(code)
        if item_code is None and _is_declared_note(layout, (family_text, intermediate_text, code)):
            continue
        if family_text:
            # A linha de família traz texto sem valor nas outras colunas: consome a linha
            # inteira em vez de tentar lê-las.
            family_header = parse_family_header(family_text)
            if family_header is None:
                _row_error(
                    layout.sheet_name,
                    row_number,
                    "cabeçalho de família ilegível",
                    value=family_text,
                )
            family_code, family_name = family_header
            intermediate_code = ""
            subgroup_code, subgroup_name = "", ""
            continue
        if intermediate_text:
            intermediate_header = parse_intermediate_header(intermediate_text)
            if intermediate_header is None:
                _row_error(
                    layout.sheet_name,
                    row_number,
                    "subgrupo intermediário ilegível",
                    value=intermediate_text,
                )
            if not family_code or intermediate_header[0][:2] != family_code:
                _row_error(
                    layout.sheet_name,
                    row_number,
                    "subgrupo intermediário fora da família corrente",
                    code=intermediate_header[0],
                    family=family_code,
                )
            intermediate_code = intermediate_header[0]
            subgroup_code, subgroup_name = "", ""
            continue
        if item_code is not None:
            if not description:
                _row_error(layout.sheet_name, row_number, "item sem descrição", code=item_code)
            if not family_code or not subgroup_code:
                _row_error(
                    layout.sheet_name,
                    row_number,
                    "item antes de família ou subgrupo",
                    code=item_code,
                )
            if parse_sco_code(item_code).subgroup_row_code != subgroup_code:
                _row_error(
                    layout.sheet_name,
                    row_number,
                    "item fora do subgrupo corrente",
                    code=item_code,
                    subgroup=subgroup_code,
                )
            if layout.is_unpriced(raw_price):
                continue
            unit_price = _item_unit_price(
                layout.sheet_name, row_number, item_code, raw_unit, raw_price
            )
            entries.append(
                PriceCatalogEntry(
                    code=item_code,
                    description=description,
                    unit=normalize_unit(raw_unit),
                    unit_price=unit_price,
                    family_code=family_code,
                    family_name=family_name,
                    subgroup_code=subgroup_code,
                    subgroup_name=subgroup_name,
                )
            )
            continue
        subgroup_header = parse_subgroup_header(code)
        if subgroup_header is not None:
            header_code, header_name = subgroup_header
            if (
                not family_code
                or not intermediate_code
                or header_code[:2] != family_code
                or header_code[:4] != intermediate_code
            ):
                _row_error(
                    layout.sheet_name,
                    row_number,
                    "cabeçalho de subgrupo fora da hierarquia corrente",
                    code=header_code,
                    family=family_code,
                    intermediate=intermediate_code,
                )
            subgroup_code, subgroup_name = header_code, header_name
            continue
        # O arquivo publicado às vezes escreve o nível intermediário na coluna de código
        # em vez da coluna própria. É o mesmo cabeçalho, com a mesma conferência de
        # hierarquia: pertencer à família corrente e zerar o subgrupo aberto.
        intermediate_in_code = parse_intermediate_header(code)
        if intermediate_in_code is not None:
            if not family_code or intermediate_in_code[0][:2] != family_code:
                _row_error(
                    layout.sheet_name,
                    row_number,
                    "subgrupo intermediário fora da família corrente",
                    code=intermediate_in_code[0],
                    family=family_code,
                )
            intermediate_code = intermediate_in_code[0]
            subgroup_code, subgroup_name = "", ""
            continue
        _row_error(layout.sheet_name, row_number, "código fora do formato SCO", code=code)
    return entries


def file_sha256(path: Path) -> str:
    """SHA-256 dos bytes do arquivo importado."""
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def catalog_id_for(source_sha256: str) -> UUID:
    """Id derivado do conteúdo: reimportar o mesmo arquivo devolve o mesmo catálogo."""
    return uuid5(_CATALOG_ID_NAMESPACE, source_sha256)


def read_price_catalog(
    workbook_path: Path,
    template: WorkbookTemplate,
    *,
    source_label: str,
    reference_month: str,
) -> PriceCatalog:
    """Lê a aba de catálogo do template e devolve o catálogo canônico."""
    source_sha256 = file_sha256(workbook_path)
    layout = template.catalog
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if layout.sheet_name not in workbook.sheetnames:
            raise ValuationValidationError(
                "CATALOG_SHEET_MISSING",
                "planilha de catálogo não possui a aba declarada no template",
                {"sheet": layout.sheet_name, "available": list(workbook.sheetnames)},
            )
        worksheet = workbook[layout.sheet_name]
        hierarchy = layout.hierarchy_columns
        if hierarchy is None:
            entries = _read_in_column_entries(worksheet, layout)
        else:
            entries = _read_columnar_entries(worksheet, layout, hierarchy)
    finally:
        workbook.close()

    if not entries:
        raise ValuationValidationError(
            "CATALOG_EMPTY",
            "aba de catálogo não possui nenhum item com preço",
            {"sheet": layout.sheet_name},
        )
    return PriceCatalog(
        id=catalog_id_for(source_sha256),
        source_label=source_label,
        reference_month=reference_month,
        source_sha256=source_sha256,
        entries=entries,
    )


LEXICAL_DICE_WEIGHT: Final = 0.75
LEXICAL_RATIO_WEIGHT: Final = 0.25

_LEXICAL_TOKEN_SPLIT_RE: Final = re.compile(r"[^a-z0-9]+")
_LEXICAL_MIN_TOKEN_LENGTH: Final = 2


def _lexical_normalize(text: str) -> str:
    """Minúsculo, sem acento: `casefold` seguido de decomposição NFKD sem combining marks."""
    decomposed = unicodedata.normalize("NFKD", text.casefold())
    return "".join(char for char in decomposed if not unicodedata.combining(char))


def lexical_tokens(text: str) -> tuple[str, ...]:
    """Tokens determinísticos de `text`, na ordem de ocorrência, com repetição.

    Insensível a caixa e acento; separa em qualquer caractere não alfanumérico e descarta
    tokens com menos de dois caracteres.
    """
    normalized = _lexical_normalize(text)
    return tuple(
        token
        for token in _LEXICAL_TOKEN_SPLIT_RE.split(normalized)
        if len(token) >= _LEXICAL_MIN_TOKEN_LENGTH
    )


_STEM_SUFFIXES: Final[tuple[tuple[str, str, int], ...]] = (
    # (sufixo, substituto, comprimento mínimo do TOKEN ORIGINAL para a regra valer).
    # Checados nesta ordem — do sufixo mais longo pro mais curto — porque um token como
    # "gramados" termina tanto em "ados" quanto em "s", e só a primeira regra que bater
    # decide o radical. "comprimento mínimo" é sobre `len(token)`, não sobre o radical.
    ("ados", "a", 6),  # "intertravados" (13>6) -> "intertrava"
    ("adas", "a", 6),
    ("ado", "a", 6),  # "gramado" (7>6) -> "grama"
    ("ada", "a", 6),
    ("oes", "ao", 4),  # "licitacoes" (10>4) -> "licitacao" (mesmo radical de "licitacao")
    ("es", "", 5),  # "pilares" (7>5) -> "pilar"
    ("s", "", 4),  # "gramados" já caiu em "ados"; "bancos" (6>4) -> "banco"
)
"""Sufixos de radicalização conservadora pt-BR, e SÓ eles.

Cobre plural comum (`-s`, `-es`, `-ões→ão`) e a família participial/derivacional mais
frequente do catálogo de praças (`-ado/-ada/-ados/-adas→-a`: "gramado"→"grama",
"intertravado"→"intertrava"). Radical agressivo (removendo mais sufixos, ou stemmers
completos como o de Porter/RSLP) destruiria precisão num catálogo de 5 mil descrições
técnicas — duas palavras que **parecem** aparentadas ("concreto"/"concretado") só
compartilham radical aqui se uma delas literalmente virar a outra por uma destas regras, o
que não é o caso (nenhuma delas termina nos sufixos acima do jeito que colidiria). Ver
`tests/valuation/test_assignment.py` para os casos de não-colisão exigidos.
"""


def _stem(token: str) -> str:
    length = len(token)
    for suffix, replacement, min_length in _STEM_SUFFIXES:
        if length > min_length and token.endswith(suffix):
            return token[: -len(suffix)] + replacement
    return token


def lexical_stems(text: str) -> tuple[str, ...]:
    """Radical conservador pt-BR de cada token de `text`, na mesma ordem de `lexical_tokens`.

    Aplica `_STEM_SUFFIXES` token a token, sobre o texto já normalizado (sem acento,
    minúsculo) de `lexical_tokens`. Não é um stemmer completo: existe só para que a mesma
    palavra em singular/plural ou nas derivações mais comuns do domínio ("gramado" /
    "grama" / "gramados") produza o mesmo radical, sem tentar cobrir toda a morfologia do
    português.
    """
    return tuple(_stem(token) for token in lexical_tokens(text))


class DomainSynonyms(ValuationContractModel):
    """Tabela de sinônimos de domínio do catálogo SCO, como DADO versionado — não código.

    `terms` é `{termo: [equivalentes]}`: cada chave e cada equivalente pode ser uma frase
    ("tela de arame galvanizado"), não só uma palavra. A normalização acontece na carga —
    sem acento, minúsculo, hífen e espaço unificados — e a validação recusa termo vazio e
    termo duplicado, seja dentro do mesmo grupo, seja entre grupos diferentes (um termo só
    pode pertencer a UM grupo de sinônimos, senão a expansão vira ambígua).

    A tabela é carregada de um seed versionado
    (`packages/valuation/src/croquito_valuation/data/sco-synonyms-v1.json`,
    `default_domain_synonyms()`) ou de um `synonyms.json` opcional na rodada do worker —
    quem carrega por rodada é o worker, não este pacote (`ADR-0016`: domínio puro, sem
    filesystem de rodada).
    """

    version: str = Field(min_length=1, max_length=40)
    terms: dict[str, list[str]] = Field(min_length=1)

    @field_validator("terms")
    @classmethod
    def _normalize_and_validate_terms(cls, raw: dict[str, list[str]]) -> dict[str, list[str]]:
        normalized: dict[str, list[str]] = {}
        seen: dict[str, str] = {}  # termo normalizado -> grupo (chave) onde apareceu primeiro
        for raw_key, raw_values in raw.items():
            key = _normalize_synonym_term(raw_key)
            if not key:
                raise ValuationValidationError(
                    "SYNONYMS_TERM_EMPTY", "termo de sinônimo vazio", {"raw_term": raw_key}
                )
            if not raw_values:
                raise ValuationValidationError(
                    "SYNONYMS_GROUP_EMPTY",
                    "grupo de sinônimos sem nenhum equivalente",
                    {"term": key},
                )
            values: list[str] = []
            for raw_value in raw_values:
                value = _normalize_synonym_term(raw_value)
                if not value:
                    raise ValuationValidationError(
                        "SYNONYMS_TERM_EMPTY", "termo de sinônimo vazio", {"raw_term": raw_value}
                    )
                if value == key or value in values:
                    raise ValuationValidationError(
                        "SYNONYMS_TERM_DUPLICATE",
                        "termo repetido dentro do mesmo grupo de sinônimos",
                        {"term": value, "group": key},
                    )
                values.append(value)
            if key in seen:
                raise ValuationValidationError(
                    "SYNONYMS_TERM_DUPLICATE",
                    "termo aparece em mais de um grupo de sinônimos",
                    {"term": key, "groups": [seen[key], key]},
                )
            for term in (key, *values):
                if term in seen:
                    raise ValuationValidationError(
                        "SYNONYMS_TERM_DUPLICATE",
                        "termo aparece em mais de um grupo de sinônimos",
                        {"term": term, "groups": [seen[term], key]},
                    )
                seen[term] = key
            normalized[key] = values
        return normalized


def _normalize_synonym_term(text: str) -> str:
    """Canonicaliza um termo de sinônimo: sem acento, minúsculo, um espaço entre palavras.

    Hífen e outra pontuação viram espaço (`"meio-fio"` -> `"meio fio"`), para que o termo
    canonizado se comporte como as frases de `expand_terms` — comparado por RADICAIS de
    cada palavra, nunca pela grafia exata.
    """
    normalized = _lexical_normalize(text)
    normalized = re.sub(r"[^a-z0-9]+", " ", normalized).strip()
    return normalized


_SYNONYMS_SEED_FILENAME: Final = "sco-synonyms-v1.json"


def default_domain_synonyms() -> DomainSynonyms:
    """Seed versionado de sinônimos de domínio, empacotado com o pacote (`data/`).

    Curadoria conferida contra o catálogo real do SCO (praças/urbanização, ~5 mil itens):
    cada par tem as duas pontas com presença real no catálogo — não é lista inventada. O
    caso que motivou a Fase 1 do M7 é `refletor -> projetor`: a legenda da Toca cita
    "REFLETOR EXISTENTE" e o código correspondente do catálogo descreve um "Projetor
    PRJ-01..."; sem o sinônimo, nenhum token literal em comum existe entre os dois textos.
    """
    payload = (
        resources.files("croquito_valuation")
        .joinpath("data", _SYNONYMS_SEED_FILENAME)
        .read_text(encoding="utf-8")
    )
    return DomainSynonyms.model_validate_json(payload)


class LegendNoiseList(ValuationContractModel):
    """Lista de termos de ESTADO da legenda, como DADO versionado — espelho de
    `DomainSynonyms`, mesma família de seed curado, mesma normalização na carga.

    O rótulo de uma prancha descreve o ESTADO do elemento além do serviço ("existente", "a
    ser recuperar"), e essas palavras entram na cobertura ponderada por IDF como qualquer
    outra — só que, sendo raras num catálogo que descreve serviço e não estado, ganham peso
    ALTO (contrapartida medida na docstring de `weighted_query_coverage_score`). `terms` é
    uma lista simples, não um dicionário de grupos como `DomainSynonyms.terms`: aqui não há
    equivalência para expandir, só um radical para amortecer.

    A tabela é carregada de um seed versionado
    (`packages/valuation/src/croquito_valuation/data/sco-legend-noise-v1.json`,
    `default_legend_noise()`). Termo novo só entra com melhora medida no golden real
    (`tests/valuation/test_matcher_golden.py`) — esta não é uma lista de "toda palavra de
    estado imaginável".
    """

    version: str = Field(min_length=1, max_length=40)
    terms: list[str] = Field(min_length=1)

    @field_validator("terms")
    @classmethod
    def _normalize_and_validate_terms(cls, raw: list[str]) -> list[str]:
        normalized: list[str] = []
        seen: set[str] = set()
        for raw_term in raw:
            term = _normalize_synonym_term(raw_term)
            if not term:
                raise ValuationValidationError(
                    "LEGEND_NOISE_TERM_EMPTY",
                    "termo de ruído de legenda vazio",
                    {"raw_term": raw_term},
                )
            if term in seen:
                raise ValuationValidationError(
                    "LEGEND_NOISE_TERM_DUPLICATE",
                    "termo repetido na lista de ruído de legenda",
                    {"term": term},
                )
            seen.add(term)
            normalized.append(term)
        return normalized

    def noise_stems(self) -> frozenset[str]:
        """Radicais de ruído: cada termo — palavra ou frase — contribui TODOS os seus radicais.

        Um termo-frase ("a ser recuperar") passa por `lexical_stems` inteiro, então "ser" e
        "recuperar" amortecem os dois, não só a frase casada literalmente — o amortecimento
        em `weighted_query_coverage_score` opera por radical ORIGINAL da consulta.
        """
        stems: set[str] = set()
        for term in self.terms:
            stems.update(lexical_stems(term))
        return frozenset(stems)


_LEGEND_NOISE_SEED_FILENAME: Final = "sco-legend-noise-v1.json"


def default_legend_noise() -> LegendNoiseList:
    """Seed versionado da lista de ruído de legenda, empacotado com o pacote (`data/`).

    Espelho de `default_domain_synonyms()`. Conteúdo mínimo hoje: as duas medições citadas
    na docstring de `weighted_query_coverage_score` — as formas de "existente" e de "a ser
    recuperar" —, porque um termo novo só entra com melhora medida no golden real.
    """
    payload = (
        resources.files("croquito_valuation")
        .joinpath("data", _LEGEND_NOISE_SEED_FILENAME)
        .read_text(encoding="utf-8")
    )
    return LegendNoiseList.model_validate_json(payload)


@dataclass(frozen=True, slots=True)
class ExpandedTerms:
    """Termos originais mais os expandidos por sinônimo, com a origem de cada expansão.

    `terms` preserva os originais primeiro, na ordem recebida e sem repetição, seguidos dos
    termos que a expansão acrescentou. `origins` mapeia cada termo EXPANDIDO (nunca um
    termo original) para os termos de entrada que o trouxeram — quem exibe decide se mostra
    "projetor (via refletor)" ou só o termo; este dataclass não formata nada.
    """

    terms: tuple[str, ...]
    origins: dict[str, tuple[str, ...]]


def expand_terms(terms: Sequence[str], synonyms: DomainSynonyms | None) -> ExpandedTerms:
    """Expande `terms` (já radicalizados, tipicamente por `lexical_stems`) com sinônimos.

    Um grupo de sinônimos (`{chave: [equivalentes]}`) pode conter frases inteiras
    ("tela de arame galvanizado"); o casamento é por SUBCONJUNTO de radicais: se todos os
    radicais de uma frase do grupo estão presentes em `terms`, os radicais de TODAS as
    outras frases do mesmo grupo entram na expansão, com a origem apontando para a frase
    que casou. A busca é feita nos dois sentidos — declarar `refletor: [projetor]` expande
    tanto uma consulta que cita "refletor" quanto uma que cita "projetor" —, então o seed
    só precisa de uma entrada por par, não duas.

    Sem `synonyms` (ou sem `terms`), devolve os termos originais sem expansão nenhuma.
    """
    original = tuple(dict.fromkeys(terms))
    if synonyms is None or not original:
        return ExpandedTerms(terms=original, origins={})

    term_set = set(original)
    extra: list[str] = []
    origins: dict[str, set[str]] = {}
    for key, equivalents in synonyms.terms.items():
        group = (key, *equivalents)
        group_stems = [(phrase, lexical_stems(phrase)) for phrase in group]
        for phrase, stems in group_stems:
            if not stems or not set(stems).issubset(term_set):
                continue
            for other_phrase, other_stems in group_stems:
                if other_phrase == phrase:
                    continue
                for stem in other_stems:
                    if stem not in term_set and stem not in extra:
                        extra.append(stem)
                    origins.setdefault(stem, set()).add(phrase)

    return ExpandedTerms(
        terms=(*original, *extra),
        origins={term: tuple(sorted(sources)) for term, sources in origins.items()},
    )


def lexical_similarity(left: str, right: str, *, synonyms: DomainSynonyms | None = None) -> float:
    """Similaridade determinística entre dois textos, combinando Dice de RADICAIS e razão de
    sequência (`difflib`, sempre com `autojunk=False`: descrição de catálogo chega a 1356
    caracteres e a heurística de autojunk mudaria o score por popularidade de caractere).

    O Dice compara `lexical_stems` dos dois lados — não os tokens crus —, para que
    singular/plural e as derivações mínimas do domínio ("gramado"/"grama") contem como o
    mesmo termo. A razão de sequência continua sobre o texto normalizado tal como está
    (sem radicalizar): ela mede proximidade de GRAFIA, e radicalizar os dois lados antes
    dela destruiria justamente essa informação. Com `synonyms`, os radicais dos dois lados
    são expandidos (`expand_terms`) antes do Dice — é o que permite "REFLETOR EXISTENTE"
    aproximar de uma descrição que só diz "Projetor...".

    Devolve 0.0 quando um dos lados não produz nenhum token; senão um valor em [0, 1].
    """
    left_stems = set(lexical_stems(left))
    right_stems = set(lexical_stems(right))
    if synonyms is not None:
        left_stems = set(expand_terms(tuple(left_stems), synonyms).terms)
        right_stems = set(expand_terms(tuple(right_stems), synonyms).terms)
    if not left_stems or not right_stems:
        return 0.0
    dice = 2 * len(left_stems & right_stems) / (len(left_stems) + len(right_stems))
    norm_left = " ".join(lexical_tokens(left))
    norm_right = " ".join(lexical_tokens(right))
    ratio = difflib.SequenceMatcher(None, norm_left, norm_right, autojunk=False).ratio()
    score = LEXICAL_DICE_WEIGHT * dice + LEXICAL_RATIO_WEIGHT * ratio
    return round(score, 4)


def query_coverage_score(
    query: str, description: str, *, synonyms: DomainSynonyms | None = None
) -> float:
    """Fração dos radicais da CONSULTA presentes nos radicais da descrição, em [0, 1].

    É o braço léxico da fusão híbrida (M7 Fase 2), e existe por causa do limite medido na
    Fase 1: `lexical_similarity` combina Dice, cujo denominador `|A|+|B|` penaliza a
    descrição longa — dois códigos que contêm a consulta inteira pontuam diferente só
    porque um deles é um parágrafo e o outro é uma linha. A cobertura não tem denominador
    do lado da descrição: "PISO INTERTRAVADO" contido num parágrafo de 521 caracteres vale
    1.0, igual a "PISO INTERTRAVADO" sozinho.

    Isso a torna deliberadamente **cega a precisão**: uma descrição que fala de tudo cobre
    qualquer consulta. Por isso ela NÃO substitui `lexical_similarity`, que segue como o
    desempate dentro do mesmo braço (e como o score de compatibilidade que o domínio já
    publica) — a cobertura ordena, a similaridade desempata.

    Com `synonyms`, cada radical da consulta é expandido isoladamente (o mesmo critério de
    `local_server._catalog_search`: a origem de um termo expandido é sempre a palavra que o
    trouxe) e os radicais da descrição são expandidos em bloco, para que uma frase de
    sinônimo com várias palavras possa casar. Um radical conta como presente quando ele —
    ou qualquer equivalente dele — está na descrição.

    Consulta sem nenhum radical utilizável devolve 0.0, nunca divide por zero.

    Todo radical pesa igual aqui. Quando o catálogo inteiro está disponível para dizer
    quais palavras discriminam e quais não, use `weighted_query_coverage_score` — é ela que
    o braço léxico da fusão usa; esta continua sendo o número simples que a resposta mostra
    e o que os testes de contrato exercitam.
    """
    groups = _query_stem_groups(query, synonyms)
    if not groups:
        return 0.0
    description_stems = _expanded_stems(description, synonyms)
    if not description_stems:
        return 0.0
    covered = sum(1 for _stem, group in groups if group & description_stems)
    return round(covered / len(groups), 4)


def _expanded_stems(text: str, synonyms: DomainSynonyms | None) -> set[str]:
    """Radicais de `text` mais os equivalentes de sinônimo, expandidos em bloco.

    Em bloco porque uma frase de sinônimo com várias palavras ("tela de arame galvanizado")
    só casa quando todos os radicais dela estão presentes ao mesmo tempo.
    """
    stems = set(lexical_stems(text))
    if not stems or synonyms is None:
        return stems
    return stems | set(expand_terms(tuple(stems), synonyms).terms)


def _query_stem_groups(
    query: str, synonyms: DomainSynonyms | None
) -> tuple[tuple[str, frozenset[str]], ...]:
    """Um par (radical ORIGINAL, grupo de equivalentes) por radical da consulta, sem repetição.

    A expansão é feita radical a radical, e não com a consulta inteira de uma vez, para que
    cada grupo pertença sem ambiguidade à palavra que o trouxe — é o mesmo critério da busca
    do servidor local, e é o que permite pesar cada palavra separadamente. O radical ORIGINAL
    viaja ao lado do grupo expandido porque `weighted_query_coverage_score` precisa dele para
    decidir ruído: a regra é sobre a palavra que a consulta de fato escreveu, nunca sobre um
    sinônimo que a expansão trouxe.
    """
    groups: list[tuple[str, frozenset[str]]] = []
    for stem in dict.fromkeys(lexical_stems(query)):
        group = {stem}
        if synonyms is not None:
            group |= set(expand_terms((stem,), synonyms).terms)
        groups.append((stem, frozenset(group)))
    return tuple(groups)


@dataclass(frozen=True, slots=True)
class StemIdfTable:
    """Quanto cada radical DISCRIMINA dentro de um catálogo (IDF suavizado).

    `weights` guarda `log(1 + N/df)` por radical, e `default_weight` é `log(1 + N)` — o
    valor de um radical que não aparece em nenhuma descrição, que é o mais discriminativo
    possível. A tabela é derivada de um catálogo específico: usá-la com outro produziria
    pesos que não descrevem aquele vocabulário.
    """

    weights: dict[str, float]
    default_weight: float

    def weight(self, stem: str) -> float:
        return self.weights.get(stem, self.default_weight)


def build_stem_idf_table(
    descriptions: Sequence[str], *, synonyms: DomainSynonyms | None = None
) -> StemIdfTable:
    """Peso de cada radical no vocabulário do catálogo, para a cobertura ponderada.

    A contagem é por DOCUMENTO (em quantas descrições o radical aparece), sobre os radicais
    já expandidos por sinônimo — assim "refletor" e "projetor" contam como o mesmo conceito
    em vez de dois termos raros. Construir custa uma varredura do catálogo inteiro; quem
    chama é responsável por reaproveitar a tabela entre consultas.
    """
    document_frequency: dict[str, int] = {}
    total = 0
    for description in descriptions:
        total += 1
        for stem in _expanded_stems(description, synonyms):
            document_frequency[stem] = document_frequency.get(stem, 0) + 1
    if total == 0:
        return StemIdfTable(weights={}, default_weight=0.0)
    return StemIdfTable(
        weights={stem: math.log(1 + total / count) for stem, count in document_frequency.items()},
        default_weight=math.log(1 + total),
    )


LEGEND_NOISE_DAMP_FACTOR: Final = 0.25
"""Quanto o peso de um radical de ESTADO da legenda é reduzido na cobertura ponderada.

Amortecimento, nunca remoção: a palavra continua contando nos dois lados (numerador e
denominador de `weighted_query_coverage_score`), só que com o peso multiplicado por este
fator em vez do peso IDF cheio.

Calibrado por JUNK-SHARE, não por rank do alvo (rodada 2.2, 2026-08-13) — a métrica certa
depende da pergunta certa. A primeira calibração mediu só o rank do código aceitável nos 12
casos reais gateados do golden e não viu ganho nenhum em `damp ∈ {0.1, 0.25, 0.5}` (ranks
byte-idênticos ao baseline): só 2 dos 12 rótulos contêm um termo de ruído, e nos dois casos
o candidato certo já estava perto do topo antes do amortecimento, então reescalar o peso não
o move. Só que o rank do alvo não é o que a orçamentista vê primeiro — ela lê a LISTA
inteira. Medindo a COMPOSIÇÃO do top-20 ("quantos itens têm como único mérito cobrir um
radical de ruído, sem cobrir nenhum radical de serviço" — `query_covered_only_by_noise`),
"REFLETOR EXISTENTE" tinha 10/20 itens desse tipo (caiação, carpete, chapiscado — todos só
porque a descrição diz "existente") SEM `noise`; COM `noise`, em qualquer um dos três
fatores testados (0.1/0.25/0.5), caem para 0/20 e a lista inteira vira refletor/luminária/
projetor. Os três fatores empataram nessa métrica (0/20 nos três); o critério de desempate
do spec (recall → top-3 → top-1 → rank médio, todos empatados também) mantém o ponto de
partida, 0.25. Ver `weighted_query_coverage_score`, `query_covered_only_by_noise` e
`note_phase_2_2` no golden para a tabela completa. O parâmetro é passado pelos call sites de
produção (`local_server.py`, `cli.py`) desde esta rodada."""


def weighted_query_coverage_score(
    query: str,
    description: str,
    *,
    idf: StemIdfTable,
    synonyms: DomainSynonyms | None = None,
    noise: LegendNoiseList | None = None,
) -> float:
    """Cobertura da consulta ponderada pelo poder de discriminação de cada palavra, em [0, 1].

    Mesma pergunta de `query_coverage_score` — que fração da consulta a descrição cobre? —
    com uma correção medida: cobrir "intertravado" não vale o mesmo que cobrir "piso". A
    fração simples empata o código certo de "PISO INTERTRAVADO" (cuja descrição diz
    "Revestimento intertravado" e nunca "piso") com centenas de serviços de limpeza de
    piso, todos em 0,5; ponderando por IDF o alvo sai do rank 127 para o 3 no braço léxico
    do catálogo real (medição de 2026-08-13).

    O peso de um grupo de sinônimos é o do membro MAIS discriminativo: o grupo é um
    conceito só, e deixar o membro comum mandar diluiria justamente o termo raro que
    identifica o item.

    Contrapartida medida, e não escondida: palavra-ruído RARA no catálogo ganha peso alto.
    O rótulo de prancha descreve o ESTADO do elemento além do serviço ("existente", "a ser
    recuperar"), e em "REFLETOR EXISTENTE" o peso de "existente" (0,532) passa o de
    "refletor" (0,468), enchendo o topo de itens cujo único mérito é conter "existente".
    Ainda assim o saldo medido no recall@20 do golden real é positivo (8/12 para 9/12 no
    oráculo de código único), e por isso a ponderação entrou.

    Com `noise`, o peso de um grupo cujo radical ORIGINAL da consulta (não o expandido por
    sinônimo) está em `noise.noise_stems()` é multiplicado por `LEGEND_NOISE_DAMP_FACTOR`
    — igualmente no numerador e no denominador, então a palavra continua contando, só vale
    menos. A regra olha o radical ORIGINAL de propósito: expansão de sinônimo não pode
    contaminar o julgamento de ruído (um sinônimo que colidisse por acidente com um termo de
    ruído não amortece o radical legítimo que o trouxe).

    Rodada 2.2 (2026-08-13): implementado COM ganho medido, mas não no lugar que a primeira
    calibração olhou. O rank do código aceitável nos 12 casos reais gateados do golden fica
    byte-idêntico ao baseline (a métrica de rank não enxergava o ganho) — porém a COMPOSIÇÃO
    do top-20 muda muito: "REFLETOR EXISTENTE" caiu de 10/20 para 0/20 itens cujo único
    mérito era conter "existente" (`query_covered_only_by_noise`, ver `LEGEND_NOISE_DAMP_FACTOR`
    e `note_phase_2_2` no golden para a tabela e o raciocínio completo). É essa lista, não só
    o rank do alvo, que a orçamentista revisa — por isso o parâmetro é passado pelos call
    sites de produção (`local_server.py`, `cli.py`) desde esta rodada.
    """
    groups = _query_stem_groups(query, synonyms)
    if not groups:
        return 0.0
    description_stems = _expanded_stems(description, synonyms)
    if not description_stems:
        return 0.0
    noise_stems: frozenset[str] = noise.noise_stems() if noise is not None else frozenset()
    total = 0.0
    covered = 0.0
    for original_stem, group in groups:
        weight = max(idf.weight(stem) for stem in group)
        if original_stem in noise_stems:
            weight *= LEGEND_NOISE_DAMP_FACTOR
        total += weight
        if group & description_stems:
            covered += weight
    if total <= 0:
        return 0.0
    return round(covered / total, 4)


def query_covered_only_by_noise(
    query: str,
    description: str,
    *,
    synonyms: DomainSynonyms | None = None,
    noise: LegendNoiseList | None = None,
) -> bool:
    """Verdadeiro quando `description` cobre algum radical de RUÍDO da consulta e NENHUM
    radical de SERVIÇO — o item cujo único mérito na cobertura ponderada é uma palavra de
    ESTADO da legenda ("existente", "a ser recuperar"), não o serviço em si.

    É a definição de "junk" da rodada 2.2 (2026-08-13): o rank do código aceitável não
    enxerga esse defeito (ele já estava perto do topo), mas a COMPOSIÇÃO do top-20 sim — sem
    `noise`, "REFLETOR EXISTENTE" tem 10 desses 20 itens (caiação, carpete, chapiscado,
    revestimentos, todos alheios a refletor); com `noise`, 0. Ver `LEGEND_NOISE_DAMP_FACTOR`
    e `note_phase_2_2` no golden para a medição completa; `hybrid_junk_max` no golden é o
    gate desta função.

    Sem `noise`, todo radical coberto conta como SERVIÇO (não existe distinção de ruído
    nenhuma) — devolve sempre `False`. Descrição sem nenhum radical utilizável também
    devolve `False` (não há cobertura de espécie alguma para classificar).
    """
    if noise is None:
        return False
    description_stems = _expanded_stems(description, synonyms)
    if not description_stems:
        return False
    noise_stems = noise.noise_stems()
    covered_service = False
    covered_noise = False
    for original_stem, group in _query_stem_groups(query, synonyms):
        if not (group & description_stems):
            continue
        if original_stem in noise_stems:
            covered_noise = True
        else:
            covered_service = True
    return covered_noise and not covered_service
