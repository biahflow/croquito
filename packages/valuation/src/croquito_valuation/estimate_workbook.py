"""Render do orçamento-base em planilha: uma aba própria, sem VLOOKUP a catálogo (ADR-0038).

Adaptador do escritor da medição (`workbook_writer.py`), NÃO generalização dele: o
`Estimate` não tem obra múltipla, `Valuation` nem `ContractWorkbook`, e o layout que ele
imprime é o do `EstimateLayout` (`template.py`), seção própria e opcional do
`WorkbookTemplate`. As peças pequenas e sem estado (`_text`/`_number`/`_formula`/
`_atomic_save`/`_sha256`) são cópias deliberadas do desenho de `workbook_writer.py` —
mesmo padrão de duplicação de helper trivial que já existe entre `workbook_writer.py` e
`canonical.py` (`_sha256` nos dois) — em vez de importar nomes privados de outro módulo.

O escritor planeja célula a célula (`plan_estimate_workbook`) antes de gravar, e o mesmo
plano é o oráculo do auditor: nenhuma das duas pontas inventa um valor que a outra não
compartilhe. O auditor reabre o arquivo com `openpyxl.load_workbook` (por dentro de
`canonical.canonicalize_workbook`, reusada como biblioteca do mini-avaliador da gramática
fechada de fórmulas — `GRAMMAR_PATTERNS` — sem estender `canonical.audit_workbook`, que é
da medição) e recomputa cada célula em `Decimal`; falha do auditor não publica nada, gate
exercido por quem chama `write_estimate_workbook`/`audit_estimate_workbook` em sequência
(o mesmo desenho de `run_export_valuation`, no worker).

Conteúdo impresso, por decisão do pacote aprovado (ADR-0038, decisão 5) e do contrato de
T2:

- uma linha por `EstimateLine`, com FONTE = origem + data-base numa célula só e as duas
  colunas de preço (sem e com BDI);
- BDI percentual declarado uma vez, não repetido por linha;
- o valor do BDI impresso é a diferença entre os totais truncados
  (`estimate.total_amount - estimate.total_amount_without_bdi`), nunca o percentual
  aplicado ao total geral — os dois totais já vêm prontos e validados do `Estimate`;
- bloco próprio "itens sem preço na cascata" com os ids declarados em
  `unpriced_item_ids`; o `Estimate` não carrega descrição desses itens (eles nunca viram
  linha), então só o id é impresso.

Ao lado disso, e sem substituí-lo, mora o escritor do GABARITO (F-043):
`plan_estimate_grid_workbook`/`write_estimate_grid_workbook`/`audit_estimate_grid_workbook`
publicam o documento que a prefeitura recebe — a ordem é a do gabarito declarado
(`template.estimate_grid`), nunca a de `estimate.lines`, TODA linha do gabarito é impressa
(inclusive as de quantidade zero) e a aba de memória de cálculo vai ao lado, renderizada
pelo MESMO `workbook_writer.plan_calc_block` da medição. Quem não declara `estimate_grid`
continua na rodada de hoje, byte a byte.

O grupo é uma LINHA própria, com apenas o número, intercalada antes da primeira linha de
cada grupo — não é coluna. O documento da prefeitura tem sete colunas
(`ITEM | COD. | DESCRIÇÃO | UN | VALOR UNIT | QUANT | TOTAL`) e 21 linhas de grupo entre as
433 de código; imprimir o grupo como oitava coluna faria o arquivo divergir do documento em
21 lugares, cada um deles visível quando os dois são abertos lado a lado.
"""

from __future__ import annotations

import hashlib
import os
import tempfile
from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path
from typing import Final, Literal

from openpyxl import Workbook
from openpyxl.styles import Font
from pydantic import Field, model_validator

from croquito_valuation.canonical import canonicalize_workbook
from croquito_valuation.errors import ValuationValidationError
from croquito_valuation.estimate import Estimate, EstimateLine
from croquito_valuation.models import CalcSheet, ExactDecimal, ValuationContractModel
from croquito_valuation.rounding import quantity_round, trunc_divergence
from croquito_valuation.template import (
    EstimateLayout,
    EstimateTemplateLayout,
    WorkbookTemplate,
)
from croquito_valuation.workbook_writer import PlannedCell, plan_calc_block

MAX_DECIMAL_PLACES: Final = 2
PINNED_REASON: Final = "TRUNC_DOUBLE_DIVERGENCE"
_FIXED_TIMESTAMP: Final = datetime(2026, 1, 1, tzinfo=UTC)

CellKind = Literal["text", "number", "formula"]


class EstimatePinnedCell(ValuationContractModel):
    """Célula de total de linha fixada por divergência de truncamento (espelho de
    `workbook_writer.PinnedCell`, sem o campo `sheet` — a planilha do orçamento tem uma
    aba só)."""

    ref: str
    reason: Literal["TRUNC_DOUBLE_DIVERGENCE"] = PINNED_REASON
    item_number: str
    quantity: ExactDecimal
    unit_price_with_bdi: ExactDecimal
    value: ExactDecimal


class EstimatePlannedCell(ValuationContractModel):
    """Uma célula planejada da aba do orçamento, com o valor exato que ela deve representar."""

    ref: str = Field(pattern=r"^[A-Z]{1,2}\d+$")
    kind: CellKind
    role: str = Field(min_length=1, max_length=60)
    item_number: str | None = None
    text: str | None = None
    number: ExactDecimal | None = None
    formula: str | None = None
    number_format: str | None = None
    bold: bool = False

    @model_validator(mode="after")
    def validate_payload(self) -> EstimatePlannedCell:
        if self.kind == "text" and self.text is None:
            raise ValuationValidationError(
                "PLANNED_CELL_INVALID", "célula de texto sem texto", {"ref": self.ref}
            )
        if self.kind == "number" and self.number is None:
            raise ValuationValidationError(
                "PLANNED_CELL_INVALID", "célula numérica sem valor", {"ref": self.ref}
            )
        if self.kind == "formula" and (self.formula is None or self.number is None):
            raise ValuationValidationError(
                "PLANNED_CELL_INVALID",
                "célula de fórmula precisa de fórmula e do valor exato esperado",
                {"ref": self.ref},
            )
        return self


class EstimatePlannedSheet(ValuationContractModel):
    """Uma aba planejada da planilha do orçamento."""

    name: str = Field(min_length=1, max_length=31)
    cells: list[EstimatePlannedCell] = Field(min_length=1)
    column_widths: dict[str, int] = Field(default_factory=dict)


class EstimateWorkbookPlan(ValuationContractModel):
    """Plano completo da planilha do orçamento, usado para gravar e para auditar.

    `sheet_name`/`cells`/`column_widths` são a aba do orçamento — a única que existia
    antes do gabarito. `memory_sheet` é a aba de memória de cálculo e só nasce na rodada
    com gabarito declarado (F-043); sem ela o plano é exatamente o de antes.
    """

    sheet_name: str
    cells: list[EstimatePlannedCell] = Field(min_length=1)
    column_widths: dict[str, int] = Field(default_factory=dict)
    pinned_cells: list[EstimatePinnedCell] = Field(default_factory=list)
    memory_sheet: EstimatePlannedSheet | None = None

    @property
    def sheets(self) -> tuple[EstimatePlannedSheet, ...]:
        """Abas planejadas, na ordem em que a pasta as grava."""
        estimate_sheet = EstimatePlannedSheet(
            name=self.sheet_name, cells=self.cells, column_widths=self.column_widths
        )
        if self.memory_sheet is None:
            return (estimate_sheet,)
        return (estimate_sheet, self.memory_sheet)

    @property
    def planned_cells(self) -> int:
        """Quantas células o plano prevê, somando todas as abas."""
        return sum(len(sheet.cells) for sheet in self.sheets)

    @property
    def formula_cells(self) -> int:
        """Quantas células de fórmula o plano prevê."""
        return sum(1 for sheet in self.sheets for cell in sheet.cells if cell.kind == "formula")


class EstimateWriteReport(ValuationContractModel):
    """Resultado da gravação da planilha do orçamento."""

    output_path: str
    workbook_sha256: str = Field(pattern=r"^[a-f0-9]{64}$")
    sheet_name: str
    memory_sheet_name: str | None = None
    written_cells: int = Field(ge=1)
    formula_cells: int = Field(ge=0)
    pinned_cells: list[EstimatePinnedCell] = Field(default_factory=list)


class EstimateAuditFinding(ValuationContractModel):
    """Uma divergência entre o arquivo reaberto e o orçamento que a planilha deveria imprimir."""

    code: str = Field(min_length=1, max_length=60)
    sheet: str
    ref: str | None = None
    expected: str | None = None
    found: str | None = None
    detail: str | None = None


class EstimateAuditReport(ValuationContractModel):
    """Resultado da auditoria de round-trip da planilha do orçamento-base."""

    status: Literal["ok", "divergent"]
    workbook_sha256: str = Field(pattern=r"^[a-f0-9]{64}$")
    sheet_name: str
    memory_sheet_name: str | None = None
    checked_cells: int = Field(ge=0)
    formula_cells: int = Field(ge=0)
    total_amount: ExactDecimal
    pinned_cells: list[EstimatePinnedCell] = Field(default_factory=list)
    findings: list[EstimateAuditFinding] = Field(default_factory=list)


def _checked_number(value: Decimal, ref: str, role: str) -> Decimal:
    exponent = value.as_tuple().exponent
    if isinstance(exponent, int) and exponent < -MAX_DECIMAL_PLACES:
        raise ValuationValidationError(
            "DECIMAL_SCALE_UNSUPPORTED",
            "planilha só representa duas casas decimais sem perder o valor exato",
            {"ref": ref, "role": role, "value": str(value)},
        )
    return value


def _text(
    ref: str,
    role: str,
    value: str,
    *,
    bold: bool = False,
    item_number: str | None = None,
) -> EstimatePlannedCell:
    return EstimatePlannedCell(
        ref=ref, kind="text", role=role, text=value, bold=bold, item_number=item_number
    )


def _number(
    ref: str,
    role: str,
    value: Decimal,
    *,
    number_format: str | None = None,
    bold: bool = False,
    item_number: str | None = None,
) -> EstimatePlannedCell:
    return EstimatePlannedCell(
        ref=ref,
        kind="number",
        role=role,
        item_number=item_number,
        number=_checked_number(value, ref, role),
        number_format=number_format,
        bold=bold,
    )


def _formula(
    ref: str,
    role: str,
    formula: str,
    value: Decimal,
    *,
    number_format: str | None = None,
    bold: bool = False,
    item_number: str | None = None,
) -> EstimatePlannedCell:
    return EstimatePlannedCell(
        ref=ref,
        kind="formula",
        role=role,
        item_number=item_number,
        formula=formula,
        number=_checked_number(value, ref, role),
        number_format=number_format,
        bold=bold,
    )


def _require_layout(template: WorkbookTemplate) -> EstimateLayout:
    if template.estimate is None:
        raise ValuationValidationError(
            "TEMPLATE_ESTIMATE_LAYOUT_MISSING",
            "o template não declara a seção do orçamento-base (WorkbookTemplate.estimate)",
            {},
        )
    return template.estimate


def _require_grid_layout(template: WorkbookTemplate) -> EstimateTemplateLayout:
    if template.estimate_grid is None:
        raise ValuationValidationError(
            "TEMPLATE_ESTIMATE_GRID_MISSING",
            "o template não declara o gabarito da prefeitura (WorkbookTemplate.estimate_grid)",
            {},
        )
    return template.estimate_grid


def _duplicated_refs(cells: list[EstimatePlannedCell]) -> list[str]:
    seen: set[str] = set()
    duplicated: list[str] = []
    for cell in cells:
        if cell.ref in seen:
            duplicated.append(cell.ref)
        seen.add(cell.ref)
    return duplicated


def _ensure_no_collision(cells: list[EstimatePlannedCell], sheet_name: str) -> None:
    duplicated = _duplicated_refs(cells)
    if duplicated:
        raise ValuationValidationError(
            "PLAN_CELL_COLLISION",
            "o plano da planilha do orçamento gravaria duas vezes a mesma célula",
            {"sheet": sheet_name, "refs": duplicated},
        )


def plan_estimate_workbook(estimate: Estimate, template: WorkbookTemplate) -> EstimateWorkbookPlan:
    """Planeja a aba inteira do orçamento, com o valor exato de cada célula."""
    layout = _require_layout(template)
    columns = layout.columns
    pinned: list[EstimatePinnedCell] = []
    cells: list[EstimatePlannedCell] = [
        _text(f"{layout.label_column}1", "title", layout.title, bold=True)
    ]

    pairs: list[tuple[str, str]] = [(layout.intervention_label, estimate.worksite_name)]
    if estimate.address is not None:
        pairs.append((layout.address_label, estimate.address))
    pairs.append((layout.bdi_label, f"{quantity_round(estimate.bdi_percent)}%"))
    if layout.header_row < len(pairs) + 3:
        raise ValuationValidationError(
            "TEMPLATE_HEADER_ROW_TOO_SMALL",
            "linha de cabeçalho do orçamento não deixa espaço para o bloco de identificação",
            {"header_row": layout.header_row, "required_rows": len(pairs) + 3},
        )
    for index, (label, value) in enumerate(pairs):
        row = 2 + index
        cells.append(_text(f"{layout.label_column}{row}", "header_label", label, bold=True))
        cells.append(_text(f"{layout.value_column}{row}", "header_value", value))

    for column in columns.ordered:
        cells.append(
            _text(f"{column.letter}{layout.header_row}", "column_label", column.label, bold=True)
        )

    first_line_row = layout.header_row + 1
    for index, line in enumerate(estimate.lines):
        row = first_line_row + index
        item = line.item_number
        cells.append(_text(f"{columns.item.letter}{row}", "line_item", item, item_number=item))
        cells.append(_text(f"{columns.code.letter}{row}", "line_code", line.code, item_number=item))
        cells.append(
            _text(
                f"{columns.source.letter}{row}",
                "line_source",
                f"{line.price_origin.value.upper()} {line.reference_month}",
                item_number=item,
            )
        )
        cells.append(
            _text(
                f"{columns.description.letter}{row}",
                "line_description",
                line.description,
                item_number=item,
            )
        )
        cells.append(_text(f"{columns.unit.letter}{row}", "line_unit", line.unit, item_number=item))
        cells.append(
            _number(
                f"{columns.unit_price.letter}{row}",
                "line_unit_price",
                line.unit_price,
                number_format=layout.money_number_format,
                item_number=item,
            )
        )
        cells.append(
            _number(
                f"{columns.unit_price_with_bdi.letter}{row}",
                "line_unit_price_with_bdi",
                line.unit_price_with_bdi,
                number_format=layout.money_number_format,
                item_number=item,
            )
        )
        cells.append(
            _number(
                f"{columns.quantity.letter}{row}",
                "line_quantity",
                line.quantity,
                number_format=layout.quantity_number_format,
                item_number=item,
            )
        )
        quantity_ref = f"{columns.quantity.letter}{row}"
        price_ref = f"{columns.unit_price_with_bdi.letter}{row}"
        total_ref = f"{columns.total.letter}{row}"
        if trunc_divergence(line.quantity, line.unit_price_with_bdi):
            pinned.append(
                EstimatePinnedCell(
                    ref=total_ref,
                    item_number=item,
                    quantity=line.quantity,
                    unit_price_with_bdi=line.unit_price_with_bdi,
                    value=line.total,
                )
            )
            cells.append(
                _number(
                    total_ref,
                    "line_total_pinned",
                    line.total,
                    number_format=layout.money_number_format,
                    item_number=item,
                )
            )
        else:
            cells.append(
                _formula(
                    total_ref,
                    "line_total",
                    f"=TRUNC({quantity_ref}*{price_ref},2)",
                    line.total,
                    number_format=layout.money_number_format,
                    item_number=item,
                )
            )

    last_line_row = first_line_row + len(estimate.lines) - 1
    total_col = columns.total.letter
    total_without_bdi_row = last_line_row + 2
    bdi_amount_row = total_without_bdi_row + 1
    total_row = bdi_amount_row + 1

    total_without_bdi_ref = f"{total_col}{total_without_bdi_row}"
    cells.append(
        _text(
            f"{columns.description.letter}{total_without_bdi_row}",
            "total_without_bdi_label",
            layout.total_without_bdi_label,
            bold=True,
        )
    )
    cells.append(
        _number(
            total_without_bdi_ref,
            "estimate_total_without_bdi",
            estimate.total_amount_without_bdi,
            number_format=layout.money_number_format,
            bold=True,
        )
    )

    total_ref = f"{total_col}{total_row}"
    bdi_amount_label = f"{layout.bdi_label} ({quantity_round(estimate.bdi_percent)}%)"
    cells.append(
        _text(
            f"{columns.description.letter}{bdi_amount_row}",
            "bdi_amount_label",
            bdi_amount_label,
            bold=True,
        )
    )
    cells.append(
        _formula(
            f"{total_col}{bdi_amount_row}",
            "estimate_bdi_amount",
            f"={total_ref}-{total_without_bdi_ref}",
            estimate.total_amount - estimate.total_amount_without_bdi,
            number_format=layout.money_number_format,
            bold=True,
        )
    )

    cells.append(
        _text(
            f"{columns.description.letter}{total_row}",
            "total_label",
            layout.total_label,
            bold=True,
        )
    )
    cells.append(
        _formula(
            total_ref,
            "estimate_total",
            f"=SUM({total_col}{first_line_row}:{total_col}{last_line_row})",
            estimate.total_amount,
            number_format=layout.money_number_format,
            bold=True,
        )
    )

    if estimate.unpriced_item_ids:
        unpriced_header_row = total_row + 2
        cells.append(
            _text(
                f"{columns.item.letter}{unpriced_header_row}",
                "unpriced_section_label",
                layout.unpriced_section_label,
                bold=True,
            )
        )
        for offset, item_id in enumerate(estimate.unpriced_item_ids):
            row = unpriced_header_row + 1 + offset
            cells.append(_text(f"{columns.item.letter}{row}", "unpriced_item", item_id))

    _ensure_no_collision(cells, layout.sheet_name)

    return EstimateWorkbookPlan(
        sheet_name=layout.sheet_name,
        cells=cells,
        column_widths={column.letter: column.width for column in columns.ordered},
        pinned_cells=pinned,
    )


def _adapt_planned_cell(cell: PlannedCell) -> EstimatePlannedCell:
    """Converte a célula planejada da medição na do orçamento, campo a campo.

    Os dois modelos têm a mesma forma e vidas separadas de propósito (a docstring do
    módulo explica por quê). Este conversor existe para que a memória do orçamento possa
    reusar `workbook_writer.plan_calc_block` — o único render de bloco do repositório —
    sem que nenhum dos dois modelos precise virar o outro.
    """
    return EstimatePlannedCell(
        ref=cell.ref,
        kind=cell.kind,
        role=cell.role,
        item_number=cell.item_number,
        text=cell.text,
        number=cell.number,
        formula=cell.formula,
        number_format=cell.number_format,
        bold=cell.bold,
    )


def _ensure_estimate_fits_the_grid(estimate: Estimate, grid: EstimateTemplateLayout) -> None:
    """Recusa ANTES de qualquer escrita: nada é gravado se o mapeamento não fecha.

    Código do orçamento ausente do gabarito é `ESTIMATE_GRID_CODE_ABSENT` nomeando o
    código; a linha jamais é acrescentada ao fim do arquivo, porque o documento entregue
    à prefeitura é o gabarito dela e uma linha inventada no fim é justamente o que ele
    não admite.

    Dois `EstimateLine` com o mesmo código também recusam
    (`ESTIMATE_GRID_CODE_DUPLICATE`): o gabarito tem UMA linha por código, e escolher qual
    das duas quantidades imprimir seria a máquina decidindo em silêncio.
    """
    index = grid.row_index_by_code
    absent = sorted({line.code for line in estimate.lines if line.code not in index})
    if absent:
        raise ValuationValidationError(
            "ESTIMATE_GRID_CODE_ABSENT",
            "o orçamento tem código que o gabarito da prefeitura não declara",
            {"sheet": grid.sheet_name, "revision_label": grid.revision_label, "codes": absent},
        )
    codes = [line.code for line in estimate.lines]
    duplicated = sorted({code for code in codes if codes.count(code) > 1})
    if duplicated:
        raise ValuationValidationError(
            "ESTIMATE_GRID_CODE_DUPLICATE",
            "o orçamento repete um código que o gabarito imprime numa linha só",
            {"sheet": grid.sheet_name, "codes": duplicated},
        )


def _plan_grid_sheet(
    estimate: Estimate,
    grid: EstimateTemplateLayout,
    lines_by_code: dict[str, EstimateLine],
) -> tuple[EstimatePlannedSheet, list[EstimatePinnedCell]]:
    """Planeja a aba do gabarito: a ordem é a dele, e toda linha dele é impressa."""
    columns = grid.columns
    pinned: list[EstimatePinnedCell] = []
    cells: list[EstimatePlannedCell] = [
        _text(f"{grid.label_column}1", "title", grid.title, bold=True)
    ]

    pairs: list[tuple[str, str]] = [(grid.intervention_label, estimate.worksite_name)]
    if estimate.address is not None:
        pairs.append((grid.address_label, estimate.address))
    pairs.append((grid.bdi_label, f"{quantity_round(estimate.bdi_percent)}%"))
    pairs.append((grid.revision_row_label, grid.revision_label))
    if grid.header_row < len(pairs) + 3:
        raise ValuationValidationError(
            "TEMPLATE_HEADER_ROW_TOO_SMALL",
            "linha de cabeçalho do gabarito não deixa espaço para o bloco de identificação",
            {"header_row": grid.header_row, "required_rows": len(pairs) + 3},
        )
    for index, (label, value) in enumerate(pairs):
        row = 2 + index
        cells.append(_text(f"{grid.label_column}{row}", "header_label", label, bold=True))
        cells.append(_text(f"{grid.value_column}{row}", "header_value", value))

    for column in columns.printed:
        cells.append(
            _text(f"{column.letter}{grid.header_row}", "column_label", column.label, bold=True)
        )

    first_line_row = grid.header_row + 1
    row = first_line_row
    previous_group: str | None = None
    for grid_row in grid.rows:
        # O documento imprime o grupo como LINHA própria, com apenas o número, e nunca como
        # coluna (pacote aprovado da F-043, revisões 2 e 3). Grupo que reaparece depois de
        # outro imprime a linha de novo: a ordem impressa é a que o gabarito declara, e
        # validar contiguidade seria o escritor decidindo sobre o documento do cliente.
        if grid_row.group != previous_group:
            cells.append(_text(f"{columns.item.letter}{row}", "group_row", grid_row.group))
            row += 1
        previous_group = grid_row.group

        item = grid_row.item
        cells.append(_text(f"{columns.item.letter}{row}", "line_item", item, item_number=item))
        cells.append(
            _text(f"{columns.code.letter}{row}", "line_code", grid_row.code, item_number=item)
        )
        cells.append(
            _text(
                f"{columns.description.letter}{row}",
                "line_description",
                grid_row.description,
                item_number=item,
            )
        )
        cells.append(
            _text(f"{columns.unit.letter}{row}", "line_unit", grid_row.unit, item_number=item)
        )

        line = lines_by_code.get(grid_row.code)
        quantity = line.quantity if line is not None else Decimal("0.00")
        cells.append(
            _number(
                f"{columns.quantity.letter}{row}",
                "line_quantity" if line is not None else "line_quantity_zero",
                quantity,
                number_format=grid.quantity_number_format,
                item_number=item,
            )
        )

        # Preço impresso: manda o do orçamento (a mesma base do total de hoje,
        # `unit_price_with_bdi`); sem linha no orçamento vale o preço declarado no
        # gabarito, se houver. Os dois NUNCA são comparados e divergência entre eles não
        # recusa nada: qual preço vale é decisão de quem monta o orçamento, e o gabarito
        # é o documento, não uma segunda fonte de preço.
        unit_price = line.unit_price_with_bdi if line is not None else grid_row.unit_price
        price_ref = f"{columns.unit_price.letter}{row}"
        if unit_price is not None:
            cells.append(
                _number(
                    price_ref,
                    "line_unit_price",
                    unit_price,
                    number_format=grid.money_number_format,
                    item_number=item,
                )
            )

        total_ref = f"{columns.total.letter}{row}"
        total = line.total if line is not None else Decimal("0.00")
        if unit_price is None:
            # Sem célula de preço a fórmula operaria sobre célula vazia e a auditoria
            # recusaria a pasta inteira (`FORMULA_REFERENCE_EMPTY`). A linha continua
            # presente e zerada, que é o que o documento exige.
            cells.append(
                _number(
                    total_ref,
                    "line_total_zero",
                    total,
                    number_format=grid.money_number_format,
                    item_number=item,
                )
            )
        elif trunc_divergence(quantity, unit_price):
            pinned.append(
                EstimatePinnedCell(
                    ref=total_ref,
                    item_number=item,
                    quantity=quantity,
                    unit_price_with_bdi=unit_price,
                    value=total,
                )
            )
            cells.append(
                _number(
                    total_ref,
                    "line_total_pinned",
                    total,
                    number_format=grid.money_number_format,
                    item_number=item,
                )
            )
        else:
            cells.append(
                _formula(
                    total_ref,
                    "line_total",
                    f"=TRUNC({columns.quantity.letter}{row}*{price_ref},2)",
                    total,
                    number_format=grid.money_number_format,
                    item_number=item,
                )
            )

        row += 1

    # Última linha IMPRESSA — o cursor já passou dela. O intervalo do total começa na
    # primeira linha impressa e inclui as linhas de grupo, que não têm célula na coluna do
    # total e por isso não entram na soma: é assim que o documento do cliente soma
    # (`G10:G463`), e o mini-avaliador da gramática ignora célula vazia do mesmo jeito.
    last_line_row = row - 1
    total_col = columns.total.letter
    total_without_bdi_row = last_line_row + 2
    bdi_amount_row = total_without_bdi_row + 1
    total_row = bdi_amount_row + 1

    total_without_bdi_ref = f"{total_col}{total_without_bdi_row}"
    cells.append(
        _text(
            f"{columns.description.letter}{total_without_bdi_row}",
            "total_without_bdi_label",
            grid.total_without_bdi_label,
            bold=True,
        )
    )
    cells.append(
        _number(
            total_without_bdi_ref,
            "estimate_total_without_bdi",
            estimate.total_amount_without_bdi,
            number_format=grid.money_number_format,
            bold=True,
        )
    )

    total_ref = f"{total_col}{total_row}"
    cells.append(
        _text(
            f"{columns.description.letter}{bdi_amount_row}",
            "bdi_amount_label",
            f"{grid.bdi_label} ({quantity_round(estimate.bdi_percent)}%)",
            bold=True,
        )
    )
    # ADR-0038, decisão 4: o BDI impresso é a DIFERENÇA entre os totais truncados, nunca o
    # percentual aplicado ao total geral.
    cells.append(
        _formula(
            f"{total_col}{bdi_amount_row}",
            "estimate_bdi_amount",
            f"={total_ref}-{total_without_bdi_ref}",
            estimate.total_amount - estimate.total_amount_without_bdi,
            number_format=grid.money_number_format,
            bold=True,
        )
    )
    cells.append(
        _text(
            f"{columns.description.letter}{total_row}",
            "total_label",
            grid.total_label,
            bold=True,
        )
    )
    cells.append(
        _formula(
            total_ref,
            "estimate_total",
            f"=SUM({total_col}{first_line_row}:{total_col}{last_line_row})",
            estimate.total_amount,
            number_format=grid.money_number_format,
            bold=True,
        )
    )

    if estimate.unpriced_item_ids:
        unpriced_header_row = total_row + 2
        cells.append(
            _text(
                f"{columns.item.letter}{unpriced_header_row}",
                "unpriced_section_label",
                grid.unpriced_section_label,
                bold=True,
            )
        )
        for offset, item_id in enumerate(estimate.unpriced_item_ids):
            row = unpriced_header_row + 1 + offset
            cells.append(_text(f"{columns.item.letter}{row}", "unpriced_item", item_id))

    _ensure_no_collision(cells, grid.sheet_name)
    return (
        EstimatePlannedSheet(
            name=grid.sheet_name,
            cells=cells,
            # Só as colunas impressas ganham largura: declarar largura para a coluna de
            # grupo criaria, no arquivo entregue, uma coluna vazia que o documento da
            # prefeitura não tem.
            column_widths={column.letter: column.width for column in columns.printed},
        ),
        pinned,
    )


def _plan_grid_memory(
    estimate: Estimate,
    template: WorkbookTemplate,
    grid: EstimateTemplateLayout,
    ordered_lines: list[EstimateLine],
) -> EstimatePlannedSheet:
    """Planeja a memória de cálculo do orçamento reusando o render de bloco da medição.

    O bloco (rótulo, operandos nomeados, dedução e subtotal) sai inteiro de
    `workbook_writer.plan_calc_block`, que consome `CalcBlock` + `template.memory` e não
    sabe nada de `Valuation`. O que este planejador faz por conta própria é o que é
    específico do orçamento: iterar as memórias do `Estimate` (`calc_sheets`), na ordem em
    que o gabarito imprime as linhas, para que as duas abas contem a mesma história na
    mesma sequência.

    A memória imprime QUANTIDADE, não dinheiro: o dinheiro está no gabarito, e repeti-lo
    aqui duplicaria a regra de truncamento em dois lugares que teriam de concordar.
    """
    layout = template.memory
    columns = layout.columns
    if layout.header_row < 3:
        raise ValuationValidationError(
            "TEMPLATE_HEADER_ROW_TOO_SMALL",
            "linha de cabeçalho da memória não deixa espaço para o título",
            {"header_row": layout.header_row, "required_rows": 3},
        )
    cells: list[EstimatePlannedCell] = [
        _text(f"{layout.label_column}1", "title", layout.title, bold=True),
        _text(f"{layout.label_column}2", "header_label", layout.intervention_label, bold=True),
        _text(f"{layout.value_column}2", "header_value", estimate.worksite_name),
    ]
    printed_columns = (
        columns.item,
        columns.code,
        columns.description,
        columns.unit,
        columns.quantity,
    )
    for column in printed_columns:
        cells.append(
            _text(f"{column.letter}{layout.header_row}", "column_label", column.label, bold=True)
        )

    sheets_by_item: dict[str, CalcSheet] = {
        sheet.item_number: sheet for sheet in estimate.calc_sheets
    }
    row = layout.header_row + 1
    for line in ordered_lines:
        sheet = sheets_by_item[line.item_number]
        item = line.item_number
        summary_row = row
        cells.append(
            _text(f"{columns.item.letter}{summary_row}", "line_item", item, item_number=item)
        )
        cells.append(
            _text(f"{columns.code.letter}{summary_row}", "line_code", line.code, item_number=item)
        )
        cells.append(
            _text(
                f"{columns.description.letter}{summary_row}",
                "line_description",
                line.description,
                item_number=item,
            )
        )
        cells.append(
            _text(f"{columns.unit.letter}{summary_row}", "line_unit", line.unit, item_number=item)
        )
        cells.append(
            _number(
                f"{columns.quantity.letter}{summary_row}",
                "line_quantity",
                line.quantity,
                number_format=layout.quantity_number_format,
                item_number=item,
            )
        )
        row = summary_row + 1
        subtotal_rows: list[int] = []
        for block in sheet.blocks:
            block_cells, value_row = plan_calc_block(block, template, row)
            cells.extend(_adapt_planned_cell(cell) for cell in block_cells)
            subtotal_rows.append(value_row)
            row = value_row + 1
        cells.append(
            _text(
                f"{layout.block_label_column}{row}",
                "sheet_total_label",
                layout.total_label,
                bold=True,
            )
        )
        cells.append(
            _formula(
                f"{layout.subtotal_column}{row}",
                "sheet_total",
                f"=SUM({layout.subtotal_column}{subtotal_rows[0]}"
                f":{layout.subtotal_column}{subtotal_rows[-1]})",
                sheet.total_quantity,
                number_format=layout.quantity_number_format,
                bold=True,
            )
        )
        row += 2

    _ensure_no_collision(cells, grid.memory_sheet_name)
    return EstimatePlannedSheet(
        name=grid.memory_sheet_name,
        cells=cells,
        column_widths={column.letter: column.width for column in columns.ordered},
    )


def plan_estimate_grid_workbook(
    estimate: Estimate, template: WorkbookTemplate
) -> EstimateWorkbookPlan:
    """Planeja as duas abas do documento da prefeitura, com o valor exato de cada célula.

    A ordem das linhas é a do GABARITO (`template.estimate_grid.rows`), nunca a de
    `estimate.lines`: a linha da planilha é `header_row + 1 + índice no gabarito`, e o
    cursor sequencial do escritor sem gabarito não vale aqui. Toda linha do gabarito é
    impressa — as que o orçamento não preenche saem com quantidade e total ZERADOS, nunca
    ausentes, porque as 390 linhas de zero do documento real fazem parte da entrega.
    """
    grid = _require_grid_layout(template)
    _ensure_estimate_fits_the_grid(estimate, grid)
    lines_by_code = {line.code: line for line in estimate.lines}
    index = grid.row_index_by_code
    ordered_lines = sorted(estimate.lines, key=lambda line: index[line.code])

    estimate_sheet, pinned = _plan_grid_sheet(estimate, grid, lines_by_code)
    memory_sheet = _plan_grid_memory(estimate, template, grid, ordered_lines)
    return EstimateWorkbookPlan(
        sheet_name=estimate_sheet.name,
        cells=estimate_sheet.cells,
        column_widths=estimate_sheet.column_widths,
        pinned_cells=pinned,
        memory_sheet=memory_sheet,
    )


def _atomic_save(workbook: Workbook, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        dir=output_path.parent, prefix=f".{output_path.name}.", suffix=".xlsx"
    )
    os.close(descriptor)
    temporary_path = Path(temporary_name)
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, output_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _write_plan(plan: EstimateWorkbookPlan, output_path: Path) -> EstimateWriteReport:
    """Grava as abas planejadas e devolve o relatório; nenhum valor nasce aqui."""
    workbook = Workbook()
    # Mesma razão da medição: a data de criação é fixa para não carimbar o relógio local
    # no artefato; a idempotência que importa é a do conteúdo lógico, não a dos bytes.
    workbook.properties.created = _FIXED_TIMESTAMP
    active = workbook.active
    assert active is not None  # sempre existe: Workbook() nasce com uma aba ativa
    for index, planned_sheet in enumerate(plan.sheets):
        worksheet = active if index == 0 else workbook.create_sheet()
        worksheet.title = planned_sheet.name
        for letter, width in planned_sheet.column_widths.items():
            worksheet.column_dimensions[letter].width = width
        for planned_cell in planned_sheet.cells:
            cell = worksheet[planned_cell.ref]
            if planned_cell.kind == "text":
                cell.value = planned_cell.text
            elif planned_cell.kind == "number":
                cell.value = planned_cell.number
            else:
                cell.value = planned_cell.formula
            if planned_cell.number_format is not None:
                cell.number_format = planned_cell.number_format
            if planned_cell.bold:
                cell.font = Font(bold=True)
    _atomic_save(workbook, output_path)
    return EstimateWriteReport(
        output_path=str(output_path),
        workbook_sha256=_sha256(output_path),
        sheet_name=plan.sheet_name,
        memory_sheet_name=None if plan.memory_sheet is None else plan.memory_sheet.name,
        written_cells=plan.planned_cells,
        formula_cells=plan.formula_cells,
        pinned_cells=plan.pinned_cells,
    )


def write_estimate_workbook(
    estimate: Estimate, template: WorkbookTemplate, output_path: Path
) -> EstimateWriteReport:
    """Grava a planilha do orçamento-base e devolve o relatório da gravação."""
    return _write_plan(plan_estimate_workbook(estimate, template), output_path)


def write_estimate_grid_workbook(
    estimate: Estimate, template: WorkbookTemplate, output_path: Path
) -> EstimateWriteReport:
    """Grava o documento da prefeitura: gabarito e memória de cálculo, nessa ordem.

    Planeja antes de abrir arquivo nenhum, então a recusa `ESTIMATE_GRID_CODE_ABSENT`
    acontece com o disco intocado.
    """
    return _write_plan(plan_estimate_grid_workbook(estimate, template), output_path)


def _expected_text(planned_cell: EstimatePlannedCell) -> str:
    """Valor canônico que o plano prevê para a célula, no mesmo formato do arquivo."""
    if planned_cell.kind == "text":
        return str(planned_cell.text)
    number = planned_cell.number
    if number is None:  # pragma: no cover - impedido pelo validador de EstimatePlannedCell
        raise ValuationValidationError(
            "PLANNED_CELL_INVALID",
            "célula numérica planejada sem valor",
            {"ref": planned_cell.ref},
        )
    return str(quantity_round(number))


def _canonical_sheet_names(canonical: dict[str, object]) -> list[str]:
    sheets = canonical["sheets"]
    assert isinstance(sheets, list)
    names: list[str] = []
    for sheet in sheets:
        assert isinstance(sheet, dict)
        names.append(str(sheet["name"]))
    return names


def _canonical_sheet_cells(
    canonical: dict[str, object], sheet_name: str
) -> dict[str, dict[str, object]] | None:
    sheets = canonical["sheets"]
    assert isinstance(sheets, list)
    for sheet in sheets:
        assert isinstance(sheet, dict)
        if str(sheet["name"]) != sheet_name:
            continue
        raw_cells = sheet["cells"]
        assert isinstance(raw_cells, list)
        indexed: dict[str, dict[str, object]] = {}
        for cell in raw_cells:
            assert isinstance(cell, dict)
            indexed[str(cell["ref"])] = cell
        return indexed
    return None


def _compare_plan(
    plan: EstimateWorkbookPlan, canonical: dict[str, object]
) -> list[EstimateAuditFinding]:
    """Confronta cada aba planejada com o arquivo reaberto, aba a aba e célula a célula."""
    findings: list[EstimateAuditFinding] = []
    for planned_sheet in plan.sheets:
        found_cells = _canonical_sheet_cells(canonical, planned_sheet.name)
        if found_cells is None:
            findings.append(
                EstimateAuditFinding(
                    code="SHEET_MISSING",
                    sheet=planned_sheet.name,
                    detail="aba do orçamento não existe no arquivo gravado",
                )
            )
            found_cells = {}

        planned_refs: set[str] = set()
        for planned_cell in planned_sheet.cells:
            planned_refs.add(planned_cell.ref)
            found = found_cells.get(planned_cell.ref)
            if found is None:
                findings.append(
                    EstimateAuditFinding(
                        code="CELL_MISSING",
                        sheet=planned_sheet.name,
                        ref=planned_cell.ref,
                        expected=_expected_text(planned_cell),
                        detail=planned_cell.role,
                    )
                )
                continue
            expected_kind = planned_cell.kind
            found_kind = str(found["kind"])
            if expected_kind != found_kind:
                findings.append(
                    EstimateAuditFinding(
                        code="CELL_KIND_MISMATCH",
                        sheet=planned_sheet.name,
                        ref=planned_cell.ref,
                        expected=expected_kind,
                        found=found_kind,
                        detail=planned_cell.role,
                    )
                )
                continue
            expected_value = _expected_text(planned_cell)
            found_value = str(found["value"])
            if expected_value != found_value:
                findings.append(
                    EstimateAuditFinding(
                        code="CELL_VALUE_MISMATCH",
                        sheet=planned_sheet.name,
                        ref=planned_cell.ref,
                        expected=expected_value,
                        found=found_value,
                        detail=planned_cell.role,
                    )
                )
            if expected_kind == "formula":
                found_formula = str(found["formula"])
                if planned_cell.formula != found_formula:
                    findings.append(
                        EstimateAuditFinding(
                            code="CELL_FORMULA_MISMATCH",
                            sheet=planned_sheet.name,
                            ref=planned_cell.ref,
                            expected=planned_cell.formula,
                            found=found_formula,
                            detail=planned_cell.role,
                        )
                    )

        for ref in sorted(set(found_cells) - planned_refs):
            findings.append(
                EstimateAuditFinding(
                    code="CELL_UNEXPECTED",
                    sheet=planned_sheet.name,
                    ref=ref,
                    found=str(found_cells[ref]["value"]),
                    detail="célula presente no arquivo e ausente do plano",
                )
            )

    planned_names = {planned_sheet.name for planned_sheet in plan.sheets}
    for sheet_name in sorted(set(_canonical_sheet_names(canonical)) - planned_names):
        findings.append(
            EstimateAuditFinding(
                code="SHEET_UNEXPECTED",
                sheet=sheet_name,
                detail="aba presente no arquivo e ausente do plano",
            )
        )
    return findings


def _report_of(
    workbook_path: Path,
    estimate: Estimate,
    plan: EstimateWorkbookPlan,
    findings: list[EstimateAuditFinding],
) -> EstimateAuditReport:
    """Relatório da auditoria: os dois caminhos publicam a mesma forma de evidência."""
    return EstimateAuditReport(
        status="ok" if not findings else "divergent",
        workbook_sha256=_sha256(workbook_path),
        sheet_name=plan.sheet_name,
        memory_sheet_name=None if plan.memory_sheet is None else plan.memory_sheet.name,
        checked_cells=plan.planned_cells,
        formula_cells=plan.formula_cells,
        total_amount=estimate.total_amount,
        pinned_cells=plan.pinned_cells,
        findings=findings,
    )


def audit_estimate_workbook(
    workbook_path: Path, estimate: Estimate, template: WorkbookTemplate
) -> EstimateAuditReport:
    """Reabre a planilha do orçamento, recomputa as fórmulas e compara centavo a centavo.

    Reusa `canonical.canonicalize_workbook` como biblioteca — a gramática fechada de
    fórmulas (`GRAMMAR_PATTERNS`) é a mesma da medição, e é ela que recomputa cada célula
    de fórmula em `Decimal` a partir do arquivo reaberto. A comparação célula a célula é
    própria: o `Estimate` não tem `Valuation`/`WorksiteBulletin`, então
    `canonical.audit_workbook` (que os exige) não serve aqui sem ser estendido — o que o
    contrato desta tarefa proíbe.
    """
    plan = plan_estimate_workbook(estimate, template)
    findings = _compare_plan(plan, canonicalize_workbook(workbook_path, template))
    return _report_of(workbook_path, estimate, plan, findings)


def audit_estimate_grid_workbook(
    workbook_path: Path, estimate: Estimate, template: WorkbookTemplate
) -> EstimateAuditReport:
    """Reabre o documento da prefeitura e confere as DUAS abas contra o mesmo plano.

    `canonicalize_workbook` ignora só a aba do catálogo, então tanto o gabarito quanto a
    memória entram na canonicalização — e por isso as duas precisam estar no plano, sob
    pena de `SHEET_UNEXPECTED`/`CELL_UNEXPECTED`. Divergência não publica: quem chama
    grava com nome pendente e só renomeia com `status == "ok"`.
    """
    plan = plan_estimate_grid_workbook(estimate, template)
    findings = _compare_plan(plan, canonicalize_workbook(workbook_path, template))
    return _report_of(workbook_path, estimate, plan, findings)
