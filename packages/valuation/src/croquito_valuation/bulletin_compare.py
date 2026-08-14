"""Leitura do BM real do cliente e comparação centavo a centavo com o boletim gerado.

Ferramenta **local** de diagnóstico, da mesma família do `parity`
(`croquito_worker.valuation.parity`): fora da cadeia de medição, fora do CI, não
recusa nem publica nada — **relata**. É a peça do aceite do M5 ("BM da Toca gerado x
real = zero centavo").

Duas decisões de desenho:

- O lado gerado entra pelo `valuation.json` (a fonte de verdade canônica; o xlsx gerado
  já é provado idêntico a ela pela auditoria de round-trip em `canonical.py`). Este
  módulo não lê xlsx gerado.
- O lado real entra por um `.xlsx` do cliente aberto com `data_only=True` — valores em
  cache, como o `parity` faz —, nunca por `canonicalize_workbook`: fórmula fora da
  gramática fechada recusaria a leitura, e o objetivo aqui é ler o que o cliente
  publicou, não validar as fórmulas dele (isso é o `parity`).

Números são convertidos com `canonical_number` (`croquito_valuation.canonical`), a
mesma conversão que a auditoria de round-trip usa: ela arredonda em `ROUND_HALF_UP` e
recusa (`NUMBER_SCALE_UNSUPPORTED`) qualquer célula cujo valor não se resolva em duas
casas — inclusive ruído de ponto flutuante do cache de fórmula. Isso é proposital e é o
oposto de "esperto": o módulo não absorve nem normaliza nada por conta própria (não há
`money_trunc` aqui — dinheiro real não é truncado de novo, é comparado como está
escrito); uma célula que não converte de forma exata em duas casas é recusa fechada, não
palpite.
"""

from __future__ import annotations

from collections.abc import Iterator, Sequence
from decimal import Decimal
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import Field

from croquito_valuation.canonical import canonical_number
from croquito_valuation.catalog import file_sha256
from croquito_valuation.errors import ValuationValidationError
from croquito_valuation.models import ExactDecimal, Valuation, ValuationContractModel
from croquito_valuation.template import BulletinLayout, SheetColumns, WorkbookTemplate


class ReferenceBulletinLine(ValuationContractModel):
    """Uma linha do boletim real, como a planilha do cliente a mostra.

    `*_ref` é a célula A1 de onde o valor foi lido, para o achado do comparador apontar
    direto na planilha real em vez de só citar o código.
    """

    code: str = Field(min_length=1)
    description: str
    unit: str
    quantity: ExactDecimal
    unit_price: ExactDecimal
    total: ExactDecimal
    quantity_ref: str
    unit_price_ref: str
    total_ref: str


class ReferenceBulletin(ValuationContractModel):
    """O que a leitura do BM real extraiu: linhas, total declarado e o que foi pulado.

    `skipped_rows` são as linhas separadoras da aba — puladas, mas nunca em silêncio.
    """

    workbook_sha256: str = Field(pattern=r"^[a-f0-9]{64}$")
    sheet_name: str
    lines: list[ReferenceBulletinLine] = Field(default_factory=list)
    declared_total: ExactDecimal
    declared_total_ref: str
    skipped_rows: list[int] = Field(default_factory=list)


class BulletinLineDiff(ValuationContractModel):
    """Divergência de um valor de linha entre o boletim gerado e o boletim real."""

    code: str
    generated: str
    reference: str
    cell: str | None = None


class BulletinTotalDiff(ValuationContractModel):
    """Divergência do total da obra entre o gerado e o total declarado no arquivo real."""

    generated: str
    reference: str
    cell: str | None = None


class UnitNote(ValuationContractModel):
    """Unidade divergente entre gerado e real com os números batendo — nota, não diff."""

    code: str
    generated: str
    reference: str


class BulletinComparisonReport(ValuationContractModel):
    """Resultado da comparação centavo a centavo do boletim de uma obra.

    `zero_cent` é `True` só quando nada aqui aponta dinheiro diferente: nenhum diff
    numérico (quantidade, preço, total de linha, total da obra) e nenhum código presente
    de um lado e ausente do outro — código ausente de um lado é, por definição, dinheiro
    que os dois arquivos não têm em comum, então ele entra na conta do zero centavo tanto
    quanto um diff explícito. `unit_notes` fica de fora de propósito: é nota, não diff de
    centavo.
    """

    worksite_key: str
    missing_in_reference: list[str] = Field(default_factory=list)
    missing_in_generated: list[str] = Field(default_factory=list)
    quantity_diffs: list[BulletinLineDiff] = Field(default_factory=list)
    unit_price_diffs: list[BulletinLineDiff] = Field(default_factory=list)
    line_total_diffs: list[BulletinLineDiff] = Field(default_factory=list)
    bulletin_total_diff: BulletinTotalDiff | None = None
    unit_notes: list[UnitNote] = Field(default_factory=list)
    zero_cent: bool


def _is_blank(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    return False


def _column_value(cells: Sequence[object], letter: str) -> object:
    index = column_index_from_string(letter) - 1
    if 0 <= index < len(cells):
        return cells[index]
    return None


def _text_at(cells: Sequence[object], letter: str) -> str:
    value = _column_value(cells, letter)
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _normalize_label(text: str) -> str:
    return " ".join(text.split()).casefold()


def _is_total_row(cells: Sequence[object], layout: BulletinLayout) -> bool:
    """A linha de total é a que carrega `total_label` no rótulo ou na coluna de código.

    O gerador do sistema escreve o rótulo na coluna de descrição, mas o BM real observado
    o traz na coluna de rótulo (`label_column`) ou na de código — os dois lugares contam.
    """
    target = _normalize_label(layout.total_label)
    return (
        _normalize_label(_text_at(cells, layout.label_column)) == target
        or _normalize_label(_text_at(cells, layout.columns.code.letter)) == target
    )


def _has_no_numeric_data(cells: Sequence[object], columns: SheetColumns) -> bool:
    return all(
        _is_blank(_column_value(cells, letter))
        for letter in (columns.unit_price.letter, columns.quantity.letter, columns.total.letter)
    )


def _numeric_at(cells: Sequence[object], letter: str, sheet_name: str, row: int) -> Decimal:
    value = _column_value(cells, letter)
    ref = f"{letter}{row}"
    if isinstance(value, bool) or not isinstance(value, int | float | Decimal):
        raise ValuationValidationError(
            "BULLETIN_CELL_NOT_NUMERIC",
            "célula do boletim real não traz um número onde o template espera um",
            {"sheet": sheet_name, "ref": ref, "type": type(value).__name__},
        )
    return canonical_number(value, sheet_name, ref)


def _iter_rows(worksheet: Any, first_row: int) -> Iterator[tuple[int, list[object]]]:
    row_number = first_row - 1
    for row in worksheet.iter_rows(min_row=first_row, values_only=True):
        row_number += 1
        yield row_number, list(row)


def _open_reference_workbook(workbook_path: Path) -> Any:
    try:
        return load_workbook(workbook_path, read_only=True, data_only=True)
    except (OSError, ValueError, KeyError, BadZipFile, InvalidFileException) as error:
        raise ValuationValidationError(
            "BULLETIN_WORKBOOK_UNREADABLE",
            "arquivo do boletim real não pôde ser aberto como planilha",
            {"path": str(workbook_path), "reason": type(error).__name__},
        ) from error


def read_bulletin_lines(
    workbook_path: Path, template: WorkbookTemplate, *, sheet_name: str
) -> ReferenceBulletin:
    """Lê a aba do BM real do cliente, do jeito que ela está escrita.

    A leitura para na linha de total (`BulletinLayout.total_label`) e devolve o total
    declarado ali; linha sem código e sem nenhum valor numérico é separadora, pulada e
    contada em `skipped_rows`. Código repetido na mesma aba recusa
    (`COMPARE_DUPLICATE_CODE`): o aceite decide o que fazer se isso acontecer de verdade.
    """
    layout = template.bulletin
    columns = layout.columns
    workbook_sha256 = file_sha256(workbook_path)
    workbook = _open_reference_workbook(workbook_path)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValuationValidationError(
                "BULLETIN_SHEET_MISSING",
                "planilha real não possui a aba do boletim informada",
                {"sheet": sheet_name, "available": list(workbook.sheetnames)},
            )
        worksheet = workbook[sheet_name]
        lines: list[ReferenceBulletinLine] = []
        skipped_rows: list[int] = []
        seen_codes: set[str] = set()
        declared_total: Decimal | None = None
        declared_total_ref: str | None = None
        first_row = layout.header_row + 1
        for row_number, cells in _iter_rows(worksheet, first_row):
            if _is_total_row(cells, layout):
                declared_total_ref = f"{columns.total.letter}{row_number}"
                declared_total = _numeric_at(cells, columns.total.letter, sheet_name, row_number)
                break
            code = _text_at(cells, columns.code.letter)
            if not code:
                if _has_no_numeric_data(cells, columns):
                    skipped_rows.append(row_number)
                    continue
                raise ValuationValidationError(
                    "BULLETIN_ROW_UNPARSEABLE",
                    "linha do boletim real não tem código e não é separadora reconhecível",
                    {"sheet": sheet_name, "row": row_number},
                )
            if code in seen_codes:
                raise ValuationValidationError(
                    "COMPARE_DUPLICATE_CODE",
                    "código repetido na aba do boletim real",
                    {"sheet": sheet_name, "code": code, "row": row_number},
                )
            seen_codes.add(code)
            lines.append(
                ReferenceBulletinLine(
                    code=code,
                    description=_text_at(cells, columns.description.letter),
                    unit=_text_at(cells, columns.unit.letter),
                    quantity=_numeric_at(cells, columns.quantity.letter, sheet_name, row_number),
                    unit_price=_numeric_at(
                        cells, columns.unit_price.letter, sheet_name, row_number
                    ),
                    total=_numeric_at(cells, columns.total.letter, sheet_name, row_number),
                    quantity_ref=f"{columns.quantity.letter}{row_number}",
                    unit_price_ref=f"{columns.unit_price.letter}{row_number}",
                    total_ref=f"{columns.total.letter}{row_number}",
                )
            )
        else:
            raise ValuationValidationError(
                "BULLETIN_TOTAL_ROW_MISSING",
                "planilha real não declara a linha de total do boletim",
                {"sheet": sheet_name, "total_label": layout.total_label},
            )
    finally:
        workbook.close()
    assert declared_total is not None
    assert declared_total_ref is not None
    return ReferenceBulletin(
        workbook_sha256=workbook_sha256,
        sheet_name=sheet_name,
        lines=lines,
        declared_total=declared_total,
        declared_total_ref=declared_total_ref,
        skipped_rows=skipped_rows,
    )


def compare_bulletin(
    valuation: Valuation, worksite_key: str, reference: ReferenceBulletin
) -> BulletinComparisonReport:
    """Casa o boletim gerado da obra com o boletim real por código, centavo a centavo.

    Sem tolerância: igualdade exata de `Decimal`, ou a divergência vai para a lista da
    classe certa. Nenhum caminho normaliza ou arredonda de novo o que já foi lido.
    """
    bulletin = next(
        (candidate for candidate in valuation.bulletins if candidate.worksite_key == worksite_key),
        None,
    )
    if bulletin is None:
        raise ValuationValidationError(
            "COMPARE_WORKSITE_NOT_FOUND",
            "medição não possui boletim para a obra informada",
            {
                "worksite_key": worksite_key,
                "available": sorted(item.worksite_key for item in valuation.bulletins),
            },
        )
    generated_codes = [line.code for line in bulletin.lines]
    duplicated_generated = sorted(
        {code for code in generated_codes if generated_codes.count(code) > 1}
    )
    if duplicated_generated:
        # O boletim gerado permite o mesmo código em linhas distintas (a unicidade do
        # modelo é por item_number); indexar por código colapsaria uma delas em silêncio.
        raise ValuationValidationError(
            "COMPARE_DUPLICATE_CODE",
            "boletim gerado repete código; a comparação por código exige um por linha",
            {"worksite_key": worksite_key, "codes": duplicated_generated},
        )
    generated_by_code = {line.code: line for line in bulletin.lines}
    reference_by_code = {line.code: line for line in reference.lines}

    missing_in_reference = sorted(set(generated_by_code) - set(reference_by_code))
    missing_in_generated = sorted(set(reference_by_code) - set(generated_by_code))

    quantity_diffs: list[BulletinLineDiff] = []
    unit_price_diffs: list[BulletinLineDiff] = []
    line_total_diffs: list[BulletinLineDiff] = []
    unit_notes: list[UnitNote] = []

    for code in sorted(set(generated_by_code) & set(reference_by_code)):
        generated_line = generated_by_code[code]
        reference_line = reference_by_code[code]
        quantity_equal = generated_line.quantity == reference_line.quantity
        price_equal = generated_line.unit_price == reference_line.unit_price
        total_equal = generated_line.total == reference_line.total
        if not quantity_equal:
            quantity_diffs.append(
                BulletinLineDiff(
                    code=code,
                    generated=str(generated_line.quantity),
                    reference=str(reference_line.quantity),
                    cell=reference_line.quantity_ref,
                )
            )
        if not price_equal:
            unit_price_diffs.append(
                BulletinLineDiff(
                    code=code,
                    generated=str(generated_line.unit_price),
                    reference=str(reference_line.unit_price),
                    cell=reference_line.unit_price_ref,
                )
            )
        if not total_equal:
            line_total_diffs.append(
                BulletinLineDiff(
                    code=code,
                    generated=str(generated_line.total),
                    reference=str(reference_line.total),
                    cell=reference_line.total_ref,
                )
            )
        if (
            generated_line.unit != reference_line.unit
            and quantity_equal
            and price_equal
            and total_equal
        ):
            unit_notes.append(
                UnitNote(code=code, generated=generated_line.unit, reference=reference_line.unit)
            )

    bulletin_total_diff: BulletinTotalDiff | None = None
    if bulletin.total_amount != reference.declared_total:
        bulletin_total_diff = BulletinTotalDiff(
            generated=str(bulletin.total_amount),
            reference=str(reference.declared_total),
            cell=reference.declared_total_ref,
        )

    zero_cent = not (
        quantity_diffs
        or unit_price_diffs
        or line_total_diffs
        or bulletin_total_diff is not None
        or missing_in_reference
        or missing_in_generated
    )

    return BulletinComparisonReport(
        worksite_key=worksite_key,
        missing_in_reference=missing_in_reference,
        missing_in_generated=missing_in_generated,
        quantity_diffs=quantity_diffs,
        unit_price_diffs=unit_price_diffs,
        line_total_diffs=line_total_diffs,
        bulletin_total_diff=bulletin_total_diff,
        unit_notes=unit_notes,
        zero_cent=zero_cent,
    )
