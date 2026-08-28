"""Render da medição em planilha: PLANILHA GERAL, RE-RA e o par BM/MEMÓRIA de cada obra.

O escritor primeiro planeja cada célula (`plan_workbook`) e só então grava. O plano
carrega, para toda célula, o valor exato em `Decimal` que a planilha deve mostrar —
inclusive para as células de fórmula. É esse plano que a auditoria confronta com o
arquivo reaberto, de modo que layout e oráculo nunca divergem.

Regras que valem em toda a aba:

- A pasta é autocontida: descrição, unidade e preço são literais vindos do modelo.
  Nenhum `VLOOKUP` para o catálogo, que pode não acompanhar o arquivo.
- Total de linha é `=TRUNC(QUANT*VALOR UNIT,2)`, exceto quando o produto em ponto
  flutuante divergiria do cálculo exato: aí o valor é gravado literal e a célula entra
  em `pinned_cells`.
- Só existem seis formas de fórmula, as mesmas que o avaliador da auditoria aceita:
  `=TRUNC(<ref>*<ref>,2)`, `=ROUND(PRODUCT(<range>),2)`, a mesma com `-<ref>`,
  `=SUM(<range>)`, `=SUM(<ref>,<ref>[,...])` e `=<ref>-<ref>`.

A PLANILHA GERAL só é planejada quando o consolidado contratual é informado: ela é o
consolidado **depois** desta medição, com os pares históricos copiados literais, o par
corrente calculado, o acumulado somado célula a célula e o saldo como diferença viva.
"""

from __future__ import annotations

import hashlib
import os
import tempfile
from collections.abc import Sequence
from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path
from typing import Final, Literal

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string, get_column_letter
from pydantic import Field, model_validator

from croquito_valuation.contract import AmendmentLine, ContractLine, ContractWorkbook
from croquito_valuation.errors import ValuationValidationError
from croquito_valuation.models import (
    CalcBlock,
    ExactDecimal,
    PriceCatalog,
    PriceOrigin,
    Valuation,
    ValuationContractModel,
    WorksiteBulletin,
)
from croquito_valuation.rounding import money_trunc, trunc_divergence
from croquito_valuation.template import (
    AmendmentColumns,
    AmendmentLayout,
    GeneralLayout,
    WorkbookTemplate,
)

MAX_DECIMAL_PLACES: Final = 2
PINNED_REASON: Final = "TRUNC_DOUBLE_DIVERGENCE"
_FIXED_TIMESTAMP: Final = datetime(2026, 1, 1, tzinfo=UTC)

CellKind = Literal["text", "number", "formula"]


class PinnedCell(ValuationContractModel):
    """Célula cujo valor foi fixado por divergência de truncamento."""

    sheet: str
    ref: str
    reason: Literal["TRUNC_DOUBLE_DIVERGENCE"] = PINNED_REASON
    item_number: str
    quantity: ExactDecimal
    unit_price: ExactDecimal
    value: ExactDecimal


class PlannedCell(ValuationContractModel):
    """Uma célula planejada, com o valor exato que ela deve representar."""

    ref: str = Field(pattern=r"^[A-Z]{1,2}\d+$")
    kind: CellKind
    role: str = Field(min_length=1, max_length=60)
    item_number: str | None = None
    text: str | None = None
    number: ExactDecimal | None = None
    formula: str | None = None
    number_format: str | None = None
    bold: bool = False

    @model_validator(mode="after")
    def validate_payload(self) -> PlannedCell:
        if self.kind == "text" and self.text is None:
            raise ValuationValidationError(
                "PLANNED_CELL_INVALID", "célula de texto sem texto", {"ref": self.ref}
            )
        if self.kind == "number" and self.number is None:
            raise ValuationValidationError(
                "PLANNED_CELL_INVALID", "célula numérica sem valor", {"ref": self.ref}
            )
        if self.kind == "formula" and (self.formula is None or self.number is None):
            raise ValuationValidationError(
                "PLANNED_CELL_INVALID",
                "célula de fórmula precisa de fórmula e do valor exato esperado",
                {"ref": self.ref},
            )
        return self


class PlannedSheet(ValuationContractModel):
    """Uma aba planejada."""

    name: str = Field(min_length=1, max_length=31)
    kind: Literal["bulletin", "memory", "general", "amendment"]
    cells: list[PlannedCell] = Field(min_length=1)
    column_widths: dict[str, int] = Field(default_factory=dict)


class PlannedWorksite(ValuationContractModel):
    """O par de abas que a medição gera para uma obra."""

    worksite_key: str
    bulletin_sheet: str
    memory_sheet: str


class WorkbookPlan(ValuationContractModel):
    """Plano completo da pasta, usado para gravar e para auditar."""

    worksites: list[PlannedWorksite] = Field(min_length=1)
    general_sheet: str | None = None
    amendment_sheet: str | None = None
    sheets: list[PlannedSheet] = Field(min_length=1)
    pinned_cells: list[PinnedCell] = Field(default_factory=list)

    @property
    def formula_cells(self) -> int:
        """Quantas células de fórmula o plano prevê."""
        return sum(1 for sheet in self.sheets for cell in sheet.cells if cell.kind == "formula")


class WriteReport(ValuationContractModel):
    """Resultado da gravação da pasta."""

    output_path: str
    workbook_sha256: str = Field(pattern=r"^[a-f0-9]{64}$")
    worksites: list[PlannedWorksite] = Field(min_length=1)
    general_sheet: str | None = None
    amendment_sheet: str | None = None
    written_cells: int = Field(ge=1)
    formula_cells: int = Field(ge=0)
    pinned_cells: list[PinnedCell] = Field(default_factory=list)


def _checked_number(value: Decimal, ref: str, role: str) -> Decimal:
    exponent = value.as_tuple().exponent
    if isinstance(exponent, int) and exponent < -MAX_DECIMAL_PLACES:
        raise ValuationValidationError(
            "DECIMAL_SCALE_UNSUPPORTED",
            "planilha só representa duas casas decimais sem perder o valor exato",
            {"ref": ref, "role": role, "value": str(value)},
        )
    return value


def _text(
    ref: str,
    role: str,
    value: str,
    *,
    bold: bool = False,
    item_number: str | None = None,
) -> PlannedCell:
    return PlannedCell(
        ref=ref, kind="text", role=role, text=value, bold=bold, item_number=item_number
    )


def _number(
    ref: str,
    role: str,
    value: Decimal,
    *,
    number_format: str | None = None,
    bold: bool = False,
    item_number: str | None = None,
) -> PlannedCell:
    return PlannedCell(
        ref=ref,
        kind="number",
        role=role,
        item_number=item_number,
        number=_checked_number(value, ref, role),
        number_format=number_format,
        bold=bold,
    )


def _formula(
    ref: str,
    role: str,
    formula: str,
    value: Decimal,
    *,
    number_format: str | None = None,
    bold: bool = False,
    item_number: str | None = None,
) -> PlannedCell:
    return PlannedCell(
        ref=ref,
        kind="formula",
        role=role,
        item_number=item_number,
        formula=formula,
        number=_checked_number(value, ref, role),
        number_format=number_format,
        bold=bold,
    )


def _validate_against_catalog(bulletin: WorksiteBulletin, catalog: PriceCatalog) -> None:
    """A pasta é autocontida, mas o preço impresso continua sendo o do catálogo.

    Segunda linha de defesa do guardrail da licitada (`calc.py::build_worksite_bulletin`
    já recusa antes): o catálogo que o escritor recebe na hora de gravar a pasta também
    precisa ser o SCO — preço de EMOP ou composição nunca chega à planilha da medição.
    """
    if catalog.origin != PriceOrigin.SCO:
        raise ValuationValidationError(
            "BULLETIN_PRICE_ORIGIN_FORBIDDEN",
            "medição de obra licitada não aceita catálogo cuja origem não seja o SCO; "
            "item fora do contrato vira dossiê de aditivo, nunca preço de outra tabela",
            {"origin": catalog.origin.value},
        )
    for line in bulletin.lines:
        if not catalog.has_code(line.code):
            raise ValuationValidationError(
                "LINE_CODE_NOT_IN_CATALOG",
                "linha do boletim usa código ausente do catálogo importado",
                {"item_number": line.item_number, "code": line.code},
            )
        entry = catalog.entry_for(line.code)
        if entry.unit_price != line.unit_price:
            raise ValuationValidationError(
                "LINE_PRICE_NOT_IN_CATALOG",
                "preço da linha diverge do preço do catálogo importado",
                {
                    "item_number": line.item_number,
                    "code": line.code,
                    "catalog": str(entry.unit_price),
                    "line": str(line.unit_price),
                },
            )
        if entry.unit != line.unit:
            raise ValuationValidationError(
                "LINE_UNIT_NOT_IN_CATALOG",
                "unidade da linha diverge da unidade do catálogo importado",
                {
                    "item_number": line.item_number,
                    "code": line.code,
                    "catalog": entry.unit,
                    "line": line.unit,
                },
            )


def _total_cell(
    *,
    sheet_name: str,
    ref: str,
    role: str,
    item_number: str,
    quantity: Decimal,
    unit_price: Decimal,
    quantity_ref: str,
    unit_price_ref: str,
    number_format: str,
    pinned: list[PinnedCell],
) -> PlannedCell:
    """Total monetário: fórmula viva quando ela reproduz o exato, literal quando não."""
    exact = money_trunc(quantity * unit_price)
    if trunc_divergence(quantity, unit_price):
        pinned.append(
            PinnedCell(
                sheet=sheet_name,
                ref=ref,
                item_number=item_number,
                quantity=quantity,
                unit_price=unit_price,
                value=exact,
            )
        )
        return _number(
            ref,
            f"{role}_pinned",
            exact,
            number_format=number_format,
            item_number=item_number,
        )
    return _formula(
        ref,
        role,
        f"=TRUNC({quantity_ref}*{unit_price_ref},2)",
        exact,
        number_format=number_format,
        item_number=item_number,
    )


def _plan_bulletin(
    valuation: Valuation,
    bulletin: WorksiteBulletin,
    template: WorkbookTemplate,
    pinned: list[PinnedCell],
) -> PlannedSheet:
    layout = template.bulletin
    columns = layout.columns
    sheet_name = template.bulletin_sheet_name(bulletin.worksite_name)
    cells: list[PlannedCell] = [_text(f"{layout.label_column}1", "title", layout.title, bold=True)]

    pairs: list[tuple[str, str]] = [(layout.intervention_label, bulletin.worksite_name)]
    if bulletin.address is not None:
        pairs.append((layout.address_label, bulletin.address))
    if bulletin.contract_label is not None:
        pairs.append((layout.contract_label, bulletin.contract_label))
    pairs.append(
        (layout.period_label, f"Nº {valuation.period_number} — {valuation.reference_label}")
    )
    if layout.header_row < len(pairs) + 3:
        raise ValuationValidationError(
            "TEMPLATE_HEADER_ROW_TOO_SMALL",
            "linha de cabeçalho do BM não deixa espaço para o bloco de identificação",
            {"header_row": layout.header_row, "required_rows": len(pairs) + 3},
        )
    for index, (label, value) in enumerate(pairs):
        row = 2 + index
        cells.append(_text(f"{layout.label_column}{row}", "header_label", label, bold=True))
        cells.append(_text(f"{layout.value_column}{row}", "header_value", value))

    for column in columns.ordered:
        cells.append(
            _text(f"{column.letter}{layout.header_row}", "column_label", column.label, bold=True)
        )

    first_line_row = layout.header_row + 1
    for index, line in enumerate(bulletin.lines):
        row = first_line_row + index
        item = line.item_number
        cells.append(_text(f"{columns.item.letter}{row}", "line_item", item, item_number=item))
        cells.append(_text(f"{columns.code.letter}{row}", "line_code", line.code, item_number=item))
        cells.append(
            _text(
                f"{columns.description.letter}{row}",
                "line_description",
                line.description,
                item_number=item,
            )
        )
        cells.append(_text(f"{columns.unit.letter}{row}", "line_unit", line.unit, item_number=item))
        cells.append(
            _number(
                f"{columns.unit_price.letter}{row}",
                "line_unit_price",
                line.unit_price,
                number_format=layout.money_number_format,
                item_number=item,
            )
        )
        cells.append(
            _number(
                f"{columns.quantity.letter}{row}",
                "line_quantity",
                line.quantity,
                number_format=layout.quantity_number_format,
                item_number=item,
            )
        )
        cells.append(
            _total_cell(
                sheet_name=sheet_name,
                ref=f"{columns.total.letter}{row}",
                role="line_total",
                item_number=line.item_number,
                quantity=line.quantity,
                unit_price=line.unit_price,
                quantity_ref=f"{columns.quantity.letter}{row}",
                unit_price_ref=f"{columns.unit_price.letter}{row}",
                number_format=layout.money_number_format,
                pinned=pinned,
            )
        )

    last_line_row = first_line_row + len(bulletin.lines) - 1
    total_row = last_line_row + 1
    cells.append(
        _text(
            f"{columns.description.letter}{total_row}",
            "total_label",
            layout.total_label,
            bold=True,
        )
    )
    cells.append(
        _formula(
            f"{columns.total.letter}{total_row}",
            "bulletin_total",
            f"=SUM({columns.total.letter}{first_line_row}:{columns.total.letter}{last_line_row})",
            bulletin.total_amount,
            number_format=layout.money_number_format,
            bold=True,
        )
    )
    return PlannedSheet(
        name=sheet_name,
        kind="bulletin",
        cells=cells,
        column_widths={column.letter: column.width for column in columns.ordered},
    )


def plan_calc_block(
    block: CalcBlock,
    template: WorkbookTemplate,
    label_row: int,
) -> tuple[list[PlannedCell], int]:
    """Planeja rótulo, cabeçalho de operandos e valores de um bloco de cálculo.

    Público (F-043) porque é o ÚNICO render de bloco de memória do repositório: ele
    consome `CalcBlock` e `template.memory` e não sabe nada de `Valuation` nem de
    `WorksiteBulletin`, então a memória do orçamento-base o reusa em vez de nascer com um
    segundo render que divergiria no primeiro ajuste. Devolve as células do bloco e a
    linha dos valores, para quem empilha blocos saber onde continuar.

    `_plan_memory` (a memória da MEDIÇÃO) continua privada e específica: ela exige
    `Valuation`/`WorksiteBulletin` e generalizá-la seria outra coisa.
    """
    layout = template.memory
    slots = list(layout.operand_columns)
    if len(block.deductions) > 1:
        raise ValuationValidationError(
            "MEMORY_DEDUCTIONS_UNSUPPORTED",
            "a memória do M1 imprime no máximo uma dedução por bloco",
            {"label": block.label, "deductions": len(block.deductions)},
        )
    needed = len(block.operands) + len(block.deductions)
    if needed > len(slots):
        raise ValuationValidationError(
            "MEMORY_BLOCK_TOO_WIDE",
            "bloco de cálculo não cabe nas colunas de operando do template",
            {"label": block.label, "needed": needed, "available": len(slots)},
        )
    header_row = label_row + 1
    value_row = label_row + 2
    cells: list[PlannedCell] = [
        _text(f"{layout.block_label_column}{label_row}", "block_label", block.label, bold=True)
    ]
    for index, operand in enumerate(block.operands):
        column = slots[index]
        cells.append(_text(f"{column}{header_row}", "operand_label", operand.name, bold=True))
        cells.append(
            _number(
                f"{column}{value_row}",
                "operand_value",
                operand.value,
                number_format=layout.quantity_number_format,
            )
        )
    first_operand_column = slots[0]
    last_operand_column = slots[len(block.operands) - 1]
    product_range = f"{first_operand_column}{value_row}:{last_operand_column}{value_row}"
    formula = f"=ROUND(PRODUCT({product_range}),2)"
    if block.deductions:
        deduction = block.deductions[0]
        deduction_column = slots[len(block.operands)]
        cells.append(
            _text(
                f"{deduction_column}{header_row}",
                "deduction_label",
                layout.deduction_label,
                bold=True,
            )
        )
        cells.append(
            _number(
                f"{deduction_column}{value_row}",
                "deduction_value",
                deduction.value,
                number_format=layout.quantity_number_format,
            )
        )
        formula = f"{formula}-{deduction_column}{value_row}"
    cells.append(
        _text(
            f"{layout.subtotal_column}{header_row}",
            "subtotal_label",
            layout.subtotal_label,
            bold=True,
        )
    )
    cells.append(
        _formula(
            f"{layout.subtotal_column}{value_row}",
            "block_subtotal",
            formula,
            block.subtotal,
            number_format=layout.quantity_number_format,
        )
    )
    return cells, value_row


def _plan_memory(
    valuation: Valuation,
    bulletin: WorksiteBulletin,
    template: WorkbookTemplate,
    pinned: list[PinnedCell],
) -> PlannedSheet:
    layout = template.memory
    columns = layout.columns
    sheet_name = template.memory_sheet_name(bulletin.worksite_name)
    cells: list[PlannedCell] = [
        _text(f"{layout.label_column}1", "title", layout.title, bold=True),
        _text(f"{layout.label_column}2", "header_label", layout.intervention_label, bold=True),
        _text(f"{layout.value_column}2", "header_value", bulletin.worksite_name),
    ]
    if layout.header_row < 3:
        raise ValuationValidationError(
            "TEMPLATE_HEADER_ROW_TOO_SMALL",
            "linha de cabeçalho da memória não deixa espaço para o título",
            {"header_row": layout.header_row, "required_rows": 3},
        )
    for column in columns.ordered:
        cells.append(
            _text(f"{column.letter}{layout.header_row}", "column_label", column.label, bold=True)
        )

    row = layout.header_row + 1
    for line in bulletin.lines:
        sheet = valuation.calc_sheet_for(bulletin.worksite_key, line.item_number)
        summary_row = row
        item = line.item_number
        cells.append(
            _text(f"{columns.item.letter}{summary_row}", "line_item", item, item_number=item)
        )
        cells.append(
            _text(f"{columns.code.letter}{summary_row}", "line_code", line.code, item_number=item)
        )
        cells.append(
            _text(
                f"{columns.description.letter}{summary_row}",
                "line_description",
                line.description,
                item_number=item,
            )
        )
        cells.append(
            _text(f"{columns.unit.letter}{summary_row}", "line_unit", line.unit, item_number=item)
        )
        cells.append(
            _number(
                f"{columns.unit_price.letter}{summary_row}",
                "line_unit_price",
                line.unit_price,
                number_format=layout.money_number_format,
                item_number=item,
            )
        )
        cells.append(
            _number(
                f"{columns.quantity.letter}{summary_row}",
                "line_quantity",
                line.quantity,
                number_format=layout.quantity_number_format,
                item_number=item,
            )
        )
        cells.append(
            _total_cell(
                sheet_name=sheet_name,
                ref=f"{columns.total.letter}{summary_row}",
                role="line_total",
                item_number=line.item_number,
                quantity=line.quantity,
                unit_price=line.unit_price,
                quantity_ref=f"{columns.quantity.letter}{summary_row}",
                unit_price_ref=f"{columns.unit_price.letter}{summary_row}",
                number_format=layout.money_number_format,
                pinned=pinned,
            )
        )
        row = summary_row + 1
        subtotal_rows: list[int] = []
        for block in sheet.blocks:
            block_cells, value_row = plan_calc_block(block, template, row)
            cells.extend(block_cells)
            subtotal_rows.append(value_row)
            row = value_row + 1
        cells.append(
            _text(
                f"{layout.block_label_column}{row}",
                "sheet_total_label",
                layout.total_label,
                bold=True,
            )
        )
        cells.append(
            _formula(
                f"{layout.subtotal_column}{row}",
                "sheet_total",
                f"=SUM({layout.subtotal_column}{subtotal_rows[0]}"
                f":{layout.subtotal_column}{subtotal_rows[-1]})",
                sheet.total_quantity,
                number_format=layout.quantity_number_format,
                bold=True,
            )
        )
        row += 2
    return PlannedSheet(
        name=sheet_name,
        kind="memory",
        cells=cells,
        column_widths={column.letter: column.width for column in columns.ordered},
    )


def _pair_columns(layout: GeneralLayout, index: int) -> tuple[str, str]:
    """Letras do par QUANTIDADE|VALOR na posição `index`, contada do primeiro período.

    É a mesma derivação que o leitor faz do cabeçalho: nenhuma coluna de par é constante.
    """
    first = column_index_from_string(layout.first_period_column)
    quantity = first + 2 * index
    return get_column_letter(quantity), get_column_letter(quantity + 1)


def _sum_of_refs(refs: Sequence[str]) -> str:
    """Forma 5 da gramática; com uma célula só o `SUM` volta a ser o de intervalo.

    O acumulado soma células alternadas (QUANTIDADE e VALOR se intercalam), então o
    intervalo contíguo não serve. Na primeira medição do contrato existe um par só, e a
    lista de um item está fora da gramática — aí o intervalo degenerado é a forma válida.
    """
    if len(refs) == 1:
        return f"=SUM({refs[0]}:{refs[0]})"
    return f"=SUM({','.join(refs)})"


def _measured_by_code(valuation: Valuation) -> dict[str, Decimal]:
    """Quanto cada código somou nesta medição, atravessando todas as obras."""
    measured: dict[str, Decimal] = {}
    for bulletin in valuation.bulletins:
        for line in bulletin.lines:
            measured[line.code] = measured.get(line.code, Decimal("0.00")) + line.quantity
    return measured


def _check_measured_codes(contract: ContractWorkbook, measured: dict[str, Decimal]) -> None:
    """Código medido fora do consolidado não teria linha na GERAL para receber o valor."""
    unknown = sorted(code for code in measured if not _has_code(contract, code))
    if unknown:
        raise ValuationValidationError(
            "GENERAL_CONSOLIDATION_MISMATCH",
            "a medição lança código que não existe no consolidado contratual e a planilha "
            "geral não teria onde lançá-lo",
            {
                "reason": "CODE_NOT_IN_CONTRACT",
                "codes": unknown,
                "contract": contract.source_label,
            },
        )


def _has_code(contract: ContractWorkbook, code: str) -> bool:
    return any(line.code == code for line in contract.lines)


def _check_consolidated_total(
    valuation: Valuation, contract: ContractWorkbook, measured: dict[str, Decimal]
) -> None:
    """O valor corrente da GERAL é por código; o da medição é por linha de boletim.

    Quando o mesmo código é medido em mais de uma obra, `TRUNC(Σq x preço)` pode ficar um
    centavo acima de `Σ TRUNC(q_i x preço)`. Não existe ainda decisão do orçamentista
    sobre qual dos dois é o valor consolidado correto, então a pasta não é gerada: fixar
    um dos dois em silêncio publicaria um total que a medição não declara.
    """
    divergences: list[dict[str, object]] = []
    general_total = Decimal("0.00")
    for line in contract.lines:
        quantity = measured.get(line.code)
        if quantity is None:
            continue
        general_amount = money_trunc(quantity * line.unit_price)
        general_total += general_amount
        bulletin_amount = sum(
            (
                bulletin_line.total
                for bulletin in valuation.bulletins
                for bulletin_line in bulletin.lines
                if bulletin_line.code == line.code
            ),
            Decimal("0.00"),
        )
        if general_amount != bulletin_amount:
            divergences.append(
                {
                    "code": line.code,
                    "quantity": str(quantity),
                    "general": str(general_amount),
                    "bulletins": str(bulletin_amount),
                }
            )
    if general_total != valuation.total_amount:
        raise ValuationValidationError(
            "GENERAL_CONSOLIDATION_MISMATCH",
            "o valor consolidado da planilha geral não fecha com o total da medição: o "
            "mesmo código medido em mais de uma obra trunca de forma diferente somado "
            "(TRUNC da soma) e linha a linha (soma dos TRUNC)",
            {
                "reason": "TRUNC_CONSOLIDATION_DRIFT",
                "general_total": str(general_total),
                "valuation_total": str(valuation.total_amount),
                "codes": divergences,
            },
        )


def _general_header_cells(
    layout: GeneralLayout,
    amended_column: str,
    period_numbers: Sequence[int],
    pairs: Sequence[tuple[str, str]],
    accumulated: tuple[str, str],
    balance_column: str,
) -> list[PlannedCell]:
    """Cabeçalho da GERAL nas posições exatas que o leitor do consolidado valida."""
    row = layout.header_row
    cells = [
        _text(f"{letter}{row}", "general_column_label", label, bold=True)
        for letter, label in (
            (layout.group_column, layout.group_label),
            (layout.item_column, layout.item_label),
            (layout.code_column, layout.code_label),
            (layout.description_column, layout.description_label),
            (layout.unit_column, layout.unit_label),
            (layout.contract_quantity_column, layout.contract_quantity_label),
            (layout.unit_price_column, layout.unit_price_label),
            (amended_column, layout.amended_quantity_label),
        )
    ]
    for number, (quantity_column, _) in zip(period_numbers, pairs, strict=True):
        cells.append(
            _text(
                f"{quantity_column}{row}",
                "general_period_label",
                layout.period_label(number),
                bold=True,
            )
        )
    cells.append(
        _text(
            f"{accumulated[0]}{row}",
            "general_accumulated_label",
            layout.accumulated_label,
            bold=True,
        )
    )
    cells.append(
        _text(f"{balance_column}{row}", "general_balance_label", layout.balance_label, bold=True)
    )
    sublabel_row = layout.pair_sublabel_row
    if sublabel_row is not None:
        for quantity_column, amount_column in (*pairs, accumulated):
            cells.append(
                _text(
                    f"{quantity_column}{sublabel_row}",
                    "general_pair_sublabel",
                    layout.quantity_pair_label,
                    bold=True,
                )
            )
            cells.append(
                _text(
                    f"{amount_column}{sublabel_row}",
                    "general_pair_sublabel",
                    layout.amount_pair_label,
                    bold=True,
                )
            )
    return cells


def _general_line_cells(
    *,
    layout: GeneralLayout,
    amended_column: str,
    sheet_name: str,
    line: ContractLine,
    row: int,
    current_quantity: Decimal,
    contract_current_quantity: Decimal,
    period_numbers: Sequence[int],
    pairs: Sequence[tuple[str, str]],
    accumulated: tuple[str, str],
    balance_column: str,
    pinned: list[PinnedCell],
) -> list[PlannedCell]:
    """Uma linha de item da GERAL: histórico literal, corrente calculado, saldo vivo."""
    item = line.item_number
    money = layout.money_number_format
    quantity_format = layout.quantity_number_format
    cells: list[PlannedCell] = [
        _text(f"{layout.item_column}{row}", "general_item", item, item_number=item),
        _text(f"{layout.code_column}{row}", "general_code", line.code, item_number=item),
        _text(
            f"{layout.description_column}{row}",
            "general_description",
            line.description,
            item_number=item,
        ),
        _text(f"{layout.unit_column}{row}", "general_unit", line.unit, item_number=item),
        _number(
            f"{layout.contract_quantity_column}{row}",
            "general_contract_quantity",
            line.contract_quantity,
            number_format=quantity_format,
            item_number=item,
        ),
        _number(
            f"{layout.unit_price_column}{row}",
            "general_unit_price",
            line.unit_price,
            number_format=money,
            item_number=item,
        ),
        _number(
            f"{amended_column}{row}",
            "general_amended_quantity",
            # Vigente DERIVADO (ADR-0056, decisão 3): contratado + RE-RA. Para o MAPÃO histórico
            # é bit a bit igual ao `amended_quantity` que a planilha declarava; para o
            # consolidado nascido do orçamento, é o número que não tem segundo dono.
            contract_current_quantity,
            number_format=quantity_format,
            item_number=item,
        ),
    ]
    periods = {period.period_number: period for period in line.periods}
    quantity_refs: list[str] = []
    amount_refs: list[str] = []
    accumulated_quantity = Decimal("0.00")
    accumulated_amount = Decimal("0.00")
    for number, (quantity_column, amount_column) in zip(period_numbers, pairs[:-1], strict=True):
        period = periods.get(number)
        quantity = Decimal("0.00") if period is None else period.quantity
        amount = Decimal("0.00") if period is None else period.amount
        quantity_refs.append(f"{quantity_column}{row}")
        amount_refs.append(f"{amount_column}{row}")
        accumulated_quantity += quantity
        accumulated_amount += amount
        cells.append(
            _number(
                f"{quantity_column}{row}",
                "general_period_quantity",
                quantity,
                number_format=quantity_format,
                item_number=item,
            )
        )
        cells.append(
            _number(
                f"{amount_column}{row}",
                "general_period_amount",
                amount,
                number_format=money,
                item_number=item,
            )
        )

    current_quantity_column, current_amount_column = pairs[-1]
    current_quantity_ref = f"{current_quantity_column}{row}"
    current_amount_ref = f"{current_amount_column}{row}"
    quantity_refs.append(current_quantity_ref)
    amount_refs.append(current_amount_ref)
    accumulated_quantity += current_quantity
    accumulated_amount += money_trunc(current_quantity * line.unit_price)
    cells.append(
        _number(
            current_quantity_ref,
            "general_current_quantity",
            current_quantity,
            number_format=quantity_format,
            item_number=item,
        )
    )
    cells.append(
        _total_cell(
            sheet_name=sheet_name,
            ref=current_amount_ref,
            role="general_current_amount",
            item_number=item,
            quantity=current_quantity,
            unit_price=line.unit_price,
            quantity_ref=current_quantity_ref,
            unit_price_ref=f"{layout.unit_price_column}{row}",
            number_format=money,
            pinned=pinned,
        )
    )
    accumulated_quantity_ref = f"{accumulated[0]}{row}"
    cells.append(
        _formula(
            accumulated_quantity_ref,
            "general_accumulated_quantity",
            _sum_of_refs(quantity_refs),
            accumulated_quantity,
            number_format=quantity_format,
            item_number=item,
        )
    )
    cells.append(
        _formula(
            f"{accumulated[1]}{row}",
            "general_accumulated_amount",
            _sum_of_refs(amount_refs),
            accumulated_amount,
            number_format=money,
            item_number=item,
        )
    )
    cells.append(
        _formula(
            f"{balance_column}{row}",
            "general_balance",
            f"={amended_column}{row}-{accumulated_quantity_ref}",
            contract_current_quantity - accumulated_quantity,
            number_format=quantity_format,
            item_number=item,
        )
    )
    return cells


def _plan_general(
    valuation: Valuation,
    contract: ContractWorkbook,
    template: WorkbookTemplate,
    pinned: list[PinnedCell],
) -> PlannedSheet:
    """Planeja a PLANILHA GERAL como ela fica **depois** desta medição.

    Os pares já lançados são cópia literal do que o consolidado importado declarava, com
    o número real de cada medição no rótulo; o par novo é esta medição; acumulado e saldo
    são fórmulas vivas sobre as duas coisas.
    """
    layout = template.general
    amended_column = layout.amended_quantity_column
    if amended_column is None:
        raise ValuationValidationError(
            "PLAN_GENERAL_NOT_REPRESENTABLE",
            "o template não declara a coluna de quantidade vigente e a planilha geral "
            "gerada não teria onde escrevê-la",
            {"sheet": layout.sheet_name, "reason": "AMENDED_QUANTITY_COLUMN_MISSING"},
        )
    measured = _measured_by_code(valuation)
    _check_measured_codes(contract, measured)
    _check_consolidated_total(valuation, contract, measured)

    period_numbers = [*contract.period_numbers, contract.next_period_number]
    pairs = [_pair_columns(layout, index) for index in range(len(period_numbers))]
    accumulated = _pair_columns(layout, len(period_numbers))
    balance_column = get_column_letter(column_index_from_string(accumulated[1]) + 1)
    current_amount_column = pairs[-1][1]

    cells: list[PlannedCell] = [_text(f"{layout.group_column}1", "title", layout.title, bold=True)]
    cells.extend(
        _general_header_cells(
            layout, amended_column, period_numbers, pairs, accumulated, balance_column
        )
    )

    row = layout.data_first_row
    group_label = ""
    total = Decimal("0.00")
    for line in contract.lines:
        if line.group_label != group_label:
            cells.append(
                _text(f"{layout.group_column}{row}", "general_group", line.group_label, bold=True)
            )
            group_label = line.group_label
            row += 1
        current_quantity = measured.get(line.code, Decimal("0.00"))
        total += money_trunc(current_quantity * line.unit_price)
        cells.extend(
            _general_line_cells(
                layout=layout,
                amended_column=amended_column,
                sheet_name=layout.sheet_name,
                line=line,
                row=row,
                current_quantity=current_quantity,
                contract_current_quantity=contract.current_quantity(line),
                period_numbers=contract.period_numbers,
                pairs=pairs,
                accumulated=accumulated,
                balance_column=balance_column,
                pinned=pinned,
            )
        )
        row += 1

    last_data_row = row - 1
    cells.append(
        _text(f"{layout.group_column}{row}", "general_total_label", layout.total_label, bold=True)
    )
    cells.append(
        _formula(
            f"{current_amount_column}{row}",
            "general_total",
            f"=SUM({current_amount_column}{layout.data_first_row}"
            f":{current_amount_column}{last_data_row})",
            total,
            number_format=layout.money_number_format,
            bold=True,
        )
    )
    return PlannedSheet(name=layout.sheet_name, kind="general", cells=cells)


def _block_anchor(block: AmendmentColumns) -> str:
    """Primeira coluna declarada do bloco; é onde o rótulo da RE-RA é ancorado."""
    for column in (block.reduced_column, block.added_column, block.new_item_column):
        if column is not None:
            return column
    raise ValuationValidationError(  # pragma: no cover - impedido por AmendmentColumns
        "TEMPLATE_AMENDMENT_BLOCK_EMPTY",
        "bloco de RE-RA sem nenhuma coluna declarada",
        {"label": block.label},
    )


def _amendment_delta_cell(
    *,
    block: AmendmentColumns,
    amendment_line: AmendmentLine,
    item_number: str,
    row: int,
    number_format: str,
) -> PlannedCell:
    """Efeito da RE-RA sobre o código, na coluna que o bloco declara para aquele efeito."""
    delta = amendment_line.quantity_delta
    if amendment_line.is_new_item:
        column, role, value = block.new_item_column, "amendment_new_item", delta
    elif delta < 0:
        column, role, value = block.reduced_column, "amendment_reduced", -delta
    else:
        column, role, value = block.added_column, "amendment_added", delta
    if column is None:
        raise ValuationValidationError(
            "PLAN_AMENDMENT_NOT_REPRESENTABLE",
            "o template não declara a coluna que a RE-RA importada usaria para este código",
            {
                "label": block.label,
                "code": amendment_line.code,
                "effect": role,
                "delta": str(delta),
            },
        )
    return _number(
        f"{column}{row}", role, value, number_format=number_format, item_number=item_number
    )


def _amendment_blocks(
    layout: AmendmentLayout, contract: ContractWorkbook
) -> dict[str, AmendmentColumns]:
    """Casa cada RE-RA importada com o bloco de colunas de mesmo rótulo no template."""
    blocks = {block.label: block for block in layout.blocks}
    missing = sorted({amendment.label for amendment in contract.amendments} - set(blocks))
    if missing:
        raise ValuationValidationError(
            "PLAN_AMENDMENT_NOT_REPRESENTABLE",
            "o template não declara o bloco de colunas de uma RE-RA do consolidado",
            {"labels": missing, "template": sorted(blocks)},
        )
    return blocks


def _plan_amendment_sheet(
    contract: ContractWorkbook, template: WorkbookTemplate
) -> PlannedSheet | None:
    """Carrega adiante a aba de RE-RA: é leitura preservada, não revisão criada aqui.

    Sem bloco declarado no template a aba não é escrita e nada é inventado: o efeito das
    RE-RA continua íntegro no consolidado JSON que acompanha o pacote.
    """
    layout = template.amendment
    if layout is None or not contract.amendments:
        return None
    blocks = _amendment_blocks(layout, contract)
    number_format = template.general.quantity_number_format
    changed = {
        amendment.label: {line.code: line for line in amendment.lines}
        for amendment in contract.amendments
    }
    cells: list[PlannedCell] = [
        _text(
            f"{layout.code_column}{layout.header_row}",
            "amendment_column_label",
            layout.code_label,
            bold=True,
        )
    ]
    if layout.amended_quantity_column is not None:
        cells.append(
            _text(
                f"{layout.amended_quantity_column}{layout.header_row}",
                "amendment_column_label",
                layout.amended_quantity_label,
                bold=True,
            )
        )
    for amendment in contract.amendments:
        cells.append(
            _text(
                f"{_block_anchor(blocks[amendment.label])}{layout.header_row}",
                "amendment_block_label",
                amendment.label,
                bold=True,
            )
        )

    row = layout.data_first_row
    for line in contract.lines:
        effects = [
            (label, by_code[line.code])
            for label, by_code in changed.items()
            if line.code in by_code
        ]
        if not effects:
            continue
        cells.append(
            _text(
                f"{layout.code_column}{row}",
                "amendment_code",
                line.code,
                item_number=line.item_number,
            )
        )
        if layout.amended_quantity_column is not None:
            cells.append(
                _number(
                    f"{layout.amended_quantity_column}{row}",
                    "amendment_amended_quantity",
                    # Vigente derivado (ADR-0056, decisão 3), como na GERAL.
                    contract.current_quantity(line),
                    number_format=number_format,
                    item_number=line.item_number,
                )
            )
        for label, amendment_line in effects:
            cells.append(
                _amendment_delta_cell(
                    block=blocks[label],
                    amendment_line=amendment_line,
                    item_number=line.item_number,
                    row=row,
                    number_format=number_format,
                )
            )
        row += 1
    return PlannedSheet(name=layout.sheet_name, kind="amendment", cells=cells)


def _check_sheet_names(sheets: Sequence[PlannedSheet]) -> None:
    """Duas obras com o mesmo nome — ou uma obra com o nome da GERAL — sobrescreveriam."""
    names = [sheet.name for sheet in sheets]
    duplicated = sorted({name for name in names if names.count(name) > 1})
    if duplicated:
        raise ValuationValidationError(
            "PLAN_SHEET_NAME_COLLISION",
            "duas abas do plano da pasta usariam o mesmo nome",
            {"names": duplicated},
        )


def plan_workbook(
    valuation: Valuation,
    catalog: PriceCatalog,
    template: WorkbookTemplate,
    contract: ContractWorkbook | None = None,
) -> WorkbookPlan:
    """Planeja a pasta inteira, com o valor exato de cada célula.

    Sem consolidado contratual a pasta é só o par BM/MEMÓRIA de cada obra; com ele, a
    PLANILHA GERAL abre a pasta e a aba de RE-RA vem logo atrás.
    """
    for bulletin in valuation.bulletins:
        _validate_against_catalog(bulletin, catalog)
    pinned: list[PinnedCell] = []
    sheets: list[PlannedSheet] = []
    general_sheet: str | None = None
    amendment_sheet: str | None = None
    if contract is not None:
        general = _plan_general(valuation, contract, template, pinned)
        general_sheet = general.name
        sheets.append(general)
        amendment = _plan_amendment_sheet(contract, template)
        if amendment is not None:
            amendment_sheet = amendment.name
            sheets.append(amendment)
    worksites: list[PlannedWorksite] = []
    for bulletin in valuation.bulletins:
        planned_bulletin = _plan_bulletin(valuation, bulletin, template, pinned)
        planned_memory = _plan_memory(valuation, bulletin, template, pinned)
        sheets.extend((planned_bulletin, planned_memory))
        worksites.append(
            PlannedWorksite(
                worksite_key=bulletin.worksite_key,
                bulletin_sheet=planned_bulletin.name,
                memory_sheet=planned_memory.name,
            )
        )
    _check_sheet_names(sheets)
    duplicated = [ref for sheet in sheets for ref in _duplicated_refs(sheet)]
    if duplicated:
        raise ValuationValidationError(
            "PLAN_CELL_COLLISION",
            "o plano da pasta gravaria duas vezes a mesma célula",
            {"refs": duplicated},
        )
    return WorkbookPlan(
        worksites=worksites,
        general_sheet=general_sheet,
        amendment_sheet=amendment_sheet,
        sheets=sheets,
        pinned_cells=pinned,
    )


def _duplicated_refs(sheet: PlannedSheet) -> list[str]:
    seen: set[str] = set()
    duplicated: list[str] = []
    for cell in sheet.cells:
        if cell.ref in seen:
            duplicated.append(f"{sheet.name}!{cell.ref}")
        seen.add(cell.ref)
    return duplicated


def _atomic_save(workbook: Workbook, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        dir=output_path.parent, prefix=f".{output_path.name}.", suffix=".xlsx"
    )
    os.close(descriptor)
    temporary_path = Path(temporary_name)
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, output_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_valuation_workbook(
    valuation: Valuation,
    catalog: PriceCatalog,
    template: WorkbookTemplate,
    output_path: Path,
    contract: ContractWorkbook | None = None,
) -> WriteReport:
    """Grava a pasta da medição e devolve o relatório da gravação."""
    plan = plan_workbook(valuation, catalog, template, contract)
    workbook = Workbook()
    # A data de criação é fixa para não carimbar o relógio local no artefato. A data de
    # modificação é reescrita pelo openpyxl na gravação, então o digest do arquivo muda a
    # cada execução mesmo com conteúdo idêntico: a idempotência da demonstração é do
    # conteúdo lógico (JSON canônico), não dos bytes do zip.
    workbook.properties.created = _FIXED_TIMESTAMP
    for index, planned_sheet in enumerate(plan.sheets):
        if index == 0:
            worksheet = workbook.active
            worksheet.title = planned_sheet.name
        else:
            worksheet = workbook.create_sheet(planned_sheet.name)
        for letter, width in planned_sheet.column_widths.items():
            worksheet.column_dimensions[letter].width = width
        for planned_cell in planned_sheet.cells:
            cell = worksheet[planned_cell.ref]
            if planned_cell.kind == "text":
                cell.value = planned_cell.text
            elif planned_cell.kind == "number":
                cell.value = planned_cell.number
            else:
                cell.value = planned_cell.formula
            if planned_cell.number_format is not None:
                cell.number_format = planned_cell.number_format
            if planned_cell.bold:
                cell.font = Font(bold=True)
    _atomic_save(workbook, output_path)
    return WriteReport(
        output_path=str(output_path),
        workbook_sha256=_sha256(output_path),
        worksites=plan.worksites,
        general_sheet=plan.general_sheet,
        amendment_sheet=plan.amendment_sheet,
        written_cells=sum(len(sheet.cells) for sheet in plan.sheets),
        formula_cells=plan.formula_cells,
        pinned_cells=plan.pinned_cells,
    )
