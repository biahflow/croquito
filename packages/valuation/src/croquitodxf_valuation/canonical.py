"""Canonicalização e auditoria da planilha gerada.

A planilha é um render; o JSON é a fonte de verdade. Para provar que o render não mente,
o arquivo é reaberto e cada célula de fórmula é RECOMPUTADA em `Decimal` por um
mini-avaliador que só entende a gramática fechada emitida pelo escritor:

1. `=TRUNC(<ref>*<ref>,2)`
2. `=ROUND(PRODUCT(<range>),2)`
3. `=ROUND(PRODUCT(<range>),2)-<ref>`
4. `=SUM(<range>)`
5. `=SUM(<ref>,<ref>[,<ref>...])` — acumulado da PLANILHA GERAL, cujas células não são
   contíguas (QUANTIDADE e VALOR se alternam ao longo dos pares de medição).
6. `=<ref>-<ref>` — saldo da PLANILHA GERAL (vigente menos acumulado).

Qualquer outra fórmula encontrada no arquivo é `FORMULA_UNSUPPORTED`: a auditoria não
adivinha semântica de planilha. Como no Excel, texto e célula vazia dentro de um
intervalo de `SUM`/`PRODUCT` — ou da lista de refs do `SUM` — não entram na conta; já a
subtração exige os dois operandos numéricos.
"""

from __future__ import annotations

import hashlib
import re
from decimal import Decimal
from pathlib import Path
from typing import Final, Literal, NoReturn

from openpyxl import load_workbook
from pydantic import Field

from croquitodxf_valuation.contract import ContractWorkbook
from croquitodxf_valuation.errors import ValuationValidationError
from croquitodxf_valuation.models import (
    ExactDecimal,
    PriceCatalog,
    Valuation,
    ValuationContractModel,
    WorksiteBulletin,
)
from croquitodxf_valuation.rounding import money_trunc, quantity_round
from croquitodxf_valuation.template import WorkbookTemplate
from croquitodxf_valuation.workbook_writer import (
    MAX_DECIMAL_PLACES,
    PinnedCell,
    PlannedCell,
    WorkbookPlan,
    plan_workbook,
)

CANONICAL_SCHEMA_VERSION: Final = "1.0.0"
AUDIT_SCHEMA_VERSION: Final = "2.0.0"

_REF: Final = r"[A-Z]{1,2}[0-9]+"
_TRUNC_RE: Final = re.compile(rf"^=TRUNC\(({_REF})\*({_REF}),2\)$")
_ROUND_PRODUCT_RE: Final = re.compile(rf"^=ROUND\(PRODUCT\(({_REF}):({_REF})\),2\)$")
_ROUND_PRODUCT_MINUS_RE: Final = re.compile(rf"^=ROUND\(PRODUCT\(({_REF}):({_REF})\),2\)-({_REF})$")
_SUM_RE: Final = re.compile(rf"^=SUM\(({_REF}):({_REF})\)$")
_SUM_LIST_RE: Final = re.compile(r"^=SUM\(([^():]+)\)$")
_DIFFERENCE_RE: Final = re.compile(rf"^=({_REF})-({_REF})$")

CELL_REF_RE: Final = re.compile(rf"^({_REF})$")
"""Referência de célula no formato A1, sem aba e sem cifrão."""

GRAMMAR_PATTERNS: Final[tuple[tuple[str, re.Pattern[str]], ...]] = (
    ("trunc_product", _TRUNC_RE),
    ("round_product_minus", _ROUND_PRODUCT_MINUS_RE),
    ("round_product", _ROUND_PRODUCT_RE),
    ("sum_range", _SUM_RE),
    ("sum_list", _SUM_LIST_RE),
    ("difference", _DIFFERENCE_RE),
)
"""As seis formas na ordem de precedência que `_SheetEvaluator.evaluate` aplica.

Exportada para que a ferramenta local de diagnóstico (`croquitodxf-valuation parity`)
reconheça exatamente a mesma gramática de uma pasta que o sistema não gerou, em vez de
manter uma segunda cópia dela que poderia divergir em silêncio.
"""


class AuditFinding(ValuationContractModel):
    """Uma divergência entre o arquivo reaberto e o que a medição prevê."""

    code: str = Field(min_length=1, max_length=60)
    sheet: str
    ref: str | None = None
    expected: str | None = None
    found: str | None = None
    detail: str | None = None


class AuditedWorksite(ValuationContractModel):
    """As abas auditadas de uma obra e o total que ela declara."""

    worksite_key: str
    bulletin_sheet: str
    memory_sheet: str
    total_amount: ExactDecimal


class AuditReport(ValuationContractModel):
    """Resultado da auditoria de round-trip da planilha."""

    schema_version: Literal["2.0.0"] = AUDIT_SCHEMA_VERSION
    status: Literal["ok", "divergent"]
    workbook_sha256: str = Field(pattern=r"^[a-f0-9]{64}$")
    worksites: list[AuditedWorksite] = Field(min_length=1)
    general_sheet: str | None = None
    amendment_sheet: str | None = None
    checked_cells: int = Field(ge=0)
    formula_cells: int = Field(ge=0)
    total_amount: ExactDecimal
    pinned_cells: list[PinnedCell] = Field(default_factory=list)
    findings: list[AuditFinding] = Field(default_factory=list)


def _column_index(letter: str) -> int:
    index = 0
    for char in letter:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index


def _column_letter(index: int) -> str:
    letters = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        letters = chr(ord("A") + remainder) + letters
    return letters


def _split_ref(ref: str) -> tuple[str, int]:
    match = CELL_REF_RE.fullmatch(ref)
    if match is None:
        raise ValuationValidationError(
            "FORMULA_UNSUPPORTED", "referência de célula fora do formato A1", {"ref": ref}
        )
    letters = "".join(char for char in ref if char.isalpha())
    digits = "".join(char for char in ref if char.isdigit())
    return letters, int(digits)


def expand_range(start: str, end: str) -> list[str]:
    """Refs de um intervalo A1:B2, linha a linha; intervalo invertido não é da gramática."""
    start_column, start_row = _split_ref(start)
    end_column, end_row = _split_ref(end)
    first_column = _column_index(start_column)
    last_column = _column_index(end_column)
    if first_column > last_column or start_row > end_row:
        raise ValuationValidationError(
            "FORMULA_UNSUPPORTED",
            "intervalo invertido na fórmula",
            {"range": f"{start}:{end}"},
        )
    refs: list[str] = []
    for row in range(start_row, end_row + 1):
        for column in range(first_column, last_column + 1):
            refs.append(f"{_column_letter(column)}{row}")
    return refs


def canonical_number(value: object, sheet_name: str, ref: str) -> Decimal:
    """Converte o valor lido da planilha em `Decimal` de duas casas.

    Fronteira declarada de leitura: a planilha devolve `float`, e a conversão passa uma
    única vez por `Decimal(str(...))`. Número com mais de duas casas é recusado, porque a
    planilha não o representaria sem perder o valor exato.
    """
    number = Decimal(str(value))
    exponent = number.as_tuple().exponent
    if isinstance(exponent, int) and exponent < -MAX_DECIMAL_PLACES:
        raise ValuationValidationError(
            "NUMBER_SCALE_UNSUPPORTED",
            "planilha traz número com mais de duas casas decimais",
            {"sheet": sheet_name, "ref": ref, "value": str(number)},
        )
    return quantity_round(number)


class _SheetEvaluator:
    """Avaliador da gramática fechada sobre os valores brutos de uma aba."""

    def __init__(self, sheet_name: str, raw_values: dict[str, object]) -> None:
        self.sheet_name = sheet_name
        self.raw_values = raw_values
        self._cache: dict[str, Decimal | None] = {}
        self._resolving: set[str] = set()

    def numeric(self, ref: str) -> Decimal | None:
        """Valor numérico canônico da célula, ou `None` quando é texto ou vazio."""
        if ref in self._cache:
            return self._cache[ref]
        if ref in self._resolving:
            raise ValuationValidationError(
                "FORMULA_CYCLE",
                "fórmula da planilha referencia a si mesma",
                {"sheet": self.sheet_name, "ref": ref},
            )
        raw = self.raw_values.get(ref)
        self._resolving.add(ref)
        try:
            value = self._numeric_of(raw, ref)
        finally:
            self._resolving.discard(ref)
        self._cache[ref] = value
        return value

    def _numeric_of(self, raw: object, ref: str) -> Decimal | None:
        if raw is None:
            return None
        if isinstance(raw, str):
            if raw.startswith("="):
                return self.evaluate(raw, ref)
            return None
        if isinstance(raw, bool):
            raise ValuationValidationError(
                "CELL_TYPE_UNSUPPORTED",
                "planilha traz booleano onde a medição só admite número ou texto",
                {"sheet": self.sheet_name, "ref": ref},
            )
        if isinstance(raw, int | float | Decimal):
            return canonical_number(raw, self.sheet_name, ref)
        raise ValuationValidationError(
            "CELL_TYPE_UNSUPPORTED",
            "tipo de célula não suportado pela medição",
            {"sheet": self.sheet_name, "ref": ref, "type": type(raw).__name__},
        )

    def _numbers_in(self, start: str, end: str) -> list[Decimal]:
        values: list[Decimal] = []
        for ref in expand_range(start, end):
            value = self.numeric(ref)
            if value is not None:
                values.append(value)
        return values

    def evaluate(self, formula: str, ref: str) -> Decimal:
        """Recomputa uma fórmula da gramática fechada em `Decimal`.

        A ordem das tentativas é a precedência da gramática: o subtraendo do bloco de
        cálculo (`=ROUND(PRODUCT(...),2)-<ref>`) é reconhecido antes da diferença simples
        do saldo, e o `SUM` de intervalo antes do `SUM` de lista.
        """
        trunc = _TRUNC_RE.fullmatch(formula)
        if trunc is not None:
            left = self._required(trunc.group(1), ref)
            right = self._required(trunc.group(2), ref)
            return money_trunc(left * right)
        minus = _ROUND_PRODUCT_MINUS_RE.fullmatch(formula)
        if minus is not None:
            product = self._product(minus.group(1), minus.group(2))
            deduction_value = self.numeric(minus.group(3))
            deduction = Decimal("0.00") if deduction_value is None else deduction_value
            return quantity_round(product) - deduction
        rounded = _ROUND_PRODUCT_RE.fullmatch(formula)
        if rounded is not None:
            return quantity_round(self._product(rounded.group(1), rounded.group(2)))
        summed = _SUM_RE.fullmatch(formula)
        if summed is not None:
            values = self._numbers_in(summed.group(1), summed.group(2))
            return sum(values, Decimal("0.00"))
        listed = _SUM_LIST_RE.fullmatch(formula)
        if listed is not None:
            return self._sum_of_refs(listed.group(1), formula, ref)
        difference = _DIFFERENCE_RE.fullmatch(formula)
        if difference is not None:
            minuend = self._required(difference.group(1), ref)
            subtrahend = self._required(difference.group(2), ref)
            return minuend - subtrahend
        self._unsupported(formula, ref)

    def _sum_of_refs(self, arguments: str, formula: str, ref: str) -> Decimal:
        """Forma 5: soma de refs avulsas; ref vazia ou de texto contribui zero.

        A lista é dividida e cada item é validado à parte porque um grupo repetido na
        expressão regular só devolveria a última ocorrência.
        """
        tokens = arguments.split(",")
        if len(tokens) < 2:
            self._unsupported(formula, ref)
        values: list[Decimal] = []
        for token in tokens:
            if CELL_REF_RE.fullmatch(token) is None:
                self._unsupported(formula, ref)
            value = self.numeric(token)
            if value is not None:
                values.append(value)
        return sum(values, Decimal("0.00"))

    def _unsupported(self, formula: str, ref: str) -> NoReturn:
        raise ValuationValidationError(
            "FORMULA_UNSUPPORTED",
            "fórmula fora da gramática aceita pela auditoria da medição",
            {"sheet": self.sheet_name, "ref": ref, "formula": formula},
        )

    def _product(self, start: str, end: str) -> Decimal:
        values = self._numbers_in(start, end)
        if not values:
            return Decimal("0.00")
        result = Decimal(1)
        for value in values:
            result *= value
        return result

    def _required(self, ref: str, source_ref: str) -> Decimal:
        value = self.numeric(ref)
        if value is None:
            raise ValuationValidationError(
                "FORMULA_REFERENCE_EMPTY",
                "fórmula opera sobre célula vazia ou de texto",
                {"sheet": self.sheet_name, "ref": source_ref, "target": ref},
            )
        return value


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def canonicalize_workbook(workbook_path: Path, template: WorkbookTemplate) -> dict[str, object]:
    """Reabre a planilha e devolve a estrutura canônica com as fórmulas recomputadas.

    A aba de catálogo declarada no template é ignorada: ela é insumo, não medição.
    """
    workbook = load_workbook(workbook_path, data_only=False)
    try:
        sheets: list[dict[str, object]] = []
        for worksheet in workbook.worksheets:
            sheet_name = str(worksheet.title)
            if sheet_name == template.catalog.sheet_name:
                continue
            raw_values: dict[str, object] = {}
            positions: dict[str, tuple[int, int]] = {}
            for row in worksheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if value is None or (isinstance(value, str) and not value.strip()):
                        continue
                    ref = str(cell.coordinate)
                    raw_values[ref] = value
                    positions[ref] = (int(cell.row), int(cell.column))
            evaluator = _SheetEvaluator(sheet_name, raw_values)
            cells: list[dict[str, object]] = []
            for ref in sorted(raw_values, key=lambda item: positions[item]):
                raw = raw_values[ref]
                if isinstance(raw, str) and raw.startswith("="):
                    cells.append(
                        {
                            "ref": ref,
                            "kind": "formula",
                            "formula": raw,
                            "value": str(evaluator.evaluate(raw, ref)),
                        }
                    )
                elif isinstance(raw, str):
                    cells.append({"ref": ref, "kind": "text", "value": raw})
                else:
                    number = evaluator.numeric(ref)
                    if number is None:
                        raise ValuationValidationError(
                            "CELL_TYPE_UNSUPPORTED",
                            "célula numérica não pôde ser canonicalizada",
                            {"sheet": sheet_name, "ref": ref},
                        )
                    cells.append({"ref": ref, "kind": "number", "value": str(number)})
            sheets.append({"name": sheet_name, "cells": cells})
    finally:
        workbook.close()
    return {"schema_version": CANONICAL_SCHEMA_VERSION, "sheets": sheets}


def _index_canonical(canonical: dict[str, object]) -> dict[str, dict[str, dict[str, object]]]:
    indexed: dict[str, dict[str, dict[str, object]]] = {}
    sheets = canonical["sheets"]
    assert isinstance(sheets, list)
    for sheet in sheets:
        assert isinstance(sheet, dict)
        name = str(sheet["name"])
        cells = sheet["cells"]
        assert isinstance(cells, list)
        indexed[name] = {str(cell["ref"]): cell for cell in cells}
    return indexed


def _compare_plan(
    plan: WorkbookPlan, indexed: dict[str, dict[str, dict[str, object]]]
) -> list[AuditFinding]:
    findings: list[AuditFinding] = []
    for planned_sheet in plan.sheets:
        found_sheet = indexed.get(planned_sheet.name)
        if found_sheet is None:
            findings.append(
                AuditFinding(
                    code="SHEET_MISSING",
                    sheet=planned_sheet.name,
                    detail="aba prevista não existe no arquivo gravado",
                )
            )
            continue
        planned_refs: set[str] = set()
        for planned_cell in planned_sheet.cells:
            planned_refs.add(planned_cell.ref)
            found = found_sheet.get(planned_cell.ref)
            if found is None:
                findings.append(
                    AuditFinding(
                        code="CELL_MISSING",
                        sheet=planned_sheet.name,
                        ref=planned_cell.ref,
                        expected=_expected_text(planned_cell),
                        detail=planned_cell.role,
                    )
                )
                continue
            expected_kind = planned_cell.kind
            found_kind = str(found["kind"])
            if expected_kind != found_kind:
                findings.append(
                    AuditFinding(
                        code="CELL_KIND_MISMATCH",
                        sheet=planned_sheet.name,
                        ref=planned_cell.ref,
                        expected=expected_kind,
                        found=found_kind,
                        detail=planned_cell.role,
                    )
                )
                continue
            expected_value = _expected_text(planned_cell)
            found_value = str(found["value"])
            if expected_value != found_value:
                findings.append(
                    AuditFinding(
                        code="CELL_VALUE_MISMATCH",
                        sheet=planned_sheet.name,
                        ref=planned_cell.ref,
                        expected=expected_value,
                        found=found_value,
                        detail=planned_cell.role,
                    )
                )
            if expected_kind == "formula":
                found_formula = str(found["formula"])
                if planned_cell.formula != found_formula:
                    findings.append(
                        AuditFinding(
                            code="CELL_FORMULA_MISMATCH",
                            sheet=planned_sheet.name,
                            ref=planned_cell.ref,
                            expected=planned_cell.formula,
                            found=found_formula,
                            detail=planned_cell.role,
                        )
                    )
        for ref in sorted(set(found_sheet) - planned_refs):
            findings.append(
                AuditFinding(
                    code="CELL_UNEXPECTED",
                    sheet=planned_sheet.name,
                    ref=ref,
                    found=str(found_sheet[ref]["value"]),
                    detail="célula presente no arquivo e ausente do plano",
                )
            )
    for sheet_name in sorted(set(indexed) - {sheet.name for sheet in plan.sheets}):
        findings.append(
            AuditFinding(
                code="SHEET_UNEXPECTED",
                sheet=sheet_name,
                detail="aba presente no arquivo e ausente do plano",
            )
        )
    return findings


def _expected_text(planned_cell: PlannedCell) -> str:
    """Valor canônico que o plano prevê para a célula, no mesmo formato do arquivo."""
    if planned_cell.kind == "text":
        return str(planned_cell.text)
    number = planned_cell.number
    if number is None:  # pragma: no cover - impedido pelo validador de PlannedCell
        raise ValuationValidationError(
            "PLANNED_CELL_INVALID",
            "célula numérica planejada sem valor",
            {"ref": planned_cell.ref},
        )
    return str(quantity_round(number))


def _bulletin_by_sheet(plan: WorkbookPlan, valuation: Valuation) -> dict[str, WorksiteBulletin]:
    """Qual boletim responde por cada aba de obra; GERAL e RE-RA ficam de fora."""
    bulletins = {bulletin.worksite_key: bulletin for bulletin in valuation.bulletins}
    by_sheet: dict[str, WorksiteBulletin] = {}
    for worksite in plan.worksites:
        bulletin = bulletins[worksite.worksite_key]
        by_sheet[worksite.bulletin_sheet] = bulletin
        by_sheet[worksite.memory_sheet] = bulletin
    return by_sheet


def _catalog_findings(
    plan: WorkbookPlan,
    indexed: dict[str, dict[str, dict[str, object]]],
    valuation: Valuation,
    catalog: PriceCatalog,
) -> list[AuditFinding]:
    """O preço impresso continua sendo o do catálogo, mesmo sem `VLOOKUP` no arquivo."""
    findings: list[AuditFinding] = []
    by_sheet = _bulletin_by_sheet(plan, valuation)
    for planned_sheet in plan.sheets:
        bulletin = by_sheet.get(planned_sheet.name)
        if bulletin is None:
            continue
        lines = {line.item_number: line for line in bulletin.lines}
        found_sheet = indexed.get(planned_sheet.name, {})
        for planned_cell in planned_sheet.cells:
            if planned_cell.role != "line_unit_price" or planned_cell.item_number is None:
                continue
            line = lines[planned_cell.item_number]
            found = found_sheet.get(planned_cell.ref)
            if found is None:
                continue
            if not catalog.has_code(line.code):
                findings.append(
                    AuditFinding(
                        code="CATALOG_CODE_MISSING",
                        sheet=planned_sheet.name,
                        ref=planned_cell.ref,
                        expected=line.code,
                        detail="código impresso não existe no catálogo importado",
                    )
                )
                continue
            entry = catalog.entry_for(line.code)
            expected = str(quantity_round(entry.unit_price))
            if expected != str(found["value"]):
                findings.append(
                    AuditFinding(
                        code="CATALOG_PRICE_MISMATCH",
                        sheet=planned_sheet.name,
                        ref=planned_cell.ref,
                        expected=expected,
                        found=str(found["value"]),
                        detail=line.code,
                    )
                )
    return findings


def audit_workbook(
    workbook_path: Path,
    valuation: Valuation,
    catalog: PriceCatalog,
    template: WorkbookTemplate,
    contract: ContractWorkbook | None = None,
) -> AuditReport:
    """Reabre a planilha, recomputa as fórmulas e compara centavo a centavo.

    Informado o consolidado contratual, a PLANILHA GERAL e a aba de RE-RA entram na
    conferência pelo mesmo caminho das demais: toda célula delas está no plano.
    """
    plan = plan_workbook(valuation, catalog, template, contract)
    canonical = canonicalize_workbook(workbook_path, template)
    indexed = _index_canonical(canonical)
    findings = _compare_plan(plan, indexed)
    findings.extend(_catalog_findings(plan, indexed, valuation, catalog))
    bulletins = {bulletin.worksite_key: bulletin for bulletin in valuation.bulletins}
    return AuditReport(
        status="ok" if not findings else "divergent",
        workbook_sha256=_sha256(workbook_path),
        worksites=[
            AuditedWorksite(
                worksite_key=worksite.worksite_key,
                bulletin_sheet=worksite.bulletin_sheet,
                memory_sheet=worksite.memory_sheet,
                total_amount=bulletins[worksite.worksite_key].total_amount,
            )
            for worksite in plan.worksites
        ],
        general_sheet=plan.general_sheet,
        amendment_sheet=plan.amendment_sheet,
        checked_cells=sum(len(sheet.cells) for sheet in plan.sheets),
        formula_cells=plan.formula_cells,
        total_amount=valuation.total_amount,
        pinned_cells=plan.pinned_cells,
        findings=findings,
    )
